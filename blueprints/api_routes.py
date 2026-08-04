from flask import Blueprint, jsonify, request, current_app, render_template, send_file, Response, stream_with_context, session, g
from blueprints.auth_routes import has_permission, require_permission
from database import get_db
from services.shopify_service import (
    get_bulk_product_details, 
    clear_product_cache, 
    find_shopify_product_gid, 
    update_shopify_inci_metafield,
    update_stock_status_in_shopify,
    get_single_product_details_for_display,
    get_all_products_for_name_sync,
    get_all_shopify_stores,
    sync_all_stock_metafields,
    sync_inci_from_shopify
)
from services.s3_service import upload_order_image, upload_order_image_bytes, get_order_images, delete_order_image, setup_s3_cors
from services.s3_service import get_s3_client, generate_presigned_url, generate_presigned_post_for_order_image
from services.s3_service import upload_instruction_image, upload_returned_damaged_image, get_returned_damaged_images
from services.email_service import poslji_email_s_pdf
from services.mk_service import mk_find_bill_by_title, mk_is_published
from services.mk_service import mk_find_bill_any, mk_get_document, mk_get_document_bill, mk_find_declaration_attachment_ts
from services.mk_service import sync_mk_declaration_uploads
from services.mk_service import mk_log_stock_event, mk_import_by_skus
from services.pdf_service import ustvari_pdf, generate_purchase_order_pdf
from services.email_service import poslji_email_s_pdf, poslji_obvestilo_o_napaki, send_new_user_welcome_email, send_purchase_order_admin_email
from services.excel_service import generate_purchase_order_excel
from services.search_service import normalize_query, upsert_synonym, upsert_inspo_target
from services.sync_parfumi_service import (
    DEFAULT_SYNC_STORE,
    sync_parfumi_from_shopify,
)
from services.restore_parfumi_names import restore_parfumi_names_from_excel
from services.import_parfumi_names import import_parfumi_names_from_csv
from services.fix_amour_parfums_names import apply_fix_amour_parfums_names, preview_fix_amour_parfums_names
import io
from openpyxl import Workbook
from openpyxl.styles import Font
import traceback
import json
import json as _json
import logging
from psycopg.rows import dict_row

# Setup logger
logger = logging.getLogger(__name__)
from datetime import date, datetime, timedelta
import os
import requests
import threading
import tempfile

import time
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash

api_bp = Blueprint('api', __name__, url_prefix='/api')

from api.utils.responses import make_ok, make_err
from api.utils.validate import validate_int, validate_str

@api_bp.before_app_request
def assign_request_id():
    try:
        import uuid
        g.request_id = str(uuid.uuid4())
    except Exception:
        g.request_id = 'unknown'

@api_bp.after_app_request
def add_request_id_header(response):
    rid = getattr(g, 'request_id', None)
    if rid:
        response.headers['X-Request-ID'] = rid
    return response

@api_bp.errorhandler(400)
def handle_400(e):
    return make_err('BAD_REQUEST', 'Neveljavna zahteva', status=400)

@api_bp.errorhandler(403)
def handle_403(e):
    return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)

@api_bp.errorhandler(404)
def handle_404(e):
    return make_err('NOT_FOUND', 'Vir ni najden', status=404)

@api_bp.errorhandler(409)
def handle_409(e):
    return make_err('CONFLICT', 'Konflikt', status=409)

@api_bp.errorhandler(429)
def handle_429(e):
    return make_err('RATE_LIMITED', 'Preveč zahtev', status=429)

@api_bp.errorhandler(500)
def handle_500(e):
    return make_err('SERVER_ERROR', 'Napaka na strežniku', status=500)

# ---------------------- PROCUREMENT API ----------------------
@api_bp.route('/procurement/suppliers', methods=['GET'])
def get_suppliers():
    return make_ok(['FLORGARDEN', 'MISTRAL'])

@api_bp.route('/procurement/stock', methods=['GET'])
def procurement_stock():
    supplier = request.args.get('supplier', '').upper()
    if supplier not in ('FLORGARDEN', 'MISTRAL'):
        return make_err('BAD_REQUEST', 'supplier mora biti FLORGARDEN ali MISTRAL', status=400)
    q = (request.args.get('q') or '').strip()
    db = get_db()
    c = db.cursor()
    where_extra = ''
    params = [supplier]
    if q:
        where_extra = " AND (ps.product_no ILIKE %s OR p.ime_parfuma ILIKE %s)"
        like = f"%{q}%"
        params.extend([like, like])
    c.execute(
        f"""
        SELECT ps.*, p.ime_parfuma, pr.ime AS proizvajalec
        FROM perfumes_stock ps
        JOIN proizvajalci pr ON ps.proizvajalec_id = pr.id
        JOIN parfumi p ON p.product_no = ps.product_no AND p.proizvajalec_id = ps.proizvajalec_id
        WHERE pr.ime = %s{where_extra}
        ORDER BY
          (ps.product_no ~ '^[0-9]+$') DESC,                                      -- najprej čisto numerični product_no
          CASE WHEN ps.product_no ~ '^[0-9]+$' THEN ps.product_no::int END ASC,   -- naravno (1,2,10,100)
          ps.product_no ASC,                                                      -- nato alfanumerični leksikografsko
          p.ime_parfuma ASC
        """,
        tuple(params)
    )
    rows = [dict(r) for r in c.fetchall()]
    c.close()
    return make_ok(rows)

# ===================== PROCUREMENT-ONLY (non-perfume) SUPPORT =====================

def _ensure_procurement_only_tables():
    db = get_db(); c = db.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS proc_suppliers (
            id SERIAL PRIMARY KEY,
            name TEXT UNIQUE NOT NULL
        );
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS proc_products (
            id SERIAL PRIMARY KEY,
            supplier_id INTEGER NOT NULL REFERENCES proc_suppliers(id) ON DELETE CASCADE,
            sku TEXT NOT NULL,
            name TEXT NOT NULL,
            unit TEXT DEFAULT 'kos',
            price NUMERIC(12,2) DEFAULT 0,
            min_on_hand INTEGER DEFAULT 0,
            on_hand INTEGER DEFAULT 0,
            pending INTEGER DEFAULT 0,
            committed INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(supplier_id, sku)
        );
    """)
    db.commit(); c.close()

@api_bp.route('/procurement2/products/<supplier>/<sku>', methods=['DELETE'])
def procurement2_delete_product(supplier: str, sku: str):
    perm = required_permission_for('DELETE', '/api/procurement/vendors/import')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    try:
        _ensure_procurement_only_tables()
        db = get_db(); c = db.cursor()
        sup = (supplier or '').strip().upper(); code = (sku or '').strip()
        if not sup or not code:
            return make_err('BAD_REQUEST', 'Manjkajo podatki', status=400)
        c.execute("SELECT id FROM proc_suppliers WHERE name = %s", (sup,))
        row = c.fetchone()
        if not row:
            return make_err('NOT_FOUND', 'Dobavitelj ne obstaja', status=404)
        supplier_id = row[0] if isinstance(row, tuple) else row['id']
        c.execute("DELETE FROM proc_products WHERE supplier_id = %s AND sku = %s", (supplier_id, code))
        db.commit(); c.close()
        return make_ok({'deleted': True})
    except Exception as e:
        try:
            get_db().rollback()
        except Exception:
            pass
        current_app.logger.error(f"procurement2_delete_product error: {e}\n{traceback.format_exc()}")
        return make_err('SERVER_ERROR', 'Napaka pri brisanju', status=500)

@api_bp.route('/procurement/suppliers/all', methods=['GET'])
def procurement_suppliers_all():
    try:
        _ensure_procurement_only_tables()
    except Exception:
        pass
    base = ['FLORGARDEN', 'MISTRAL']
    try:
        db = get_db(); c = db.cursor()
        c.execute("SELECT name FROM proc_suppliers ORDER BY name ASC")
        extra = [r[0] if isinstance(r, tuple) else (r['name'] if isinstance(r, dict) else r) for r in c.fetchall()]
        c.close()
        return make_ok(base + extra)
    except Exception as e:
        current_app.logger.error(f"suppliers/all error: {e}")
        return make_ok(base)

@api_bp.route('/procurement/vendors/template', methods=['GET'])
def procurement_vendors_template():
    # Generate minimal Excel template
    wb = Workbook(); ws = wb.active; ws.title = 'products'
    ws.append(['supplier', 'sku', 'name', 'unit', 'price', 'min_on_hand', 'on_hand'])
    ws.append(['MY_SUPPLIER', 'SKU-001', 'Primer izdelka', 'kos', 1.99, 5, 10])
    ws.append(['MY_SUPPLIER', 'SKU-002', 'Drugi izdelek', 'kos', 3.49, 2, 0])
    f = io.BytesIO(); wb.save(f); f.seek(0)
    return send_file(f, as_attachment=True, download_name='procurement_template.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@api_bp.route('/procurement/vendors/import', methods=['POST'])
def procurement_vendors_import():
    # Permission gate (admin-only)
    perm = required_permission_for('POST', '/api/procurement/vendors/import')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    try:
        _ensure_procurement_only_tables()
        if 'file' not in request.files:
            return make_err('BAD_REQUEST', 'Manjka datoteka', status=400)
        file = request.files['file']
        if not (file.filename.lower().endswith('.xlsx') or file.filename.lower().endswith('.xls')):
            return make_err('BAD_REQUEST', 'Podprte so samo Excel datoteke (.xlsx, .xls)', status=400)
        import pandas as pd
        df = pd.read_excel(file.stream)
        # Normaliziraj prazne vrednosti, da se izognemo NaN → None
        try:
            df = df.where(pd.notna(df), None)
        except Exception:
            pass
        if df is None or df.empty:
            return make_err('BAD_REQUEST', 'Prazna datoteka', status=400)
        db = get_db(); c = db.cursor()
        imported = 0
        def _to_float(val, default=0.0):
            try:
                if val is None:
                    return default
                if isinstance(val, str):
                    v = val.replace(',', '.').strip()
                    if v == '':
                        return default
                    return float(v)
                return float(val)
            except Exception:
                return default

        def _to_int(val, default=0):
            try:
                if val is None:
                    return default
                if isinstance(val, str):
                    v = val.replace(',', '.').strip()
                    if v == '':
                        return default
                    return int(float(v))
                return int(float(val))
            except Exception:
                return default

        for _, row in df.iterrows():
            supplier_name = str((row.get('supplier') or '')).strip().upper()
            sku = str((row.get('sku') or '')).strip()
            name = str((row.get('name') or '')).strip()
            unit = str((row.get('unit') or 'kos')).strip()
            price = _to_float(row.get('price'), 0.0)
            min_on_hand = _to_int(row.get('min_on_hand'), 0)
            on_hand = _to_int(row.get('on_hand'), 0)
            if not supplier_name or not sku or not name:
                continue
            c.execute("INSERT INTO proc_suppliers (name) VALUES (%s) ON CONFLICT (name) DO NOTHING RETURNING id", (supplier_name,))
            row_id = c.fetchone()
            if row_id:
                supplier_id = row_id[0] if isinstance(row_id, tuple) else (row_id['id'] if isinstance(row_id, dict) else row_id)
            else:
                c.execute("SELECT id FROM proc_suppliers WHERE name = %s", (supplier_name,))
                got = c.fetchone()
                supplier_id = got[0] if isinstance(got, tuple) else (got['id'] if isinstance(got, dict) else got)
            c.execute(
                """
                INSERT INTO proc_products (supplier_id, sku, name, unit, price, min_on_hand, on_hand, pending, committed)
                VALUES (%s,%s,%s,%s,%s,%s,%s,0,0)
                ON CONFLICT (supplier_id, sku) DO UPDATE SET
                    name=EXCLUDED.name,
                    unit=EXCLUDED.unit,
                    price=EXCLUDED.price,
                    min_on_hand=EXCLUDED.min_on_hand,
                    on_hand=EXCLUDED.on_hand,
                    updated_at=NOW()
                """,
                (supplier_id, sku, name, unit, price, min_on_hand, on_hand)
            )
            imported += 1
        db.commit(); c.close()
        return make_ok({'imported_count': imported})
    except Exception as e:
        current_app.logger.error(f"vendors/import error: {e}\n{traceback.format_exc()}")
        try:
            get_db().rollback()
        except Exception:
            pass
        return make_err('SERVER_ERROR', 'Napaka pri uvozu', status=500)

def _supplier_is_proc_only(supplier: str) -> bool:
    return supplier.upper() not in ('FLORGARDEN', 'MISTRAL')

@api_bp.route('/procurement2/stock', methods=['GET'])
def procurement2_stock():
    supplier = (request.args.get('supplier') or '').strip().upper()
    if not supplier:
        return make_err('BAD_REQUEST', 'Manjka supplier', status=400)
    try:
        _ensure_procurement_only_tables()
        db = get_db(); c = db.cursor()
        c.execute("SELECT id FROM proc_suppliers WHERE name = %s", (supplier,))
        row = c.fetchone()
        if not row:
            return make_ok([])
        supplier_id = row[0] if isinstance(row, tuple) else row['id']
        q = (request.args.get('q') or '').strip()
        params = [supplier_id]
        where = "WHERE supplier_id = %s"
        if q:
            where += " AND (sku ILIKE %s OR name ILIKE %s)"
            like = f"%{q}%"; params.extend([like, like])
        c.execute(f"SELECT id, sku, name, unit, price, min_on_hand, on_hand, pending, committed FROM proc_products {where} ORDER BY name ASC", tuple(params))
        rows = [dict(r) for r in c.fetchall()]
        c.close();
        return make_ok(rows)
    except Exception as e:
        current_app.logger.error(f"procurement2_stock error: {e}")
        return make_err('SERVER_ERROR', 'Napaka', status=500)

@api_bp.route('/procurement2/search', methods=['GET'])
def procurement2_search():
    supplier = (request.args.get('supplier') or '').strip().upper()
    if not supplier:
        return make_err('BAD_REQUEST', 'Manjka supplier', status=400)
    q = (request.args.get('q') or '').strip()
    if len(q) < 1:
        return make_ok([])
    try:
        _ensure_procurement_only_tables()
        db = get_db(); c = db.cursor()
        c.execute("SELECT id FROM proc_suppliers WHERE name = %s", (supplier,))
        row = c.fetchone()
        if not row:
            return make_ok([])
        supplier_id = row[0] if isinstance(row, tuple) else row['id']
        # Fuzzy multi-token: vsi tokeni morajo biti prisotni v sku ali name
        tokens = [t for t in q.split() if t]
        where_clauses = ["supplier_id = %s"]
        params = [supplier_id]
        for t in tokens:
            where_clauses.append("(sku ILIKE %s OR name ILIKE %s)")
            like = f"%{t}%"; params.extend([like, like])
        where_sql = " AND ".join(where_clauses)
        sql = f"""
            SELECT id, sku, name, unit, price, min_on_hand, on_hand, pending, committed
            FROM proc_products
            WHERE {where_sql}
            ORDER BY name ASC
        """
        c.execute(sql, tuple(params))
        rows = [dict(r) for r in c.fetchall()]
        c.close();
        return make_ok(rows)
    except Exception as e:
        current_app.logger.error(f"procurement2_search error: {e}")
        return make_err('SERVER_ERROR', 'Napaka', status=500)

@api_bp.route('/procurement2/cart/add', methods=['POST'])
def procurement2_cart_add():
    """Increment pending for a procurement-only SKU under a supplier."""
    try:
        data = request.get_json(force=True) or {}
    except Exception:
        data = {}
    supplier = (data.get('supplier') or '').strip().upper()
    sku = str((data.get('sku') or '')).strip()
    qty = int(data.get('qty') or 0)
    if not supplier or not sku or qty <= 0:
        return make_err('BAD_REQUEST', 'Manjkajo podatki', status=400)
    try:
        _ensure_procurement_only_tables()
        db = get_db(); c = db.cursor()
        c.execute("SELECT id FROM proc_suppliers WHERE name = %s", (supplier,))
        row = c.fetchone()
        if not row:
            return make_err('NOT_FOUND', 'Dobavitelj ne obstaja', status=404)
        supplier_id = row[0] if isinstance(row, tuple) else row['id']
        # Validate product exists for this supplier
        c.execute("SELECT 1 FROM proc_products WHERE supplier_id = %s AND sku = %s", (supplier_id, sku))
        exists = c.fetchone()
        if not exists:
            c.close()
            return make_err('NOT_FOUND', 'Izdelek (SKU) ne obstaja pri izbranem dobavitelju', status=404)
        c.execute("""
            UPDATE proc_products
            SET pending = GREATEST(0, pending + %s), updated_at = NOW()
            WHERE supplier_id = %s AND sku = %s
        """, (qty, supplier_id, sku))
        db.commit(); c.close()
        return make_ok({'added': qty})
    except Exception as e:
        try:
            get_db().rollback()
        except Exception:
            pass
        current_app.logger.error(f"procurement2_cart_add error: {e}\n{traceback.format_exc()}")
        return make_err('SERVER_ERROR', 'Napaka pri dodajanju na naročilo', status=500)

@api_bp.route('/procurement2/stock/add-onhand', methods=['POST'])
def procurement2_add_onhand():
    """Increment on_hand for a procurement-only SKU under a supplier."""
    try:
        data = request.get_json(force=True) or {}
    except Exception:
        data = {}
    supplier = (data.get('supplier') or '').strip().upper()
    sku = str((data.get('sku') or '')).strip()
    qty = int(data.get('qty') or 0)
    if not supplier or not sku or qty <= 0:
        return make_err('BAD_REQUEST', 'Manjkajo podatki', status=400)
    try:
        _ensure_procurement_only_tables()
        db = get_db(); c = db.cursor()
        c.execute("SELECT id FROM proc_suppliers WHERE name = %s", (supplier,))
        row = c.fetchone()
        if not row:
            return make_err('NOT_FOUND', 'Dobavitelj ne obstaja', status=404)
        supplier_id = row[0] if isinstance(row, tuple) else row['id']
        # Validate product exists for this supplier
        c.execute("SELECT 1 FROM proc_products WHERE supplier_id = %s AND sku = %s", (supplier_id, sku))
        exists = c.fetchone()
        if not exists:
            c.close()
            return make_err('NOT_FOUND', 'Izdelek (SKU) ne obstaja pri izbranem dobavitelju', status=404)
        c.execute("""
            UPDATE proc_products
            SET on_hand = GREATEST(0, on_hand + %s), updated_at = NOW()
            WHERE supplier_id = %s AND sku = %s
        """, (qty, supplier_id, sku))
        db.commit(); c.close()
        return make_ok({'added': qty})
    except Exception as e:
        try:
            get_db().rollback()
        except Exception:
            pass
        current_app.logger.error(f"procurement2_add_onhand error: {e}\n{traceback.format_exc()}")
        return make_err('SERVER_ERROR', 'Napaka pri dodajanju v predal', status=500)

@api_bp.route('/procurement2/products/quick-add', methods=['POST'])
def procurement2_quick_add_product():
    """Quickly create a procurement-only SKU under a supplier with minimal fields."""
    try:
        data = request.get_json(force=True) or {}
    except Exception:
        data = {}
    supplier = (data.get('supplier') or '').strip().upper()
    sku = str((data.get('sku') or '')).strip()
    name = str((data.get('name') or '')).strip()
    try:
        price = float(data.get('price') or 0)
    except (TypeError, ValueError):
        price = 0.0
    unit = str((data.get('unit') or 'kos')).strip() or 'kos'
    if not supplier or not sku or not name:
        return make_err('BAD_REQUEST', 'supplier, sku in name so obvezni', status=400)
    if supplier in ('FLORGARDEN', 'MISTRAL'):
        return make_err('BAD_REQUEST', 'Parfumski dobavitelji uporabljajo katalog parfumov, ne procurement-only', status=400)
    try:
        _ensure_procurement_only_tables()
        db = get_db(); c = db.cursor()
        c.execute(
            """
            INSERT INTO proc_suppliers (name) VALUES (%s)
            ON CONFLICT (name) DO NOTHING
            """,
            (supplier,)
        )
        c.execute("SELECT id FROM proc_suppliers WHERE name = %s", (supplier,))
        row = c.fetchone()
        if not row:
            return make_err('SERVER_ERROR', 'Dobavitelja ni mogoče ustvariti', status=500)
        supplier_id = row[0] if isinstance(row, tuple) else row['id']
        c.execute(
            """
            INSERT INTO proc_products (supplier_id, sku, name, unit, price, min_on_hand, on_hand, pending, committed)
            VALUES (%s, %s, %s, %s, %s, 0, 0, 0, 0)
            ON CONFLICT (supplier_id, sku) DO UPDATE SET
                name = EXCLUDED.name,
                price = EXCLUDED.price,
                updated_at = NOW()
            RETURNING id
            """,
            (supplier_id, sku, name, unit, price)
        )
        new = c.fetchone()
        db.commit(); c.close()
        return make_ok({'id': (new[0] if isinstance(new, tuple) else new.get('id')), 'supplier': supplier, 'sku': sku, 'name': name})
    except Exception as e:
        try:
            get_db().rollback()
        except Exception:
            pass
        current_app.logger.error(f"procurement2_quick_add_product error: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri dodajanju artikla', status=500)


@api_bp.route('/procurement2/stock/min', methods=['POST'])
def procurement2_set_min_on_hand():
    """Update min_on_hand threshold for a procurement-only SKU."""
    try:
        data = request.get_json(force=True) or {}
    except Exception:
        data = {}
    supplier = (data.get('supplier') or '').strip().upper()
    sku = str((data.get('sku') or '')).strip()
    try:
        min_val = max(0, int(data.get('min_on_hand') or 0))
    except (TypeError, ValueError):
        min_val = 0
    if not supplier or not sku:
        return make_err('BAD_REQUEST', 'Manjkajo podatki', status=400)
    try:
        _ensure_procurement_only_tables()
        db = get_db(); c = db.cursor()
        c.execute("SELECT id FROM proc_suppliers WHERE name = %s", (supplier,))
        row = c.fetchone()
        if not row:
            return make_err('NOT_FOUND', 'Dobavitelj ne obstaja', status=404)
        supplier_id = row[0] if isinstance(row, tuple) else row['id']
        c.execute(
            """
            UPDATE proc_products
            SET min_on_hand = %s, updated_at = NOW()
            WHERE supplier_id = %s AND sku = %s
            """,
            (min_val, supplier_id, sku)
        )
        db.commit(); c.close()
        return make_ok({'updated': True, 'min_on_hand': min_val})
    except Exception as e:
        try:
            get_db().rollback()
        except Exception:
            pass
        current_app.logger.error(f"procurement2_set_min_on_hand error: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri shranjevanju', status=500)


@api_bp.route('/procurement2/cart/bulk-set', methods=['POST'])
def procurement2_cart_bulk_set():
    try:
        data = request.get_json(force=True)
    except Exception:
        data = {}
    supplier = (data.get('supplier') or '').strip().upper()
    items = data.get('items') or []
    if not supplier or not isinstance(items, list):
        return make_err('BAD_REQUEST', 'Manjkajo podatki', status=400)
    try:
        _ensure_procurement_only_tables()
        db = get_db(); c = db.cursor()
        c.execute("SELECT id FROM proc_suppliers WHERE name = %s", (supplier,))
        row = c.fetchone()
        if not row:
            return make_err('NOT_FOUND', 'Dobavitelj ne obstaja', status=404)
        supplier_id = row[0] if isinstance(row, tuple) else row['id']
        for it in items:
            sku = str(it.get('sku','')).strip()
            qty = int(it.get('qty', 0))
            if not sku or qty < 0:
                continue
            c.execute("""
                UPDATE proc_products
                SET pending = %s, updated_at = NOW()
                WHERE supplier_id = %s AND sku = %s
            """, (qty, supplier_id, sku))
        db.commit(); c.close()
        return make_ok(True)
    except Exception as e:
        current_app.logger.error(f"procurement2_cart_bulk_set error: {e}")
        try:
            get_db().rollback()
        except Exception:
            pass
        return make_err('SERVER_ERROR', 'Napaka', status=500)
@api_bp.route('/procurement/cart/add', methods=['POST'])
def procurement_cart_add():
    # Sprejmi prazne body-je varno, ne sproži 400, če je Content-Type napačen
    try:
        data = request.get_json(silent=True) or {}
    except Exception:
        data = {}
    product_no = str(data.get('product_no', '')).strip()
    proizvajalec_id = data.get('proizvajalec_id')
    qty = int(data.get('qty', 0))
    if not product_no or not proizvajalec_id or qty <= 0:
        return make_err('BAD_REQUEST', 'Manjkajo podatki', status=400)
    db = get_db(); c = db.cursor()
    c.execute("""
        INSERT INTO perfumes_stock (product_no, proizvajalec_id, on_hand, on_order_pending, on_order_committed)
        VALUES (%s, %s, 0, 0, 0)
        ON CONFLICT (product_no, proizvajalec_id) DO NOTHING
    """, (product_no, proizvajalec_id))
    c.execute("""
        UPDATE perfumes_stock
        SET on_order_pending = GREATEST(0, on_order_pending + %s), updated_at = NOW()
        WHERE product_no = %s AND proizvajalec_id = %s
    """, (qty, product_no, proizvajalec_id))
    db.commit(); c.close()
    return make_ok({'product_no': product_no, 'proizvajalec_id': proizvajalec_id, 'delta': qty})

@api_bp.route('/procurement/cart/clear', methods=['POST'])
def procurement_cart_clear():
    supplier = (request.args.get('supplier') or '').upper()
    if supplier not in ('FLORGARDEN', 'MISTRAL'):
        return make_err('BAD_REQUEST', 'supplier mora biti FLORGARDEN ali MISTRAL', status=400)
    # Admin-only
    current_user = get_current_user() if 'get_current_user' in globals() else None
    if not current_user or (current_user.get('role') or '').lower() != 'admin':
        return make_err('FORBIDDEN', 'Čiščenje naročila je dovoljeno le administratorju', status=403)
    db = get_db(); c = db.cursor()
    c.execute("""
        UPDATE perfumes_stock ps
        SET on_order_pending = 0, updated_at = NOW()
        FROM proizvajalci pr
        WHERE ps.proizvajalec_id = pr.id AND pr.ime = %s
    """, (supplier,))
    db.commit(); c.close()
    return make_ok({'cleared_supplier': supplier})

@api_bp.route('/procurement2/cart/clear', methods=['POST'])
def procurement2_cart_clear():
    supplier = (request.args.get('supplier') or '').strip().upper()
    if not supplier:
        return make_err('BAD_REQUEST', 'Manjka supplier', status=400)
    # Admin-only via permissions (reuse same permission as import/vendors)
    perm = required_permission_for('POST', '/api/procurement/vendors/import')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Čiščenje naročila je dovoljeno le administratorju', status=403)
    try:
        _ensure_procurement_only_tables()
        db = get_db(); c = db.cursor()
        c.execute("SELECT id FROM proc_suppliers WHERE name = %s", (supplier,))
        row = c.fetchone()
        if not row:
            c.close(); return make_ok({'cleared_supplier': supplier, 'note': 'supplier not found'})
        supplier_id = row[0] if isinstance(row, tuple) else row['id']
        c.execute("UPDATE proc_products SET pending = 0, updated_at = NOW() WHERE supplier_id = %s", (supplier_id,))
        db.commit(); c.close()
        return make_ok({'cleared_supplier': supplier})
    except Exception as e:
        current_app.logger.error(f"procurement2_cart_clear error: {e}")
        try:
            get_db().rollback()
        except Exception:
            pass
        return make_err('SERVER_ERROR', 'Napaka pri čiščenju', status=500)

@api_bp.route('/procurement/cart/bulk-set', methods=['POST'])
def procurement_cart_bulk_set():
    data = request.get_json(force=True) or {}
    supplier = (data.get('supplier') or '').upper()
    items = data.get('items') or []
    if supplier not in ('FLORGARDEN','MISTRAL'):
        return make_err('BAD_REQUEST', 'Neveljaven dobavitelj', status=400)
    if not isinstance(items, list) or not items:
        return make_err('BAD_REQUEST', 'Ni podatkov', status=400)
    db = get_db(); c = db.cursor()
    try:
        for it in items:
            pn = str(it.get('product_no') or '').strip(); pid = it.get('proizvajalec_id'); qty = int(it.get('qty') or 0)
            if not pn or not pid or qty < 0:
                continue
            # enforce supplier match
            c.execute("SELECT ime FROM proizvajalci WHERE id = %s", (pid,))
            r = c.fetchone(); sup = (r['ime'] if isinstance(r, dict) else (r[0] if r else '')).upper()
            if sup != supplier:
                continue
            c.execute(
                """
                UPDATE perfumes_stock
                SET on_order_pending = GREATEST(0, %s), updated_at = NOW()
                WHERE product_no = %s AND proizvajalec_id = %s
                """,
                (qty, pn, pid)
            )
        db.commit()
        return make_ok({'updated': True})
    except Exception as e:
        db.rollback(); current_app.logger.error(f"bulk-set pending error: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri shranjevanju', status=500)
    finally:
        c.close()

@api_bp.route('/procurement/stock/bulk-onhand', methods=['POST'])
def procurement_bulk_onhand():
    data = request.get_json(force=True) or {}
    supplier = (data.get('supplier') or '').upper()
    updates = data.get('updates') or []
    if supplier not in ('FLORGARDEN','MISTRAL'):
        return make_err('BAD_REQUEST', 'Neveljaven dobavitelj', status=400)
    if not isinstance(updates, list) or not updates:
        return make_err('BAD_REQUEST', 'Ni podatkov', status=400)
    db = get_db(); c = db.cursor()
    try:
        for u in updates:
            pn = str(u.get('product_no') or '').strip(); pid = u.get('proizvajalec_id'); onh = u.get('on_hand')
            if not pn or not pid or onh is None:
                continue
            c.execute("""
                UPDATE perfumes_stock
                SET on_hand = GREATEST(0, %s), updated_at = NOW()
                WHERE product_no = %s AND proizvajalec_id = %s
            """, (int(onh), pn, int(pid)))
        db.commit()
        return make_ok({'updated': True})
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"bulk-onhand error: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri shranjevanju', status=500)
    finally:
        c.close()

@api_bp.route('/procurement/stock/add-onhand', methods=['POST'])
def procurement_add_onhand():
    data = request.get_json(force=True) or {}
    product_no = str((data.get('product_no') or '').strip())
    proizvajalec_id = data.get('proizvajalec_id')
    qty = int(data.get('qty') or 0)
    if not product_no or not proizvajalec_id or qty <= 0:
        return make_err('BAD_REQUEST', 'Neveljavni podatki', status=400)
    db = get_db(); c = db.cursor()
    try:
        c.execute("""
            INSERT INTO perfumes_stock (product_no, proizvajalec_id, on_hand, on_order_pending, on_order_committed)
            VALUES (%s, %s, 0, 0, 0)
            ON CONFLICT (product_no, proizvajalec_id) DO NOTHING
        """, (product_no, int(proizvajalec_id)))
        c.execute("""
            UPDATE perfumes_stock
            SET on_hand = GREATEST(0, on_hand + %s), updated_at = NOW()
            WHERE product_no = %s AND proizvajalec_id = %s
        """, (qty, product_no, int(proizvajalec_id)))
        db.commit()
        return make_ok({'added': qty})
    except Exception as e:
        db.rollback(); current_app.logger.error(f"add-onhand error: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri dodajanju v predal', status=500)
    finally:
        c.close()

@api_bp.route('/procurement/stock/bulk-min', methods=['POST'])
def procurement_bulk_min_on_hand():
    data = request.get_json(force=True) or {}
    supplier = (data.get('supplier') or '').upper()
    updates = data.get('updates') or []
    if supplier not in ('FLORGARDEN','MISTRAL'):
        return make_err('BAD_REQUEST', 'Neveljaven dobavitelj', status=400)
    db = get_db(); c = db.cursor()
    try:
        for u in updates:
            pn = str(u.get('product_no') or '').strip(); pid = u.get('proizvajalec_id'); thresh = u.get('min_on_hand')
            if not pn or not pid or thresh is None:
                continue
            c.execute(
                """
                UPDATE perfumes_stock
                SET min_on_hand = GREATEST(0, %s), updated_at = NOW()
                WHERE product_no = %s AND proizvajalec_id = %s
                """,
                (int(thresh), pn, int(pid))
            )
        db.commit()
        return make_ok({'updated': True})
    except Exception as e:
        db.rollback(); current_app.logger.error(f"bulk-min error: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri shranjevanju', status=500)
    finally:
        c.close()

# --- Stock movements (audit log) -------------------------------------------------
# For perfume suppliers (FLORGARDEN, MISTRAL) "last sale" = last `serije` row
# (pouring event). For procurement-only suppliers it is the last
# `proc_stock_movements` row with negative delta.

@api_bp.route('/procurement/stock-movements/last', methods=['GET'])
def procurement_last_movements_perfumes():
    """Map of product_no -> {last_at, source} for a perfume supplier."""
    supplier = (request.args.get('supplier') or '').upper()
    if supplier not in ('FLORGARDEN', 'MISTRAL'):
        return make_err('BAD_REQUEST', 'supplier mora biti FLORGARDEN ali MISTRAL', status=400)
    db = get_db(); c = db.cursor()
    try:
        c.execute(
            """
            SELECT p.product_no AS key,
                   MAX(COALESCE(s.created_at, s.created_at_original)) AS last_at
            FROM serije s
            JOIN parfumi p ON p.id = s.parfum_id
            JOIN proizvajalci pr ON pr.id = p.proizvajalec_id
            WHERE pr.ime = %s
            GROUP BY p.product_no
            """,
            (supplier,)
        )
        rows = c.fetchall() or []
        out = {}
        for r in rows:
            key = r['key'] if isinstance(r, dict) else r[0]
            last_at = r['last_at'] if isinstance(r, dict) else r[1]
            if key and last_at:
                out[str(key)] = {'last_at': last_at.isoformat() if hasattr(last_at, 'isoformat') else str(last_at), 'source': 'serija'}
        return make_ok(out)
    except Exception as e:
        current_app.logger.error(f"procurement_last_movements_perfumes error: {e}")
        return make_ok({})
    finally:
        c.close()


@api_bp.route('/procurement2/stock-movements/last', methods=['GET'])
def procurement_last_movements_proc():
    """Map of sku -> {last_at, source} for a procurement-only supplier.

    Considers only DECREMENT events (delta < 0) -- those are "sales". Inserts
    via order receive (positive delta) are not counted as sales.
    """
    supplier = (request.args.get('supplier') or '').upper()
    if not supplier:
        return make_err('BAD_REQUEST', 'supplier je obvezen', status=400)
    db = get_db(); c = db.cursor()
    try:
        # Ensure tables exist (idempotent in case migration race)
        c.execute("""
            CREATE TABLE IF NOT EXISTS proc_stock_movements (
                id BIGSERIAL PRIMARY KEY,
                supplier_id INTEGER NOT NULL,
                sku TEXT NOT NULL,
                delta INTEGER NOT NULL,
                on_hand_before INTEGER NOT NULL,
                on_hand_after INTEGER NOT NULL,
                source TEXT NOT NULL,
                source_ref TEXT NULL,
                note TEXT NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
        """)
        c.execute(
            """
            SELECT m.sku AS key,
                   MAX(m.created_at) AS last_at,
                   (SELECT source FROM proc_stock_movements m2
                      WHERE m2.supplier_id = ps.id AND m2.sku = m.sku AND m2.delta < 0
                      ORDER BY m2.created_at DESC LIMIT 1) AS source
            FROM proc_stock_movements m
            JOIN proc_suppliers ps ON ps.id = m.supplier_id
            WHERE ps.name = %s AND m.delta < 0
            GROUP BY m.sku, ps.id
            """,
            (supplier,)
        )
        rows = c.fetchall() or []
        out = {}
        for r in rows:
            key = r['key'] if isinstance(r, dict) else r[0]
            last_at = r['last_at'] if isinstance(r, dict) else r[1]
            source = r['source'] if isinstance(r, dict) else r[2]
            if key and last_at:
                out[str(key)] = {
                    'last_at': last_at.isoformat() if hasattr(last_at, 'isoformat') else str(last_at),
                    'source': source or 'unknown',
                }
        return make_ok(out)
    except Exception as e:
        current_app.logger.error(f"procurement_last_movements_proc error: {e}")
        return make_ok({})
    finally:
        c.close()


@api_bp.route('/procurement/stock-movements', methods=['GET'])
def procurement_movements_perfume_history():
    """History of last N serije rows for a perfume."""
    supplier = (request.args.get('supplier') or '').upper()
    product_no = (request.args.get('product_no') or '').strip()
    try:
        limit = max(1, min(int(request.args.get('limit') or 20), 100))
    except (TypeError, ValueError):
        limit = 20
    if supplier not in ('FLORGARDEN', 'MISTRAL') or not product_no:
        return make_err('BAD_REQUEST', 'supplier in product_no sta obvezna', status=400)
    db = get_db(); c = db.cursor()
    try:
        c.execute(
            """
            SELECT s.id,
                   COALESCE(s.created_at, s.created_at_original) AS at,
                   s.serijska_stevilka,
                   s.vnesel_uporabnik,
                   s.rok_uporabe,
                   s.stanje
            FROM serije s
            JOIN parfumi p ON p.id = s.parfum_id
            JOIN proizvajalci pr ON pr.id = p.proizvajalec_id
            WHERE pr.ime = %s AND p.product_no = %s
            ORDER BY at DESC NULLS LAST
            LIMIT %s
            """,
            (supplier, product_no, limit)
        )
        rows = c.fetchall() or []
        out = []
        for r in rows:
            d = dict(r) if not isinstance(r, dict) else r
            if d.get('at') and hasattr(d['at'], 'isoformat'):
                d['at'] = d['at'].isoformat()
            if d.get('rok_uporabe') and hasattr(d['rok_uporabe'], 'isoformat'):
                d['rok_uporabe'] = d['rok_uporabe'].isoformat()
            d['source'] = 'serija'
            out.append(d)
        return make_ok(out)
    except Exception as e:
        current_app.logger.error(f"procurement_movements_perfume_history error: {e}")
        return make_ok([])
    finally:
        c.close()


@api_bp.route('/procurement2/stock-movements', methods=['GET'])
def procurement_movements_proc_history():
    """History of last N proc_stock_movements rows for a SKU."""
    supplier = (request.args.get('supplier') or '').upper()
    sku = (request.args.get('sku') or '').strip()
    try:
        limit = max(1, min(int(request.args.get('limit') or 20), 100))
    except (TypeError, ValueError):
        limit = 20
    if not supplier or not sku:
        return make_err('BAD_REQUEST', 'supplier in sku sta obvezna', status=400)
    db = get_db(); c = db.cursor()
    try:
        c.execute(
            """
            SELECT m.id, m.created_at AS at, m.delta, m.on_hand_before, m.on_hand_after,
                   m.source, m.source_ref, m.note
            FROM proc_stock_movements m
            JOIN proc_suppliers ps ON ps.id = m.supplier_id
            WHERE ps.name = %s AND m.sku = %s
            ORDER BY m.created_at DESC
            LIMIT %s
            """,
            (supplier, sku, limit)
        )
        rows = c.fetchall() or []
        out = []
        for r in rows:
            d = dict(r) if not isinstance(r, dict) else r
            if d.get('at') and hasattr(d['at'], 'isoformat'):
                d['at'] = d['at'].isoformat()
            out.append(d)
        return make_ok(out)
    except Exception as e:
        current_app.logger.error(f"procurement_movements_proc_history error: {e}")
        return make_ok([])
    finally:
        c.close()


@api_bp.route('/procurement/cart/remove', methods=['POST'])
def procurement_cart_remove():
    data = request.get_json(force=True) or {}
    product_no = str(data.get('product_no', '')).strip()
    proizvajalec_id = data.get('proizvajalec_id')
    qty = int(data.get('qty', 0))
    remove_all = bool(data.get('remove_all'))
    if not product_no or not proizvajalec_id:
        return make_err('BAD_REQUEST', 'Manjkajo podatki', status=400)
    db = get_db(); c = db.cursor()
    if remove_all:
        c.execute("""
            UPDATE perfumes_stock
            SET on_order_pending = 0, updated_at = NOW()
            WHERE product_no = %s AND proizvajalec_id = %s
        """, (product_no, proizvajalec_id))
    else:
        if qty <= 0:
            return make_err('BAD_REQUEST', 'qty mora biti > 0', status=400)
        c.execute("""
            UPDATE perfumes_stock
            SET on_order_pending = GREATEST(0, on_order_pending - %s), updated_at = NOW()
            WHERE product_no = %s AND proizvajalec_id = %s
        """, (qty, product_no, proizvajalec_id))
    db.commit(); c.close()
    return make_ok({'product_no': product_no, 'proizvajalec_id': proizvajalec_id})

@api_bp.route('/procurement/search-perfumes', methods=['GET'])
def procurement_search_perfumes():
    q = (request.args.get('q') or '').strip()
    supplier = (request.args.get('supplier') or '').strip().upper()
    limit = int(request.args.get('limit') or 20)
    if not q:
        return make_ok(data=[])
    db = get_db(); c = db.cursor()
    params = [f'%{q}%', f'%{q}%', limit]
    supplier_sql = ''
    if supplier in ('FLORGARDEN','MISTRAL'):
        supplier_sql = 'AND UPPER(pr.ime) = %s'
        params = [f'%{q}%', f'%{q}%', supplier, limit]
    c.execute(
        f"""
        SELECT p.id as parfum_id, p.product_no, p.ime_parfuma, pr.id as proizvajalec_id, pr.ime as proizvajalec
        FROM parfumi p
        JOIN proizvajalci pr ON pr.id = p.proizvajalec_id
        WHERE (p.product_no ILIKE %s OR p.ime_parfuma ILIKE %s)
        {supplier_sql}
        ORDER BY p.ime_parfuma
        LIMIT %s
        """,
        tuple(params)
    )
    rows = c.fetchall(); c.close()
    return make_ok([dict(r) for r in rows])

# Purchase Orders CRUD (minimal MVP)
@api_bp.route('/procurement/orders/create', methods=['POST'])
def procurement_orders_create():
    # Zahteva: { supplier: 'FLORGARDEN'|'MISTRAL', notes?: str, items?:[{product_no,proizvajalec_id,qty}] }
    perm = required_permission_for('POST', '/api/procurement/orders')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    data = request.get_json(force=True) or {}
    supplier = (data.get('supplier') or '').upper()
    notes = data.get('notes')
    if supplier not in ('FLORGARDEN', 'MISTRAL'):
        return make_err('BAD_REQUEST', 'Neveljaven supplier', status=400)
    db = get_db(); c = db.cursor()
    items = (data.get('items') or [])
    if items:
        # filtriraj na istem supplierju (varnostno)
        pending = []
        for it in items:
            pn = str(it.get('product_no') or '').strip(); pid = it.get('proizvajalec_id'); qty = int(it.get('qty') or 0)
            if not pn or not pid or qty <= 0:
                continue
            # preveri supplier
            c.execute("SELECT ime FROM proizvajalci WHERE id = %s", (pid,))
            r = c.fetchone(); sup = (r['ime'] if isinstance(r, dict) else (r[0] if r else '')).upper()
            if sup != supplier:
                continue
            pending.append({'product_no': pn, 'proizvajalec_id': pid, 'on_order_pending': qty})
    else:
        # Zberi vse pending vrstice za dobavitelja iz perfumes_stock
        c.execute(
            """
            SELECT ps.product_no, ps.proizvajalec_id, ps.on_order_pending
            FROM perfumes_stock ps
            JOIN proizvajalci pr ON pr.id = ps.proizvajalec_id
            WHERE pr.ime = %s AND ps.on_order_pending > 0
            ORDER BY ps.product_no
            """,
            (supplier,)
        )
        pending = c.fetchall()
    if not pending:
        c.close()
        return make_err('BAD_REQUEST', 'Košarica je prazna', status=400)
    # Ustvari purchase_orders zapis
    created_by = session.get('user_id')
    c.execute("""
        INSERT INTO purchase_orders (supplier, status, created_by, notes)
        VALUES (%s, 'DRAFT', %s, %s)
        RETURNING id
    """, (supplier, created_by, notes))
    po_id = c.fetchone()['id']
    # Dodaj postavke
    for row in pending:
        c.execute("""
            INSERT INTO purchase_order_items (purchase_order_id, product_no, proizvajalec_id, requested_qty)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (purchase_order_id, product_no, proizvajalec_id) DO UPDATE SET
                requested_qty = EXCLUDED.requested_qty,
                updated_at = NOW()
        """, (po_id, row['product_no'], row['proizvajalec_id'], row['on_order_pending']))
    db.commit(); c.close()
    return make_ok({'purchase_order_id': po_id, 'status': 'DRAFT'})
@api_bp.route('/procurement/orders/submit/<int:po_id>', methods=['POST'])
def procurement_orders_submit(po_id: int):
    perm = required_permission_for('POST', '/api/procurement/orders/submit')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    db = get_db(); c = db.cursor()
    # Preveri obstoj in status DRAFT
    c.execute("SELECT id, supplier, status FROM purchase_orders WHERE id = %s", (po_id,))
    po = c.fetchone()
    if not po:
        c.close(); return make_err('NOT_FOUND', 'Naročilo ne obstaja', status=404)
    if po['status'] != 'DRAFT':
        c.close(); return make_err('BAD_REQUEST', 'Naročilo ni v statusu DRAFT', status=400)
    # Prenesi pending v committed in počisti pending
    c.execute("""
        SELECT product_no, proizvajalec_id, requested_qty FROM purchase_order_items WHERE purchase_order_id = %s
    """, (po_id,))
    items = c.fetchall()
    for it in items:
        c.execute("""
            UPDATE perfumes_stock
            SET on_order_pending = GREATEST(0, on_order_pending - %s),
                on_order_committed = on_order_committed + %s,
                updated_at = NOW()
            WHERE product_no = %s AND proizvajalec_id = %s
        """, (it['requested_qty'], it['requested_qty'], it['product_no'], it['proizvajalec_id']))
    c.execute("UPDATE purchase_orders SET status = 'SUBMITTED', submitted_at = NOW(), updated_at = NOW() WHERE id = %s", (po_id,))
    db.commit()

    # Preberi PO + items z metapodatki za PDF/CSV
    c = db.cursor()
    c.execute("SELECT * FROM purchase_orders WHERE id = %s", (po_id,))
    po = c.fetchone() or {}
    c.execute(
        """
        SELECT i.product_no, i.proizvajalec_id, i.requested_qty, i.received_qty,
               p.ime_parfuma, pr.ime AS proizvajalec
        FROM purchase_order_items i
        JOIN parfumi p ON p.product_no = i.product_no AND p.proizvajalec_id = i.proizvajalec_id
        JOIN proizvajalci pr ON pr.id = i.proizvajalec_id
        WHERE i.purchase_order_id = %s
        ORDER BY i.product_no
        """,
        (po_id,)
    )
    items = c.fetchall() or []
    c.close()

    # Ustvari PDF in XLSX
    pdf_path, _ = generate_purchase_order_pdf(po, items)
    xlsx_path, _ = generate_purchase_order_excel(po, items)

    # Pošlji adminu
    try:
        if pdf_path:
            # Preberi XLSX v bytes
            xlsx_bytes = b''
            if xlsx_path:
                with open(xlsx_path, 'rb') as xf:
                    xlsx_bytes = xf.read()
            # Preberi PDF v bytes
            with open(pdf_path, 'rb') as pf:
                pdf_bytes = pf.read()
            # Pošlji
            send_purchase_order_admin_email(po, items, pdf_bytes, xlsx_bytes)
    except Exception as e:
        current_app.logger.error(f'PO submit mail send failed: {e}')

    c = db.cursor(); c.close()
    return make_ok({'purchase_order_id': po_id, 'status': 'SUBMITTED'})

@api_bp.route('/procurement/orders/<int:po_id>', methods=['GET'])
def procurement_orders_get(po_id: int):
    perm = required_permission_for('GET', '/api/procurement/orders')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    db = get_db(); c = db.cursor()
    c.execute("SELECT * FROM purchase_orders WHERE id = %s", (po_id,))
    po = c.fetchone()
    if not po:
        c.close(); return make_err('NOT_FOUND', 'Naročilo ne obstaja', status=404)
    # Prinesi postavke skupaj z imenom parfuma za prikaz v UI
    c.execute(
        """
        SELECT i.*, p.ime_parfuma
        FROM purchase_order_items i
        LEFT JOIN parfumi p
          ON p.product_no = i.product_no AND p.proizvajalec_id = i.proizvajalec_id
        WHERE i.purchase_order_id = %s
        ORDER BY i.id
        """,
        (po_id,)
    )
    items = c.fetchall(); c.close()
    return make_ok({'order': po, 'items': items})

@api_bp.route('/procurement/orders', methods=['GET'])
def procurement_orders_list():
    perm = required_permission_for('GET', '/api/procurement/orders')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    supplier = (request.args.get('supplier') or '').upper()
    status = (request.args.get('status') or '').upper()
    limit = int(request.args.get('limit') or 50)
    offset = int(request.args.get('offset') or 0)
    db = get_db(); c = db.cursor()
    where = []
    params = []
    if supplier:
        where.append('supplier = %s'); params.append(supplier)
    if status in ('DRAFT','SUBMITTED','PARTIAL_RECEIVED','RECEIVED','ARCHIVED','CANCELLED'):
        where.append('status = %s'); params.append(status)
    where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''
    c.execute(
        f"""
        SELECT po.id, po.supplier, po.status, po.submitted_at, po.received_at, po.created_at, po.updated_at, po.email_pdf_url,
               (SELECT COUNT(*) FROM purchase_order_items i WHERE i.purchase_order_id = po.id) AS items_count,
               (SELECT COALESCE(SUM(i.requested_qty),0) FROM purchase_order_items i WHERE i.purchase_order_id = po.id) AS total_requested,
               (SELECT COALESCE(SUM(i.received_qty),0) FROM purchase_order_items i WHERE i.purchase_order_id = po.id) AS total_received
        FROM purchase_orders po
        {where_sql}
        ORDER BY po.id DESC
        LIMIT %s OFFSET %s
        """,
        (*params, limit, offset)
    )
    orders = [dict(r) for r in c.fetchall()]
    c.close()
    return make_ok({'orders': orders})

@api_bp.route('/procurement/orders/<int:po_id>/print', methods=['GET'])
def procurement_orders_print(po_id: int):
    """Vrne PDF povzetek naročila (za 'Natisni seznam')."""
    perm = required_permission_for('GET', '/api/procurement/orders')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    db = get_db(); c = db.cursor()
    c.execute("SELECT * FROM purchase_orders WHERE id = %s", (po_id,))
    po = c.fetchone()
    if not po:
        c.close(); return make_err('NOT_FOUND', 'Naročilo ne obstaja', status=404)
    c.execute(
        """
        SELECT i.product_no, i.proizvajalec_id, i.requested_qty, i.received_qty,
               p.ime_parfuma, pr.ime AS proizvajalec
        FROM purchase_order_items i
        JOIN parfumi p ON p.product_no = i.product_no AND p.proizvajalec_id = i.proizvajalec_id
        JOIN proizvajalci pr ON pr.id = i.proizvajalec_id
        WHERE i.purchase_order_id = %s
        ORDER BY i.product_no
        """,
        (po_id,)
    )
    items = c.fetchall() or []
    c.close()
    pdf_path, msg = generate_purchase_order_pdf(po, items)
    if not pdf_path:
        return make_err('SERVER_ERROR', f'Napaka pri generiranju PDF: {msg}', status=500)
    return send_file(pdf_path, mimetype='application/pdf', as_attachment=False, download_name=f'PO_{po_id}.pdf')
@api_bp.route('/procurement/orders/<int:po_id>/xlsx', methods=['GET'])
def procurement_orders_xlsx(po_id: int):
    perm = required_permission_for('GET', '/api/procurement/orders')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    db = get_db(); c = db.cursor()
    c.execute("SELECT * FROM purchase_orders WHERE id = %s", (po_id,))
    po = c.fetchone()
    if not po:
        c.close(); return make_err('NOT_FOUND', 'Naročilo ne obstaja', status=404)
    c.execute(
        """
        SELECT i.product_no, i.proizvajalec_id, i.requested_qty, i.received_qty,
               p.ime_parfuma, pr.ime AS proizvajalec
        FROM purchase_order_items i
        JOIN parfumi p ON p.product_no = i.product_no AND p.proizvajalec_id = i.proizvajalec_id
        JOIN proizvajalci pr ON pr.id = i.proizvajalec_id
        WHERE i.purchase_order_id = %s
        ORDER BY i.product_no
        """,
        (po_id,)
    )
    items = c.fetchall() or []
    c.close()
    from services.excel_service import generate_purchase_order_excel
    xlsx_path, msg = generate_purchase_order_excel(po, items)
    if not xlsx_path:
        return make_err('SERVER_ERROR', f'Napaka pri generiranju XLSX: {msg}', status=500)
    return send_file(xlsx_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'PO_{po_id}.xlsx')

@api_bp.route('/procurement/orders/<int:po_id>/duplicate', methods=['POST'])
def procurement_orders_duplicate(po_id: int):
    perm = required_permission_for('POST', '/api/procurement/orders/duplicate')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    db = get_db(); c = db.cursor()
    try:
        c.execute("SELECT supplier, notes FROM purchase_orders WHERE id = %s", (po_id,))
        row = c.fetchone()
        if not row:
            return make_err('NOT_FOUND', 'Naročilo ne obstaja', status=404)
        supplier = row['supplier'] if isinstance(row, dict) else row[0]
        notes = row['notes'] if isinstance(row, dict) else (row[1] if len(row)>1 else None)
        created_by = session.get('user_id')
        c.execute("INSERT INTO purchase_orders (supplier, status, created_by, notes) VALUES (%s,'DRAFT',%s,%s) RETURNING id", (supplier, created_by, notes))
        new_id = c.fetchone()['id']
        c.execute("SELECT product_no, proizvajalec_id, requested_qty FROM purchase_order_items WHERE purchase_order_id = %s", (po_id,))
        for it in c.fetchall():
            pno = it['product_no'] if isinstance(it, dict) else it[0]
            pid = it['proizvajalec_id'] if isinstance(it, dict) else it[1]
            qty = it['requested_qty'] if isinstance(it, dict) else it[2]
            c.execute("INSERT INTO purchase_order_items (purchase_order_id, product_no, proizvajalec_id, requested_qty) VALUES (%s,%s,%s,%s)", (new_id, pno, pid, qty))
        db.commit(); c.close()
        return make_ok({'purchase_order_id': new_id, 'status': 'DRAFT'})
    except Exception as e:
        db.rollback(); current_app.logger.error(f"duplicate PO error: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri podvajanju', status=500)

@api_bp.route('/procurement/orders/<int:po_id>/receive', methods=['POST'])
def procurement_orders_receive(po_id: int):
    perm = required_permission_for('POST', '/api/procurement/orders/receive')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    data = request.get_json(silent=True) or {}
    items = data.get('items') or []  # [{product_no, proizvajalec_id, received_qty, notes?}]
    require_image = bool(data.get('require_image'))
    all_received = bool(data.get('all_received'))
    db = get_db(); c = db.cursor()
    c.execute("SELECT id, status FROM purchase_orders WHERE id = %s", (po_id,))
    po = c.fetchone()
    if not po:
        c.close(); return make_err('NOT_FOUND', 'Naročilo ne obstaja', status=404)
    if po['status'] not in ('SUBMITTED', 'PARTIAL_RECEIVED'):
        c.close(); return make_err('BAD_REQUEST', 'Naročilo ni oddano ali v delnem prejemu', status=400)
    # Zahtevaj vsaj eno postavko
    if not items:
        c.close(); return make_err('BAD_REQUEST', 'Ni postavk za prejem', status=400)
    # Če je zahtevana slika, preveri pred spremembami
    if require_image:
        c2 = db.cursor()
        c2.execute("SELECT COUNT(1) FROM order_images WHERE purchase_order_id = %s", (po_id,))
        cnt = c2.fetchone(); c2.close()
        num = cnt['count'] if isinstance(cnt, dict) else (cnt[0] if cnt else 0)
        if (num or 0) <= 0:
            c.close(); return make_err('BAD_REQUEST', 'Za prejem je obvezna vsaj ena slika računa/dobavnice', status=400)
    # Posodobi postavke in zalogo
    new_status = 'RECEIVED' if all_received else 'PARTIAL_RECEIVED'
    for it in items:
        pn = str(it.get('product_no') or '').strip(); pid = it.get('proizvajalec_id'); rcv = int(it.get('received_qty') or 0)
        if not pn or not pid or rcv < 0:
            db.rollback(); c.close(); return make_err('BAD_REQUEST', 'Neveljavni podatki postavk', status=400)
        # Preberi obstoječo postavko
        c.execute("""
            SELECT requested_qty, received_qty
            FROM purchase_order_items
            WHERE purchase_order_id = %s AND product_no = %s AND proizvajalec_id = %s
        """, (po_id, pn, pid))
        row = c.fetchone()
        if not row:
            db.rollback(); c.close(); return make_err('BAD_REQUEST', 'Postavka ne obstaja', status=400)
        max_add = max(0, row['requested_qty'] - row['received_qty'])
        if rcv > max_add:
            db.rollback(); c.close(); return make_err('BAD_REQUEST', 'Prejeta količina presega preostanek', status=400)
        # Posodobi postavko
        c.execute("""
            UPDATE purchase_order_items
            SET received_qty = received_qty + %s,
                backordered_qty = GREATEST(0, requested_qty - (received_qty + %s)),
                updated_at = NOW()
            WHERE purchase_order_id = %s AND product_no = %s AND proizvajalec_id = %s
        """, (rcv, rcv, po_id, pn, pid))
        # Posodobi zalogo: committed se zmanjša, on_hand se poveča
        c.execute("""
            UPDATE perfumes_stock
            SET on_order_committed = GREATEST(0, on_order_committed - %s),
                on_hand = on_hand + %s,
                updated_at = NOW()
            WHERE product_no = %s AND proizvajalec_id = %s
        """, (rcv, rcv, pn, pid))
    # Posodobi status naročila
    c.execute("UPDATE purchase_orders SET status = %s, received_at = CASE WHEN %s = 'RECEIVED' THEN NOW() ELSE received_at END, updated_at = NOW() WHERE id = %s",
              (new_status, new_status, po_id))
    db.commit(); c.close()
    return make_ok({'purchase_order_id': po_id, 'status': new_status})

@api_bp.route('/procurement/orders/<int:po_id>/archive', methods=['POST'])
def procurement_orders_archive(po_id: int):
    perm = required_permission_for('POST', '/api/procurement/orders/archive')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    db = get_db(); c = db.cursor()
    archived_by = session.get('user_id')
    # Ensure columns exist (no-op if already there)
    try:
        c.execute("ALTER TABLE purchase_orders ADD COLUMN IF NOT EXISTS archived_by INTEGER")
    except Exception:
        pass
    c.execute("UPDATE purchase_orders SET status = 'ARCHIVED', archived_by = %s, updated_at = NOW() WHERE id = %s", (archived_by, po_id))
    if c.rowcount == 0:
        db.rollback(); c.close(); return make_err('NOT_FOUND', 'Naročilo ne obstaja', status=404)
    db.commit(); c.close()
    return make_ok({'purchase_order_id': po_id, 'status': 'ARCHIVED'})

@api_bp.route('/procurement/orders/<int:po_id>', methods=['DELETE'])
def procurement_orders_delete(po_id: int):
    # Le admin lahko izbriše naročilo
    current_user = get_current_user()
    if not current_user or (current_user.get('role') or '').lower() != 'admin':
        return make_err('FORBIDDEN', 'Brisanje je dovoljeno le administratorju', status=403)
    db = get_db(); c = db.cursor()
    # Najprej vrni zalogo nazaj, če je SUBMITTED/COMMITTED
    c.execute("SELECT status FROM purchase_orders WHERE id = %s", (po_id,))
    row = c.fetchone()
    if not row:
        c.close(); return make_err('NOT_FOUND', 'Naročilo ne obstaja', status=404)
    status = row['status'] if isinstance(row, dict) else row[0]
    if status in ('SUBMITTED', 'PARTIAL_RECEIVED'):
        # Vrni committed v pending? Za varnost zmanjšamo committed in ne povečamo pending
        c.execute("SELECT product_no, proizvajalec_id, requested_qty FROM purchase_order_items WHERE purchase_order_id = %s", (po_id,))
        for it in c.fetchall():
            pno = it['product_no'] if isinstance(it, dict) else it[0]
            pid = it['proizvajalec_id'] if isinstance(it, dict) else it[1]
            qty = it['requested_qty'] if isinstance(it, dict) else it[2]
            c.execute(
                """
                UPDATE perfumes_stock
                SET on_order_committed = GREATEST(0, on_order_committed - %s),
                    updated_at = NOW()
                WHERE product_no = %s AND proizvajalec_id = %s
                """,
                (qty, pno, pid)
            )
    # Izbriši postavke in order
    c.execute("DELETE FROM purchase_order_items WHERE purchase_order_id = %s", (po_id,))
    # Track deleter (best-effort, table column may be absent)
    deleter = session.get('user_id')
    try:
        c.execute("ALTER TABLE purchase_orders ADD COLUMN IF NOT EXISTS deleted_by INTEGER")
        c.execute("UPDATE purchase_orders SET deleted_by = %s WHERE id = %s", (deleter, po_id))
    except Exception:
        pass
    c.execute("DELETE FROM purchase_orders WHERE id = %s", (po_id,))
    db.commit(); c.close()
    return make_ok({'deleted': True, 'purchase_order_id': po_id})

# ----------------------
# Manual receive (without prior submitted order)
# ----------------------

@api_bp.route('/procurement/receive/manual/create', methods=['POST'])
def procurement_manual_receive_create():
    perm = required_permission_for('POST', '/api/procurement/receive')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    payload = request.get_json(silent=True) or {}
    supplier = (payload.get('supplier') or '').upper()
    if supplier not in ('FLORGARDEN', 'MISTRAL'):
        return make_err('BAD_REQUEST', 'Neveljaven supplier', status=400)
    db = get_db(); c = db.cursor()
    created_by = session.get('user_id')
    c.execute(
        """
        INSERT INTO purchase_orders (supplier, status, created_by, notes)
        VALUES (%s, 'DRAFT', %s, %s)
        RETURNING id
        """,
        (supplier, created_by, 'MANUAL_RECEIVE')
    )
    po_id = c.fetchone()['id']
    db.commit(); c.close()
    return make_ok({'purchase_order_id': po_id, 'status': 'DRAFT'})

@api_bp.route('/procurement/receive/manual/commit/<int:po_id>', methods=['POST'])
def procurement_manual_receive_commit(po_id: int):
    perm = required_permission_for('POST', '/api/procurement/receive')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    payload = request.get_json(silent=True) or {}
    items = payload.get('items') or []
    require_image = bool(payload.get('require_image'))
    if not items:
        return make_err('BAD_REQUEST', 'Ni postavk za prejem', status=400)
    db = get_db(); c = db.cursor()
    # Validate PO exists and is DRAFT (created by manual)
    c.execute("SELECT id, supplier, status FROM purchase_orders WHERE id = %s", (po_id,))
    po = c.fetchone()
    if not po:
        c.close(); return make_err('NOT_FOUND', 'Naročilo ne obstaja', status=404)
    if po['status'] not in ('DRAFT', 'PARTIAL_RECEIVED'):
        c.close(); return make_err('BAD_REQUEST', 'Naročilo ni v ustreznem statusu', status=400)
    # Optionally enforce at least one image exists
    if require_image:
        c2 = db.cursor()
        c2.execute("SELECT COUNT(1) FROM order_images WHERE purchase_order_id = %s", (po_id,))
        cnt = c2.fetchone(); c2.close()
        num = cnt['count'] if isinstance(cnt, dict) else (cnt[0] if cnt else 0)
        if (num or 0) <= 0:
            c.close(); return make_err('BAD_REQUEST', 'Za prejem je obvezna vsaj ena slika računa/dobavnice', status=400)
    # Insert or update items, and update stock
    for it in items:
        pn = str(it.get('product_no') or '').strip(); pid = it.get('proizvajalec_id'); rcv = int(it.get('received_qty') or 0)
        if not pn or not pid or rcv <= 0:
            db.rollback(); c.close(); return make_err('BAD_REQUEST', 'Neveljavni podatki postavk', status=400)
        # ensure perfume exists
        c.execute("SELECT 1 FROM parfumi WHERE product_no = %s AND proizvajalec_id = %s", (pn, pid))
        if not c.fetchone():
            db.rollback(); c.close(); return make_err('BAD_REQUEST', 'Parfum ne obstaja', status=400)
        # upsert item
        c.execute(
            """
            INSERT INTO purchase_order_items (purchase_order_id, product_no, proizvajalec_id, requested_qty, received_qty)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (purchase_order_id, product_no, proizvajalec_id) DO UPDATE SET
                requested_qty = purchase_order_items.requested_qty + EXCLUDED.requested_qty,
                received_qty = purchase_order_items.received_qty + EXCLUDED.received_qty,
                updated_at = NOW()
            """,
            (po_id, pn, pid, rcv, rcv)
        )
        # update stock (only on_hand increases; no committed decrease for manual receive)
        c.execute(
            """
            UPDATE perfumes_stock
            SET on_hand = on_hand + %s,
                updated_at = NOW()
            WHERE product_no = %s AND proizvajalec_id = %s
            """,
            (rcv, pn, pid)
        )
    # finalize PO as RECEIVED
    c.execute("UPDATE purchase_orders SET status = 'RECEIVED', received_at = NOW(), updated_at = NOW() WHERE id = %s", (po_id,))
    db.commit(); c.close()
    return make_ok({'purchase_order_id': po_id, 'status': 'RECEIVED'})

# Purchase order receipt images (list, upload, delete, proxy)
@api_bp.route('/purchase-orders/<int:po_id>/images', methods=['GET'])
def list_po_images(po_id: int):
    perm = required_permission_for('GET', '/api/purchase-orders')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    db = get_db(); c = db.cursor()
    c.execute("SELECT id, s3_key, uploaded_at, uploaded_by, user_id FROM order_images WHERE purchase_order_id = %s ORDER BY uploaded_at DESC", (po_id,))
    rows = c.fetchall(); c.close()
    # Proxy URL-ji
    from flask import url_for
    images = []
    for r in rows:
        s3_key = r['s3_key'] if isinstance(r, dict) else r[1]
        url = url_for('api.proxy_po_image', s3_key=s3_key, _external=True)
        images.append({
            'id': r['id'] if isinstance(r, dict) else r[0],
            'url': url,
            's3_key': s3_key,
            'uploaded_at': (r['uploaded_at'] if isinstance(r, dict) else r[2])
        })
    # Vrni neposredno seznam slik, da se ujema z ostalimi APIji (json.data je array)
    return make_ok(images)

@api_bp.route('/purchase-orders/<int:po_id>/images', methods=['POST'])
def upload_po_image(po_id: int):
    perm = required_permission_for('POST', '/api/purchase-orders')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    try:
        if 'image' not in request.files:
            return make_err('BAD_REQUEST', "Manjka datoteka 'image'", status=400)
        file = request.files['image']
        file_bytes = file.read()
        if not file_bytes:
            return make_err('BAD_REQUEST', 'Prazna datoteka', status=400)
        from services.s3_service import upload_purchase_order_image
        res = upload_purchase_order_image(file_bytes, po_id, session.get('user_id'))
        # Zapiši v order_images s purchase_order_id; če stolpec ne obstaja (starejša shema), ga dodaj
        db = get_db(); c = db.cursor()
        try:
            c.execute("""
                SELECT 1 FROM information_schema.columns
                WHERE table_name = 'order_images' AND column_name = 'purchase_order_id'
            """)
            if not c.fetchone():
                c.execute("ALTER TABLE order_images ADD COLUMN purchase_order_id INTEGER NULL")
        except Exception:
            # nadaljuj – če obstaja, ignoriraj napako
            db.rollback(); c = db.cursor()
        # Nekatere instalacije imajo NOT NULL na order_number → uporabimo sintetičen "PO-<id>"
        order_number_value = f"PO-{po_id}"
        c.execute(
            """
            INSERT INTO order_images (order_number, s3_key, s3_url, uploaded_by, user_id, purchase_order_id, uploaded_at)
            VALUES (%s, %s, %s, %s, %s, %s, NOW())
            RETURNING id
            """,
            (order_number_value, res['key'], res.get('url',''), session.get('username'), session.get('user_id'), po_id)
        )
        img_id = c.fetchone()['id']
        db.commit(); c.close()
        return make_ok({'id': img_id, 'url': res['url'], 's3_key': res['key']})
    except Exception as e:
        current_app.logger.error(f"PO image upload error for PO {po_id}: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri nalaganju slike', status=500)

@api_bp.route('/purchase-orders/images/<int:image_id>', methods=['DELETE'])
def delete_po_image(image_id: int):
    perm = required_permission_for('DELETE', '/api/purchase-orders')
    if perm and not has_permission(perm):
        return make_err('FORBIDDEN', 'Dostop zavrnjen', status=403)
    db = get_db(); c = db.cursor()
    c.execute("SELECT s3_key FROM order_images WHERE id = %s", (image_id,))
    row = c.fetchone()
    if not row:
        c.close(); return make_err('NOT_FOUND', 'Slika ne obstaja', status=404)
    s3_key = row['s3_key'] if isinstance(row, dict) else row[0]
    from services.s3_service import delete_order_image
    try:
        delete_order_image(s3_key)
    except Exception:
        current_app.logger.warning(f'PO image delete S3 failed for {s3_key}')
    c.execute("DELETE FROM order_images WHERE id = %s", (image_id,))
    db.commit(); c.close()
    return make_ok({'deleted': True})

@api_bp.route('/purchase-orders/proxy/<path:s3_key>', methods=['GET'])
def proxy_po_image(s3_key: str):
    try:
        if not s3_key.startswith('po_receipts/'):
            return jsonify({'error': 'Neveljavna pot slike'}), 403
        from services.s3_service import get_s3_client
        s3_client = get_s3_client()
        bucket_name = current_app.config['S3_BUCKET_NAME']
        response = s3_client.get_object(Bucket=bucket_name, Key=s3_key)
        image_data = response['Body'].read()
        content_type = response.get('ContentType', 'image/jpeg')
        return Response(image_data, mimetype=content_type, headers={'Cache-Control': 'public, max-age=3600', 'Content-Disposition':'inline'})
    except Exception as e:
        current_app.logger.error(f"Napaka pri proxy_po_image {s3_key}: {e}")
        return jsonify({'error': 'Slika ni dostopna'}), 404
# Globalna spremenljivka za sledenje zadnjega osveževanja
last_order_check = None

# ----------------------
# Centraliziran nadzor dovoljenj za API
# ----------------------

ALLOWED_PUBLIC_PATHS = {
    ('POST', '/api/login'),
    ('GET', '/api/health'),
    ('HEAD', '/api/health'),  # JS pinga /api/health preko HEAD za online/offline preverjanje
    ('OPTIONS', '/api/health'),
    ('GET', '/api/current-user'),  # vrne podatke o trenutnem uporabniku po seji
    # MetaKocka stock webhook (avtorizirano preko X-MK-Secret — fail-closed)
    ('POST', '/api/mk/webhook/stock'),
    ('POST', '/api/mk/sync-bills/secret'),
    ('GET', '/api/mk/sync-bills/status/secret'),
    ('POST', '/api/mk/retail/import-by-ids/secret'),
    ('POST', '/api/mk/retail/delta/secret'),
    ('POST', '/api/mk/retail/change-delta/secret'),
    ('POST', '/api/mk/retail/search/secret'),
    ('GET', '/api/mk/retail/status/secret'),
    ('GET', '/api/mk/retail/inspect/secret'),
    ('GET', '/api/mk/retail/inspect-items/secret'),
}
def required_permission_for(method: str, path: str):
    # Metakocka računi
    if path == '/api/mk/send-invoice' and method == 'POST':
        return 'send_invoice'
    """Vrne potrebno dovoljenje glede na metodo in pot."""
    # Naročila
    if method == 'GET' and path.startswith('/api/narocila'):
        return 'view_orders'

    # Proizvajalci
    # Procurement
    if path.startswith('/api/procurement'):
        if method == 'GET':
            return 'view_orders'  # videnje zalog/osnovnih podatkov
        if method in ('POST', 'PUT', 'DELETE'):
            return 'order_procurement'
        return 'order_procurement'
    # Purchase order images and actions
    if path.startswith('/api/purchase-orders'):
        if method == 'GET':
            return 'view_orders'
        return 'order_procurement'
    if path == '/api/proizvajalci' and method == 'GET':
        return 'view_proizvajalci'
    if path == '/api/proizvajalci' and method == 'POST':
        return 'add_proizvajalci'
    if path.startswith('/api/proizvajalci/') and method == 'DELETE':
        return 'delete_proizvajalci'

    # Parfumi
    if path == '/api/parfumi' and method == 'GET':
        return 'view_perfumes'
    if path == '/api/parfumi' and method == 'POST':
        # Shranjevanje parfuma: običajno 'edit_perfumes'. Dodatna logika v enforce_permissions dovoli delna dovoljenja pri posodobitvah.
        return 'edit_perfumes'
    if path.startswith('/api/parfum/') and method == 'GET':
        return 'view_perfumes'
    if path.startswith('/api/parfumi_by_proizvajalec/') and method == 'GET':
        return 'view_perfumes'
    if path.endswith('/stock-status') and path.startswith('/api/parfum/') and method == 'POST':
        return 'edit_perfumes'
    if path.endswith('/sync-status') and path.startswith('/api/parfum/') and method == 'POST':
        return 'edit_perfumes'

    # Serije
    if path == '/api/serije' and method == 'GET':
        return 'view_perfumes'
    if path == '/api/serije' and method == 'POST':
        return 'add_serije'
    if path.startswith('/api/serije/') and method == 'PUT':
        return 'edit_serije'
    if path.startswith('/api/serije/') and method == 'DELETE':
        return 'delete_serije'

    # PDF in e-pošta
    if path in (
        '/api/generiraj-deklaracijo-za-tisk',
        '/api/generiraj-pdf-rocno',
    ) and method == 'POST':
        return 'generate_pdf'
    if path in (
        '/api/generiraj_in_poslji',
        '/api/ponovno_poslji_deklaracijo',
    ) and method == 'POST':
        return 'send_auto_declarations'
    if path in (
        '/api/poslji-rocno',
    ) and method == 'POST':
        return 'send_email'
    if path == '/api/email-logs' and method == 'GET':
        return 'send_email'
    if path.startswith('/api/email-details/') and method == 'GET':
        return 'send_email'

    # Shopify sinhronizacije in migracije
    if path in (
        '/api/sync-new-orders',
        '/api/sync-fulfilled-status',
        '/api/sync-new-perfumes',
        '/api/sync-names',
        '/api/sync-stock-status',
        '/api/sync-inci-from-shopify',
        '/api/sync-all-inci-from-shopify',
        '/api/auto-enable-shopify-sync',
        '/api/auto-disable-shopify-sync',
        '/api/check-new-orders',
        '/api/migrate-onedrive',
        '/api/migrate-local-excel',
        '/api/migrate-local-file',
        '/api/restore-parfumi-names',
        '/api/export-parfumi-imena',
        '/api/import-parfumi-imena',
        '/api/preview-fix-amour-parfums-names',
        '/api/apply-fix-amour-parfums-names',
        '/api/run-migration',
        '/api/register-webhooks',
        '/api/list-webhooks',
    ) and method in ('GET', 'POST'):
        return 'shopify_sync'

    if path == '/api/shopify-stores' and method == 'GET':
        return 'edit_perfumes'

    if path.startswith('/api/search-synonyms') and method in ('GET', 'POST', 'DELETE'):
        return 'edit_perfumes'

    # Slike naročil
    if path.startswith('/api/order-images/') and method == 'DELETE':
        return 'view_orders'  # Ista pravica kot za ogled naročil
    if path.startswith('/api/order-images/') and method == 'GET':
        return 'view_orders'  # Get images for order
    if path.startswith('/api/order-images') and method == 'POST':
        return 'view_orders'  # Upload slik
    if path.endswith('/set-nalivalec') and method == 'POST':
        return 'view_orders'
    if path.endswith('/set-prepared-by') and method == 'POST':
        return 'view_orders'
    if path.endswith('/reset-preparation') and method == 'POST':
        return 'view_orders'  # Reset preparation fields
    if path == '/api/mk/sync-declaration-uploads' and method == 'POST':
        return 'admin'
    if path == '/api/cleanup-orders-without-images' and method == 'POST':
        return 'admin'  # Admin only cleanup

    # Uporabniki
    if path == '/api/users' and method == 'GET':
        return 'view_users'
    if path == '/api/users' and method == 'POST':
        return 'manage_users'
    if path.startswith('/api/users/') and path.endswith('/permissions') and method == 'PUT':
        return 'manage_users'
    if path.startswith('/api/users/') and method == 'DELETE':
        return 'manage_users'

    # Ostalo
    if path == '/api/expiring-perfumes' and method == 'GET':
        return 'view_perfumes'

    # Navodila
    if path == '/api/instructions' and method == 'POST':
        return 'edit_users'
    if path == '/api/instruction-categories' and method == 'POST':
        return 'edit_users'
    if path.startswith('/api/instructions/') and method in ('PUT', 'DELETE', 'POST'):
        # PUT za posodobitve, DELETE za brisanje, POST za npr. upload slike pod navodilom
        return 'edit_users'

    # Fix user permissions endpoint naj bo samo za admina prek manage_users
    if path == '/api/fix-user-permissions' and method == 'POST':
        return 'manage_users'

    # Metakocka računi
    if path == '/api/mk/send-invoice' and method == 'POST':
        return 'send_invoice'

    return None

@api_bp.before_request
def enforce_permissions():
    """Globalno uveljavi dovoljenja za API poti."""
    method = request.method
    path = request.path

    # Dovoli javne poti
    if (method, path) in ALLOWED_PUBLIC_PATHS:
        return None

    # Vsi ostali API klici zahtevajo, da je uporabnik prijavljen
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Uporabnik ni prijavljen'}), 401

    # Admin samo po vlogi v seji (NE po uporabniškem imenu — to je bil bypass).
    role_norm_user_obj = str(session.get('user', {}).get('role', '')).strip().lower()
    role_norm_flat = str(session.get('role', '')).strip().lower()
    if role_norm_user_obj == 'admin' or role_norm_flat == 'admin':
        return None

    # Posebna obravnava za shranjevanje parfuma: pri posodobitvi dovoli tudi delna dovoljenja
    if path == '/api/parfumi' and method == 'POST':
        data = request.get_json(silent=True) or {}
        is_update = bool(data.get('id'))
        if is_update:
            # Posodobitev obstoječega parfuma je dovoljena, če ima uporabnik eno od teh dovoljenj
            if has_permission('edit_perfumes') or has_permission('edit_perfume_stock') or has_permission('edit_perfume_sync'):
                return None
            current_app.logger.warning(
                f"Dostop zavrnjen: {session.get('username')} potrebuje eno izmed ['edit_perfumes','edit_perfume_stock','edit_perfume_sync'] za {method} {path} (update)"
            )
            return jsonify({'success': False, 'error': 'Nimate dovoljenja za to akcijo'}), 403

    # Eksplicitno nevarne admin akcije, ki so bile prej odprte vsakemu prijavljenemu.
    DANGEROUS_UNMAPPED = {
        '/api/cleanup-duplicate-perfumes',
        '/api/merge-perfumes',
        '/api/fix-user-permissions',
        '/api/init-db',
    }
    if path in DANGEROUS_UNMAPPED or path.startswith('/api/admin/'):
        if role_norm_user_obj != 'admin' and role_norm_flat != 'admin':
            return jsonify({'success': False, 'error': 'Samo administrator'}), 403

    perm = required_permission_for(method, path)
    if not perm:
        # Še ni mapirano — pusti (kompatibilnost). Nevarne poti so zgoraj.
        return None

    if perm and not has_permission(perm):
        current_app.logger.warning(f"Dostop zavrnjen: {session.get('username')} potrebuje '{perm}' za {method} {path}")
        return jsonify({'success': False, 'error': 'Nimate dovoljenja za to akcijo'}), 403

    return None
def _mk_send_invoice_in_background(app_obj, order_number: str, recipient_email: str | None = None):
    """Težko delo za MK pošiljanje računa v ozadju, da se izognemo H12 timeoutom."""
    with app_obj.app_context():
        db = get_db(); c = db.cursor()
        # Audit table ensure
        try:
            c.execute(
                """
                CREATE TABLE IF NOT EXISTS invoice_email_log (
                  id BIGSERIAL PRIMARY KEY,
                  order_number TEXT,
                  mk_id TEXT,
                  mk_type TEXT,
                  recipient TEXT,
                  status TEXT,
                  error TEXT,
                  created_at TIMESTAMP DEFAULT NOW()
                )
                """
            ); db.commit()
        except Exception:
            db.rollback()
        def _audit(status: str, recipient: str = None, mk_id: str = None, mk_type: str = None, error: str = None):
            try:
                cc = db.cursor()
                cc.execute(
                    """
                    INSERT INTO invoice_email_log (order_number, mk_id, mk_type, recipient, status, error)
                    VALUES (%s,%s,%s,%s,%s,%s)
                    """,
                    (order_number, mk_id, mk_type, recipient, status, error)
                ); db.commit(); cc.close()
            except Exception:
                db.rollback()
        try:
            c.execute(
                """
                SELECT customer_email, customer_name, country_code, status_url, line_items, shopify_order_id,
                       fulfilled_at, shopify_fulfilled_at
                FROM orders
                WHERE order_number = %s OR order_number = %s
                """,
                (order_number, f"#{order_number}")
            )
            row = c.fetchone()
            if not row:
                current_app.logger.error(f"MK send-invoice: order {order_number} not found")
                _audit('not_found')
                return
            customer_email = row['customer_email'] if isinstance(row, dict) else row[0]
            customer_name  = row['customer_name']  if isinstance(row, dict) else row[1]
            country_code   = row['country_code']   if isinstance(row, dict) else row[2]
            status_url     = row['status_url']     if isinstance(row, dict) else row[3]
            line_items     = row['line_items']     if isinstance(row, dict) else row[4]
            shopify_order_id = row['shopify_order_id'] if isinstance(row, dict) else row[5]
            fulfilled_at   = row['fulfilled_at'] if isinstance(row, dict) else row[6]
            shopify_fulfilled_at = row['shopify_fulfilled_at'] if isinstance(row, dict) else row[7]
            shop_url = current_app.config.get('SHOP_URL') or current_app.config.get('APP_BASE_URL') or 'https://deklaracije.eu'
            if recipient_email:
                customer_email = recipient_email
            if not fulfilled_at and not shopify_fulfilled_at:
                current_app.logger.error(f"MK send-invoice: order {order_number} is not fulfilled")
                _audit('not_fulfilled', recipient=customer_email)
                return
            if not customer_email:
                current_app.logger.error(f"MK send-invoice: order {order_number} has no customer_email")
                _audit('no_recipient')
                return

            # Najprej poskusi iz naše baze mk_bills; nato query + strogi post-filter
            from services.mk_service import mk_find_bill_in_db, mk_find_bill_any, mk_is_published, mk_get_document, mk_get_document_bill, mk_doc_matches_customer
            db_hit = mk_find_bill_in_db(order_number)
            bill = None
            if db_hit:
                bill = mk_get_document(db_hit['doc_type'], db_hit['mk_id'])
                if bill:
                    bill['_doc_type'] = db_hit['doc_type']
            if not bill:
                bill = mk_find_bill_any(order_number)
            if not bill and shopify_order_id:
                # Fallback: nekateri računi imajo v buyer_order ali title zapisano Shopify ID namesto #name
                bill = mk_find_bill_any(str(shopify_order_id))
            if not bill:
                try:
                    from services.mk_service import mk_search_bills
                    candidates = mk_search_bills(['sales_bill_domestic','sales_bill_foreign','sales_bill_retail','sales_bill_prepaid'], order_number)
                    current_app.logger.error(f"MK send-invoice: bill not found for title {order_number}. Candidates: {candidates[:5]}")
                except Exception:
                    current_app.logger.error("MK send-invoice: diagnostic search failed")
                _audit('bill_not_found', recipient=customer_email)
                return
            # Osveži dokument z get_document/get_document_bill za polna polja (customer/publish_ts)
            try:
                mk_id = bill.get('mk_id')
                doc_type = bill.get('_doc_type') or bill.get('doc_type')
                detailed = None
                if mk_id and doc_type:
                    detailed = mk_get_document(doc_type, mk_id)
                if not detailed and mk_id:
                    detailed = mk_get_document_bill(mk_id)
                if detailed:
                    detailed['_doc_type'] = doc_type or detailed.get('_doc_type')
                    bill = detailed
            except Exception:
                pass

            if customer_name and not mk_doc_matches_customer(bill, customer_name):
                current_app.logger.error(f"MK send-invoice: customer mismatch for {order_number} ({customer_name})")
                _audit('customer_mismatch', recipient=customer_email, mk_id=bill.get('mk_id'), mk_type=bill.get('_doc_type'))
                return
            if not mk_is_published(bill):
                current_app.logger.error(f"MK send-invoice: bill {bill.get('mk_id')} not published")
                _audit('not_published', recipient=customer_email, mk_id=bill.get('mk_id'), mk_type=bill.get('_doc_type'))
                return

            pdf_path = None
            try:
                # Poskusi najprej uradni MK PDF (OBVEZNO za pošiljanje računa)
                from services.mk_service import mk_print_bill_pdf
                found_type = bill.get('_doc_type') or 'sales_bill_domestic'
                # Map country code to MK supported locale
                cc_map = {
                    'SI': ('sl', 'si'),
                    'EN': ('en', 'gb'),
                    'GB': ('en', 'gb'),
                    'US': ('en', 'us'),
                    'DE': ('de', 'de'),
                    'AT': ('de', 'at'),
                    'IT': ('it', 'it'),
                    'HR': ('hr', 'hr'),
                    'SR': ('sr_RS', 'rs'),
                    'RS': ('sr_RS', 'rs'),
                    'BA': ('sr_ba', 'ba'),
                    'RO': ('ro_RO', 'ro'),
                    'PT': ('pt_PT', 'pt'),
                    'ES': ('es_ES', 'es'),
                    'CZ': ('cz_CZ', 'cz'),
                    'SK': ('sk_SK', 'sk'),
                    'HU': ('hu_HU', 'hu'),
                    'PL': ('pl_PL', 'pl'),
                    'MK': ('mk_MK', 'mk'),
                    'NL': ('nl', 'nl'),
                    'GR': ('gr', 'gr'),
                    'FR': ('fr', 'fr'),
                    'BG': ('bg', 'bg'),
                }
                try:
                    locale_pair = cc_map.get((country_code or 'SI').upper(), ('sl', 'si'))
                except Exception:
                    locale_pair = ('sl', 'si')
                official_pdf = mk_print_bill_pdf(found_type, bill.get('mk_id'), locale=locale_pair[0], country=locale_pair[1])
                if not official_pdf:
                    current_app.logger.error(f"MK send-invoice: official MK PDF not found for order {order_number} (mk_id={bill.get('mk_id')}). Email will NOT be sent.")
                    _audit('no_pdf', recipient=customer_email, mk_id=bill.get('mk_id'), mk_type=bill.get('_doc_type'))
                    return
                import tempfile
                with tempfile.NamedTemporaryFile(suffix=f"_{order_number}.pdf", delete=False) as tmp:
                    tmp.write(official_pdf)
                    pdf_path = tmp.name
            except Exception as e:
                current_app.logger.error(f"MK send-invoice: official PDF generation failed for order {order_number}: {e}")
                pdf_path = None
            if not pdf_path:
                _audit('no_pdf', recipient=customer_email, mk_id=bill.get('mk_id'), mk_type=bill.get('_doc_type'))
                return

            try:
                import json as _json
                try:
                    items = _json.loads(line_items) if isinstance(line_items, str) else (line_items or [])
                except Exception:
                    items = []
                # Pošlji RAČUN (uradni MK PDF je obvezen)
                from services.email_service import send_invoice_email
                ok = send_invoice_email(
                    customer_email,
                    f"#{order_number}",
                    pdf_path,
                    country_code=country_code,
                    status_url=status_url,
                    store_url=shop_url,
                    items=items,
                    skip_test_redirect=True,
                )
                _audit('sent' if ok else 'send_failed', recipient=customer_email, mk_id=bill.get('mk_id'), mk_type=bill.get('_doc_type'))
            finally:
                try:
                    import os as _os
                    if pdf_path and _os.path.exists(pdf_path):
                        _os.remove(pdf_path)
                except Exception:
                    pass

            try:
                mk_id = bill.get('mk_id')
                if mk_id:
                    # Stolpci (mk_bill_id/type/publish_ts) so del sheme — brez ALTER
                    # TABLE na poti zahteve (ACCESS EXCLUSIVE lock → lock pileup).
                    c.execute(
                        "UPDATE orders SET mk_bill_id = %s, mk_bill_type = %s, mk_publish_ts = COALESCE(%s, mk_publish_ts) WHERE order_number = %s OR order_number = %s",
                        (str(mk_id), bill.get('_doc_type') or 'sales_bill_domestic', bill.get('publish_ts'), f"#{order_number}", order_number)
                    )
                    db.commit()
            except Exception as e:
                current_app.logger.warning(f"MK persist bill id failed: {e}")
        except Exception as e:
            current_app.logger.error(f"MK send-invoice BG error for order {order_number}: {e}\n{traceback.format_exc()}")
            try:
                _audit('error', recipient=customer_email if 'customer_email' in locals() else None, mk_id=bill.get('mk_id') if 'bill' in locals() and bill else None, mk_type=(bill.get('_doc_type') if 'bill' in locals() and bill else None), error=str(e))
            except Exception:
                pass
        finally:
            try:
                c.close()
            except Exception:
                pass


# --- Metakocka: pošlji račun ---
@api_bp.route('/mk/send-invoice', methods=['POST'])
def mk_send_invoice():
    try:
        data = request.get_json(silent=True) or {}
        order_number = str(data.get('order_number', '')).lstrip('#').strip()
        recipient_email = (data.get('recipient_email') or '').strip() or None
        if not order_number:
            return jsonify({'success': False, 'error': 'Manjka order_number'}), 400

        if not has_permission('send_invoice'):
            return jsonify({'success': False, 'error': 'Nimate dovoljenja'}), 403

        # Preveri, da je naročilo fulfilled (tudi partial)
        try:
            db = get_db()
            c = db.cursor()
            c.execute(
                """
                SELECT fulfilled_at, shopify_fulfilled_at, customer_name, shopify_order_id
                FROM orders
                WHERE order_number = %s OR order_number = %s
                """,
                (order_number, f"#{order_number}")
            )
            row = c.fetchone()
            c.close()
            if not row:
                return jsonify({'success': False, 'error': 'Naročilo ne obstaja'}), 404
            fulfilled_at = row['fulfilled_at'] if isinstance(row, dict) else row[0]
            shopify_fulfilled_at = row['shopify_fulfilled_at'] if isinstance(row, dict) else row[1]
            customer_name = row['customer_name'] if isinstance(row, dict) else row[2]
            shopify_order_id = row['shopify_order_id'] if isinstance(row, dict) else row[3]
            if not fulfilled_at and not shopify_fulfilled_at:
                return jsonify({'success': False, 'error': 'Naročilo ni fulfilled. Računa ni mogoče poslati.'}), 400
        except Exception as e:
            current_app.logger.warning(f"MK send-invoice precheck failed: {e}")

        # Hitra preverba (ne blokira več odgovora; le zalogira opozorilo)
        try:
            from services.mk_service import mk_find_bill_quick, mk_is_published, mk_find_bill_in_db, mk_find_bill_any, mk_get_document, mk_get_document_bill, mk_doc_matches_customer
            quick_db = mk_find_bill_in_db(order_number)
            quick_bill = None
            if quick_db:
                # če imamo v bazi, uporabi neposredno get_document
                from services.mk_service import mk_get_document
                quick_bill = mk_get_document(quick_db['doc_type'], quick_db['mk_id'])
            if not quick_bill:
                quick_bill = mk_find_bill_any(order_number)
            if not quick_bill and shopify_order_id:
                quick_bill = mk_find_bill_any(str(shopify_order_id))
            if not quick_bill:
                return jsonify({'success': False, 'error': 'Račun v MK ni najden za to naročilo.'}), 404
            try:
                mk_id = quick_bill.get('mk_id')
                doc_type = quick_bill.get('_doc_type') or quick_bill.get('doc_type')
                detailed = None
                if mk_id and doc_type:
                    detailed = mk_get_document(doc_type, mk_id)
                if not detailed and mk_id:
                    detailed = mk_get_document_bill(mk_id)
                if detailed:
                    detailed['_doc_type'] = doc_type or detailed.get('_doc_type')
                    quick_bill = detailed
            except Exception:
                pass
            if customer_name and not mk_doc_matches_customer(quick_bill, customer_name):
                return jsonify({'success': False, 'error': 'Račun v MK se ne ujema z imenom kupca.'}), 400
            if not mk_is_published(quick_bill):
                return jsonify({'success': False, 'error': 'Račun v MK še ni objavljen.'}), 400
        except Exception:
            current_app.logger.warning("MK quick check failed, proceeding in background")

        app_obj = current_app._get_current_object()
        threading.Thread(
            target=_mk_send_invoice_in_background,
            args=(app_obj, order_number, recipient_email),
            daemon=True
        ).start()
        return jsonify({'success': True, 'sporocilo': 'Pošiljanje računa sproženo. Prejeli boste e‑pošto v nekaj trenutkih.'})
    except Exception as e:
        current_app.logger.error(f"mk_send_invoice error: {e}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': 'Napaka pri pošiljanju računa.'}), 500


def _mk_secret_ok(provided) -> bool:
    """Fail-closed primerjava MK_WEBHOOK_SECRET (timing-safe)."""
    import hmac as hmac_lib
    secret_cfg = (
        current_app.config.get('MK_WEBHOOK_SECRET') or os.environ.get('MK_WEBHOOK_SECRET') or ''
    ).strip()
    if not secret_cfg or provided is None:
        return False
    try:
        return hmac_lib.compare_digest(str(provided), str(secret_cfg))
    except (TypeError, ValueError):
        return False


# Ročni sprožilec: MK sync bills (admin‑only via permission)
@api_bp.route('/mk/sync-bills', methods=['POST'])
def mk_sync_bills_endpoint():
    # allow secret override (same as webhook secret)
    body_try = request.get_json(silent=True) or {}
    provided = request.headers.get('X-MK-Secret') or request.args.get('secret') or body_try.get('secret')
    secret_ok = _mk_secret_ok(provided)
    perm = required_permission_for('POST', '/api/procurement/vendors/import')  # reuse admin-like perm
    if not secret_ok:
        if perm and not has_permission(perm):
            return jsonify({'success': False, 'error': 'Dostop zavrnjen'}), 403
    try:
        body_json = (request.get_json(silent=True) or {})
        days = int(body_json.get('days', 1))
        from services.mk_service import mk_sync_bills, mk_sync_bills_from_orders, import_retail_window
        app_obj = current_app._get_current_object()

        def _run_sync(app_obj_ref, days_ref: int, types_filter: list[str] | None):
            try:
                with app_obj_ref.app_context():
                    # reset cancel flag & progress
                    current_app.config['MK_SYNC_CANCEL'] = False
                    current_app.config['MK_SYNC_PROGRESS'] = {'phase': 'starting'}
                    # Strict retail path when only sales_bill_retail is selected
                    if types_filter and set([t.strip() for t in types_filter]) == {'sales_bill_retail'}:
                        # compute window
                        to_date = datetime.utcnow().date()
                        from_date = (to_date - timedelta(days=days_ref))
                        res = import_retail_window(from_date.strftime('%Y-%m-%d'), to_date.strftime('%Y-%m-%d'))
                        imported_total = int(res.get('upserted', 0))
                        current_app.config['MK_SYNC_LAST_RESULT'] = {
                            'ts': datetime.utcnow().isoformat(),
                            'days': days_ref,
                            'imported': imported_total,
                            'details': {
                                'retail_strict': res,
                                'filtered_types': types_filter or []
                            },
                            'status': 'done'
                        }
                        current_app.config['MK_SYNC_PROGRESS'] = {'phase': 'finished', 'imported_total': imported_total}
                    else:
                        imported1 = mk_sync_bills(days=days_ref, doc_types=types_filter, seed_mk_ids=body_json.get('seed_mk_ids'))
                        # Če je uporabnik izbral točno določen doc_type, preskoči uvoz iz naročil (ta lahko vnese druge tipe)
                        imported2 = 0
                        if not types_filter:
                            imported2 = mk_sync_bills_from_orders(days=days_ref)
                        # Pull per-type counts if present
                        progress = current_app.config.get('MK_SYNC_PROGRESS') or {}
                        counts_by_type = progress.get('counts_by_type') or {}
                        current_app.config['MK_SYNC_LAST_RESULT'] = {
                            'ts': datetime.utcnow().isoformat(),
                            'days': days_ref,
                            'imported': imported1 + imported2,
                            'details': {
                                'search_tail': imported1,
                                'from_orders': imported2,
                                'filtered_types': types_filter or [],
                                'by_type': counts_by_type
                            },
                            'status': 'done'
                        }
                        current_app.config['MK_SYNC_PROGRESS'] = {'phase': 'finished', 'imported_total': imported1 + imported2}
            except Exception as e:
                try:
                    current_app.config['MK_SYNC_LAST_RESULT'] = {
                        'ts': datetime.utcnow().isoformat(),
                        'days': days_ref,
                        'error': str(e),
                        'status': 'error'
                    }
                    current_app.config['MK_SYNC_PROGRESS'] = {'phase': 'error', 'error': str(e)}
                except Exception:
                    pass

        # optional doc_types from body (e.g., ["sales_bill_retail"]) 
        types_filter = None
        try:
            body = request.get_json(silent=True) or {}
            t = body.get('doc_types')
            if isinstance(t, list):
                types_filter = [str(x).strip() for x in t if str(x).strip()]
        except Exception:
            types_filter = None

        # init status
        current_app.config['MK_SYNC_LAST_RESULT'] = {'ts': datetime.utcnow().isoformat(), 'days': days, 'status': 'running'}
        current_app.config['MK_SYNC_CANCEL'] = False
        current_app.config['MK_SYNC_PROGRESS'] = {'phase': 'queued'}
        threading.Thread(target=_run_sync, args=(app_obj, days, types_filter), daemon=True).start()
        return jsonify({'success': True, 'started': True, 'message': 'Sinhronizacija zagnana v ozadju. Preverite status prek /api/mk/sync-bills/status.'})
    except Exception as e:
        current_app.logger.error(f"mk_sync_bills_endpoint error: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri sinhronizaciji'}), 500

@api_bp.route('/mk/sync-bills/secret', methods=['POST'])
def mk_sync_bills_secret():
    try:
        body = request.get_json(silent=True) or {}
        provided = request.headers.get('X-MK-Secret') or request.args.get('secret') or body.get('secret')
        if not _mk_secret_ok(provided):
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        days = int(body.get('days', 7))
        types = body.get('doc_types') if isinstance(body.get('doc_types'), list) else ['sales_bill_retail']
        seeds = body.get('seed_mk_ids') if isinstance(body.get('seed_mk_ids'), list) else []
        from services.mk_service import mk_sync_bills
        # Run in background to avoid router timeouts; expose status via secret-status
        app_obj = current_app._get_current_object()
        def _bg(app_obj_ref, d, t, s):
            with app_obj_ref.app_context():
                try:
                    current_app.config['MK_SYNC_CANCEL'] = False
                    current_app.config['MK_SYNC_PROGRESS'] = {'phase': 'starting'}
                    imported = mk_sync_bills(days=d, doc_types=t, seed_mk_ids=s)
                    progress = current_app.config.get('MK_SYNC_PROGRESS') or {}
                    counts_by_type = progress.get('counts_by_type') or {}
                    current_app.config['MK_SYNC_LAST_RESULT'] = {
                        'ts': datetime.utcnow().isoformat(),
                        'days': d,
                        'imported': imported,
                        'details': {'by_type': counts_by_type, 'filtered_types': t, 'seeds': s},
                        'status': 'done'
                    }
                    current_app.config['MK_SYNC_PROGRESS'] = {'phase': 'finished', 'imported_total': imported}
                except Exception as e:
                    current_app.config['MK_SYNC_LAST_RESULT'] = {'status': 'error', 'error': str(e)}
                    current_app.config['MK_SYNC_PROGRESS'] = {'phase': 'error', 'error': str(e)}
        threading.Thread(target=_bg, args=(app_obj, days, types, seeds), daemon=True).start()
        return jsonify({'success': True, 'started': True, 'types': types, 'seeds': seeds})
    except Exception as e:
        current_app.logger.error(f"mk_sync_bills_secret fatal: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri sinhronizaciji'}), 500

@api_bp.route('/mk/sync-bills/status/secret', methods=['GET'])
def mk_sync_bills_status_secret():
    try:
        provided = request.headers.get('X-MK-Secret') or request.args.get('secret')
        if not _mk_secret_ok(provided):
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        res = current_app.config.get('MK_SYNC_LAST_RESULT') or {'status': 'idle'}
        progress = current_app.config.get('MK_SYNC_PROGRESS') or {}
        cancel_flag = bool(current_app.config.get('MK_SYNC_CANCEL') or False)
        return jsonify({'success': True, 'status': res, 'progress': progress, 'cancelled': cancel_flag})
    except Exception as e:
        current_app.logger.error(f"mk_sync_bills_status_secret error: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri statusu'}), 500


@api_bp.route('/mk/retail/delta/secret', methods=['POST'])
def mk_retail_delta_secret():
    try:
        body = request.get_json(silent=True) or {}
        provided = request.headers.get('X-MK-Secret') or request.args.get('secret') or body.get('secret')
        if not _mk_secret_ok(provided):
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        try:
            hours = int(body.get('hours', 24))
        except Exception:
            hours = 24
        try:
            scan_window = int(body.get('scan_window', 5000))
        except Exception:
            scan_window = 5000

        from services.mk_service import mk_import_retail_bills_delta
        app_obj = current_app._get_current_object()

        def _bg(app_obj_ref, h, w):
            with app_obj_ref.app_context():
                try:
                    imported = mk_import_retail_bills_delta(hours=h, scan_window=w)
                    current_app.logger.info(f"Retail delta import done: imported={imported} (hours={h}, window={w})")
                    # expose a simple last result for diagnostics
                    current_app.config['MK_RETAIL_DELTA_LAST_RESULT'] = {'status': 'done', 'imported': imported, 'hours': h, 'scan_window': w}
                except Exception as e:
                    current_app.logger.error(f"mk_retail_delta_secret bg error: {e}")
                    current_app.config['MK_RETAIL_DELTA_LAST_RESULT'] = {'status': 'error', 'error': str(e)}

        threading.Thread(target=_bg, args=(app_obj, hours, scan_window), daemon=True).start()
        return jsonify({'success': True, 'started': True, 'hours': hours, 'scan_window': scan_window})
    except Exception as e:
        current_app.logger.error(f"mk_retail_delta_secret fatal: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri retail delta uvozu'}), 500


@api_bp.route('/mk/retail/change-delta/secret', methods=['POST'])
def mk_retail_change_delta_secret():
    try:
        body = request.get_json(silent=True) or {}
        provided = request.headers.get('X-MK-Secret') or request.args.get('secret') or body.get('secret')
        if not _mk_secret_ok(provided):
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        try:
            days_back = int(body.get('days_back', 3))
        except Exception:
            days_back = 3
        from services.mk_service import sync_retail_bills_delta
        app_obj = current_app._get_current_object()
        def _bg(app_obj_ref, d):
            with app_obj_ref.app_context():
                try:
                    res = sync_retail_bills_delta(days_back=d)
                    current_app.config['MK_RETAIL_CHANGE_DELTA_LAST_RESULT'] = {'status': 'done', **(res or {})}
                except Exception as e:
                    current_app.logger.error(f"mk_retail_change_delta_secret bg error: {e}")
                    current_app.config['MK_RETAIL_CHANGE_DELTA_LAST_RESULT'] = {'status': 'error', 'error': str(e)}
        threading.Thread(target=_bg, args=(app_obj, days_back), daemon=True).start()
        return jsonify({'success': True, 'started': True, 'days_back': days_back})
    except Exception as e:
        current_app.logger.error(f"mk_retail_change_delta_secret fatal: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri change-delta uvozu'}), 500


@api_bp.route('/mk/retail/import-by-ids/secret', methods=['POST'])
def mk_retail_import_by_ids_secret():
    try:
        body = request.get_json(silent=True) or {}
        provided = request.headers.get('X-MK-Secret') or request.args.get('secret') or body.get('secret')
        if not _mk_secret_ok(provided):
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        mk_ids = body.get('mk_ids') if isinstance(body.get('mk_ids'), list) else []
        from services.mk_service import mk_import_retail_bills_by_ids
        app_obj = current_app._get_current_object()

        def _bg(app_obj_ref, ids_list):
            with app_obj_ref.app_context():
                try:
                    imported = mk_import_retail_bills_by_ids(ids_list)
                    current_app.logger.info(f"Retail import-by-ids done: imported={imported} ids={ids_list}")
                    current_app.config['MK_RETAIL_IDS_LAST_RESULT'] = {'status': 'done', 'imported': imported, 'ids': ids_list}
                except Exception as e:
                    current_app.logger.error(f"mk_retail_import_by_ids_secret bg error: {e}")
                    current_app.config['MK_RETAIL_IDS_LAST_RESULT'] = {'status': 'error', 'error': str(e)}

        threading.Thread(target=_bg, args=(app_obj, mk_ids), daemon=True).start()
        return jsonify({'success': True, 'started': True, 'ids': mk_ids})
    except Exception as e:
        current_app.logger.error(f"mk_retail_import_by_ids_secret fatal: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri importu po ID-jih'}), 500

@api_bp.route('/mk/retail/search/secret', methods=['POST'])
def mk_retail_search_secret():
    try:
        body = request.get_json(silent=True) or {}
        provided = request.headers.get('X-MK-Secret') or request.args.get('secret') or body.get('secret')
        if not _mk_secret_ok(provided):
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        hours = int(body.get('hours', 24))
        max_pages = int(body.get('max_pages', 50))
        page_size = int(body.get('page_size', 100))
        ignore_last_ts = bool(body.get('ignore_last_ts', False))

        from services.mk_service import mk_import_retail_bills_search
        app_obj = current_app._get_current_object()

        def _bg(app_obj_ref, h, mp, ps, ig):
            with app_obj_ref.app_context():
                try:
                    imported = mk_import_retail_bills_search(hours=h, max_pages=mp, page_size=ps, ignore_last_ts=ig)
                    current_app.logger.info(f"Retail search import done: imported={imported} (hours={h}, pages={mp}, size={ps}, ignore_last_ts={ig})")
                    current_app.config['MK_RETAIL_SEARCH_LAST_RESULT'] = {'status': 'done', 'imported': imported, 'hours': h, 'ignore_last_ts': ig}
                except Exception as e:
                    current_app.logger.error(f"mk_retail_search_secret bg error: {e}")
                    current_app.config['MK_RETAIL_SEARCH_LAST_RESULT'] = {'status': 'error', 'error': str(e)}

        threading.Thread(target=_bg, args=(app_obj, hours, max_pages, page_size, ignore_last_ts), daemon=True).start()
        return jsonify({'success': True, 'started': True, 'hours': hours, 'max_pages': max_pages, 'page_size': page_size, 'ignore_last_ts': ignore_last_ts})
    except Exception as e:
        current_app.logger.error(f"mk_retail_search_secret fatal: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri search uvozu retail računov'}), 500

@api_bp.route('/mk/retail/status/secret', methods=['GET'])
def mk_retail_status_secret():
    try:
        provided = request.headers.get('X-MK-Secret') or request.args.get('secret')
        if not _mk_secret_ok(provided):
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        res_delta = current_app.config.get('MK_RETAIL_DELTA_LAST_RESULT') or {'status': 'idle'}
        res_ids = current_app.config.get('MK_RETAIL_IDS_LAST_RESULT') or {'status': 'idle'}
        res_search = current_app.config.get('MK_RETAIL_SEARCH_LAST_RESULT') or {'status': 'idle'}
        res_change = current_app.config.get('MK_RETAIL_CHANGE_DELTA_LAST_RESULT') or {'status': 'idle'}
        return jsonify({'success': True, 'delta': res_delta, 'ids': res_ids, 'search': res_search, 'change': res_change})
    except Exception as e:
        current_app.logger.error(f"mk_retail_status_secret error: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri statusu retail importov'}), 500

@api_bp.route('/mk/retail/inspect/secret', methods=['GET'])
def mk_retail_inspect_secret():
    try:
        provided = request.headers.get('X-MK-Secret') or request.args.get('secret')
        if not _mk_secret_ok(provided):
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        limit = max(1, min(50, int(request.args.get('limit', '20'))))
        from database import get_db
        db = get_db(); c = db.cursor(row_factory=dict_row)
        c.execute(
            """
            SELECT mk_id, document_number, issue_date, publish_ts, currency_code, sum_eur
            FROM mk_bill
            ORDER BY COALESCE(publish_ts, updated_at) DESC
            LIMIT %s
            """,
            (limit,)
        )
        rows = c.fetchall() or []
        return jsonify({'success': True, 'rows': rows})
    except Exception as e:
        current_app.logger.error(f"mk_retail_inspect_secret error: {e}")
        return jsonify({'success': False}), 500

@api_bp.route('/mk/retail/inspect-items/secret', methods=['GET'])
def mk_retail_inspect_items_secret():
    try:
        provided = request.headers.get('X-MK-Secret') or request.args.get('secret')
        if not _mk_secret_ok(provided):
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        mk_id = (request.args.get('mk_id') or '').strip()
        if not mk_id:
            return jsonify({'success': False, 'error': 'Manjka mk_id'}), 400
        from database import get_db
        db = get_db(); c = db.cursor(row_factory=dict_row)
        c.execute(
            """
            SELECT mk_id, row_no, product_id, product_code, title, qty, unit_price, tax_rate, discount
            FROM mk_bill_item
            WHERE mk_id = %s
            ORDER BY row_no ASC
            """,
            (mk_id,)
        )
        rows = c.fetchall() or []
        return jsonify({'success': True, 'mk_id': mk_id, 'items': rows})
    except Exception as e:
        current_app.logger.error(f"mk_retail_inspect_items_secret error: {e}")
        return jsonify({'success': False}), 500

@api_bp.route('/mk/webhook/stock', methods=['POST'])
def mk_webhook_stock():
    try:
        provided = (
            request.headers.get('X-MK-Secret')
            or request.args.get('secret')
            or (request.get_json(silent=True) or {}).get('secret')
            or ''
        )
        # Fail-closed: brez nastavljenega secret-a ali ob neujemanju → 401
        if not _mk_secret_ok(provided):
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        body = request.get_json(silent=True) or {}
        # Log receive
        try:
            from services.mk_service import app_log
            app_log('webhook.mk_stock', 'info', 'Received MK stock webhook', {'payload': body})
        except Exception:
            pass
        mk_log_stock_event(body)
        # extract SKUs
        skus_set = set()
        if isinstance(body, dict):
            # Newer spec: events list
            if isinstance(body.get('events'), list):
                for ev in body['events']:
                    code = (str(ev.get('code') or ev.get('sku') or ev.get('product_code') or '')).strip()
                    if code:
                        skus_set.add(code)
            # Warehouse stock sync spec: stock_list
            if isinstance(body.get('stock_list'), list):
                for it in body['stock_list']:
                    code = (str(it.get('code') or it.get('count_code') or '')).strip()
                    if code:
                        skus_set.add(code)
            # Fallback: single code fields at root
            root_code = (str(body.get('code') or body.get('sku') or body.get('product_code') or '')).strip()
            if root_code:
                skus_set.add(root_code)
        skus = list(skus_set)

        def _bg(app_obj, sku_list):
            with app_obj.app_context():
                # Scan all sales bill types by SKU (days=3)
                try:
                    from services.mk_service import app_log, mk_apply_procurement_from_stock_list, mk_apply_procurement_from_mk_ids
                    # First, import relevant bills for SKU history window
                    imported = mk_import_by_skus(sku_list, days=7)
                    app_log('import.mk_sku', 'info', 'Imported by SKUs', {'skus': sku_list, 'imported_total': imported})
                    # Then, apply procurement increments by exact bill product_list using mk_id
                    stock_list = (body or {}).get('stock_list') or []
                    try:
                        by_bill = mk_apply_procurement_from_mk_ids(stock_list)
                        app_log('procurement.apply', 'info', 'Applied from mk_ids', {'bills_processed': by_bill})
                    except Exception as apperr:
                        current_app.logger.error(f"mk_webhook_stock mk_ids apply error: {apperr}")
                    # Fallback: sales-only negative deltas directly from stock_list (kept for resilience)
                    try:
                        updated = mk_apply_procurement_from_stock_list(stock_list)
                        app_log('procurement.apply', 'info', 'Applied from stock_list (delta)', {'updated_rows': updated})
                    except Exception as apperr2:
                        current_app.logger.error(f"mk_webhook_stock delta apply error: {apperr2}")
                except Exception as e:
                    current_app.logger.error(f"mk_webhook_stock bg error: {e}")

        if skus:
            app_obj = current_app._get_current_object()
            threading.Thread(target=_bg, args=(app_obj, skus), daemon=True).start()
        return jsonify({'success': True})
    except Exception as e:
        current_app.logger.error(f"mk_webhook_stock error: {e}")
        return jsonify({'success': False}), 500

@api_bp.route('/mk/sync-bills/status', methods=['GET'])
def mk_sync_bills_status():
    perm = required_permission_for('POST', '/api/procurement/vendors/import')  # admin-like
    if perm and not has_permission(perm):
        return jsonify({'success': False, 'error': 'Dostop zavrnjen'}), 403
    try:
        res = current_app.config.get('MK_SYNC_LAST_RESULT') or {'status': 'idle'}
        progress = current_app.config.get('MK_SYNC_PROGRESS') or {}
        cancel_flag = bool(current_app.config.get('MK_SYNC_CANCEL') or False)
        return jsonify({'success': True, 'status': res, 'progress': progress, 'cancelled': cancel_flag})
    except Exception as e:
        current_app.logger.error(f"mk_sync_bills_status error: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri branju statusa'}), 500


# --- Admin: pregled pošiljanja računov (audit log) ---
@api_bp.route('/mk/invoice-email-log', methods=['GET'])
def mk_invoice_email_log():
    try:
        if not has_permission('view_orders'):
            return jsonify({'success': False, 'error': 'Dostop zavrnjen'}), 403

        order_number = request.args.get('order_number', '').lstrip('#').strip()
        try:
            limit = max(1, min(200, int(request.args.get('limit', '50'))))
        except Exception:
            limit = 50

        # Unified source: app_logs (category=email.send_invoice). Fallback to legacy if empty.
        db = get_db(); c = db.cursor(row_factory=dict_row)
        where = ["category = 'email.send_invoice'"]
        params = []
        if order_number:
            like = f"%{order_number}%"
            where.append('(message ILIKE %s OR CAST(data AS TEXT) ILIKE %s)')
            params.extend([like, like])
        where_sql = 'WHERE ' + ' AND '.join(where)
        c.execute(
            f"""
            SELECT id, ts, category, level, message, data
            FROM app_logs
            {where_sql}
            ORDER BY id DESC
            LIMIT %s
            """,
            tuple(params + [limit])
        )
        rows = c.fetchall() or []
        if not rows:
            # legacy fallback
            cc = db.cursor()
            if order_number:
                cc.execute(
                    """
                    SELECT id, order_number, mk_id, mk_type, recipient, status, error, created_at
                    FROM invoice_email_log
                    WHERE order_number = %s OR order_number = %s
                    ORDER BY id DESC
                    LIMIT %s
                    """,
                    (order_number, f"#{order_number}", limit)
                )
            else:
                cc.execute(
                    """
                    SELECT id, order_number, mk_id, mk_type, recipient, status, error, created_at
                    FROM invoice_email_log
                    ORDER BY id DESC
                    LIMIT %s
                    """,
                    (limit,)
                )
            legacy = cc.fetchall() or []
            cc.close()
            results = []
            for r in legacy:
                if isinstance(r, dict):
                    rec = r
                else:
                    rec = {
                        'id': r[0],
                        'order_number': r[1],
                        'mk_id': r[2],
                        'mk_type': r[3],
                        'recipient': r[4],
                        'status': r[5],
                        'error': r[6],
                        'created_at': r[7].isoformat() if hasattr(r[7], 'isoformat') else str(r[7])
                    }
                results.append(rec)
            c.close(); return jsonify({'success': True, 'data': results, 'count': len(results)})

        # normalize from app_logs
        results = []
        for r in rows:
            d = r.get('data') or {}
            results.append({
                'id': r.get('id'),
                'order_number': d.get('order_number'),
                'mk_id': d.get('mk_id'),
                'mk_type': d.get('mk_type'),
                'recipient': d.get('recipient'),
                'status': d.get('status'),
                'error': d.get('error'),
                'created_at': r.get('ts')
            })
        c.close(); return jsonify({'success': True, 'data': results, 'count': len(results)})
    except Exception as e:
        current_app.logger.error(f"invoice_email_log error: {e}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': 'Napaka pri branju loga.'}), 500


# --- Generic app logs ---
@api_bp.route('/logs', methods=['GET'])
def app_logs_list():
    try:
        # view_orders is enough to read logs
        if not has_permission('view_orders'):
            return jsonify({'success': False, 'error': 'Dostop zavrnjen'}), 403
        category = (request.args.get('category') or '').strip()
        level = (request.args.get('level') or '').strip()
        q = (request.args.get('q') or '').strip()
        try:
            limit = max(1, min(500, int(request.args.get('limit', '100'))))
        except Exception:
            limit = 100
        db = get_db(); c = db.cursor(row_factory=dict_row)
        where = []
        params = []
        if category:
            where.append('category = %s')
            params.append(category)
        if level:
            where.append('level = %s')
            params.append(level)
        if q:
            like = f"%{q}%"
            where.append('(message ILIKE %s OR CAST(data AS TEXT) ILIKE %s)')
            params.extend([like, like])
        where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''
        c.execute(
            f"""
            SELECT id, ts, category, level, message, data
            FROM app_logs
            {where_sql}
            ORDER BY id DESC
            LIMIT %s
            """,
            tuple(params + [limit])
        )
        rows = [dict(r) for r in c.fetchall()]
        c.close()
        return jsonify({'success': True, 'rows': rows, 'count': len(rows)})
    except Exception as e:
        current_app.logger.error(f"app_logs_list error: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri branju logov'}), 500

@api_bp.route('/mk/sync-bills/cancel', methods=['POST'])
def mk_sync_bills_cancel():
    perm = required_permission_for('POST', '/api/procurement/vendors/import')  # admin-like
    if perm and not has_permission(perm):
        return jsonify({'success': False, 'error': 'Dostop zavrnjen'}), 403
    try:
        current_app.config['MK_SYNC_CANCEL'] = True
        return jsonify({'success': True, 'message': 'Preklicano. Trenutna iteracija se bo ustavila v kratkem.'})
    except Exception as e:
        current_app.logger.error(f"mk_sync_bills_cancel error: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri preklicu'}), 500


# Diagnostic: inspect MetaKocka search results for a given query/title
@api_bp.route('/mk/debug-search', methods=['POST'])
def mk_debug_search():
    perm = required_permission_for('POST', '/api/procurement/vendors/import')  # admin-like
    if perm and not has_permission(perm):
        return jsonify({'success': False, 'error': 'Dostop zavrnjen'}), 403
    try:
        body = request.get_json(silent=True) or {}
        title = str(body.get('q') or body.get('title') or body.get('order_number') or '').lstrip('#')
        per_mode_limit = int(body.get('limit', 50))
        max_scan = int(body.get('max_scan', 200))
        types = ['sales_bill_domestic','sales_bill_foreign','sales_bill_retail','sales_bill_prepaid']
        if title:
            from services.mk_service import mk_search_bills
            rows = mk_search_bills(types, title, per_mode_limit=per_mode_limit, max_scan=max_scan)
            summary = {}
            for r in rows:
                dt = r.get('doc_type')
                summary[dt] = summary.get(dt, 0) + 1
            return jsonify({'success': True, 'count': len(rows), 'by_type': summary, 'rows': rows[:100]})
        # If no title provided: return the first page for each type directly from MK /search for diagnostics
        base = current_app.config.get('MK_API_BASE') or os.environ.get('MK_API_BASE') or 'https://main.metakocka.si/rest/eshop/v1'
        base = base.rstrip('/')
        company_id = os.environ.get('MK_COMPANY_ID') or str(current_app.config.get('MK_COMPANY_ID', ''))
        secret = os.environ.get('MK_API_KEY') or os.environ.get('MK_SECRET_KEY') or current_app.config.get('MK_API_KEY') or current_app.config.get('MK_SECRET_KEY')
        url = f"{base}/search"
        out = []
        for dt in types:
            payload = {
                'company_id': str(company_id),
                'secret_key': str(secret),
                'doc_type': dt,
                'offset': 0,
                'limit': min(50, per_mode_limit)
            }
            # Do not force result_type for retail
            if dt not in ('sales_bill_retail','bill'):
                payload['result_type'] = 'doc'
            try:
                r = requests.post(url, json=payload, timeout=int(current_app.config.get('MK_TIMEOUT', 15)))
                if not r.ok:
                    out.append({'doc_type': dt, 'http': r.status_code, 'error': r.text[:200]})
                    continue
                data = r.json() if r.headers.get('Content-Type','').startswith('application/json') else {}
                rows = []
                if isinstance(data, list):
                    rows = data
                elif isinstance(data, dict):
                    rows = data.get('rows') or data.get('result') or data.get('documents') or []
                sample = rows[0] if rows else {}
                out.append({'doc_type': dt, 'count_page': len(rows), 'sample_keys': list(sample.keys())[:20] if isinstance(sample, dict) else None})
            except Exception as e:
                out.append({'doc_type': dt, 'error': str(e)})
        return jsonify({'success': True, 'diagnostic': out})
    except Exception as e:
        current_app.logger.error(f"mk_debug_search error: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri debug iskanju'}), 500


@api_bp.route('/mk/bills', methods=['GET'])
def mk_bills_list():
    try:
        q = (request.args.get('q') or '').strip()
        doc_type = (request.args.get('doc_type') or '').strip()
        published = (request.args.get('published') or '').strip().lower()  # 'true' | 'false' | ''
        limit = max(1, min(int(request.args.get('limit', 50)), 200))
        offset = max(0, int(request.args.get('offset', 0)))

        db = get_db(); c = db.cursor(row_factory=dict_row)
        where = []
        params = []
        if q:
            like = f"%{q}%"
            where.append("(mk_id ILIKE %s OR title ILIKE %s OR buyer_order ILIKE %s OR count_code ILIKE %s)")
            params.extend([like, like, like, like])
        if doc_type:
            where.append("doc_type = %s")
            params.append(doc_type)
        if published == 'true':
            where.append("publish_ts IS NOT NULL")
        elif published == 'false':
            where.append("publish_ts IS NULL")
        where_sql = ("WHERE " + " AND ".join(where)) if where else ""

        # total
        c.execute(f"SELECT COUNT(*) AS total FROM mk_bills {where_sql}", tuple(params))
        total_row = c.fetchone()
        try:
            total_count = int(total_row['total'])
        except Exception:
            # Fallback for non-dict row
            try:
                total_count = int(list(total_row.values())[0])
            except Exception:
                total_count = 0

        # rows
        c.execute(
            f"""
            SELECT mk_id, doc_type, title, buyer_order, count_code, publish_ts, furs_zoi, furs_eor, total, created_ts, updated_at
            FROM mk_bills
            {where_sql}
            ORDER BY COALESCE(publish_ts, created_ts) DESC NULLS LAST
            LIMIT %s OFFSET %s
            """,
            tuple(params + [limit, offset])
        )
        rows = [dict(r) for r in c.fetchall()]
        c.close()
        return jsonify({'success': True, 'total': total_count, 'rows': rows, 'limit': limit, 'offset': offset})
    except Exception as e:
        current_app.logger.error(f"mk_bills_list error: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri branju mk_bills'}), 500


@api_bp.route('/mk/bills/<mk_id>', methods=['GET'])
def mk_bills_get(mk_id: str):
    try:
        db = get_db(); c = db.cursor(row_factory=dict_row)
        c.execute(
            """
            SELECT mk_id, doc_type, title, buyer_order, count_code, publish_ts, furs_zoi, furs_eor, total, created_ts, updated_at
            FROM mk_bills WHERE mk_id = %s
            """,
            (str(mk_id),)
        )
        row = c.fetchone()
        c.close()
        if not row:
            return jsonify({'success': False, 'error': 'Ni najdeno'}), 404
        data = dict(row)
        return jsonify({'success': True, 'row': data})
    except Exception as e:
        current_app.logger.error(f"mk_bills_get error: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri branju zapisa'}), 500
def check_serija_permissions(serija_id, action='read'):
    """
    Preveri, ali ima trenutni uporabnik dovoljenja za operacijo na seriji.
    
    Args:
        serija_id: ID serije
        action: 'read', 'update', 'delete'
    
    Returns:
        dict: {'allowed': bool, 'reason': str, 'serija': dict}
    """
    try:
        # Preveri, ali je uporabnik prijavljen
        if 'user_id' not in session:
            return {'allowed': False, 'reason': 'Uporabnik ni prijavljen', 'serija': None}
        
        current_user_id = session['user_id']

        # Admin samo po vlogi (ne po username — to je bil bypass)
        role = str(
            session.get('user', {}).get('role') or session.get('role') or ''
        ).strip().lower()
        is_admin = role == 'admin'
        
        db = get_db()
        cursor = db.cursor()
        
        # Pridobi podatke o seriji
        cursor.execute("""
            SELECT s.id, s.vnesel_uporabnik, u.username, u.id as user_id
            FROM serije s
            LEFT JOIN users u ON s.vnesel_uporabnik = CONCAT(u.first_name, ' ', u.last_name)
            WHERE s.id = %s
        """, (serija_id,))
        
        serija_data = cursor.fetchone()
        
        if not serija_data:
            return {'allowed': False, 'reason': 'Serija ne obstaja', 'serija': None}
        
        # Admin lahko vse
        if is_admin:
            return {'allowed': True, 'reason': 'Admin dovoljenja', 'serija': serija_data}

        # Če ima uporabnik sistemsko dovoljenje za urejanje serij, dovoli
        if action in ('update', 'delete') and has_permission('edit_serije'):
            return {'allowed': True, 'reason': 'Dovoljenje edit_serije', 'serija': serija_data}
        
        # Preveri, ali je serija vnesla trenutni uporabnik
        if serija_data['user_id'] == current_user_id:
            return {'allowed': True, 'reason': 'Lastnik serije', 'serija': serija_data}
        
        # Preveri, ali se vnesel_uporabnik ujema z trenutnim uporabnikom
        current_user = session.get('user', {})
        current_user_full_name = f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}".strip()
        
        if serija_data['vnesel_uporabnik'] == current_user_full_name:
            return {'allowed': True, 'reason': 'Lastnik serije (po imenu)', 'serija': serija_data}
        
        # Preveri, ali se username ujema
        if serija_data['vnesel_uporabnik'] == current_username:
            return {'allowed': True, 'reason': 'Lastnik serije (po username)', 'serija': serija_data}
        
        return {'allowed': False, 'reason': 'Nimate dovoljenj za to operacijo', 'serija': serija_data}
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri preverjanju dovoljenj: {e}")
        return {'allowed': False, 'reason': 'Napaka pri preverjanju dovoljenj', 'serija': None}
    finally:
        if 'cursor' in locals():
            cursor.close()
class DictAsObject:
    """Pomožni razred, ki omogoča dostop do ključev slovarja kot atributov (npr. obj.key namesto obj['key'])."""
    def __init__(self, data):
        if data:
            self.__dict__.update({str(k): v for k, v in data.items()})

    def __getattr__(self, name):
        # Vrne None, če atribut ne obstaja, da se preprečijo nadaljnje napake.
        return self.__dict__.get(name)

# --- Pomožne funkcije, specifične za API ---

def preveri_rok_uporabe(rok_uporabe_obj):
    if not rok_uporabe_obj or not isinstance(rok_uporabe_obj, (date, datetime)): return False, None
    rok_uporabe_datum = rok_uporabe_obj.date() if isinstance(rok_uporabe_obj, datetime) else rok_uporabe_obj
    danes = date.today()
    meja_opozorila = danes + timedelta(days=60)
    if rok_uporabe_datum < danes: return True, f"Rok uporabe ({rok_uporabe_datum.strftime('%d.%m.%Y')}) je že potekel."
    if rok_uporabe_datum < meja_opozorila: return True, f"OPOZORILO: Rok uporabe ({rok_uporabe_datum.strftime('%d.%m.%Y')}) poteče v manj kot 60 dneh."
    return False, None
def _pridobi_podatke_za_deklaracijo(items, db_cursor):
    declaration_items, missing_data_details, expiration_warnings = [], [], []
    for item in items:
        db_cursor.execute("SELECT p.*, pr.ime as ime_proizvajalca FROM parfumi p JOIN proizvajalci pr ON p.proizvajalec_id = pr.id WHERE p.product_no = %s AND pr.ime = %s", (item['product_no'], item['proizvajalec_ime']))
        parfum_data = db_cursor.fetchone()
        
        if not parfum_data or not parfum_data.get('sestava_inci'):
            missing_data_details.append(f"Manjka INCI za '{item['title']}'")
            continue
        
        db_cursor.execute("SELECT rok_uporabe, serijska_stevilka FROM serije WHERE parfum_id = %s AND rok_uporabe >= CURRENT_DATE ORDER BY id DESC LIMIT 1", (parfum_data['id'],))
        serija_data = db_cursor.fetchone()
        if not serija_data:
            missing_data_details.append(f"Ni razpoložljive serije (z veljavnim rokom) za '{item['title']}'")
            continue
            
        is_problem, problem_message = preveri_rok_uporabe(serija_data['rok_uporabe'])
        if is_problem: expiration_warnings.append(f"{item['title']}: {problem_message}")
        declaration_items.append({
            "product_no": parfum_data['product_no'],
            "proizvajalec_ime": parfum_data['ime_proizvajalca'],
            "sestava_inci": parfum_data.get('sestava_inci', ''),
            "rok_uporabe": serija_data['rok_uporabe'].strftime('%d.%m.%Y') if hasattr(serija_data['rok_uporabe'], 'strftime') and serija_data['rok_uporabe'] else (str(serija_data['rok_uporabe']) if serija_data['rok_uporabe'] else None),
            "serijska_stevilka": serija_data.get("serijska_stevilka", "N/A"), 
            "title": item['title'], 
            "product_id": parfum_data['product_no']
        })
    return declaration_items, missing_data_details, expiration_warnings
def _pridobi_podatke_za_rocno_deklaracijo(perfume_ids, db_cursor):
    """Pridobi podatke za ročno deklaracijo iz seznama ID parfumov."""
    if not perfume_ids:
        return [], [], []
    
    placeholders = ','.join(['%s'] * len(perfume_ids))
    db_cursor.execute(f"""
        SELECT p.id, p.product_no, pr.ime as proizvajalec_ime, p.sestava_inci, p.ime_parfuma, p.na_zalogi
        FROM parfumi p
        JOIN proizvajalci pr ON p.proizvajalec_id = pr.id
        WHERE p.id IN ({placeholders})
    """, perfume_ids)
    
    perfumes = db_cursor.fetchall()
    
    # Pretvori podatke v format, ki ga pričakuje PDF funkcija
    formatted_perfumes = []
    missing_data_details = []
    expiration_warnings = []
    
    for perfume in perfumes:
        # Preveri, ali parfum ima INCI sestavo
        if not perfume.get('sestava_inci'):
            missing_data_details.append(
                f"Manjka INCI za parfum '{perfume['ime_parfuma']}' (product_no: {perfume['product_no']})"
            )
            continue

        # Pridobi rok uporabe iz serije (isti pristop kot za običajna naročila)
        db_cursor.execute(
            "SELECT rok_uporabe, serijska_stevilka FROM serije WHERE parfum_id = %s AND rok_uporabe >= CURRENT_DATE ORDER BY id DESC LIMIT 1",
            (perfume['id'],)
        )
        serija_data = db_cursor.fetchone()

        if not serija_data:
            missing_data_details.append(
                f"Ni razpoložljive serije (z veljavnim rokom) za parfum '{perfume['ime_parfuma']}'"
            )
            continue

        rok_uporabe = (
            serija_data['rok_uporabe'].strftime('%d.%m.%Y')
            if hasattr(serija_data['rok_uporabe'], 'strftime') and serija_data['rok_uporabe']
            else (str(serija_data['rok_uporabe']) if serija_data['rok_uporabe'] else None)
        )
        serijska_stevilka = serija_data.get("serijska_stevilka", "N/A")

        # Preveri rok uporabe
        is_problem, problem_message = preveri_rok_uporabe(serija_data['rok_uporabe'])
        if is_problem:
            expiration_warnings.append(f"{perfume['ime_parfuma']}: {problem_message}")

        formatted_perfume = {
            'id_parfuma': perfume['product_no'],
            'title': f"Parfum \"{perfume['product_no']}\" navdihnjen po \"{perfume['ime_parfuma']}\"",
            'product_no': perfume['product_no'],
            'proizvajalec_ime': perfume['proizvajalec_ime'],
            'proizvajalec': perfume['proizvajalec_ime'].lower(),  # Dodano za template
            'sestava_inci': perfume['sestava_inci'],
            'rok_uporabe': rok_uporabe,
            'serijska_stevilka': serijska_stevilka
        }
        formatted_perfumes.append(formatted_perfume)
    
    return formatted_perfumes, missing_data_details, expiration_warnings

def _shrani_deklaracijo_v_bazo(order_number, declaration_items, db_cursor):
    """
    Shrani podatke deklaracije v tabelo declarations.
    Vrne True, če je bilo uspešno shranjeno, False sicer.
    """
    try:
        current_app.logger.info(f"Začenjam shranjevanje v declarations za naročilo {order_number} z {len(declaration_items)} izdelki")
        
        # Najprej poiščemo order_id
        db_cursor.execute(
            """
            SELECT id FROM orders
            WHERE order_number = %s OR order_number = %s
            """,
            (order_number, f"#{str(order_number).replace('#','')}")
        )
        order_result = db_cursor.fetchone()
        if not order_result:
            current_app.logger.error(f"Order {order_number} ni najden v bazi")
            return False
        
        order_id = order_result['id']
        current_app.logger.info(f"Našel order_id: {order_id} za naročilo {order_number}")
        
        # Shranimo vsak izdelek v declarations tabelo
        for i, item in enumerate(declaration_items):
            current_app.logger.info(f"Shranjujem izdelek {i+1}/{len(declaration_items)}: {item.get('product_no')} - {item.get('proizvajalec_ime')}")
            
            # Preveri, ali deklaracija že obstaja (preveri vse polja, ki so v constraint)
            db_cursor.execute("""
                SELECT id FROM declarations
                WHERE (order_number = %s OR order_number = %s)
                  AND product_no = %s AND proizvajalec_ime = %s AND rok_uporabe = %s AND serijska_stevilka = %s
            """, (
                order_number,
                f"#{str(order_number).replace('#','')}",
                item.get('product_no', ''),
                item.get('proizvajalec_ime', ''),
                item.get('rok_uporabe'),
                item.get('serijska_stevilka', '')
            ))
            
            if db_cursor.fetchone():
                current_app.logger.info(f"Deklaracija za {item.get('product_no')} - {item.get('proizvajalec_ime')} z istimi podatki že obstaja, preskačem.")
                continue
            
            # Vstavi ali ignoriraj, če obstaja zapis za isto (order_number, product_no, proizvajalec_ime)
            db_cursor.execute("""
                INSERT INTO declarations (
                    order_number, product_no, proizvajalec_ime, sestava_inci,
                    rok_uporabe, serijska_stevilka, order_id, quantity
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (order_number, product_no, proizvajalec_ime) DO NOTHING
            """, (
                order_number,
                item.get('product_no', ''),
                item.get('proizvajalec_ime', ''),
                item.get('sestava_inci', ''),
                item.get('rok_uporabe'),
                item.get('serijska_stevilka', ''),
                order_id,
                item.get('quantity', 1)
            ))
            current_app.logger.info(f"Uspešno shranjen (ali že obstajal) izdelek {item.get('product_no')} - {item.get('proizvajalec_ime')}")

        current_app.logger.info(f"Uspešno shranjenih {len(declaration_items)} izdelkov v declarations za naročilo {order_number}")
        return True
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri shranjevanju v declarations: {e}")
        return False

def _pridobi_deklaracijo_iz_baze(order_number, db_cursor):
    """
    Pridobi podatke deklaracije iz tabele declarations.
    Vrne seznam podatkov ali None, če podatki ne obstajajo.
    """
    try:
        current_app.logger.info(f"Pridobivam deklaracijo iz baze za naročilo: '{order_number}'")
        
        # Preveri oba možna formata order_number (z in brez #)
        db_cursor.execute("""
            SELECT product_no, proizvajalec_ime, sestava_inci, rok_uporabe, serijska_stevilka, quantity
            FROM declarations 
            WHERE order_number = %s OR order_number = %s
            ORDER BY created_at
        """, (order_number, f"#{order_number}"))
        
        results = db_cursor.fetchall()
        current_app.logger.info(f"SQL poizvedba vrnila {len(results)} rezultatov za naročilo '{order_number}' (preveril tudi '#{order_number}')")
        
        if results:
            current_app.logger.info(f"Našel {len(results)} zapisov v declarations za naročilo {order_number}")
            return results
        else:
            current_app.logger.info(f"Ni najdenih zapisov v declarations za naročilo {order_number} (preveril tudi '#{order_number}')")
            return None
            
    except Exception as e:
        current_app.logger.error(f"Napaka pri pridobivanju iz declarations: {e}")
        return None
def _dodaj_podatke_iz_excel_database():
    """
    Doda podatke iz Excel Database sheet-a v declarations tabelo.
    Prenese datoteko iz OneDrive-a in jo obdela.
    """
    try:
        from openpyxl import load_workbook
        import os
        import requests
        import tempfile
        
        current_app.logger.info("Začenjam prenos Excel datoteke iz OneDrive-a...")
        
        # Pridobimo OneDrive podatke iz konfiguracije
        client_id = current_app.config.get('CLIENT_ID')
        client_secret = current_app.config.get('CLIENT_SECRET')
        tenant_id = current_app.config.get('TENANT_ID')
        excel_file_id = current_app.config.get('EXCEL_FILE_ID')
        drive_id = current_app.config.get('DRIVE_ID')
        
        current_app.logger.info(f"OneDrive konfiguracija: CLIENT_ID={'*'*10 if client_id else 'MANJKA'}, TENANT_ID={'*'*10 if tenant_id else 'MANJKA'}, EXCEL_FILE_ID={'*'*10 if excel_file_id else 'MANJKA'}, DRIVE_ID={'*'*10 if drive_id else 'MANJKA'}")
        
        if not all([client_id, client_secret, tenant_id, excel_file_id, drive_id]):
            missing = []
            if not client_id: missing.append('CLIENT_ID')
            if not client_secret: missing.append('CLIENT_SECRET')
            if not tenant_id: missing.append('TENANT_ID')
            if not excel_file_id: missing.append('EXCEL_FILE_ID')
            if not drive_id: missing.append('DRIVE_ID')
            current_app.logger.error(f"Manjkajo OneDrive konfiguracijski podatki: {missing}")
            return {"success": False, "message": f"Manjkajo OneDrive konfiguracijski podatki: {', '.join(missing)}"}
        
        # 1. Pridobimo access token
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        token_data = {
            'client_id': client_id,
            'client_secret': client_secret,
            'scope': 'https://graph.microsoft.com/.default',
            'grant_type': 'client_credentials'
        }
        
        current_app.logger.info("Pridobivam access token...")
        token_response = requests.post(token_url, data=token_data, timeout=10)
        if not token_response.ok:
            current_app.logger.error(f"Napaka pri pridobivanju tokena: {token_response.status_code} - {token_response.text}")
            return {"success": False, "message": f"Napaka pri pridobivanju OneDrive dostopa: {token_response.status_code}"}
        
        access_token = token_response.json()['access_token']
        current_app.logger.info("Access token uspešno pridobljen")
        
        # 2. Prenesemo Excel datoteko iz OneDrive-a
        download_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{excel_file_id}/content"
        headers = {'Authorization': f'Bearer {access_token}'}
        
        current_app.logger.info("Prenašam Excel datoteko iz OneDrive-a...")
        excel_response = requests.get(download_url, headers=headers, timeout=30)
        if not excel_response.ok:
            current_app.logger.error(f"Napaka pri prenosu datoteke: {excel_response.status_code} - {excel_response.text}")
            return {"success": False, "message": f"Napaka pri prenosu Excel datoteke: {excel_response.status_code}"}
        
        current_app.logger.info(f"Excel datoteka prenesena, velikost: {len(excel_response.content)} bajtov")
        
        # Shrani datoteko v tmp direktorij
        excel_path = '/tmp/DEKLARACIJE_PARFUMOV_KOPER.xlsm'
        print(f"🔍 DEBUG: Poskušam shraniti Excel datoteko v: {excel_path}")
        print(f"🔍 DEBUG: Velikost datoteke za shranjevanje: {len(excel_response.content)} bajtov")
        current_app.logger.info(f"Poskušam shraniti Excel datoteko v: {excel_path}")
        current_app.logger.info(f"Velikost datoteke za shranjevanje: {len(excel_response.content)} bajtov")
        
        try:
            with open(excel_path, 'wb') as f:
                f.write(excel_response.content)
            print(f"✅ DEBUG: Excel datoteka uspešno shranjena v: {excel_path}")
            current_app.logger.info(f"Excel datoteka uspešno shranjena v: {excel_path}")
            
            # Preverimo, ali datoteka res obstaja
            import os
            if os.path.exists(excel_path):
                file_size = os.path.getsize(excel_path)
                print(f"✅ DEBUG: Datoteka res obstaja, velikost: {file_size} bajtov")
                current_app.logger.info(f"Datoteka res obstaja, velikost: {file_size} bajtov")
            else:
                print(f"❌ DEBUG: Datoteka se ni shranila: {excel_path}")
                current_app.logger.error(f"Datoteka se ni shranila: {excel_path}")
                
        except Exception as e:
            print(f"❌ DEBUG: Napaka pri shranjevanju datoteke: {e}")
            current_app.logger.error(f"Napaka pri shranjevanju datoteke: {e}")
            return {"success": False, "message": f"Napaka pri shranjevanju datoteke: {str(e)}"}
        
        print("✅ DEBUG: Excel datoteka uspešno prenesena na strežnik")
        current_app.logger.info("Excel datoteka uspešno prenesena na strežnik")
        return {"success": True, "message": "Excel datoteka uspešno prenesena na strežnik"}
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri Excel importu: {e}")
        return {"success": False, "message": f"Napaka: {str(e)}"}

def _process_batch_insert(cursor, batch_data):
    """Izvede batch insert za hitrejše delovanje"""
    try:
        # Uporabimo UPSERT (INSERT ... ON CONFLICT DO NOTHING)
        cursor.executemany("""
            INSERT INTO declarations (
                order_number, product_no, proizvajalec_ime, sestava_inci,
                rok_uporabe, serijska_stevilka
            ) VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (order_number, product_no, proizvajalec_ime) DO NOTHING
        """, batch_data)
    except Exception as e:
        current_app.logger.error(f"Napaka pri batch insert: {e}")
        # Poskusimo individualno vstavljanje
        for data in batch_data:
            try:
                cursor.execute("""
                    INSERT INTO declarations (
                        order_number, product_no, proizvajalec_ime, sestava_inci,
                        rok_uporabe, serijska_stevilka
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (order_number, product_no, proizvajalec_ime) DO NOTHING
                """, data)
            except Exception as e2:
                current_app.logger.error(f"Napaka pri individual insert: {e2}")
                continue
def _sync_shopify_orders():
    """
    Pridobi zadnjih 250 naročil iz Shopifyja (status=any, vključno z archived).
    Doda morebitne nove v lokalno bazo in posodobi obstoječe (da doda ceno).
    """
    db = get_db()
    cursor = db.cursor()
    
    try:
        from services.shopify_service import get_all_shopify_stores
        stores = get_all_shopify_stores(include_default=True)
        api_version = os.environ.get('SHOPIFY_API_VERSION') or current_app.config.get('SHOPIFY_API_VERSION', '2025-01')

        new_orders_count = 0
        updated_orders_count = 0

        for store in stores:
            shop_domain = store.get('shop_domain')
            if not shop_domain or not store.get('access_token'):
                continue

            # Build existing IDs set per store
            if store.get('is_default'):
                cursor.execute(
                    "SELECT shopify_order_id FROM orders WHERE shopify_store_domain = %s OR shopify_store_domain IS NULL",
                    (shop_domain,),
                )
            else:
                cursor.execute(
                    "SELECT shopify_order_id FROM orders WHERE shopify_store_domain = %s",
                    (shop_domain,),
                )
            existing_ids = {str(row['shopify_order_id']) for row in cursor.fetchall()}

            url = f"https://{shop_domain}/admin/api/{api_version}/orders.json?status=any&limit=250&order=created_at+desc"
            headers = {"X-Shopify-Access-Token": store.get('access_token')}
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            shopify_orders = response.json().get('orders', [])

            for order_data in shopify_orders:
                order_id_str = str(order_data['id'])

                line_items_for_db = [{
                    'product_id': item.get('product_id'), 'variant_id': item.get('variant_id'),
                    'title': item.get('title'), 'quantity': item.get('quantity'),
                    'sku': item.get('sku'), 'vendor': item.get('vendor'),
                    'price': item.get('price'), 'image_url': item.get('image_url')
                } for item in order_data.get('line_items', [])]

                if order_id_str not in existing_ids:
                    customer_data = order_data.get('customer')
                    shipping_address = order_data.get('shipping_address')
                    if store.get('is_default'):
                        cursor.execute(
                            "UPDATE orders SET shopify_store_domain = %s WHERE shopify_order_id = %s AND shopify_store_domain IS NULL",
                            (shop_domain, order_data['id']),
                        )

                    sql = """
                        INSERT INTO orders (
                            shopify_order_id, order_number, customer_email, customer_name, 
                            created_at, fulfilled_at, line_items, country_code, status_url, shopify_store_domain
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (shopify_store_domain, shopify_order_id) DO NOTHING
                    """
                    params = (
                        order_data['id'], order_data['name'], order_data.get('email'),
                        f"{customer_data.get('first_name', '')} {customer_data.get('last_name', '')}".strip() if customer_data else 'N/A',
                        order_data['created_at'],
                        order_data.get('fulfillments')[0].get('created_at') if order_data.get('fulfillments') else None,
                        json.dumps(line_items_for_db),
                        shipping_address.get('country_code') if shipping_address else None,
                        order_data.get('order_status_url'),
                        shop_domain
                    )
                    cursor.execute(sql, params)
                    if cursor.rowcount > 0:
                        new_orders_count += 1
                        # Obvesti SSE kliente o novem naročilu
                        notify_new_order(order_data['name'])
                else:
                    sql = """
                        UPDATE orders 
                        SET line_items = %s
                        WHERE shopify_order_id = %s
                          AND (shopify_store_domain = %s OR shopify_store_domain IS NULL)
                    """
                    params = (json.dumps(line_items_for_db), order_id_str, shop_domain)
                    cursor.execute(sql, params)
                    if cursor.rowcount > 0:
                        updated_orders_count += 1

        db.commit()

        if new_orders_count > 0 and updated_orders_count > 0:
            message = f"Uspešno dodanih {new_orders_count} novih in posodobljenih {updated_orders_count} obstoječih naročil."
        elif new_orders_count > 0:
            message = f"Uspešno dodanih {new_orders_count} novih naročil."
        elif updated_orders_count > 0:
            message = f"Uspešno posodobljenih {updated_orders_count} obstoječih naročil."
        else:
            message = "Vsa naročila so že posodobljena. Ni bilo sprememb."

        return new_orders_count + updated_orders_count, message

    except requests.exceptions.RequestException as e:
        current_app.logger.error(f"Napaka pri pridobivanju naročil iz Shopify: {e}")
        return 0, f"Napaka pri komunikaciji s Shopify: {e}"
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri vstavljanju/posodabljanju naročil v bazo: {e}")
        traceback.print_exc()
        return 0, f"Napaka pri zapisovanju naročil v bazo: {e}"
    finally:
        cursor.close()

# --- API poti ---

@api_bp.route('/sync-new-orders', methods=['POST'])
def sync_new_orders_endpoint():
    try:
        count, message = _sync_shopify_orders()
        return jsonify({"message": message, "new_orders_count": count})
    except Exception as e:
        current_app.logger.error(f"Kritična napaka v /sync-new-orders: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do kritične napake pri sinhronizaciji naročil."}), 500

@api_bp.route('/sync-fulfilled-status', methods=['POST'])
def sync_fulfilled_status_endpoint():
    """Sinhronizira fulfilled status naročil iz Shopify-ja"""
    db = get_db()
    cursor = db.cursor()
    try:
        # Zagotovi, da je transakcija v pravilnem stanju
        db.rollback()
        current_app.logger.info("Začenjam sinhronizacijo fulfilled statusa naročil...")
        
        # Pridobi unfulfilled naročila iz baze — omejeno na zadnjih 30 dni.
        # Starejša naročila so pogosto izbrisana iz Shopify-ja (trajni 404) ali
        # opuščena; brez omejitve je endpoint delal 1000+ Shopify klicev na klik.
        cursor.execute(
            """
            SELECT shopify_order_id, order_number, shopify_store_domain
            FROM orders
            WHERE fulfilled_at IS NULL
              AND shopify_order_id IS NOT NULL
              AND created_at > NOW() - INTERVAL '30 days'
            ORDER BY created_at DESC
            """
        )
        unfulfilled_orders = cursor.fetchall()
        
        if not unfulfilled_orders:
            return jsonify({"message": "Ni unfulfilled naročil za preverjanje.", "updated_count": 0})
        
        current_app.logger.info(f"Preverjam fulfilled status za {len(unfulfilled_orders)} naročil")
        
        updated_count = 0
        
        # Pridobi podatke iz Shopify-ja za vsa naročila
        from services.shopify_service import get_orders_fulfillment_status
        
        for order in unfulfilled_orders:
            try:
                shopify_order_id = order['shopify_order_id']
                order_number = order['order_number']
                shop_domain = order.get('shopify_store_domain')
                
                # Preveri fulfilled status v Shopify
                is_fulfilled = get_orders_fulfillment_status(shopify_order_id, shop_domain=shop_domain)
                
                if is_fulfilled:
                    # Pridobi fulfilled čas iz Shopify-ja
                    try:
                        from services.shopify_service import get_order_fulfillment_details
                        fulfillment_details = get_order_fulfillment_details(shopify_order_id, shop_domain=shop_domain)
                        shopify_fulfilled_at = fulfillment_details.get('created_at') if fulfillment_details else None
                    except:
                        shopify_fulfilled_at = None
                    
                    # Posodobi status v bazi
                    cursor.execute(
                        """
                        UPDATE orders
                        SET fulfilled_at = NOW(), shopify_fulfilled_at = %s
                        WHERE shopify_order_id = %s
                          AND (shopify_store_domain = %s OR shopify_store_domain IS NULL)
                        """,
                        (shopify_fulfilled_at, shopify_order_id, shop_domain)
                    )
                    updated_count += 1
                    current_app.logger.info(f"Naročilo {order_number} označeno kot fulfilled")
                    
                    # AVTOMATSKO PROCESIRANJE FULFILLED NAROČILA
                    try:
                        current_app.logger.info(f"Začenjam avtomatsko procesiranje fulfilled naročila {order_number}")
                        
                        # 1. Pridobi podatke naročila
                        cursor.execute("""
                            SELECT o.*, d.id as declaration_id 
                            FROM orders o 
                            LEFT JOIN declarations d ON o.order_number = d.order_number 
                            WHERE o.shopify_order_id = %s
                              AND (o.shopify_store_domain = %s OR o.shopify_store_domain IS NULL)
                        """, (shopify_order_id, shop_domain))
                        
                        order_data = cursor.fetchone()
                        
                        if not order_data:
                            current_app.logger.error(f"Naročilo {order_number} ni najdeno v bazi")
                            continue
                        
                        # Preveri, ali je že procesirano
                        if order_data['declaration_id']:
                            current_app.logger.info(f"Naročilo {order_number} je že procesirano (ima declaration)")
                            continue
                        
                        if order_data['pdf_generated_at']:
                            current_app.logger.info(f"Naročilo {order_number} ima že generiran PDF")
                            continue
                        
                        # 2. Pridobi podatke za deklaracijo
                        _shop_domain_for_decl = order_data.get('shopify_store_domain') if isinstance(order_data, dict) else None
                        line_items_raw = order_data['line_items'] or '[]'
                        line_items = json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])
                        if not line_items:
                            current_app.logger.error(f"Naročilo {order_number} nima line_items (shop={_shop_domain_for_decl})")
                            continue
                        
                        # 3. Pridobi podatke za deklaracijo
                        declaration_items, missing, warnings = _pridobi_podatke_za_deklaracijo_iz_shopify(
                            line_items, cursor, shop_domain=_shop_domain_for_decl,
                        )
                        
                        if warnings:
                            current_app.logger.warning(
                                f"Expiry block for {order_number}: {warnings}"
                            )
                            continue
                        if not declaration_items:
                            current_app.logger.error(f"Ni bilo mogoče pridobiti podatkov za deklaracijo za naročilo {order_number}. Manjkajo: {missing}")
                            continue
                        
                        # 4. Shrani v declarations tabelo
                        success = _shrani_deklaracijo_v_bazo(order_number, declaration_items, cursor)
                        
                        if not success:
                            current_app.logger.error(f"Napaka pri shranjevanju deklaracije za naročilo {order_number}")
                            continue
                        
                        # 5. Generiraj PDF
                        # Pripravimo email_line_items iz line_items
                        email_line_items = []
                        for item in line_items:
                            if not item: continue
                            try:
                                price = float(item.get('price', 0.0))
                            except (ValueError, TypeError):
                                price = 0.0

                            email_line_items.append({
                                'title': item.get('title', 'N/A'),
                                'quantity': item.get('quantity', 1),
                                'price': price,
                                'image_url': item.get('image_url', 'https://cdn.shopify.com/s/files/1/0533/2089/files/placeholder-images-image_large.png?v=1529089297')
                            })
                        
                        from services.pdf_service import ustvari_pdf
                        pdf_path, pdf_message = ustvari_pdf(
                            declaration_items, 
                            email_line_items, 
                            order_data['country_code'], 
                            order_number
                        )
                        
                        if not pdf_path:
                            current_app.logger.error(f"Napaka pri generiranju PDF-ja za naročilo {order_number}: {pdf_message}")
                            continue

                        cursor.execute("""
                            UPDATE orders 
                            SET pdf_generated_at = NOW(), processed_at = NOW()
                            WHERE shopify_order_id = %s
                        """, (shopify_order_id,))
                        
                        current_app.logger.info(
                            f"Fulfilled naročilo {order_number}: PDF pripravljen; "
                            f"Mandrill pošiljanje prepuščeno safety net job-u"
                        )
                        
                    except Exception as e:
                        current_app.logger.error(f"Napaka pri avtomatskem procesiranju fulfilled naročila {order_number}: {e}")
                        traceback.print_exc()
                        db.rollback()  # Ponastavi transakcijo
                        continue
                    
            except Exception as e:
                current_app.logger.error(f"Napaka pri preverjanju naročila {order.get('order_number', 'N/A')}: {e}")
                continue
        
        db.commit()
        
        message = f"Uspešno posodobljen fulfilled status za {updated_count} naročil."
        current_app.logger.info(message)
        
        return jsonify({
            "message": message,
            "updated_count": updated_count,
            "total_checked": len(unfulfilled_orders)
        })
        
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Kritična napaka v /sync-fulfilled-status: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do kritične napake pri sinhronizaciji fulfilled statusa."}), 500
    finally:
        cursor.close()
@require_permission('send_email')
@api_bp.route('/process-unprocessed-fulfilled-orders', methods=['POST'])
def process_unprocessed_fulfilled_orders():
    """Procesira vsa fulfilled naročila, ki še niso bila procesirana (ni PDF-ja ali email-a)"""
    db = get_db()
    cursor = db.cursor()
    try:
        # Zagotovi, da je transakcija v pravilnem stanju
        db.rollback()
        current_app.logger.info("Začenjam procesiranje neprocesiranih fulfilled naročil...")
        
        # Pridobi vsa fulfilled naročila, ki še niso bila procesirana
        cursor.execute("""
            SELECT o.*, d.id as declaration_id 
            FROM orders o 
            LEFT JOIN declarations d ON o.order_number = d.order_number 
            WHERE o.fulfilled_at IS NOT NULL 
            AND (o.pdf_generated_at IS NULL OR o.email_sent_at IS NULL)
            ORDER BY o.fulfilled_at ASC
        """)
        
        unprocessed_orders = cursor.fetchall()
        
        if not unprocessed_orders:
            return jsonify({"message": "Ni neprocesiranih fulfilled naročil.", "processed_count": 0})
        
        current_app.logger.info(f"Našel {len(unprocessed_orders)} neprocesiranih fulfilled naročil")
        
        processed_count = 0
        
        for order_data in unprocessed_orders:
            try:
                order_number = order_data['order_number']
                shopify_order_id = order_data['shopify_order_id']
                
                current_app.logger.info(f"Procesiram neprocesirano fulfilled naročilo {order_number}")
                
                # Preveri, ali je že procesirano
                if order_data['declaration_id']:
                    current_app.logger.info(f"Naročilo {order_number} ima že declaration")
                else:
                    # 1. Pridobi podatke za deklaracijo
                    _shop_domain_for_decl = order_data.get('shopify_store_domain') if isinstance(order_data, dict) else None
                    line_items_raw = order_data['line_items'] or '[]'
                    line_items = json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])
                    if not line_items:
                        current_app.logger.error(f"Naročilo {order_number} nima line_items (shop={_shop_domain_for_decl})")
                        continue
                    
                    # 2. Pridobi podatke za deklaracijo
                    declaration_items, missing, warnings = _pridobi_podatke_za_deklaracijo_iz_shopify(
                        line_items, cursor, shop_domain=_shop_domain_for_decl,
                    )
                    
                    if not declaration_items:
                        current_app.logger.error(f"Ni bilo mogoče pridobiti podatkov za deklaracijo za naročilo {order_number}. Manjkajo: {missing}")
                        
                        # Pošljemo obvestilo administratorju
                        from services.email_service import poslji_obvestilo_o_narocilu_z_manjkajocimi_podatki
                        poslji_obvestilo_o_narocilu_z_manjkajocimi_podatki(
                            order_number=order_number,
                            missing_data_details=missing,
                            customer_email=order_data.get('customer_email'),
                            shopify_order_id=shopify_order_id
                        )
                        continue
                    
                    # 3. Shrani v declarations tabelo
                    success = _shrani_deklaracijo_v_bazo(order_number, declaration_items, cursor)
                    
                    if not success:
                        current_app.logger.error(f"Napaka pri shranjevanju deklaracije za naročilo {order_number}")
                        continue
                
                # 4. Generiraj PDF (če še ni)
                if not order_data['pdf_generated_at']:
                    _shop_domain_for_decl = order_data.get('shopify_store_domain') if isinstance(order_data, dict) else None
                    line_items_raw = order_data['line_items'] or '[]'
                    line_items = json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])
                    declaration_items, missing, warnings = _pridobi_podatke_za_deklaracijo_iz_shopify(
                        line_items, cursor, shop_domain=_shop_domain_for_decl,
                    )
                    
                    # Pripravimo email_line_items iz line_items
                    email_line_items = []
                    for item in line_items:
                        if not item: continue
                        try:
                            price = float(item.get('price', 0.0))
                        except (ValueError, TypeError):
                            price = 0.0

                        email_line_items.append({
                            'title': item.get('title', 'N/A'),
                            'quantity': item.get('quantity', 1),
                            'price': price,
                            'image_url': 'https://cdn.shopify.com/s/files/1/0533/2089/files/placeholder-images-image_large.png?v=1529089297'
                        })
                    
                    from services.pdf_service import ustvari_pdf
                    pdf_path, pdf_message = ustvari_pdf(
                        declaration_items, 
                        email_line_items, 
                        order_data['country_code'], 
                        order_number
                    )
                    
                    if not pdf_path:
                        current_app.logger.error(f"Napaka pri generiranju PDF-ja za naročilo {order_number}: {pdf_message}")
                        continue
                else:
                    pdf_path = None
                
                # 5. PDF pripravljen — pošiljanje kupcu prepuščeno Mandrill safety net-u
                update_fields = []
                update_params = []
                
                if not order_data['pdf_generated_at'] and pdf_path:
                    update_fields.append("pdf_generated_at = NOW()")
                
                if not order_data['processed_at']:
                    update_fields.append("processed_at = NOW()")
                
                if update_fields:
                    update_params.append(shopify_order_id)
                    cursor.execute(f"""
                        UPDATE orders 
                        SET {', '.join(update_fields)}
                        WHERE shopify_order_id = %s
                    """, update_params)
                
                processed_count += 1
                current_app.logger.info(
                    f"Fulfilled naročilo {order_number}: PDF OK; Mandrill po safety net-u"
                )
                
            except Exception as e:
                current_app.logger.error(f"Napaka pri procesiranju fulfilled naročila {order_data.get('order_number', 'N/A')}: {e}")
                traceback.print_exc()
                db.rollback()  # Ponastavi transakcijo
                continue
        
        db.commit()
        
        message = f"Uspešno procesiranih {processed_count} fulfilled naročil."
        current_app.logger.info(message)
        
        return jsonify({
            "message": message,
            "processed_count": processed_count,
            "total_found": len(unprocessed_orders)
        })
        
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Kritična napaka pri procesiranju fulfilled naročil: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do kritične napake pri procesiranju fulfilled naročil."}), 500
    finally:
        cursor.close()
@api_bp.route('/sync-shopify-fulfilled-times', methods=['POST'])
def sync_shopify_fulfilled_times_endpoint():
    """Posodobi Shopify fulfilled čase za že fulfilled naročila"""
    db = get_db()
    cursor = db.cursor()
    try:
        current_app.logger.info("Začenjam sinhronizacijo Shopify fulfilled časov...")
        
        # Pridobi vsa fulfilled naročila, ki nimajo Shopify časa
        cursor.execute("SELECT shopify_order_id, order_number, shopify_store_domain FROM orders WHERE fulfilled_at IS NOT NULL AND shopify_fulfilled_at IS NULL")
        orders_without_shopify_time = cursor.fetchall()
        
        current_app.logger.info(f"DEBUG: Našel {len(orders_without_shopify_time)} fulfilled naročil brez Shopify časa")
        for order in orders_without_shopify_time:
            current_app.logger.info(f"DEBUG: Naročilo {order['order_number']} (ID: {order['shopify_order_id']}) nima Shopify časa")
        
        if not orders_without_shopify_time:
            return jsonify({"message": "Vsa fulfilled naročila že imajo Shopify čas.", "updated_count": 0})
        
        current_app.logger.info(f"Posodabljam Shopify čase za {len(orders_without_shopify_time)} naročil")
        
        updated_count = 0
        
        # Pridobi podatke iz Shopify-ja za vsa naročila
        from services.shopify_service import get_order_fulfillment_details
        
        for order in orders_without_shopify_time:
            try:
                shopify_order_id = order['shopify_order_id']
                order_number = order['order_number']
                shop_domain = order.get('shopify_store_domain')
                
                current_app.logger.info(f"DEBUG: Preverjam Shopify fulfillment za naročilo {order_number} (ID: {shopify_order_id})")
                
                # Pridobi fulfilled čas iz Shopify-ja
                fulfillment_details = get_order_fulfillment_details(shopify_order_id, shop_domain=shop_domain)
                current_app.logger.info(f"DEBUG: Shopify odgovor za {order_number}: {fulfillment_details}")
                
                if fulfillment_details and fulfillment_details.get('created_at'):
                    shopify_fulfilled_at = fulfillment_details.get('created_at')
                    
                    # Posodobi Shopify čas v bazi
                    cursor.execute(
                        """
                        UPDATE orders
                        SET shopify_fulfilled_at = %s
                        WHERE shopify_order_id = %s
                          AND (shopify_store_domain = %s OR shopify_store_domain IS NULL)
                        """,
                        (shopify_fulfilled_at, shopify_order_id, shop_domain)
                    )
                    updated_count += 1
                    current_app.logger.info(f"Posodobljen Shopify čas za naročilo {order_number}: {shopify_fulfilled_at}")
                else:
                    current_app.logger.info(f"DEBUG: Ni fulfillment podrobnosti za naročilo {order_number}")
                    
            except Exception as e:
                current_app.logger.error(f"Napaka pri posodabljanju Shopify časa za naročilo {order.get('order_number', 'N/A')}: {e}")
                continue
        
        db.commit()
        
        message = f"Uspešno posodobljen Shopify čas za {updated_count} naročil."
        current_app.logger.info(message)
        
        return jsonify({
            "message": message,
            "updated_count": updated_count,
            "total_checked": len(orders_without_shopify_time)
        })
        
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Kritična napaka v /sync-shopify-fulfilled-times: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do kritične napake pri sinhronizaciji Shopify časov."}), 500
    finally:
        cursor.close()

@api_bp.route('/register-webhooks', methods=['POST'])
def register_webhooks_endpoint():
    """Registrira potrebne webhook-e v Shopify-ju"""
    try:
        from services.shopify_service import register_webhooks
        
        success = register_webhooks()
        
        if success:
            return jsonify({"message": "Webhook-i uspešno registrirani v Shopify-ju"})
        else:
            return jsonify({"error": "Napaka pri registraciji webhook-ov"}), 500
            
    except Exception as e:
        current_app.logger.error(f"Kritična napaka pri registraciji webhook-ov: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do kritične napake pri registraciji webhook-ov."}), 500

@api_bp.route('/list-webhooks', methods=['GET'])
def list_webhooks_endpoint():
    """Pridobi seznam vseh webhook-ov iz Shopify-ja"""
    try:
        from services.shopify_service import list_webhooks
        
        webhooks = list_webhooks()
        
        return jsonify({
            "webhooks": webhooks,
            "count": len(webhooks)
        })
        
    except Exception as e:
        current_app.logger.error(f"Kritična napaka pri pridobivanju webhook-ov: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do kritične napake pri pridobivanju webhook-ov."}), 500

@api_bp.route('/check-shopify-fulfillment/<shopify_order_id>', methods=['GET'])
def check_shopify_fulfillment(shopify_order_id):
    """Preveri fulfillment podrobnosti za specifično naročilo iz Shopify-ja"""
    cursor = None
    try:
        current_app.logger.info(f"Preverjam fulfillment za Shopify naročilo {shopify_order_id}")
        
        from services.shopify_service import get_order_fulfillment_details

        db = get_db()
        cursor = db.cursor()
        shop_domain = request.args.get('shop')
        if not shop_domain:
            try:
                cursor.execute(
                    "SELECT shopify_store_domain FROM orders WHERE shopify_order_id = %s ORDER BY created_at DESC LIMIT 1",
                    (shopify_order_id,),
                )
                row = cursor.fetchone()
                shop_domain = row.get('shopify_store_domain') if isinstance(row, dict) else (row[0] if row else None)
            except Exception:
                shop_domain = None
        
        # Pridobi fulfillment podrobnosti iz Shopify-ja
        fulfillment_details = get_order_fulfillment_details(shopify_order_id, shop_domain=shop_domain)
        
        if fulfillment_details and fulfillment_details.get('created_at'):
            shopify_fulfilled_at = fulfillment_details.get('created_at')
            
            # Posodobi bazo podatkov
            cursor.execute(
                """
                UPDATE orders
                SET shopify_fulfilled_at = %s
                WHERE shopify_order_id = %s
                  AND (shopify_store_domain = %s OR shopify_store_domain IS NULL)
                """,
                (shopify_fulfilled_at, shopify_order_id, shop_domain)
            )
            db.commit()
            
            current_app.logger.info(f"Posodobljen Shopify fulfilled čas za naročilo {shopify_order_id}: {shopify_fulfilled_at}")
            
            return jsonify({
                "success": True,
                "fulfilled_at": shopify_fulfilled_at,
                "message": f"Posodobljen Shopify čas: {shopify_fulfilled_at}"
            })
        else:
            return jsonify({
                "success": False,
                "message": "Naročilo ni fulfilled v Shopify-ju"
            })
    except Exception as e:
        current_app.logger.error(f"Napaka pri preverjanju fulfillment za naročilo {shopify_order_id}: {e}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass
@api_bp.route('/narocila', methods=['GET'])
def get_narocila():
    db = get_db()
    cursor = db.cursor()
    try:
        # Pridobi parametre za paginacijo, filter in iskanje
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        filter_type = request.args.get('filter', 'unfulfilled')  # Default: unfulfilled
        search_term = request.args.get('search', '').strip()
        offset = (page - 1) * per_page
        
        # Kolona nalivalec_id se doda prek migracij; ne izvajamo DDL v runtime, da se izognemo H12

        def _shopify_admin_store_handle(shop_domain: str | None, order_number: str | None = None) -> str | None:
            on = str(order_number or '').lstrip('#').upper()
            if on.startswith('SI'):
                return 'amour-parfums-2'
            if not shop_domain:
                return current_app.config.get('SHOP_NAME')
            sd = str(shop_domain).strip()
            sd = sd.replace('https://', '').replace('http://', '')
            if 'admin.shopify.com/store/' in sd:
                try:
                    return sd.split('admin.shopify.com/store/')[1].split('/')[0]
                except Exception:
                    return current_app.config.get('SHOP_NAME')
            # alias -> canonical handle
            if sd.startswith('kxugn4-mu.'):
                return 'amour-parfums-2'
            if sd.endswith('.myshopify.com'):
                return sd.split('.')[0]
            return current_app.config.get('SHOP_NAME') or sd

        # Pripravi WHERE pogoj glede na filter in iskanje
        where_conditions = []
        
        # Filter pogoj
        if filter_type == 'fulfilled':
            where_conditions.append("(o.fulfilled_at IS NOT NULL OR o.shopify_fulfilled_at IS NOT NULL)")
        elif filter_type == 'unfulfilled':
            where_conditions.append("(o.fulfilled_at IS NULL AND o.shopify_fulfilled_at IS NULL)")
        elif filter_type == 'prepared':
            # Pripravljeno = Unfulfilled AND ima slike AND prepared_by IS NOT NULL (točno kot badge)
            where_conditions.append("(o.fulfilled_at IS NULL AND o.shopify_fulfilled_at IS NULL)")
            where_conditions.append("o.prepared_by IS NOT NULL")
        elif filter_type in ('invoice_sent', 'invoice_not_sent'):
            exists_invoice = (
                "(EXISTS (SELECT 1 FROM app_logs a "
                "WHERE a.category = 'email.send_invoice' "
                "AND (COALESCE(a.data->>'order_number','') = o.order_number OR COALESCE(a.data->>'order_number','') = REPLACE(o.order_number, '#',''))) "
                "OR EXISTS (SELECT 1 FROM invoice_email_log l WHERE l.order_number = o.order_number OR l.order_number = REPLACE(o.order_number, '#','')))"
            )
            if filter_type == 'invoice_sent':
                where_conditions.append(exists_invoice)
            else:
                where_conditions.append("NOT " + exists_invoice)
        elif filter_type == 'manjkajo_podatki':
            where_conditions.append("o.status = 'manjkajo_podatki'")
        
        # Iskalni pogoj
        if search_term:
            # Išči po številki naročila (z ali brez #)
            where_conditions.append("(order_number ILIKE %s OR order_number ILIKE %s)")
            search_with_hash = f"%{search_term}%"
            search_without_hash = f"%#{search_term}%"
        
        # Sestavi WHERE klavzulo
        if where_conditions:
            where_clause = "WHERE " + " AND ".join(where_conditions)
        else:
            where_clause = ""
        
        # Najprej preštejemo skupno število naročil glede na filter in iskanje
        total_count = None
        try:
            count_query = f"SELECT COUNT(*) as total FROM orders o {where_clause}"
            # Kratek timeout za count, da se izognemo H12
            cursor.execute("SET LOCAL statement_timeout = '2000ms'")
            if search_term:
                cursor.execute(count_query, (search_with_hash, search_without_hash))
            else:
                cursor.execute(count_query)
            total_count = cursor.fetchone()['total']
        except Exception as e:
            current_app.logger.warning(f"Count query timeout/fail in /api/narocila: {e}")
            try:
                db.rollback()
            except Exception:
                pass
            total_count = None
        finally:
            try:
                cursor.execute("SET LOCAL statement_timeout = '0'")
            except Exception:
                pass
        
        # Pridobimo naročila za trenutno stran glede na filter in iskanje z podatki o uporabniku
        query = f"""
            SELECT o.*,
                   u.first_name,
                   u.last_name,
                   u.id as prepared_by_id,
                   u2.first_name as nalivalec_first,
                   u2.last_name as nalivalec_last,
                   EXISTS (
                       SELECT 1
                       FROM declarations d
                       WHERE d.order_number = o.order_number
                          OR d.order_number = REPLACE(o.order_number, '#','')
                          OR d.order_number = CONCAT('#', REPLACE(o.order_number, '#',''))
                   ) AS has_declarations
            FROM orders o 
            LEFT JOIN users u ON o.prepared_by = u.username 
            LEFT JOIN users u2 ON o.nalivalec_id = u2.id
            {where_clause} 
            ORDER BY COALESCE(o.created_at, o.email_sent_at, o.fulfilled_at) DESC 
            LIMIT %s OFFSET %s
        """
        if search_term:
            cursor.execute(query, (search_with_hash, search_without_hash, per_page, offset))
        else:
            cursor.execute(query, (per_page, offset))
        narocila = cursor.fetchall()

        # Zbere order_number vrednosti za enkratni lookup "invoice sent" iz app_logs/invoice_email_log
        try:
            order_numbers = [str(n.get('order_number') or '') for n in narocila]
            normalized = []
            for on in order_numbers:
                if not on:
                    continue
                normalized.append(on)
                if on.startswith('#'):
                    normalized.append(on.lstrip('#'))
                else:
                    normalized.append(f"#{on}")
            invoice_sent_set = set()
            if normalized:
                # Preveri app_logs (email.send_invoice)
                q_marks = ','.join(['%s'] * len(normalized))
                cursor.execute(
                    f"""
                    SELECT DISTINCT COALESCE(data->>'order_number','') AS onum
                    FROM app_logs
                    WHERE category = 'email.send_invoice'
                      AND COALESCE(data->>'order_number','') IN ({q_marks})
                    """,
                    tuple(normalized)
                )
                for r in cursor.fetchall() or []:
                    if r and (r.get('onum') or r[0]):
                        invoice_sent_set.add(str(r.get('onum') if isinstance(r, dict) else r[0]))
                # Fallback: legacy invoice_email_log
                cursor.execute(
                    f"""
                    SELECT DISTINCT order_number FROM invoice_email_log
                    WHERE order_number IN ({q_marks})
                    """,
                    tuple(normalized)
                )
                for r in cursor.fetchall() or []:
                    val = r.get('order_number') if isinstance(r, dict) else r[0]
                    if val:
                        invoice_sent_set.add(str(val))
        except Exception:
            invoice_sent_set = set()
        
        # DEBUG: Izpiši podatke o fulfilled naročilih
        fulfilled_orders = [n for n in narocila if n.get('fulfilled_at')]
        current_app.logger.info(f"DEBUG: Na strani {page} je {len(fulfilled_orders)} fulfilled naročil")
        for order in fulfilled_orders:
            current_app.logger.info(f"DEBUG: Naročilo {order.get('order_number')}: fulfilled_at={order.get('fulfilled_at')}, shopify_fulfilled_at={order.get('shopify_fulfilled_at')}")
        
        # DEBUG: Preveri tudi vsa naročila za shopify_fulfilled_at
        orders_with_shopify_time = [n for n in narocila if n.get('shopify_fulfilled_at')]
        current_app.logger.info(f"DEBUG: Na strani {page} je {len(orders_with_shopify_time)} naročil s Shopify časom")
        
        # Za hitrost: privzeto uporabljamo shranjen status/manjkajoče podatke.
        compute_missing = False
        shopify_details = {}
        if compute_missing:
            product_ids_to_fetch = set()
            for narocilo in narocila:
                line_items_raw = narocilo.get('line_items', '[]')
                try:
                    line_items = json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])
                    for item in line_items:
                        if item and item.get('product_id'):
                            product_ids_to_fetch.add(str(item['product_id']))
                except (json.JSONDecodeError, TypeError) as e:
                    current_app.logger.warning(f"Could not parse line_items for order {narocilo.get('order_number')}: {e}")

            # Ne briši cache-a pred vsakim klicem; uporabi robustno batch funkcijo
            shopify_details = get_bulk_product_details(list(product_ids_to_fetch))
        
        # Vnaprej naloži image_count za vsa naročila na strani (1 query)
        image_count_map = {}
        try:
            order_nums = [str(n.get('order_number', '')).replace('#', '') for n in narocila if n.get('order_number')]
            order_nums = [o for o in order_nums if o]
            if order_nums:
                q_marks = ','.join(['%s'] * len(order_nums))
                cursor.execute(
                    f"""
                    SELECT REPLACE(order_number, '#','') as order_number_norm,
                           COUNT(*) as image_count
                    FROM order_images
                    WHERE REPLACE(order_number, '#','') IN ({q_marks})
                    GROUP BY order_number_norm
                    """,
                    tuple(order_nums)
                )
                for r in cursor.fetchall() or []:
                    onum = r.get('order_number_norm') if isinstance(r, dict) else r[0]
                    cnt = r.get('image_count') if isinstance(r, dict) else r[1]
                    if onum:
                        image_count_map[str(onum)] = int(cnt or 0)
        except Exception as e:
            current_app.logger.warning(f"Image count preload failed: {e}")
            image_count_map = {}

        processed_narocila = []
        for narocilo in narocila:
            try:
                missing_data_details = []
                affected_perfumes = []
                status = 'neznan'
                if compute_missing:
                    line_items_raw = narocilo.get('line_items', '[]')
                    line_items = json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])

                    items_for_declaration, missing_data_details = [], []
                    affected_perfumes = []
                    
                    for item in line_items:
                        if not item or not item.get('product_id'): continue
                        product_id = str(item.get('product_id'))
                        details = shopify_details.get(product_id, {})
                        
                        product_type = details.get('product_type')
                        
                        if not product_type or product_type.strip().lower() != 'parfumi':
                            continue

                        product_no = details.get('product_no')
                        proizvajalec_ime_raw = details.get('proizvajalec_id')

                        if not product_no or not isinstance(proizvajalec_ime_raw, str):
                            missing_data_details.append(f"Manjkajo metafields za '{item.get('title', 'N/A')}'")
                            continue
                        
                        items_for_declaration.append({
                            'title': item.get('title'), 
                            'product_no': product_no, 
                            'proizvajalec_ime': proizvajalec_ime_raw.upper()
                        })
                        # Poskusi najti parfum v bazi, da omogočimo hitro urejanje manjkajočih podatkov iz UI
                        try:
                            cursor.execute(
                                """
                                SELECT p.id, p.product_no, p.ime_parfuma, pr.ime AS proizvajalec
                                FROM parfumi p 
                                JOIN proizvajalci pr ON p.proizvajalec_id = pr.id
                                WHERE p.product_no = %s AND UPPER(pr.ime) = %s
                                LIMIT 1
                                """,
                                (product_no, proizvajalec_ime_raw.upper())
                            )
                            parf_row = cursor.fetchone()
                            if parf_row:
                                # Filtriraj: obdrži le parfume, ki so dejansko problematični v tem naročilu
                                # To pomeni: manjkajoča INCI ali ni veljavne serije
                                is_problematic = False
                                # Preveri INCI in zalogo
                                db_cursor = cursor
                                db_cursor.execute("""
                                    SELECT p.id, p.sestava_inci, p.na_zalogi,
                                           (SELECT COUNT(*) FROM serije s WHERE s.parfum_id = p.id AND s.rok_uporabe >= CURRENT_DATE) AS veljavne_serije
                                    FROM parfumi p
                                    WHERE p.id = %s
                                """, (parf_row['id'],))
                                parfum_check = db_cursor.fetchone()
                                if parfum_check:
                                    if not parfum_check.get('sestava_inci'):
                                        is_problematic = True
                                    elif parfum_check.get('veljavne_serije', 0) == 0:
                                        is_problematic = True
                                if is_problematic:
                                    affected_perfumes.append({
                                        'id': parf_row['id'],
                                        'product_no': parf_row['product_no'],
                                        'ime_parfuma': parf_row['ime_parfuma'],
                                        'proizvajalec': parf_row['proizvajalec']
                                    })
                        except Exception:
                            pass
                    
                    if narocilo.get('email_sent_at'):
                        status = 'email_poslan'
                    elif not items_for_declaration:
                        status = 'brez_parfumov'
                    else:
                        _, db_missing, _ = _pridobi_podatke_za_deklaracijo(items_for_declaration, cursor)
                        missing_data_details.extend(db_missing)

                        if missing_data_details:
                            status = 'manjkajo_podatki'
                        else:
                            # Preveri, ali je naročilo fulfilled ali unfulfilled
                            if narocilo.get('fulfilled_at') or narocilo.get('shopify_fulfilled_at'):
                                status = 'fulfilled'
                            else:
                                status = 'unfulfilled'
                else:
                    # Uporabi shranjen status le za posebne primere, sicer izpelji iz fulfilled_at
                    raw_status = (narocilo.get('status') or '').strip()
                    if raw_status in ('manjkajo_podatki', 'email_poslan', 'brez_parfumov'):
                        status = raw_status
                    else:
                        status = 'fulfilled' if (narocilo.get('fulfilled_at') or narocilo.get('shopify_fulfilled_at')) else 'unfulfilled'
                    raw_missing = narocilo.get('missing_data_details')
                    if raw_missing:
                        if isinstance(raw_missing, str):
                            try:
                                missing_data_details = json.loads(raw_missing)
                            except Exception:
                                missing_data_details = raw_missing
                        else:
                            missing_data_details = raw_missing
                    raw_aff = narocilo.get('affected_perfumes')
                    if raw_aff:
                        if isinstance(raw_aff, str):
                            try:
                                affected_perfumes = json.loads(raw_aff)
                            except Exception:
                                affected_perfumes = []
                        else:
                            affected_perfumes = raw_aff

                # Preveri, ali naročilo ima slike (1 poizvedba za vse order_number)
                has_images = False
                image_count = 0
                if image_count_map:
                    image_count = image_count_map.get(narocilo['order_number'].replace('#', ''), 0)
                    has_images = image_count > 0

                # Status ostane 'unfulfilled'; 'Pripravljeno' prikazujemo kot dodatni badge v UI
                
                # Sestavi prikazno ime za prepared_by
                first_name = narocilo.get('first_name')
                last_name = narocilo.get('last_name')
                prepared_by = narocilo.get('prepared_by')
                prepared_by_display = None
                
                if first_name and last_name:
                    prepared_by_display = f"{first_name} {last_name}"
                elif first_name:
                    prepared_by_display = first_name
                elif last_name:
                    prepared_by_display = last_name
                elif prepared_by:
                    prepared_by_display = prepared_by  # Fallback na username

                # Sestavi prikazno ime za nalivalca
                nalivalec_first = narocilo.get('nalivalec_first')
                nalivalec_last = narocilo.get('nalivalec_last')
                nalivalec_display = None
                if nalivalec_first and nalivalec_last:
                    nalivalec_display = f"{nalivalec_first} {nalivalec_last}"
                elif nalivalec_first:
                    nalivalec_display = nalivalec_first
                elif nalivalec_last:
                    nalivalec_display = nalivalec_last

                # Dodaj v rezultat (ne glede na to, ali imamo prikazno ime)
                # Normalize order number for invoice_sent detection
                on_raw = str(narocilo.get('order_number') or '')
                on_no = on_raw.lstrip('#')
                on_hash = f"#{on_no}" if on_no else on_raw
                store_handle = _shopify_admin_store_handle(
                    narocilo.get('shopify_store_domain'),
                    narocilo.get('order_number'),
                )
                processed_narocila.append({
                    **narocilo,
                    'status': status,
                    'missing_data_details': missing_data_details,
                    'has_images': has_images,
                    'image_count': image_count,
                    'prepared_by_display': prepared_by_display,
                    'nalivalec_display': nalivalec_display,
                    'affected_perfumes': affected_perfumes,
                    'order_admin_url': f"https://admin.shopify.com/store/{store_handle}/orders/{narocilo['shopify_order_id']}" if store_handle else '#',
                    'invoice_sent': (on_raw in invoice_sent_set) or (on_no in invoice_sent_set) or (on_hash in invoice_sent_set)
                })

            except Exception as e:
                current_app.logger.error(f"Error processing order {narocilo.get('order_number', 'N/A')}: {e}")
                traceback.print_exc()
                processed_narocila.append({
                    'order_number': narocilo.get('order_number', 'Neznano'),
                    'customer_name': 'NAPAKA PRI OBDELAVI',
                    'created_at': narocilo.get('created_at'),
                    'fulfilled_at': None, 'email_sent_at': None, 'email_recipient': None, 'pdf_generated_at': None,
                    'status': 'napaka_pri_obdelavi',
                    'missing_data_details': [f"Strežniška napaka: {e}"],
                    'order_admin_url': '#'
                })

        # Izračunaj informacije o paginaciji
        if total_count is None:
            # Fallback: oceni count iz trenutne strani, da UI ostane odziven
            total_count = offset + len(processed_narocila)
            if len(processed_narocila) == per_page:
                total_count += per_page
        total_pages = (total_count + per_page - 1) // per_page
        
        return jsonify({
            'narocila': processed_narocila,
            'pagination': {
                'current_page': page,
                'per_page': per_page,
                'total_count': total_count,
                'total_pages': total_pages,
                'has_next': page < total_pages,
                'has_prev': page > 1
            }
        })

    except Exception as e:
        current_app.logger.error(f"Critical error in /api/narocila: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do napake na strežniku pri pridobivanju naročil."}), 500
    finally:
        cursor.close()
def sync_new_perfumes():
    """Sinhronizira nove parfume iz Shopify-ja v lokalno bazo."""
    db = get_db()
    cursor = db.cursor()
    try:
        current_app.logger.info("Starting new perfumes sync from Shopify...")
        
        # Pridobi podatke iz Shopify-ja
        current_app.logger.info("Fetching products from Shopify...")
        shopify_products = get_all_products_for_name_sync()
        
        if shopify_products is None:
            current_app.logger.error("Failed to fetch products from Shopify")
            return jsonify({"error": "Napaka pri pridobivanju podatkov iz Shopify."}), 500
        
        current_app.logger.info(f"Successfully fetched {len(shopify_products)} products from Shopify")
        
        added_count = 0
        updated_count = 0
        skipped_count = 0
        
        for i, product in enumerate(shopify_products):
            try:
                vendor = product.get('vendor')
                fragrance = product.get('product_fragrance', {}).get('value')
                product_no = product.get('product_no', {}).get('value')
                proizvajalec_id_val = product.get('proizvajalec_id', {}).get('value')
                sestava_inci = product.get('sestava_inci', {}).get('value')
                na_zalogi = product.get('na_zalogi', False)

                current_app.logger.debug(f"Processing product {i+1}/{len(shopify_products)}: vendor={vendor}, product_no={product_no}, proizvajalec={proizvajalec_id_val}")

                # Če proizvajalec_id ni na voljo, uporabimo vendor
                if not proizvajalec_id_val and vendor:
                    proizvajalec_id_val = vendor
                    current_app.logger.info(f"Using vendor '{vendor}' as proizvajalec for product {product_no}")

                # Če product_fragrance ni na voljo, poskusimo z drugimi metafield-i
                if not fragrance:
                    # Pridobi vse metafield-e za ta produkt
                    metafields = {}
                    metafields_data = product.get('metafields', {})
                    if metafields_data and 'edges' in metafields_data:
                        for metafield_edge in metafields_data.get('edges', []):
                            metafield_node = metafield_edge.get('node')
                            if metafield_node:
                                namespace = metafield_node.get('namespace', '')
                                key = metafield_node.get('key', '')
                                value = metafield_node.get('value', '')
                                metafields[f"{namespace}.{key}"] = value
                    
                    # Poskusi različne metafield ključe za fragrance
                    fragrance_keys = [
                        'custom.product_fragrance_',
                        'custom.fragrance',
                        'custom.product_name',
                        'custom.title'
                    ]
                    
                    for key in fragrance_keys:
                        if key in metafields and metafields[key]:
                            fragrance = metafields[key]
                            current_app.logger.info(f"Found fragrance in {key}: {fragrance} for product {product_no}")
                            break

                if not all([vendor, fragrance, product_no, proizvajalec_id_val]):
                    current_app.logger.warning(f"Preskačem produkt z manjkajočimi podatki: vendor={vendor}, fragrance={fragrance}, product_no={product_no}, proizvajalec={proizvajalec_id_val}")
                    skipped_count += 1
                    continue

                # Preverimo, ali proizvajalec obstaja
                cursor.execute("SELECT id FROM proizvajalci WHERE ime = %s", (proizvajalec_id_val,))
                proizvajalec = cursor.fetchone()
                
                if not proizvajalec:
                    current_app.logger.warning(f"Proizvajalec '{proizvajalec_id_val}' ne obstaja v bazi. Dodajam...")
                    cursor.execute("INSERT INTO proizvajalci (ime) VALUES (%s) ON CONFLICT (ime) DO NOTHING", (proizvajalec_id_val,))
                    db.commit()
                    cursor.execute("SELECT id FROM proizvajalci WHERE ime = %s", (proizvajalec_id_val,))
                    proizvajalec = cursor.fetchone()

                proizvajalec_id = proizvajalec['id']
                new_name = f"{vendor} - {fragrance}"
                
                # Preverimo, ali parfum že obstaja
                cursor.execute("""
                    SELECT id, ime_parfuma, sestava_inci, na_zalogi
                    FROM parfumi 
                    WHERE product_no = %s AND proizvajalec_id = %s
                """, (product_no, proizvajalec_id))
                
                existing_perfume = cursor.fetchone()
                
                if existing_perfume:
                    # Posodobimo obstoječi parfum
                    updates = []
                    params = []
                    
                    if existing_perfume['ime_parfuma'] != new_name:
                        updates.append("ime_parfuma = %s")
                        params.append(new_name)
                    
                    if sestava_inci and existing_perfume['sestava_inci'] != sestava_inci:
                        updates.append("sestava_inci = %s")
                        params.append(sestava_inci)
                    
                    if existing_perfume['na_zalogi'] != na_zalogi:
                        updates.append("na_zalogi = %s")
                        params.append(na_zalogi)
                    
                    if updates:
                        params.extend([product_no, proizvajalec_id])
                        cursor.execute(f"""
                            UPDATE parfumi 
                            SET {', '.join(updates)}
                            WHERE product_no = %s AND proizvajalec_id = %s
                        """, params)
                        updated_count += 1
                        current_app.logger.info(f"Posodobljen parfum: {new_name}")
                else:
                    # Dodamo nov parfum
                    cursor.execute("""
                        INSERT INTO parfumi (product_no, proizvajalec_id, ime_parfuma, sestava_inci, sinhroniziraj_s_shopify, na_zalogi)
                        VALUES (%s, %s, %s, %s, TRUE, %s)
                        ON CONFLICT (product_no, proizvajalec_id) DO NOTHING
                    """, (product_no, proizvajalec_id, new_name, sestava_inci, na_zalogi))
                    
                    if cursor.rowcount > 0:
                        added_count += 1
                        current_app.logger.info(f"Dodan nov parfum: {new_name}")
                        
            except Exception as e:
                current_app.logger.error(f"Error processing product {i+1}: {e}")
                skipped_count += 1
                continue
        
        db.commit()
        message = f"Sinhronizacija novih parfumov končana. Dodanih {added_count} novih, posodobljenih {updated_count} obstoječih, preskočenih {skipped_count} parfumov."
        current_app.logger.info(message)
        return jsonify({"message": message, "added_count": added_count, "updated_count": updated_count, "skipped_count": skipped_count})

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri sinhronizaciji novih parfumov: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Prišlo je do napake na strežniku: {str(e)}"}), 500
    finally:
        cursor.close()

@api_bp.route('/sync-names', methods=['POST'])
def sync_perfume_names():
    """Posodobi sestavo INCI iz Shopify za obstoječe parfume (ne prepisuje imen)."""
    data = request.get_json(silent=True) or {}
    shop_domain = (data.get('shop_domain') or DEFAULT_SYNC_STORE).strip()
    dry_run = bool(data.get('dry_run'))

    current_app.logger.info(
        "Starting INCI sync from Shopify (shop=%s)...",
        shop_domain,
    )

    result = sync_parfumi_from_shopify(
        shop_domain,
        dry_run=dry_run,
        update_existing=True,
    )
    if result.get("error") and not result.get("ok"):
        return jsonify({"error": result["error"]}), 400

    message = (
        f"Sinhronizacija INCI končana ({result.get('shop_domain')}). "
        f"Posodobljenih INCI: {result.get('updated', 0)}, "
        f"preskočenih: {result.get('skipped', 0)}."
        f"{ ' (dry run)' if dry_run else '' }"
    )
    current_app.logger.info(message)
    return jsonify({"message": message, "result": result})

def get_all_shopify_products_with_metafields():
    """Pridobi vse produkte in njihove ključne metafielde/tage v enem samem učinkovitem klicu."""
    products = {}
    hasNextPage = True
    cursor = None
    
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": current_app.config['SHOPIFY_API_PASSWORD']
    }
    api_version = os.environ.get('SHOPIFY_API_VERSION') or current_app.config.get('SHOPIFY_API_VERSION', '2025-01')
    url = f"https://{current_app.config['SHOP_NAME']}.myshopify.com/admin/api/{api_version}/graphql.json"

    while hasNextPage:
        after_cursor = f', after: "{cursor}"' if cursor else ""
        query = f"""
        {{
          products(first: 100{after_cursor}) {{
            pageInfo {{
              hasNextPage
              endCursor
            }}
            edges {{
              node {{
                id
                tags
                product_no_metafield: metafield(namespace: "custom", key: "product_no") {{
                  value
                }}
                proizvajalec_id_metafield: metafield(namespace: "custom", key: "proizvajalec_id") {{
                  value
                }}
              }}
            }}
          }}
        }}
        """
        
        try:
            response = requests.post(url, json={'query': query}, headers=headers)
            response.raise_for_status()
            data = response.json().get('data', {}).get('products', {})
            
            for edge in data.get('edges', []):
                node = edge['node']
                product_no_node = node.get('product_no_metafield')
                proizvajalec_id_node = node.get('proizvajalec_id_metafield')

                product_no = product_no_node.get('value') if product_no_node else None
                proizvajalec_id = proizvajalec_id_node.get('value') if proizvajalec_id_node else None
                
                if product_no and proizvajalec_id:
                    key = f"{product_no.strip()}_{proizvajalec_id.strip().upper()}"
                    products[key] = {
                        'tags': node.get('tags', [])
                    }

            hasNextPage = data.get('pageInfo', {}).get('hasNextPage', False)
            cursor = data.get('pageInfo', {}).get('endCursor')
        except Exception as e:
            current_app.logger.error(f"Error fetching all Shopify products: {e}")
            traceback.print_exc()
            return None

    return products

@api_bp.route('/sync-stock-status', methods=['POST'])
def sync_stock_status():
    """Posodobi na_zalogi=TRUE le za parfume, ki imajo GREEN tag v Shopify.

    App je vir resnice za zalogo — ta endpoint nikoli ne nastavi na_zalogi=FALSE
    (npr. ko je izdelek izbrisan iz Shopify-ja).
    """
    db = get_db()
    cursor = db.cursor()
    try:
        current_app.logger.info("Starting stock sync: Fetching all products from Shopify...")
        shopify_products = get_all_shopify_products_with_metafields()
        if shopify_products is None:
            return jsonify({"error": "Napaka pri pridobivanju podatkov iz Shopify."}), 500
        current_app.logger.info(f"Fetched {len(shopify_products)} products from Shopify.")

        cursor.execute("""
            SELECT p.id, p.product_no, pr.ime as ime_proizvajalca, p.sinhroniziraj_s_shopify
            FROM parfumi p 
            JOIN proizvajalci pr ON p.proizvajalec_id = pr.id
        """)
        local_perfumes = cursor.fetchall()
        
        updated_count = 0
        skipped_count = 0
        in_stock_count = 0
        
        for perfume in local_perfumes:
            lookup_key = f"{perfume['product_no'].strip()}_{perfume['ime_proizvajalca'].strip().upper()}"
            shopify_data = shopify_products.get(lookup_key)

            if not shopify_data:
                skipped_count += 1
                continue

            is_in_stock_shopify = "GREEN" in shopify_data.get('tags', [])
            if is_in_stock_shopify:
                cursor.execute(
                    "UPDATE parfumi SET na_zalogi = TRUE WHERE id = %s AND na_zalogi IS DISTINCT FROM TRUE",
                    (perfume['id'],),
                )
                if cursor.rowcount > 0:
                    updated_count += 1
                in_stock_count += 1
        
        db.commit()
        
        message = (
            f"Sinhronizacija končana. Pregledanih {len(local_perfumes)} izdelkov. "
            f"Na zalogi (GREEN v Shopify): {in_stock_count}. "
            f"Posodobljenih na TRUE: {updated_count}. "
            f"Preskočenih (ni v Shopify): {skipped_count}. "
            f"Status na_zalogi=FALSE se iz tega gumba ne nastavlja."
        )
        return jsonify({"message": message})

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri sinhronizaciji statusa zaloge: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do napake na strežniku."}), 500
    finally:
        cursor.close()

@api_bp.route('/sync-new-perfumes', methods=['POST'])
def sync_new_perfumes_endpoint():
    """Sinhronizira parfume iz izbrane Shopify trgovine (enako kot app-v2 /seznami)."""
    data = request.get_json(silent=True) or {}
    shop_domain = (data.get('shop_domain') or DEFAULT_SYNC_STORE).strip()
    dry_run = bool(data.get('dry_run'))
    update_existing = bool(data.get('update_existing'))
    product_ids = data.get('product_ids') or None
    if product_ids is not None and not isinstance(product_ids, list):
        return jsonify({"error": "product_ids mora biti seznam ID-jev Shopify izdelkov."}), 400

    current_app.logger.info(
        "Starting perfume sync from Shopify (shop=%s, dry_run=%s, update_existing=%s, targeted=%s)...",
        shop_domain,
        dry_run,
        update_existing,
        bool(product_ids),
    )

    result = sync_parfumi_from_shopify(
        shop_domain,
        dry_run=dry_run,
        update_existing=update_existing,
        product_ids=product_ids,
    )
    if result.get("error") and not result.get("ok"):
        return jsonify({"error": result["error"]}), 400

    message = (
        f"Sinhronizacija končana ({result.get('shop_domain')}). "
        f"Pridobljenih: {result.get('fetched', 0)}, "
        f"dodanih: {result.get('added', 0)}, "
        f"posodobljenih: {result.get('updated', 0)}, "
        f"preskočenih: {result.get('skipped', 0)}, "
        f"napak: {result.get('errors', 0)}"
        f"{ ' (dry run)' if dry_run else ''}."
    )
    current_app.logger.info(message)
    return jsonify({"message": message, "result": result})
@api_bp.route('/generiraj_in_poslji', methods=['POST'])
def generiraj_in_poslji():
    current_app.logger.info("=== GENERIRAJ_IN_POSLJI ENDPOINT CALLED ===")
    order_number = request.json.get('order_number')
    current_app.logger.info(f"Začenjam generiranje in pošiljanje za naročilo: {order_number}")
    db = get_db()
    cursor = db.cursor()
    try:
        # Dodaj # prefix, če ga ni, ker se v bazi shranjuje z # prefixom
        if order_number and not order_number.startswith('#'):
            order_number_with_hash = f"#{order_number}"
        else:
            order_number_with_hash = order_number
            
        current_app.logger.info(f"Iščem naročilo v bazi z: {order_number_with_hash}")
        cursor.execute("SELECT * FROM orders WHERE order_number = %s", (order_number_with_hash,))
        order = cursor.fetchone()
        if not order: return jsonify({"sporocilo": "Naročilo ni najdeno."}), 404

        from services.declaration_safety_net import should_wait_for_2100_pdf_batch
        wait_batch, wait_msg = should_wait_for_2100_pdf_batch(dict(order))
        if wait_batch:
            return jsonify({"sporocilo": wait_msg}), 409
        
        # Najprej preverimo, ali podatki že obstajajo v declarations tabeli
        declaration_data = _pridobi_deklaracijo_iz_baze(order_number_with_hash, cursor)
        
        if declaration_data:
            # Podatki že obstajajo, uporabimo jih
            current_app.logger.info(f"Uporabljam obstoječe podatke iz declarations za naročilo {order_number_with_hash}")
            
            # Pretvorimo podatke v format, ki ga pričakuje PDF servis
            declaration_items = []
            for item in declaration_data:
                declaration_items.append({
                    'title': f"{item['product_no']} - {item['proizvajalec_ime']}",
                    'product_no': item['product_no'],
                    'proizvajalec_ime': item['proizvajalec_ime'],
                    'sestava_inci': item['sestava_inci'],
                    'rok_uporabe': item['rok_uporabe'].strftime('%d.%m.%Y') if hasattr(item['rok_uporabe'], 'strftime') and item['rok_uporabe'] else (str(item['rok_uporabe']) if item['rok_uporabe'] else None),
                    'serijska_stevilka': item['serijska_stevilka'] or 'N/A'
                })
            
            # Preveri opozorila glede roka uporabe (blokada < 60 dni)
            items_for_warning_check = []
            for item in declaration_items:
                items_for_warning_check.append({
                    'title': item.get('title', 'N/A'),
                    'product_no': item.get('product_no'),
                    'proizvajalec_ime': item.get('proizvajalec_ime'),
                })
            _, _, warnings = _pridobi_podatke_za_deklaracijo(items_for_warning_check, cursor)
            if warnings:
                return jsonify({
                    "sporocilo": "PDF ni bil generiran, ker imajo nekateri izdelki rok uporabe manj kot 60 dni.",
                    "warnings": warnings
                }), 400
            
            # Pripravimo email_line_items iz obstoječih podatkov
            line_items_raw = order.get('line_items', '[]')
            line_items = json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])
            email_line_items = []
            for item in line_items:
                if not item:
                    continue
                try:
                    price = float(item.get('price', 0.0))
                except (ValueError, TypeError):
                    price = 0.0
                email_line_items.append({
                    'title': item.get('title', 'N/A'),
                    'quantity': item.get('quantity', 1),
                    'price': price,
                    'image_url': 'https://cdn.shopify.com/s/files/1/0533/2089/files/placeholder-images-image_large.png?v=1529089297'
                })

            # Generiraj PDF tudi v primeru, ko podatki obstajajo v bazi
            pdf_path, pdf_msg = ustvari_pdf(declaration_items, email_line_items, order['country_code'], order_number_with_hash, warnings)
            if not pdf_path:
                cursor.execute(
                    "UPDATE orders SET status = 'manjkajo_podatki', missing_data_details = %s WHERE order_number = %s",
                    (json.dumps([pdf_msg]), order_number_with_hash)
                )
                db.commit()
                from services.email_service import poslji_obvestilo_o_narocilu_z_manjkajocimi_podatki
                poslji_obvestilo_o_narocilu_z_manjkajocimi_podatki(
                    order_number=order_number_with_hash,
                    missing_data_details=[f"Napaka pri generiranju PDF-ja: {pdf_msg}"],
                    customer_email=order.get('customer_email'),
                    shopify_order_id=order.get('shopify_order_id')
                )
                return jsonify({"sporocilo": pdf_msg}), 500
        
        else:
            # Podatki ne obstajajo, pridobimo jih iz Shopify-ja in shranimo
            current_app.logger.info(f"Podatki ne obstajajo v declarations, pridobivam iz Shopify-ja za naročilo {order_number_with_hash}")
            
            line_items_raw = order.get('line_items', '[]')
            line_items = json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])
            
            # Pridobimo vse potrebne podatke iz Shopifyja
            product_ids = [str(item['product_id']) for item in line_items if item and item.get('product_id')]
            clear_product_cache()
            shopify_details = get_bulk_product_details(product_ids)

            # Pripravimo podatke za email predlogo (seznam slovarjev)
            email_line_items = []
            for item in line_items:
                if not item or not item.get('product_id'):
                    continue
                product_id_str = str(item.get('product_id'))
                details = shopify_details.get(product_id_str, {})
                
                try:
                    price = float(item.get('price', 0.0))
                except (ValueError, TypeError):
                    price = 0.0

                email_line_items.append({
                    'title': item.get('title', 'N/A'),
                    'quantity': item.get('quantity', 1),
                    'price': price,
                    'image_url': details.get('image_url', 'https://cdn.shopify.com/s/files/1/0533/2089/files/placeholder-images-image_large.png?v=1529089297')
                })

            # Pripravimo podatke za PDF deklaracijo (samo parfumi)
            items_for_declaration = []
            for item in line_items:
                if not item or not item.get('product_id'):
                    continue
                details = shopify_details.get(str(item.get('product_id')), {})
                product_type = details.get('product_type')
                if not product_type or product_type.strip().lower() != 'parfumi':
                    continue
                if details.get('product_no') and details.get('proizvajalec_id'):
                    items_for_declaration.append({
                        'title': item.get('title'), 
                        'product_no': details['product_no'], 
                        'proizvajalec_ime': details['proizvajalec_id'].upper()
                    })
            
            current_app.logger.info(f"Našel {len(items_for_declaration)} parfumov za deklaracijo v naročilu {order_number_with_hash}")
            if not items_for_declaration:
                return jsonify({"sporocilo": "V naročilu ni nobenega izdelka tipa 'Parfumi', zato deklaracija ni bila poslana."}), 200
            declaration_items, missing, warnings = _pridobi_podatke_za_deklaracijo(items_for_declaration, cursor)
            current_app.logger.info(f"Pridobil podatke za deklaracijo: {len(declaration_items)} izdelkov, manjkajo: {missing}")
            if missing:
                # Posodobimo status naročila na manjkajo_podatki
                cursor.execute(
                    "UPDATE orders SET status = 'manjkajo_podatki', missing_data_details = %s WHERE order_number = %s",
                    (json.dumps(missing), order_number_with_hash)
                )
                db.commit()
                # Pošljemo obvestilo administratorju
                from services.email_service import poslji_obvestilo_o_narocilu_z_manjkajocimi_podatki
                poslji_obvestilo_o_narocilu_z_manjkajocimi_podatki(
                    order_number=order_number_with_hash,
                    missing_data_details=missing,
                    customer_email=order.get('customer_email'),
                    shopify_order_id=order.get('shopify_order_id')
                )
                return jsonify({"sporocilo": f"Manjkajo podatki: {', '.join(missing)}"}), 400
            if warnings:
                return jsonify({
                    "sporocilo": "PDF ni bil generiran, ker imajo nekateri izdelki rok uporabe manj kot 60 dni.",
                    "warnings": warnings
                }), 400
        # Shranimo podatke v declarations tabelo
        current_app.logger.info(f"Shranjujem podatke v declarations tabelo za naročilo {order_number_with_hash}")
        success = _shrani_deklaracijo_v_bazo(order_number_with_hash, declaration_items, cursor)
        if not success:
            return jsonify({"sporocilo": "Napaka pri shranjevanju deklaracije."}), 500

        # Klic PDF servisa (po uspešnem shranjevanju deklaracije)
            pdf_path, pdf_msg = ustvari_pdf(declaration_items, email_line_items, order['country_code'], order_number_with_hash, warnings)
            if not pdf_path: 
                # Posodobimo status naročila na manjkajo_podatki
                cursor.execute(
                    "UPDATE orders SET status = 'manjkajo_podatki', missing_data_details = %s WHERE order_number = %s",
                    (json.dumps([pdf_msg]), order_number_with_hash)
                )
                db.commit()

                # Pošljemo obvestilo administratorju
                from services.email_service import poslji_obvestilo_o_narocilu_z_manjkajocimi_podatki
                poslji_obvestilo_o_narocilu_z_manjkajocimi_podatki(
                    order_number=order_number_with_hash,
                    missing_data_details=[f"Napaka pri generiranju PDF-ja: {pdf_msg}"],
                    customer_email=order.get('customer_email'),
                    shopify_order_id=order.get('shopify_order_id')
                )

                return jsonify({"sporocilo": pdf_msg}), 500
        
            # PDF je bil uspešno generiran - posodobimo pdf_generated_at
            cursor.execute(
                "UPDATE orders SET pdf_generated_at = NOW() WHERE order_number = %s",
                (order_number_with_hash,)
            )
            db.commit()
        
        # Preveri Mandrill — pošiljanje samo prek safety net (21:00 + MK Zaključeno)
        from services.declaration_safety_net import process_one, should_wait_for_2100_pdf_batch
        order_dict = dict(order) if not isinstance(order, dict) else order
        wait, wait_msg = should_wait_for_2100_pdf_batch(order_dict)
        if wait:
            return jsonify({"sporocilo": wait_msg}), 409

        send_result = process_one(order_dict, cursor)
        db.commit()
        if send_result.get('action') == 'uploaded_and_mandrill':
            return jsonify({
                "sporocilo": "Deklaracija poslana prek Mandrill!",
                "mandrill_message_id": send_result.get('mandrill_message_id'),
            })
        return jsonify({
            "sporocilo": send_result.get('reason') or "Pošiljanje ni uspelo.",
            "action": send_result.get('action'),
        }), 409 if send_result.get('action') in ('waiting_2100_batch', 'uploaded_mk_only') else 500
            
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri generiranju in pošiljanju: {e}")
        return jsonify({"sporocilo": "Prišlo je do napake na strežniku."}), 500
    finally:
        cursor.close()
@api_bp.route('/ponovno_poslji_deklaracijo', methods=['POST'])
def ponovno_poslji_deklaracijo():
    order_number = request.json.get('order_number')
    nov_email = request.json.get('nov_email')
    current_app.logger.info(f"Ponovno pošiljanje deklaracije za naročilo: '{order_number}' z novim emailom: '{nov_email}'")
    
    # Direktno preverjanje v bazi
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute("SELECT COUNT(*) as count FROM declarations WHERE order_number = %s", (order_number,))
        count_result = cursor.fetchone()
        current_app.logger.info(f"Direktno preverjanje: {count_result['count']} zapisov v declarations za '{order_number}'")
        
        # Preveri oba možna formata order_number (z in brez #)
        cursor.execute("SELECT * FROM orders WHERE order_number = %s OR order_number = %s", (order_number, f"#{order_number}"))
        order = cursor.fetchone()
        if not order: 
            current_app.logger.error(f"Naročilo '{order_number}' ni najdeno v orders tabeli (preveril tudi '#{order_number}')")
            return jsonify({"sporocilo": "Naročilo ni najdeno."}), 404
        
        current_app.logger.info(f"Našel naročilo '{order_number}' v orders tabeli")
        current_app.logger.info(f"Originalni customer_email: '{order['customer_email']}'")
        
        # Preveri, ali je email že bil poslan
        if order.get('email_sent_at'):
            current_app.logger.warning(f"Email je že bil poslan za naročilo {order_number} ob {order['email_sent_at']}. Pošiljam ponovno kot zahtevano.")
        
        # Pridobimo podatke iz declarations tabele
        declaration_data = _pridobi_deklaracijo_iz_baze(order_number, cursor)
        
        if not declaration_data:
            current_app.logger.error(f"Podatki deklaracije ne obstajajo za naročilo '{order_number}'")
            return jsonify({"sporocilo": "Podatki deklaracije ne obstajajo. Najprej generirajte deklaracijo."}), 400
        
        current_app.logger.info(f"Uspešno pridobil {len(declaration_data)} zapisov iz declarations za naročilo '{order_number}'")
        
        # Pretvorimo podatke v format, ki ga pričakuje PDF servis
        declaration_items = []
        for item in declaration_data:
            declaration_items.append({
                'title': f"{item['product_no']} - {item['proizvajalec_ime']}",
                'product_no': item['product_no'],
                'proizvajalec_ime': item['proizvajalec_ime'],
                'sestava_inci': item['sestava_inci'],
                'rok_uporabe': item['rok_uporabe'].strftime('%d.%m.%Y') if hasattr(item['rok_uporabe'], 'strftime') and item['rok_uporabe'] else (str(item['rok_uporabe']) if item['rok_uporabe'] else None),
                'serijska_stevilka': item['serijska_stevilka'] or 'N/A'
            })
        
        # Pridobimo line_items za email template
        line_items_raw = order.get('line_items', '[]')
        line_items = json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])
        
        # Pridobimo podatke iz Shopify-ja za email template
        product_ids = [str(item['product_id']) for item in line_items if item and item.get('product_id')]
        clear_product_cache()
        shopify_details = get_bulk_product_details(product_ids)

        # Pripravimo podatke za email predlogo
        email_line_items = []
        for item in line_items:
            if not item or not item.get('product_id'): continue
            product_id_str = str(item.get('product_id'))
            details = shopify_details.get(product_id_str, {})
            
            try:
                price = float(item.get('price', 0.0))
            except (ValueError, TypeError):
                price = 0.0

            email_line_items.append({
                'title': item.get('title', 'N/A'),
                'quantity': item.get('quantity', 1),
                'price': price,
                'image_url': details.get('image_url', 'https://cdn.shopify.com/s/files/1/0533/2089/files/placeholder-images-image_large.png?v=1529089297')
            })
        
        # Generiramo PDF
        pdf_path, pdf_msg = ustvari_pdf(declaration_items, email_line_items, order['country_code'], order_number, [])
        if not pdf_path: return jsonify({"sporocilo": pdf_msg}), 500
        
        # Pošlji prek safety net (MK Zaključeno + 21:00 batch)
        from services.declaration_safety_net import process_one
        order_dict = dict(order) if not isinstance(order, dict) else order
        if nov_email:
            order_dict = {**order_dict, 'customer_email': nov_email}
        send_result = process_one(order_dict, cursor)
        db.commit()

        if send_result.get('action') == 'uploaded_and_mandrill':
            return jsonify({
                "sporocilo": "Deklaracija ponovno poslana prek Mandrill!",
                "mandrill_message_id": send_result.get('mandrill_message_id'),
            })
        return jsonify({
            "sporocilo": send_result.get('reason') or "Pošiljanje ni uspelo.",
            "action": send_result.get('action'),
        }), 409 if send_result.get('action') in ('waiting_2100_batch', 'uploaded_mk_only') else 500

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri ponovnem pošiljanju deklaracije: {e}")
        return jsonify({"sporocilo": "Prišlo je do napake na strežniku."}), 500
    finally:
        cursor.close()


@api_bp.route('/expiring-perfumes', methods=['GET'])
def get_expiring_perfumes():
    db = get_db()
    cursor = db.cursor()
    try:
        query_date = date.today() + timedelta(days=60)
        
        sql_query = """
            SELECT
                p.id as parfum_id,
                p.product_no,
                p.ime_parfuma,
                pr.ime as ime_proizvajalca,
                s.id as serija_id,
                s.rok_uporabe
            FROM
                serije s
            JOIN
                (SELECT parfum_id, MAX(id) as max_id FROM serije GROUP BY parfum_id) as najnovejse_serije
                ON s.parfum_id = najnovejse_serije.parfum_id AND s.id = najnovejse_serije.max_id
            JOIN
                parfumi p ON s.parfum_id = p.id
            JOIN
                proizvajalci pr ON p.proizvajalec_id = pr.id
            WHERE
                p.na_zalogi = TRUE
                AND s.rok_uporabe <= %s
            ORDER BY
                s.rok_uporabe ASC;
        """
        
        cursor.execute(sql_query, (query_date,))
        
        expiring_items = []
        for item in cursor.fetchall():
            current_app.logger.info(f"Expiring item: parfum_id={item['parfum_id']}, product_no={item['product_no']}, serija_id={item['serija_id']}, ime={item['ime_parfuma']}, proizvajalec={item['ime_proizvajalca']}")
            is_problem, message = preveri_rok_uporabe(item['rok_uporabe'])
            if is_problem: 
                # Preverimo, ali je rok_uporabe datum objekt ali string
                rok_uporabe_str = item['rok_uporabe']
                if hasattr(rok_uporabe_str, 'strftime'):
                    rok_uporabe_str = rok_uporabe_str.strftime('%d.%m.%Y')
                elif isinstance(rok_uporabe_str, str):
                    # Če je že string, ga pustimo kot je
                    pass
                else:
                    rok_uporabe_str = str(rok_uporabe_str) if rok_uporabe_str else 'N/A'
                
                expiring_item = {
                    'ime': item['ime_parfuma'],
                    'proizvajalec': item['ime_proizvajalca'],
                    'parfum_id': item['parfum_id'],
                    'product_no': item['product_no'],
                    'serija_id': item['serija_id'],
                    'rok_uporabe': rok_uporabe_str, 
                    'opozorilo': message
                }
                current_app.logger.info(f"Adding expiring item to response: {expiring_item}")
                expiring_items.append(expiring_item)
        
        current_app.logger.info(f"Returning {len(expiring_items)} expiring items")
        return jsonify(expiring_items)
    except Exception as e:
        current_app.logger.error(f"Napaka v /expiring-perfumes: {e}")
        traceback.print_exc()
        return jsonify({"error": "Napaka pri pridobivanju podatkov o rokih uporabe."}), 500
    finally:
        cursor.close()

@api_bp.route('/proizvajalci', methods=['GET'])
def get_proizvajalci():
    supplier = (request.args.get('supplier') or '').upper().strip()
    db = get_db(); cursor = db.cursor()
    if supplier in ('FLORGARDEN','MISTRAL'):
        cursor.execute("SELECT id, ime FROM proizvajalci WHERE UPPER(ime) = %s ORDER BY ime ASC", (supplier,))
    else:
        cursor.execute("SELECT id, ime FROM proizvajalci ORDER BY ime ASC")
    proizvajalci = cursor.fetchall(); cursor.close()
    return jsonify(proizvajalci)

@api_bp.route('/parfumi', methods=['GET'])
def get_vsi_parfumi():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT p.id, p.product_no, p.ime_parfuma, p.na_zalogi, pr.ime as ime_proizvajalca FROM parfumi p JOIN proizvajalci pr ON p.proizvajalec_id = pr.id ORDER BY pr.ime, p.ime_parfuma")
    parfumi = cursor.fetchall()
    cursor.close()
    return jsonify(parfumi)

@api_bp.route('/parfum-search', methods=['GET'])
def parfum_search():
    q = (request.args.get('q') or '').strip()
    limit = int(request.args.get('limit') or 200)
    if len(q) < 2 and not q.isdigit():
        return jsonify([])
    db = get_db()
    cursor = db.cursor()
    try:
        import re, unicodedata
        like = f"%{q}%"
        q_norm = unicodedata.normalize("NFKD", q).encode("ascii", "ignore").decode("ascii")
        q_norm = re.sub(r"[^a-zA-Z0-9]+", "", q_norm).lower()
        like_norm = f"%{q_norm}%"
        q_digits = q if q.isdigit() else None
        q_int = int(q_digits) if q_digits is not None else None
        cursor.execute(
            """
            SELECT p.id, p.product_no, p.ime_parfuma, pr.ime AS proizvajalec
            FROM parfumi p
            JOIN proizvajalci pr ON pr.id = p.proizvajalec_id
            WHERE (p.product_no::text ILIKE %s)
               OR (p.ime_parfuma ILIKE %s)
               OR (pr.ime ILIKE %s)
               OR (p.product_no::text = %s)
               OR (%s <> '' AND regexp_replace(lower(p.ime_parfuma), '[^a-z0-9]+', '', 'g') LIKE %s)
               OR (%s <> '' AND regexp_replace(lower(pr.ime), '[^a-z0-9]+', '', 'g') LIKE %s)
            ORDER BY
                CASE WHEN p.product_no::text = %s THEN 0 ELSE 1 END,
                CASE
                    WHEN %s IS NOT NULL AND p.product_no::text ~ '^[0-9]+$'
                    THEN ABS(p.product_no::int - %s)
                    ELSE NULL
                END ASC NULLS LAST,
                p.product_no ASC,
                p.ime_parfuma ASC
            LIMIT %s
            """,
            (like, like, like, q_digits, q_norm, like_norm, q_norm, like_norm, q_digits, q_int, q_int, limit)
        )
        rows = cursor.fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        cursor.close()

@api_bp.route('/parfumi_by_proizvajalec/<int:proizvajalec_id>', methods=['GET'])
def get_parfumi_by_proizvajalec(proizvajalec_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, product_no, ime_parfuma, na_zalogi, proizvajalec_id FROM parfumi WHERE proizvajalec_id = %s ORDER BY product_no ASC", (proizvajalec_id,))
    parfumi = cursor.fetchall()
    cursor.close()
    return jsonify(parfumi)

@api_bp.route('/parfum/<int:perfume_id>', methods=['GET'])
def get_perfume_details_api(perfume_id):
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute("""
            SELECT p.*, pr.ime as ime_proizvajalca 
            FROM parfumi p 
            JOIN proizvajalci pr ON p.proizvajalec_id = pr.id 
            WHERE p.id = %s
        """, (perfume_id,))
        parfum = cursor.fetchone()
        if not parfum:
            return jsonify({"error": "Parfum ni najden."}), 404

        parfum_dict = dict(parfum)
        
        gid, find_msg = find_shopify_product_gid(parfum['product_no'], parfum['proizvajalec_id'])
        parfum_dict['exists_in_shopify'] = gid is not None
        
        parfum_dict['ime_parfuma_shopify'] = parfum_dict['ime_parfuma']

        if gid:
            shopify_details = get_single_product_details_for_display(gid)
            if shopify_details:
                vendor = shopify_details.get('vendor', '')
                fragrance = shopify_details.get('product_fragrance', '')
                if vendor and fragrance:
                    parfum_dict['ime_parfuma_shopify'] = f"{vendor} - {fragrance}"
                elif vendor:
                    parfum_dict['ime_parfuma_shopify'] = vendor
        else:
            current_app.logger.warning(f"Could not find GID for perfume {perfume_id} to fetch display name: {find_msg}")

        return jsonify(parfum_dict)

    except Exception as e:
        current_app.logger.error(f"Error in get_perfume_details_api for perfume_id {perfume_id}: {e}")
        traceback.print_exc()
        return jsonify({"error": "Napaka na strežniku pri pridobivanju podrobnosti."}), 500
    finally:
        cursor.close()
@api_bp.route('/parfumi', methods=['POST'])
def add_or_update_perfume():
    data = request.get_json()
    perfume_id = data.get('id')
    product_no = data.get('product_no')
    proizvajalec_id = data.get('proizvajalec_id')
    ime_parfuma = data.get('ime_parfuma')
    sestava_inci = data.get('sestava_inci')
    sinhroniziraj = data.get('sinhroniziraj_s_shopify', False)
    na_zalogi = data.get('na_zalogi', False)

    # Pri posodobitvi (ko obstaja ID) dopolni manjkajoča polja iz baze,
    # da lahko uporabniki z delnimi dovoljenji (npr. samo zaloga) shranijo spremembe
    try:
        conn, cursor = get_db(), get_db().cursor()
        if perfume_id:
            cursor.execute("""
                SELECT id, product_no, proizvajalec_id, ime_parfuma, sestava_inci, sinhroniziraj_s_shopify, na_zalogi
                FROM parfumi
                WHERE id = %s
            """, (perfume_id,))
            existing = cursor.fetchone()
            if not existing:
                cursor.close()
                return jsonify({"error": "Parfum ni najden."}), 404

            # Če polja manjkajo v payloadu, uporabi obstoječe vrednosti
            product_no = product_no or existing.get('product_no')
            proizvajalec_id = proizvajalec_id or existing.get('proizvajalec_id')
            ime_parfuma = ime_parfuma or existing.get('ime_parfuma')
            sestava_inci = sestava_inci if sestava_inci is not None else existing.get('sestava_inci')
            sinhroniziraj = sinhroniziraj if sinhroniziraj is not None else existing.get('sinhroniziraj_s_shopify')
            na_zalogi = na_zalogi if na_zalogi is not None else existing.get('na_zalogi')

        else:
            # Pri kreaciji zahtevaj obvezna polja
            if not all([product_no, proizvajalec_id, ime_parfuma]): 
                cursor.close()
                return jsonify({"error": "Manjkajo obvezni podatki."}), 400
    except Exception as e:
        current_app.logger.error(f"Napaka pri branju obstoječega parfuma: {e}")
        try:
            cursor.close()
        except Exception:
            pass
        return jsonify({"error": "Napaka na strežniku."}), 500
    
    try:
        
        if perfume_id: 
            cursor.execute("""
                UPDATE parfumi 
                SET product_no = %s, proizvajalec_id = %s, ime_parfuma = %s, sestava_inci = %s, sinhroniziraj_s_shopify = %s, na_zalogi = %s
                WHERE id = %s
            """, (product_no, proizvajalec_id, ime_parfuma, sestava_inci, sinhroniziraj, na_zalogi, perfume_id))
        else: 
            cursor.execute("""
                INSERT INTO parfumi (product_no, proizvajalec_id, ime_parfuma, sestava_inci, sinhroniziraj_s_shopify, na_zalogi) 
                VALUES (%s, %s, %s, %s, %s, %s) 
                ON CONFLICT (product_no, proizvajalec_id) DO UPDATE SET 
                    ime_parfuma = EXCLUDED.ime_parfuma, 
                    sestava_inci = EXCLUDED.sestava_inci, 
                    sinhroniziraj_s_shopify = EXCLUDED.sinhroniziraj_s_shopify,
                    na_zalogi = EXCLUDED.na_zalogi
                RETURNING id
            """, (product_no, proizvajalec_id, ime_parfuma, sestava_inci, sinhroniziraj, na_zalogi))
            
            # Pridobi ID novega/posodobljenega parfuma
            result = cursor.fetchone()
            perfume_id = result['id'] if result else None
        
        conn.commit()

        # Safety net: vnos/popravek INCI ali metafield-ov lahko odblokira
        # naročila, ki so bila blokirana z 'missing_inci' ali 'parfum_not_in_db'.
        # Sprosti njihove block flag-e, da jih naslednji safety net cron retry-a.
        if perfume_id and sestava_inci:
            try:
                from services.declaration_safety_net import (
                    invalidate_blocks_for_parfum,
                    CODE_MISSING_INCI,
                    CODE_PARFUM_NOT_IN_DB,
                    CODE_MISSING_METAFIELDS,
                )
                unblocked = invalidate_blocks_for_parfum(
                    perfume_id,
                    codes=[CODE_MISSING_INCI, CODE_PARFUM_NOT_IN_DB, CODE_MISSING_METAFIELDS],
                )
                if unblocked > 0:
                    current_app.logger.info(
                        f"Parfum update: invalidated {unblocked} blocked orders for perfume_id={perfume_id}"
                    )
            except Exception as e:
                current_app.logger.error(f"Parfum update: safety net invalidation failed: {e}")
        
        sync_msg = ""
        if sinhroniziraj:
            cursor.execute("SELECT ime FROM proizvajalci WHERE id = %s", (proizvajalec_id,))
            proizvajalec = cursor.fetchone()
            if not proizvajalec: 
                sync_msg = " Proizvajalec ni najden za Shopify sinhronizacijo."
            else:
                # Poskusi sinhronizacijo na vseh konfiguriranih trgovinah (vključno z default)
                stores = get_all_shopify_stores(include_default=True)
                any_found = False
                per_store_msgs: list[str] = []
                for store in stores:
                    domain = store.get('shop_domain')
                    try:
                        gid, find_msg = find_shopify_product_gid(product_no, proizvajalec_id, shop_domain=domain)
                        if gid:
                            any_found = True
                            # INCI
                            success_inci, update_msg_inci = update_shopify_inci_metafield(gid, sestava_inci, shop_domain=domain)
                            # Zaloga
                            success_stock, update_msg_stock = update_stock_status_in_shopify(gid, na_zalogi, shop_domain=domain)
                            per_store_msgs.append(
                                f"{domain}: INCI {'OK' if success_inci else 'FAIL'} ({update_msg_inci}); STOCK {'OK' if success_stock else 'FAIL'} ({update_msg_stock})"
                            )
                        else:
                            per_store_msgs.append(f"{domain}: {find_msg}")
                    except Exception as ex:
                        per_store_msgs.append(f"{domain}: napaka pri sinhronizaciji ({ex})")
                if not any_found:
                    # Če izdelek ne obstaja v nobeni trgovini, izklopi sinhronizacijo
                    cursor.execute("""
                        UPDATE parfumi 
                        SET sinhroniziraj_s_shopify = FALSE
                        WHERE id = %s
                    """, (perfume_id,))
                    conn.commit()
                    sync_msg = " Shopify sinhronizacija avtomatsko izklopljena: izdelek ni najden v nobeni Shopify trgovini."
                else:
                    sync_msg = " " + " | ".join(per_store_msgs)
        else:
            sync_msg = " Izdelek ni označen za sinhronizacijo s Shopify."

        cursor.close()
        return jsonify({"message": f"Parfum uspešno shranjen!{sync_msg}"}), 201
    except Exception as e: 
        current_app.logger.error(f"Napaka v /parfumi POST: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Napaka na strežniku: {e}"}), 500

@api_bp.route('/parfum/<int:perfume_id>/stock-status', methods=['POST'])
def update_stock_status(perfume_id):
    data = request.get_json()
    is_in_stock = data.get('na_zalogi', False)

    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute("""
            SELECT p.product_no, pr.ime as ime_proizvajalca, p.sinhroniziraj_s_shopify
            FROM parfumi p 
            JOIN proizvajalci pr ON p.proizvajalec_id = pr.id 
            WHERE p.id = %s
        """, (perfume_id,))
        parfum_info = cursor.fetchone()

        if not parfum_info:
            return jsonify({"error": "Parfum ni najden v lokalni bazi."}), 404

        if parfum_info['sinhroniziraj_s_shopify']:
            local_product_no = parfum_info.get('product_no')
            local_proizvajalec_ime = parfum_info.get('ime_proizvajalca')

            # Pridobi proizvajalec_id iz imena proizvajalca
            cursor.execute("SELECT id FROM proizvajalci WHERE ime = %s", (local_proizvajalec_ime,))
            proizvajalec_result = cursor.fetchone()
            if not proizvajalec_result:
                return jsonify({"error": f"Proizvajalec '{local_proizvajalec_ime}' ni najden v bazi."}), 400
            
            local_proizvajalec_id = proizvajalec_result['id']
            gid, find_msg = find_shopify_product_gid(str(local_product_no).strip(), local_proizvajalec_id)
            
            if not gid:
                return jsonify({"error": f"Napaka pri iskanju izdelka v Shopify: {find_msg}"}), 400

            success, sync_msg = update_stock_status_in_shopify(gid, is_in_stock)
            if not success:
                return jsonify({"error": f"Napaka pri sinhronizaciji s Shopify: {sync_msg}. Sprememba ni bila shranjena."}), 500
            
            cursor.execute("UPDATE parfumi SET na_zalogi = %s WHERE id = %s", (is_in_stock, perfume_id))
            db.commit()
            return jsonify({"message": f"Status zaloge uspešno posodobljen! {sync_msg}"})

        else:
            cursor.execute("UPDATE parfumi SET na_zalogi = %s WHERE id = %s", (is_in_stock, perfume_id))
            db.commit()
            return jsonify({"message": "Status zaloge uspešno posodobljen (samo lokalno)."})

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri posodabljanju statusa zaloge za perfume_id {perfume_id}: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do napake na strežniku."}), 500
    finally:
        cursor.close()

@api_bp.route('/parfum/<int:perfume_id>/sync-status', methods=['POST'])
def update_sync_status(perfume_id):
    data = request.get_json()
    sync_with_shopify = data.get('sinhroniziraj_s_shopify', False)

    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            "UPDATE parfumi SET sinhroniziraj_s_shopify = %s WHERE id = %s",
            (sync_with_shopify, perfume_id)
        )
        db.commit()
        
        if cursor.rowcount == 0:
            return jsonify({"error": "Parfum ni bil najden."}), 404

        message = "Sinhronizacija s Shopify je vklopljena." if sync_with_shopify else "Sinhronizacija s Shopify je izklopljena."
        return jsonify({"message": message})

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri posodabljanju statusa sinhronizacije za perfume_id {perfume_id}: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do napake na strežniku."}), 500
    finally:
        cursor.close()

def _to_iso(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    return str(val)

@api_bp.route('/parfum/<int:perfume_id>/serije', methods=['GET'])
def get_serije_for_perfume(perfume_id):
    db = get_db()
    cursor = db.cursor()
    # Cast date/timestamp columns to text to avoid driver errors on invalid dates.
    cursor.execute("""
        SELECT
            s.id,
            s.excel_row_id,
            s.parfum_id,
            s.rok_uporabe::text AS rok_uporabe,
            s.serijska_stevilka,
            s.stanje,
            s.datum_odprtja::text AS datum_odprtja,
            s.je_tester,
            s.vnesel_uporabnik,
            s.created_at_original::text AS created_at_original,
            s.created_at::text AS created_at,
            s.updated_at::text AS updated_at,
            s.updated_by,
            pr.ime as ime_proizvajalca
        FROM serije s
        JOIN parfumi p ON s.parfum_id = p.id
        JOIN proizvajalci pr ON p.proizvajalec_id = pr.id
        WHERE s.parfum_id = %s
        ORDER BY s.id DESC
    """, (perfume_id,))
    serije = cursor.fetchall()
    cursor.close()
    
    # Dodaj informacije o dovoljenjih za vsako serijo
    for serija in serije:
        if serija.get('rok_uporabe') is not None:
            serija['rok_uporabe'] = _to_iso(serija['rok_uporabe'])
        if serija.get('datum_odprtja') is not None:
            serija['datum_odprtja'] = _to_iso(serija['datum_odprtja'])
        if serija.get('updated_at') is not None:
            serija['updated_at'] = _to_iso(serija['updated_at'])
        
        # Preveri dovoljenja za to serijo posebej za update in delete
        perm_update = check_serija_permissions(serija['id'], 'update')
        perm_delete = check_serija_permissions(serija['id'], 'delete')
        serija['can_edit'] = perm_update['allowed']
        serija['can_delete'] = perm_delete['allowed']
        serija['permission_reason'] = '' if perm_update['allowed'] else perm_update['reason']
    
    return jsonify(serije)

@api_bp.route('/serije', methods=['GET'])
def get_all_serije():
    """Pridobi vse serije za Local First sinhronizacijo"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Cast date/timestamp columns to text to avoid driver errors on invalid dates.
        cursor.execute("""
            SELECT
                s.id,
                s.excel_row_id,
                s.parfum_id,
                s.rok_uporabe::text AS rok_uporabe,
                s.serijska_stevilka,
                s.stanje,
                s.datum_odprtja::text AS datum_odprtja,
                s.je_tester,
                s.vnesel_uporabnik,
                s.created_at_original::text AS created_at_original,
                s.created_at::text AS created_at,
                s.updated_at::text AS updated_at,
                s.updated_by,
                p.ime_parfuma,
                pr.ime as ime_proizvajalca
            FROM serije s
            LEFT JOIN parfumi p ON s.parfum_id = p.id
            LEFT JOIN proizvajalci pr ON p.proizvajalec_id = pr.id
            ORDER BY s.created_at DESC
        """)
        
        serije = cursor.fetchall()
        
        # Pretvori v seznam slovarjev
        serije_list = []
        for serija in serije:
            serija_dict = dict(serija)
            # Pretvori datume v ISO format
            if serija_dict.get('created_at') is not None:
                serija_dict['created_at'] = _to_iso(serija_dict['created_at'])
            if serija_dict.get('datum_odprtja') is not None:
                serija_dict['datum_odprtja'] = _to_iso(serija_dict['datum_odprtja'])
            if serija_dict.get('rok_uporabe') is not None:
                serija_dict['rok_uporabe'] = _to_iso(serija_dict['rok_uporabe'])
            if serija_dict.get('updated_at') is not None:
                serija_dict['updated_at'] = _to_iso(serija_dict['updated_at'])
            if serija_dict.get('created_at_original') is not None:
                serija_dict['created_at_original'] = _to_iso(serija_dict['created_at_original'])
            serije_list.append(serija_dict)
        
        return jsonify(serije_list)
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri pridobivanju serij: {e}")
        return jsonify({"error": "Napaka pri pridobivanju serij"}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
@api_bp.route('/serije', methods=['POST'])
def add_serija():
    data = request.get_json()
    try:
        db, cursor = get_db(), get_db().cursor()
        
        # Pridobi ime trenutnega uporabnika iz session ali localStorage
        current_user = session.get('user')
        current_app.logger.info(f"Session user data: {current_user}")
        vnesel_uporabnik = None
        
        # Poskusi iz session-a
        if current_user:
            if current_user.get('first_name') and current_user.get('last_name'):
                vnesel_uporabnik = f"{current_user['first_name']} {current_user['last_name']}"
                current_app.logger.info(f"Using session full name: {vnesel_uporabnik}")
            else:
                vnesel_uporabnik = current_user.get('username', 'Neznan uporabnik')
                current_app.logger.info(f"Using session username: {vnesel_uporabnik}")
        
        # Fallback: poskusi iz request headers (če frontend pošlje podatke)
        if not vnesel_uporabnik or vnesel_uporabnik == 'Neznan uporabnik':
            user_header = request.headers.get('X-User-Info')
            if user_header:
                try:
                    raw = user_header
                    if isinstance(raw, str) and raw.startswith('b64:'):
                        import base64
                        raw = base64.b64decode(raw[4:]).decode('utf-8', errors='ignore')
                    user_info = json.loads(raw)
                    if user_info.get('first_name') and user_info.get('last_name'):
                        vnesel_uporabnik = f"{user_info['first_name']} {user_info['last_name']}"
                        current_app.logger.info(f"Using header full name: {vnesel_uporabnik}")
                    else:
                        vnesel_uporabnik = user_info.get('username', 'Neznan uporabnik')
                        current_app.logger.info(f"Using header username: {vnesel_uporabnik}")
                except Exception:
                    pass
        
        # Final fallback
        if not vnesel_uporabnik or vnesel_uporabnik == 'Neznan uporabnik':
            vnesel_uporabnik = 'Neznan uporabnik'
            current_app.logger.warning("No user data found in session or headers")
        
        # Preveri proizvajalca in validiraj serijsko številko
        cursor.execute(
            """
            SELECT p.product_no, p.proizvajalec_id, pr.ime as proizvajalec_ime 
            FROM parfumi p 
            JOIN proizvajalci pr ON p.proizvajalec_id = pr.id 
            WHERE p.id = %s
            """,
            (data['parfum_id'],)
        )
        
        parfum_info = cursor.fetchone()
        if not parfum_info:
            return jsonify({"error": "Parfum ni bil najden"}), 404
        
        # Procurement info
        product_no = parfum_info['product_no'] if isinstance(parfum_info, dict) else parfum_info[0]
        proizvajalec_id = parfum_info['proizvajalec_id'] if isinstance(parfum_info, dict) else parfum_info[1]
        proizvajalec_ime = (parfum_info['proizvajalec_ime'] if isinstance(parfum_info, dict) else parfum_info[2]).upper()
        is_mistral = proizvajalec_ime == 'MISTRAL'
        is_florgarden = proizvajalec_ime == 'FLORGARDEN'
        serijska_stevilka = data.get('serijska_stevilka')
        
        # Za MISTRAL prepreči vnos serijske številke
        if is_mistral and serijska_stevilka and serijska_stevilka.strip():
            return jsonify({"error": "Za MISTRAL proizvajalca se serijska številka ne sme vnašati"}), 400
        
        # Za FLORGARDEN je serijska številka obvezna
        if is_florgarden and (not serijska_stevilka or not serijska_stevilka.strip()):
            return jsonify({"error": "Za FLORGARDEN proizvajalca je serijska številka obvezna"}), 400
        
        # Za FLORGARDEN preveri format serijske številke
        if is_florgarden and serijska_stevilka:
            import re
            pattern = r'^\d{2}/\d{5}\s+\d{3}/\d{4}$'
            if not re.match(pattern, serijska_stevilka.strip()):
                return jsonify({"error": "Serijska številka mora biti v formatu: YY/AAAAA BBB/DDMM (samo številke)"}), 400
        
        # Če je MISTRAL, nastavi serijsko številko na None
        if is_mistral:
            serijska_stevilka = None
        
        cursor.execute("INSERT INTO serije (parfum_id, rok_uporabe, serijska_stevilka, datum_odprtja, je_tester, vnesel_uporabnik) VALUES (%s, %s, %s, %s, %s, %s)", 
                       (data['parfum_id'], data['rok_uporabe'], serijska_stevilka, data.get('datum_odprtja'), data['je_tester'], vnesel_uporabnik))
        
        # --- Procurement: ob vnosu serije (iztočena plastenka) posodobi zalogo ---
        try:
            # zagotovi zapis v perfumes_stock
            cursor.execute(
                """
                INSERT INTO perfumes_stock (product_no, proizvajalec_id, on_hand, on_order_pending, on_order_committed)
                VALUES (%s, %s, 0, 0, 0)
                ON CONFLICT (product_no, proizvajalec_id) DO NOTHING
                """,
                (product_no, proizvajalec_id)
            )
            # iztočena plastenka: on_hand -= 1, pending += 1
            cursor.execute(
                """
                UPDATE perfumes_stock
                SET on_hand = GREATEST(0, on_hand - 1),
                    on_order_pending = on_order_pending + 1,
                    updated_at = NOW()
                WHERE product_no = %s AND proizvajalec_id = %s
                """,
                (product_no, proizvajalec_id)
            )
            # Če po zmanjšanju on_hand pade pod min_on_hand, dodatno povečaj pending do praga
            cursor.execute(
                """
                UPDATE perfumes_stock
                SET on_order_pending = on_order_pending + GREATEST(0, (min_on_hand - on_hand))
                WHERE product_no = %s AND proizvajalec_id = %s AND on_hand < min_on_hand
                """,
                (product_no, proizvajalec_id)
            )
        except Exception as pe:
            current_app.logger.error(f"Procurement stock update failed on serija insert: {pe}")
        
        db.commit()

        # Safety net: vnos nove serije lahko odblokira naročila, ki so čakala
        # na 'expired_serije'. Sprosti njihov pdf_generation_blocked_reason, da
        # jih naslednji cron retry-a.
        try:
            from services.declaration_safety_net import invalidate_blocks_for_parfum, CODE_EXPIRED_SERIJE
            unblocked = invalidate_blocks_for_parfum(data['parfum_id'], codes=[CODE_EXPIRED_SERIJE])
            if unblocked > 0:
                current_app.logger.info(
                    f"Serija add: invalidated {unblocked} blocked orders for parfum_id={data['parfum_id']}"
                )
        except Exception as e:
            current_app.logger.error(f"Serija add: safety net invalidation failed: {e}")

        try:
            _trigger_reconcile_after_series_update()
        except Exception as e:
            current_app.logger.error(f"Serija add: reconcile trigger failed: {e}")

        cursor.close()
        return jsonify({"message": "Serija uspešno dodana."}), 201
    except Exception as e: 
        db.rollback()
        current_app.logger.error(f"Napaka v /serije POST: {e}")
        return jsonify({"error": str(e)}), 500


def _trigger_reconcile_after_series_update() -> None:
    """Run reconciliation in background after series add/update."""
    app_obj = current_app._get_current_object()

    def _run(app_obj):
        try:
            with app_obj.app_context():
                from services.background_service import reconcile_missing_declarations
                reconcile_missing_declarations(hours_back=24 * 365, limit=1000)
        except Exception as e:
            current_app.logger.error(f"Series reconcile background error: {e}")

    threading.Thread(target=_run, args=(app_obj,), daemon=True).start()

@api_bp.route('/serije/<int:serija_id>', methods=['PUT'])
def update_serija(serija_id):
    data = request.get_json()
    
    # Preveri dovoljenja
    permissions = check_serija_permissions(serija_id, 'update')
    if not permissions['allowed']:
        return jsonify({"error": permissions['reason']}), 403
    
    try:
        db, cursor = get_db(), get_db().cursor()
        
        # Preveri proizvajalca in validiraj serijsko številko
        cursor.execute("""
            SELECT pr.ime as proizvajalec_ime 
            FROM serije s
            JOIN parfumi p ON s.parfum_id = p.id 
            JOIN proizvajalci pr ON p.proizvajalec_id = pr.id 
            WHERE s.id = %s
        """, (serija_id,))
        
        serija_info = cursor.fetchone()
        if not serija_info:
            return jsonify({"error": "Serija ni bila najdena"}), 404
        
        proizvajalec_ime = serija_info['proizvajalec_ime'].upper()
        is_mistral = proizvajalec_ime == 'MISTRAL'
        is_florgarden = proizvajalec_ime == 'FLORGARDEN'
        serijska_stevilka = data.get('serijska_stevilka')
        
        # Za MISTRAL prepreči vnos serijske številke
        if is_mistral and serijska_stevilka and serijska_stevilka.strip():
            return jsonify({"error": "Za MISTRAL proizvajalca se serijska številka ne sme vnašati"}), 400
        
        # Za FLORGARDEN je serijska številka obvezna
        if is_florgarden and (not serijska_stevilka or not serijska_stevilka.strip()):
            return jsonify({"error": "Za FLORGARDEN proizvajalca je serijska številka obvezna"}), 400
        
        # Za FLORGARDEN preveri format serijske številke
        if is_florgarden and serijska_stevilka:
            import re
            pattern = r'^\d{2}/\d{5}\s+\d{3}/\d{4}$'
            if not re.match(pattern, serijska_stevilka.strip()):
                return jsonify({"error": "Serijska številka mora biti v formatu: YY/AAAAA BBB/DDMM (samo številke)"}), 400
        
        # Če je MISTRAL, nastavi serijsko številko na None
        if is_mistral:
            serijska_stevilka = None
        
        # Poskrbi, da audit stolpca obstajata tudi, če migracija še ni bila izvedena
        try:
            cursor.execute("ALTER TABLE serije ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP WITH TIME ZONE")
            cursor.execute("ALTER TABLE serije ADD COLUMN IF NOT EXISTS updated_by VARCHAR(255)")
        except Exception:
            pass

        # Zabeleži kdo in kdaj je urejal
        updated_by = None
        current_user = session.get('user', {})
        if current_user.get('first_name') and current_user.get('last_name'):
            updated_by = f"{current_user['first_name']} {current_user['last_name']}"
        else:
            updated_by = session.get('username', 'Neznan uporabnik')

        cursor.execute("UPDATE serije SET rok_uporabe = %s, serijska_stevilka = %s, datum_odprtja = %s, je_tester = %s, updated_at = CURRENT_TIMESTAMP, updated_by = %s WHERE id = %s", 
                       (data['rok_uporabe'], serijska_stevilka, data.get('datum_odprtja'), data['je_tester'], updated_by, serija_id))
        db.commit()

        try:
            _trigger_reconcile_after_series_update()
        except Exception as e:
            current_app.logger.error(f"Serija update: reconcile trigger failed: {e}")

        cursor.close()
        return jsonify({"message": "Serija uspešno posodobljena."})
    except Exception as e: 
        db.rollback()
        current_app.logger.error(f"Napaka v /serije PUT: {e}")
        return jsonify({"error": str(e)}), 500

@api_bp.route('/serije/<int:serija_id>', methods=['DELETE'])
def delete_serija(serija_id):
    # Preveri dovoljenja
    permissions = check_serija_permissions(serija_id, 'delete')
    if not permissions['allowed']:
        return jsonify({"error": permissions['reason']}), 403
    
    try:
        db, cursor = get_db(), get_db().cursor()

        # Pred brisanjem preberi podatke serije za morebitno povrnitev zaloge.
        cursor.execute(
            """
            SELECT s.created_at, p.product_no, p.proizvajalec_id, pr.ime AS proizvajalec_ime
            FROM serije s
            JOIN parfumi p ON p.id = s.parfum_id
            JOIN proizvajalci pr ON pr.id = p.proizvajalec_id
            WHERE s.id = %s
            """,
            (serija_id,)
        )
        info = cursor.fetchone()

        cursor.execute("DELETE FROM serije WHERE id = %s", (serija_id,))

        # --- Procurement: brisanje serije = plastenka se vrne v predal ---
        # on_hand += 1; iz naročila (on_order_pending) odštejemo le, če je bila
        # serija vnesena PO zadnjem oddanem naročilu (sicer je že v oddanem).
        if info:
            product_no = info['product_no'] if isinstance(info, dict) else info[1]
            proizvajalec_id = info['proizvajalec_id'] if isinstance(info, dict) else info[2]
            proizvajalec_ime = (info['proizvajalec_ime'] if isinstance(info, dict) else info[3] or '').upper()
            created_at = info['created_at'] if isinstance(info, dict) else info[0]
            if proizvajalec_ime in ('MISTRAL', 'FLORGARDEN'):
                try:
                    cursor.execute(
                        """
                        INSERT INTO perfumes_stock (product_no, proizvajalec_id, on_hand, on_order_pending, on_order_committed)
                        VALUES (%s, %s, 0, 0, 0)
                        ON CONFLICT (product_no, proizvajalec_id) DO NOTHING
                        """,
                        (product_no, proizvajalec_id)
                    )
                    cursor.execute(
                        """
                        UPDATE perfumes_stock
                        SET on_hand = on_hand + 1,
                            on_order_pending = CASE
                                WHEN %s >= COALESCE(
                                    (SELECT MAX(sent_at) FROM order_sends WHERE supplier_id = %s),
                                    'epoch'::timestamptz
                                )
                                THEN GREATEST(0, on_order_pending - 1)
                                ELSE on_order_pending
                            END,
                            updated_at = NOW()
                        WHERE product_no = %s AND proizvajalec_id = %s
                        """,
                        (created_at, proizvajalec_id, product_no, proizvajalec_id)
                    )
                except Exception as pe:
                    current_app.logger.error(f"Procurement stock update failed on serija delete: {pe}")

        db.commit()
        cursor.close()
        return jsonify({"message": "Serija uspešno izbrisana."})
    except Exception as e: 
        db.rollback()
        current_app.logger.error(f"Napaka v /serije DELETE: {e}")
        return jsonify({"error": str(e)}), 500


# --- Globalna akcija: preveži serije iz starega parfuma na novega ---
@api_bp.route('/serije/rebind', methods=['POST'])
def rebind_serije():
    try:
        if not has_permission('edit_serije'):
            return jsonify({'success': False, 'error': 'Nimate dovoljenja (edit_serije)'}), 403

        data = request.get_json(silent=True) or {}
        vendor = str(data.get('vendor', '')).strip().upper()
        old_product_no = str(data.get('old_product_no', '')).strip()
        new_product_no = str(data.get('new_product_no', '')).strip()
        if not vendor or not old_product_no or not new_product_no:
            return jsonify({'success': False, 'error': 'Manjka vendor ali product_no (stara/nova)'}), 400

        db = get_db(); c = db.cursor()
        # Najdi proizvajalec_id
        c.execute("SELECT id FROM proizvajalci WHERE UPPER(ime) = %s", (vendor,))
        row = c.fetchone()
        if not row:
            return jsonify({'success': False, 'error': f'Dobavitelj {vendor} ne obstaja'}), 404
        proizvajalec_id = row[0] if not isinstance(row, dict) else row['id']

        # Najdi oba parfuma
        c.execute("SELECT id FROM parfumi WHERE product_no = %s AND proizvajalec_id = %s", (old_product_no, proizvajalec_id))
        r_old = c.fetchone()
        c.execute("SELECT id FROM parfumi WHERE product_no = %s AND proizvajalec_id = %s", (new_product_no, proizvajalec_id))
        r_new = c.fetchone()
        if not r_old:
            return jsonify({'success': False, 'error': f'Parfum {vendor} {old_product_no} ni najden'}), 404
        if not r_new:
            return jsonify({'success': False, 'error': f'Parfum {vendor} {new_product_no} ni najden'}), 404
        old_id = r_old[0] if not isinstance(r_old, dict) else r_old['id']
        new_id = r_new[0] if not isinstance(r_new, dict) else r_new['id']
        if old_id == new_id:
            return jsonify({'success': False, 'error': 'Stari in novi parfum sta enaka'}), 400

        # Preveži serije
        c.execute("UPDATE serije SET parfum_id = %s WHERE parfum_id = %s", (new_id, old_id))
        moved = c.rowcount or 0
        db.commit(); c.close()
        return jsonify({'success': True, 'moved': moved, 'old_parfum_id': old_id, 'new_parfum_id': new_id})
    except Exception as e:
        try:
            get_db().rollback()
        except Exception:
            pass
        current_app.logger.error(f"rebind_serije error: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri prevezavi serij'}), 500
def create_manual_order():
    try:
        # Dovoli adminom oz. uporabnikom z ogledom naročil (admini imajo to pravico)
        if not has_permission('view_orders'):
            return jsonify({'success': False, 'error': 'Nimate dovoljenja'}), 403

        data = request.get_json(silent=True) or {}
        customer_email = (data.get('customer_email') or '').strip() or None
        customer_name = (data.get('customer_name') or '').strip() or None
        country_code = (data.get('country_code') or 'SI').strip().upper()
        channel = (data.get('channel') or 'manual').strip().lower()
        items = data.get('items') or []  # [{product_no, vendor, quantity}]
        mk_id_opt = (data.get('mk_id') or '').strip() or None

        # Generiraj številko naročila MAN-YYYYMMDD-HHMMSS
        ts = datetime.utcnow().strftime('%Y%m%d-%H%M%S')
        order_number = f"MAN-{ts}"

        # Pretvori v line_items format za skladnost (title, quantity, price=0)
        norm_items = []
        for it in items:
            try:
                norm_items.append({
                    'title': str(it.get('product_no') or ''),
                    'vendor': (it.get('vendor') or '').upper(),
                    'quantity': int(it.get('quantity') or 1),
                    'price': 0.0,
                    '_channel': channel,
                })
            except Exception:
                continue

        db = get_db(); c = db.cursor()
        c.execute(
            """
            INSERT INTO orders (order_number, customer_email, customer_name, country_code, line_items, created_at)
            VALUES (%s, %s, %s, %s, %s::jsonb, NOW())
            RETURNING id
            """,
            (order_number, customer_email, customer_name, country_code, _json.dumps(norm_items))
        )
        oid = c.fetchone()[0]
        # opcijsko zapiši mk_bill_id
        if mk_id_opt:
            c.execute("UPDATE orders SET mk_bill_id = %s WHERE id = %s", (mk_id_opt, oid))
        db.commit(); c.close()
        return jsonify({'success': True, 'order_number': order_number, 'id': oid})
    except Exception as e:
        try:
            get_db().rollback()
        except Exception:
            pass
        current_app.logger.error(f"create_manual_order error: {e}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': 'Napaka pri ustvarjanju naročila'}), 500


# --- Globalna akcija: uvozi ročno naročilo iz MetaKocke po mk_id ---
@api_bp.route('/mk/import-order', methods=['POST'])
def mk_import_order():
    try:
        if not has_permission('view_orders'):
            return jsonify({'success': False, 'error': 'Nimate dovoljenja'}), 403

        data = request.get_json(silent=True) or {}
        mk_id = str(data.get('mk_id', '')).strip()
        if not mk_id:
            return jsonify({'success': False, 'error': 'Manjka mk_id'}), 400

        from services.mk_service import mk_get_document
        # Najprej poskusi iz lokalne tabele tip
        doc = None; found_type = None
        try:
            db = get_db(); c = db.cursor()
            c.execute("SELECT doc_type FROM mk_bills WHERE mk_id = %s", (mk_id,))
            row = c.fetchone();
            if row:
                found_type = row[0] if not isinstance(row, dict) else row['doc_type']
                d = mk_get_document(found_type, mk_id)
                if d and d.get('mk_id'):
                    doc = d
            if not doc:
                # fallback čez tipe
                for dt in ['sales_bill_domestic','sales_bill_foreign','sales_bill','sales_bill_retail','sales_bill_prepaid','bill']:
                    d2 = mk_get_document(dt, mk_id)
                    if d2 and d2.get('mk_id'):
                        doc = d2; found_type = dt; break
        except Exception:
            pass
        if not doc:
            return jsonify({'success': False, 'error': 'Račun ni najden prek API'}), 404

        # Izlušči osnovne podatke
        customer_email = doc.get('partner_email') or doc.get('email') or None
        customer_name = doc.get('partner_name') or doc.get('buyer') or None
        country_code = (doc.get('country') or doc.get('partner_country') or 'SI')
        try:
            country_code = str(country_code).upper()[:2]
        except Exception:
            country_code = 'SI'

        # Poskusi prebrati postavke
        raw_items = doc.get('items') or doc.get('rows') or doc.get('positions') or []
        norm_items = []
        if isinstance(raw_items, list) and raw_items:
            for it in raw_items:
                try:
                    title = it.get('title') or it.get('name') or it.get('product') or it.get('sku') or 'Artikel'
                    qty = int(it.get('qty') or it.get('quantity') or it.get('kolicina') or 1)
                    price = float(it.get('price') or it.get('amount') or 0)
                    vendor = it.get('vendor') or ''
                    norm_items.append({'title': str(title), 'quantity': qty, 'price': price, 'vendor': vendor})
                except Exception:
                    continue
        if not norm_items:
            # Fallback placeholder
            norm_items = [{'title': f'MK order {mk_id}', 'quantity': 1, 'price': 0}]

        # Ustvari lokalno naročilo
        order_number = f"MK-{mk_id}"
        db2 = get_db(); c2 = db2.cursor()
        c2.execute(
            """
            INSERT INTO orders (order_number, customer_email, customer_name, country_code, line_items, created_at)
            VALUES (%s, %s, %s, %s, %s::jsonb, NOW())
            RETURNING id
            """,
            (order_number, customer_email, customer_name, country_code, _json.dumps(norm_items))
        )
        oid = c2.fetchone()[0]
        # Zapiši tudi mk bill reference (stolpci so del sheme — brez ALTER tukaj).
        c2.execute("UPDATE orders SET mk_bill_id = %s, mk_bill_type = %s WHERE id = %s", (mk_id, found_type or '', oid))
        db2.commit(); c2.close()
        return jsonify({'success': True, 'order_number': order_number, 'id': oid})
    except Exception as e:
        try:
            get_db().rollback()
        except Exception:
            pass
        current_app.logger.error(f"mk_import_order error: {e}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': 'Napaka pri uvozu naročila iz MK'}), 500
@api_bp.route('/poslji-rocno', methods=['POST'])
def send_manual_declaration():
    data = request.json
    email, perfume_ids = data.get('email'), data.get('perfumes')
    if not email or not perfume_ids: return jsonify({"sporocilo": "Manjka email ali seznam parfumov."}), 400
    db = get_db()
    cursor = db.cursor()
    try:
        perfumes_data, missing, warnings = _pridobi_podatke_za_rocno_deklaracijo(perfume_ids, cursor)
        if missing: 
            return jsonify({"sporocilo": f"Ne morem poslati deklaracije. Naslednji parfumi imajo probleme:\n\n{chr(10).join(missing)}"}), 400
        
        # --- POPRAVEK: Priprava podatkov za email predlogo ---
        manual_line_items = []
        for p in perfumes_data:
            manual_line_items.append({
                'title': p.get('title', 'N/A'),
                'quantity': 1,
                'price': 0.0, # Ročno pošiljanje nima konteksta cene
                'image_url': 'https://cdn.shopify.com/s/files/1/0533/2089/files/placeholder-images-image_large.png?v=1529089297'
            })
        # --- Konec popravka ---

        pdf_path, pdf_msg = ustvari_pdf(perfumes_data, manual_line_items, 'SI', None, warnings)
        if not pdf_path: return jsonify({"sporocilo": pdf_msg}), 500
        
        email_success = poslji_email_s_pdf(
            recipient_email=email, 
            order_number=None, 
            shopify_order_id=None, 
            pdf_path=pdf_path, 
            declaration_items=perfumes_data, 
            status_url=None, 
            shop_url=f"https://{current_app.config['SHOP_NAME']}.myshopify.com", 
            country_code='SI', 
            line_items=manual_line_items, # Uporabimo nov, pravilen seznam
            skip_test_redirect=True
        )
        
        if os.path.exists(pdf_path): os.remove(pdf_path)
        
        return jsonify({"sporocilo": "Email uspešno poslan!"}) if email_success else (jsonify({"sporocilo": "Napaka pri pošiljanju emaila."}), 500)
    except Exception as e:
        current_app.logger.error(f"Napaka v /poslji-rocno: {e}")
        traceback.print_exc()
        return jsonify({"sporocilo": f"Strežniška napaka: {str(e)}"}), 500
    finally:
        cursor.close()

@api_bp.route('/mk/sync-declaration-uploads', methods=['POST'])
def mk_sync_declaration_uploads():
    """Preveri MK dokumente in posodobi mk_decl_uploaded_at."""
    data = request.get_json(silent=True) or {}
    days_back = int(data.get('days', 7))
    limit = int(data.get('limit', 200))
    include_already = bool(data.get('include_already', False))
    order_numbers = data.get('order_numbers')
    try:
        if order_numbers and not isinstance(order_numbers, list):
            order_numbers = [order_numbers]
        result = sync_mk_declaration_uploads(
            days_back=days_back,
            limit=limit,
            include_already=include_already,
            order_numbers=order_numbers,
        )
        return jsonify({"success": True, **result})
    except Exception as e:
        current_app.logger.error(f"MK sync declaration uploads error: {e}")
        current_app.logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({"error": "Napaka pri MK sync-u deklaracij"}), 500


@api_bp.route('/admin/declarations-audit', methods=['GET'])
@require_permission('admin')
def declarations_audit():
    """Return fulfilled orders missing PDF or MK attachment in recent window."""
    days = int(request.args.get('days', 4))
    limit = int(request.args.get('limit', 500))
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            """
            SELECT order_number,
                   COALESCE(fulfilled_at, shopify_fulfilled_at) AS fulfilled_at,
                   pdf_generated_at,
                   mk_decl_uploaded_at,
                   status,
                   missing_data_details
            FROM orders
            WHERE (fulfilled_at IS NOT NULL OR shopify_fulfilled_at IS NOT NULL)
              AND COALESCE(fulfilled_at, shopify_fulfilled_at) > NOW() - (%s * INTERVAL '1 day')
              AND (pdf_generated_at IS NULL OR mk_decl_uploaded_at IS NULL)
            ORDER BY COALESCE(fulfilled_at, shopify_fulfilled_at) DESC
            LIMIT %s
            """,
            (days, limit),
        )
        rows = cursor.fetchall() or []
        return jsonify({"count": len(rows), "orders": rows})
    except Exception as e:
        current_app.logger.error(f"Declarations audit error: {e}")
        return jsonify({"error": "Napaka pri pregledu deklaracij"}), 500
    finally:
        cursor.close()

@api_bp.route('/generiraj-deklaracijo-za-tisk', methods=['POST'])
def generate_declaration_for_print():
    perfume_ids = request.json.get('perfumes')
    if not perfume_ids: return jsonify({"sporocilo": "Manjka seznam parfumov."}), 400
    db = get_db()
    cursor = db.cursor()
    try:
        perfumes_data, missing, warnings = _pridobi_podatke_za_rocno_deklaracijo(perfume_ids, cursor)
        if missing: 
            return jsonify({"sporocilo": f"Ne morem ustvariti deklaracije. Naslednji parfumi imajo probleme:\n\n{chr(10).join(missing)}"}), 400
        return render_template('print_template.html', perfumes=perfumes_data, expiration_warnings=warnings)
    except Exception as e:
        current_app.logger.error(f"Napaka v /generiraj-deklaracijo-za-tisk: {e}")
        traceback.print_exc()
        return jsonify({"sporocilo": f"Strežniška napaka: {str(e)}"}), 500
    finally:
        cursor.close()

@api_bp.route('/generiraj-pdf-rocno', methods=['POST'])
def generate_manual_pdf_for_print():
    perfume_ids = request.json.get('perfumes')
    if not perfume_ids: return jsonify({"sporocilo": "Manjka seznam parfumov."}), 400
    db = get_db()
    cursor = db.cursor()
    try:
        perfumes_data, missing, warnings = _pridobi_podatke_za_rocno_deklaracijo(perfume_ids, cursor)
        if missing: 
            return jsonify({"sporocilo": f"Ne morem ustvariti PDF-ja. Naslednji parfumi imajo probleme:\n\n{chr(10).join(missing)}"}), 400
        
        pdf_path, message = ustvari_pdf(perfumes_data, None, 'SI', None, warnings)
        if not pdf_path: return jsonify({"sporocilo": message}), 500
        
        try:
            return send_file(pdf_path, as_attachment=False, mimetype='application/pdf')
        finally:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
    except Exception as e:
        current_app.logger.error(f"Napaka v /generiraj-pdf-rocno: {e}")
        traceback.print_exc()
        return jsonify({"sporocilo": f"Strežniška napaka: {str(e)}"}), 500
    finally:
        cursor.close()

@api_bp.route('/sync-data-status', methods=['POST'])
def sync_data_status_endpoint():
    """Sinhronizira status podatkov z Shopify-jem."""
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        import threading

        def run_sync_in_background(app_instance):
            with app_instance.app_context():
                _sync_shopify_orders()
        
        # Zaženemo sinhronizacijo v ozadju
        thread = threading.Thread(target=run_sync_in_background, args=(current_app._get_current_object(),))
        thread.start()
        return jsonify({"message": "Sinhronizacija se izvaja v ozadju."})
    except Exception as e:
        current_app.logger.error(f"Napaka pri sinhronizaciji: {e}")
        return jsonify({"error": "Prišlo je do napake na strežniku."}), 500
@api_bp.route('/dodaj-iz-excel', methods=['POST'])
def dodaj_iz_excel_endpoint():
    """Endpoint za dodajanje podatkov iz Excel Database sheet-a."""
    try:
        result = _dodaj_podatke_iz_excel_database()
        return result
    except Exception as e:
        current_app.logger.error(f"Napaka pri dodajanju iz Excel: {e}")
        return {"success": False, "message": f"Napaka pri dodajanju iz Excel: {str(e)}"}
@api_bp.route('/migrate-onedrive', methods=['POST'])
def migrate_onedrive_endpoint():
    """Endpoint za popolno migracijo iz OneDrive Excel datoteke."""
    try:
        from migrations import migrate_from_onedrive
        
        current_app.logger.info("Začenjam migracijo novih serij iz OneDrive...")
        success = migrate_from_onedrive()
        
        if success:
            current_app.logger.info("Migracija novih serij iz OneDrive uspešno končana")
            return {"success": True, "message": "Migracija novih serij iz OneDrive uspešno končana"}
        else:
            current_app.logger.error("Migracija iz OneDrive ni uspela")
            return {"success": False, "message": "Migracija iz OneDrive ni uspela"}
            
    except Exception as e:
        current_app.logger.error(f"Napaka pri migraciji iz OneDrive: {e}")
        return {"success": False, "message": f"Napaka pri migraciji iz OneDrive: {str(e)}"}

@api_bp.route('/migration-status', methods=['GET'])
def migration_status():
    """Preveri status migracije."""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT COUNT(*) FROM declarations")
        declarations_count = cursor.fetchone()[0]
        
        return {"success": True, "declarations_count": declarations_count}
    except Exception as e:
        current_app.logger.error(f"Napaka pri preverjanju statusa migracije: {e}")
        return {"success": False, "error": str(e)}

@api_bp.route('/download-excel', methods=['POST'])
def download_excel_endpoint():
    """Prenesi Excel datoteko iz OneDrive na strežnik."""
    try:
        import tempfile
        import os
        import requests
        
        client_id = current_app.config.get('CLIENT_ID')
        client_secret = current_app.config.get('CLIENT_SECRET')
        tenant_id = current_app.config.get('TENANT_ID')
        excel_file_id = current_app.config.get('EXCEL_FILE_ID')
        drive_id = current_app.config.get('DRIVE_ID')
        
        if not all([client_id, client_secret, tenant_id, excel_file_id, drive_id]):
            return {"success": False, "message": "Manjkajo OneDrive konfiguracijski podatki"}
        
        current_app.logger.info("Prenašam Excel datoteko iz OneDrive...")
        
        # Pridobi access token
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        token_data = {
            'client_id': client_id,
            'client_secret': client_secret,
            'scope': 'https://graph.microsoft.com/.default',
            'grant_type': 'client_credentials'
        }
        
        token_response = requests.post(token_url, data=token_data, timeout=10)
        if not token_response.ok:
            return {"success": False, "message": f"Napaka pri pridobivanju tokena: {token_response.status_code}"}
        
        access_token = token_response.json()['access_token']
        
        # Prenesi datoteko
        download_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{excel_file_id}/content"
        headers = {'Authorization': f'Bearer {access_token}'}
        
        excel_response = requests.get(download_url, headers=headers, timeout=60)
        if not excel_response.ok:
            return {"success": False, "message": f"Napaka pri prenosu datoteke: {excel_response.status_code}"}
        
        # Shrani datoteko v static direktorij
        excel_path = os.path.join(current_app.root_path, 'static', 'DEKLARACIJE_PARFUMOV_KOPER.xlsm')
        with open(excel_path, 'wb') as f:
            f.write(excel_response.content)
        
        current_app.logger.info(f"Excel datoteka uspešno prenesena: {excel_path}")
        return {"success": True, "message": "Excel datoteka uspešno prenesena na strežnik"}
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri prenosu Excel datoteke: {e}")
        return {"success": False, "message": f"Napaka pri prenosu Excel datoteke: {str(e)}"}

@api_bp.route('/migrate-local-excel', methods=['POST'])
def migrate_local_excel_endpoint():
    """Migracija novih serij iz lokalne Excel datoteke."""
    try:
        from migrations import migrate_from_local_excel
        import os
        
        # Preverimo, ali Excel datoteka obstaja
        excel_path = '/tmp/DEKLARACIJE_PARFUMOV_KOPER.xlsm'
        if not os.path.exists(excel_path):
            return {
                "success": False, 
                "message": "Excel datoteka ne obstaja na strežniku. Najprej prenesite Excel datoteko z gumbom 'Prenesi Excel iz OneDrive'."
            }
        
        current_app.logger.info("Začenjam migracijo novih serij iz lokalne Excel datoteke...")
        success = migrate_from_local_excel()
        
        if success:
            current_app.logger.info("Migracija novih serij iz lokalne Excel datoteke uspešno končana")
            return {"success": True, "message": "Migracija novih serij iz lokalne Excel datoteke uspešno končana"}
        else:
            current_app.logger.error("Migracija iz lokalne Excel datoteke ni uspela")
            return {"success": False, "message": "Migracija iz lokalne Excel datoteke ni uspela"}
            
    except Exception as e:
        current_app.logger.error(f"Napaka pri migraciji iz lokalne Excel datoteke: {e}")
        return {"success": False, "message": f"Napaka pri migraciji iz lokalne Excel datoteke: {str(e)}"}

@api_bp.route('/migrate-local-file', methods=['POST'])
def migrate_local_file_endpoint():
    """Migracija novih serij iz lokalne Excel datoteke (podobno kot prejšnje skripte)."""
    try:
        from migrations import migrate_from_local_excel_file
        
        current_app.logger.info("Začenjam migracijo novih serij iz lokalne Excel datoteke...")
        success = migrate_from_local_excel_file()
        
        if success:
            current_app.logger.info("Migracija novih serij iz lokalne Excel datoteke uspešno končana")
            return {"success": True, "message": "Migracija novih serij iz lokalne Excel datoteke uspešno končana"}
        else:
            current_app.logger.error("Migracija iz lokalne Excel datoteke ni uspela")
            return {"success": False, "message": "Migracija iz lokalne Excel datoteke ni uspela"}
            
    except Exception as e:
        current_app.logger.error(f"Napaka pri migraciji iz lokalne Excel datoteke: {e}")
        return {"success": False, "message": f"Napaka pri migraciji iz lokalne Excel datoteke: {str(e)}"}

@api_bp.route('/sync-inci-from-shopify', methods=['POST'])
def sync_inci_from_shopify_endpoint():
    """Sinhronizira INCI podatke iz Shopify-ja za določen parfum."""
    try:
        data = request.get_json()
        product_no = data.get('product_no')
        proizvajalec_id = data.get('proizvajalec_id')
        
        if not product_no or not proizvajalec_id:
            return jsonify({"success": False, "message": "Manjkajo product_no ali proizvajalec_id"}), 400
        
        success, message = sync_inci_from_shopify(product_no, proizvajalec_id)
        
        if success:
            return jsonify({"success": True, "message": message})
        else:
            return jsonify({"success": False, "message": message}), 400
            
    except Exception as e:
        current_app.logger.error(f"Napaka pri sinhronizaciji INCI: {e}")
        return jsonify({"success": False, "message": f"Napaka: {str(e)}"}), 500

@api_bp.route('/sync-all-inci-from-shopify', methods=['POST'])
def sync_all_inci_from_shopify_endpoint():
    """Sinhronizira INCI podatke iz Shopify-ja za vse parfume, ki nimajo INCI."""
    try:
        from database import get_db
        from services.shopify_service import sync_inci_from_shopify
        
        db = get_db()
        cursor = db.cursor()
        
        # Pridobi vse parfume brez INCI
        cursor.execute("""
            SELECT p.product_no, p.proizvajalec_id, p.ime_parfuma, pr.ime as proizvajalec_ime
            FROM parfumi p 
            JOIN proizvajalci pr ON p.proizvajalec_id = pr.id
            WHERE p.sestava_inci IS NULL OR p.sestava_inci = ''
        """)
        
        parfumi_brez_inci = cursor.fetchall()
        
        if not parfumi_brez_inci:
            return jsonify({"success": True, "message": "Vsi parfumi že imajo INCI podatke"})
        
        success_count = 0
        error_count = 0
        errors = []
        
        for parfum in parfumi_brez_inci:
            try:
                success, message = sync_inci_from_shopify(
                    parfum['product_no'], 
                    parfum['proizvajalec_id']
                )
                if success:
                    success_count += 1
                else:
                    error_count += 1
                    errors.append(f"{parfum['product_no']} - {parfum['ime_parfuma']}: {message}")
            except Exception as e:
                error_count += 1
                errors.append(f"{parfum['product_no']} - {parfum['ime_parfuma']}: {str(e)}")
        
        message = f"Sinhronizacija končana. Uspešno: {success_count}, Napake: {error_count}"
        if errors:
            message += f". Napake: {'; '.join(errors[:5])}"  # Prikaži prvih 5 napak
        
        return jsonify({
            "success": True, 
            "message": message,
            "success_count": success_count,
            "error_count": error_count,
            "total_processed": len(parfumi_brez_inci)
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri sinhronizaciji vseh INCI: {e}")
        return jsonify({"success": False, "message": f"Napaka: {str(e)}"}), 500
@api_bp.route('/cleanup-duplicate-perfumes', methods=['POST'])
def cleanup_duplicate_perfumes():
    """Počisti duplikate parfumov z istem product_no IN istem proizvajalcem."""
    db = get_db()
    cursor = db.cursor()
    try:
        current_app.logger.info("Starting cleanup of duplicate perfumes...")
        
        # Najdi PRAVE duplikate - parfume z istim product_no IN istim proizvajalcem
        cursor.execute("""
            SELECT product_no, proizvajalec_id, COUNT(*) as count, 
                   array_agg(id) as perfume_ids,
                   array_agg(ime_parfuma) as imena
            FROM parfumi 
            GROUP BY product_no, proizvajalec_id
            HAVING COUNT(*) > 1
            ORDER BY product_no, proizvajalec_id
        """)
        
        duplicates = cursor.fetchall()
        cleaned_count = 0
        
        for duplicate in duplicates:
            product_no = duplicate['product_no']
            proizvajalec_id = duplicate['proizvajalec_id']
            perfume_ids = duplicate['perfume_ids']
            imena = duplicate['imena']
            
            current_app.logger.info(f"Found duplicate for product_no {product_no}, proizvajalec_id {proizvajalec_id}: {len(perfume_ids)} entries")
            
            # Najdi proizvajalca
            cursor.execute("SELECT ime FROM proizvajalci WHERE id = %s", (proizvajalec_id,))
            proizvajalec = cursor.fetchone()
            proizvajalec_ime = proizvajalec['ime'] if proizvajalec else f"ID:{proizvajalec_id}"
            
            # Če je proizvajalec "vendor", izbriši vse razen prvega
            if proizvajalec_ime == 'vendor':
                keep_id = perfume_ids[0]
                delete_ids = perfume_ids[1:]
                
                cursor.execute("DELETE FROM parfumi WHERE id = ANY(%s)", (delete_ids,))
                cleaned_count += len(delete_ids)
                current_app.logger.info(f"Deleted {len(delete_ids)} duplicates with vendor proizvajalec for product_no {product_no}")
            else:
                # Za ostale proizvajalce obdrži prvega in izbriši ostale
                keep_id = perfume_ids[0]
                delete_ids = perfume_ids[1:]
                
                cursor.execute("DELETE FROM parfumi WHERE id = ANY(%s)", (delete_ids,))
                cleaned_count += len(delete_ids)
                current_app.logger.info(f"Kept first perfume and deleted {len(delete_ids)} duplicates for product_no {product_no}, proizvajalec {proizvajalec_ime}")
        
        db.commit()
        message = f"Čiščenje duplikatov končano. Odstranjenih {cleaned_count} duplikatov."
        current_app.logger.info(message)
        return jsonify({"message": message, "cleaned_count": cleaned_count})

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri čiščenju duplikatov: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Prišlo je do napake na strežniku: {str(e)}"}), 500
    finally:
        cursor.close()

@api_bp.route('/merge-perfumes', methods=['POST'])
def merge_perfumes():
    """Združi dva parfuma z istim product_no."""
    db = get_db()
    cursor = db.cursor()
    try:
        data = request.get_json()
        source_id = data.get('source_id')
        target_id = data.get('target_id')
        
        if not source_id or not target_id:
            return jsonify({"error": "Manjkata source_id in target_id"}), 400
        
        current_app.logger.info(f"Merging perfume {source_id} into {target_id}")
        
        # Preveri, če parfuma obstajata
        cursor.execute("SELECT id, product_no, proizvajalec_id, ime_parfuma FROM parfumi WHERE id IN (%s, %s)", (source_id, target_id))
        perfumes = cursor.fetchall()
        
        if len(perfumes) != 2:
            return jsonify({"error": "Eden od parfumov ne obstaja"}), 400
        
        source_perfume = next(p for p in perfumes if p['id'] == source_id)
        target_perfume = next(p for p in perfumes if p['id'] == target_id)
        
        # Preveri, če imata isti product_no
        if source_perfume['product_no'] != target_perfume['product_no']:
            return jsonify({"error": "Parfuma nimata istega product_no"}), 400
        
        # Premakni serije iz source v target
        cursor.execute("UPDATE serije SET parfum_id = %s WHERE parfum_id = %s", (target_id, source_id))
        
        # Premakni deklaracije iz source v target
        cursor.execute("UPDATE declarations SET parfum_id = %s WHERE parfum_id = %s", (target_id, source_id))
        
        # Izbriši source parfum
        cursor.execute("DELETE FROM parfumi WHERE id = %s", (source_id,))
        
        db.commit()
        message = f"Parfum {source_id} uspešno združen v {target_id}"
        current_app.logger.info(message)
        return jsonify({"message": message})

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri združevanju parfumov: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Prišlo je do napake na strežniku: {str(e)}"}), 500
    finally:
        cursor.close()

@api_bp.route('/proizvajalci', methods=['POST'])
def add_proizvajalec():
    """Dodaj nov proizvajalec."""
    db = get_db()
    cursor = db.cursor()
    try:
        data = request.get_json()
        ime = data.get('ime')
        
        if not ime:
            return jsonify({"error": "Manjka ime proizvajalca"}), 400
        
        cursor.execute("INSERT INTO proizvajalci (ime) VALUES (%s) ON CONFLICT (ime) DO NOTHING", (ime,))
        
        if cursor.rowcount > 0:
            db.commit()
            message = f"Proizvajalec '{ime}' uspešno dodan"
            current_app.logger.info(message)
            return jsonify({"message": message})
        else:
            return jsonify({"error": f"Proizvajalec '{ime}' že obstaja"}), 400

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri dodajanju proizvajalca: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Prišlo je do napake na strežniku: {str(e)}"}), 500
    finally:
        cursor.close()
@api_bp.route('/proizvajalci/<int:proizvajalec_id>', methods=['DELETE'])
def delete_proizvajalec(proizvajalec_id):
    """Izbriši proizvajalca in vse njegove parfume."""
    db = get_db()
    cursor = db.cursor()
    try:
        # Preveri, če proizvajalec obstaja
        cursor.execute("SELECT ime FROM proizvajalci WHERE id = %s", (proizvajalec_id,))
        proizvajalec = cursor.fetchone()
        
        if not proizvajalec:
            return jsonify({"error": "Proizvajalec ne obstaja"}), 404
        
        ime = proizvajalec['ime']
        
        # Najdi vse parfume tega proizvajalca
        cursor.execute("SELECT id, ime_parfuma FROM parfumi WHERE proizvajalec_id = %s", (proizvajalec_id,))
        parfumi = cursor.fetchall()
        
        if parfumi:
            # Izbriši vse serije in deklaracije za te parfume
            parfum_ids = [p['id'] for p in parfumi]
            cursor.execute("DELETE FROM serije WHERE parfum_id = ANY(%s)", (parfum_ids,))
            cursor.execute("DELETE FROM declarations WHERE parfum_id = ANY(%s)", (parfum_ids,))
            
            # Izbriši parfume
            cursor.execute("DELETE FROM parfumi WHERE proizvajalec_id = %s", (proizvajalec_id,))
        
        # Izbriši proizvajalca
        cursor.execute("DELETE FROM proizvajalci WHERE id = %s", (proizvajalec_id,))
        
        db.commit()
        message = f"Proizvajalec '{ime}' in {len(parfumi)} parfumov uspešno izbrisani"
        current_app.logger.info(message)
        return jsonify({"message": message, "deleted_perfumes": len(parfumi)})

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri brisanju proizvajalca: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Prišlo je do napake na strežniku: {str(e)}"}), 500
    finally:
        cursor.close()

@api_bp.route('/migrate-perfumes', methods=['POST'])
def migrate_perfumes_endpoint():
    """Migrira parfume iz Excel datoteke."""
    try:
        from migrations import migrate_perfumes_from_excel
        
        current_app.logger.info("Starting perfume migration from Excel...")
        
        if migrate_perfumes_from_excel():
            message = "Migracija parfumov iz Excel datoteke uspešno končana"
            current_app.logger.info(message)
            return jsonify({"message": message})
        else:
            error_msg = "Migracija parfumov ni uspela"
            current_app.logger.error(error_msg)
            return jsonify({"error": error_msg}), 500

    except Exception as e:
        current_app.logger.error(f"Napaka pri migraciji parfumov: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Prišlo je do napake na strežniku: {str(e)}"}), 500


@api_bp.route('/restore-parfumi-names', methods=['POST'])
def restore_parfumi_names_endpoint():
    """Obnovi samo ime_parfuma iz Excel zvezka Parfumi (upload ali privzeta datoteka)."""
    dry_run = True
    try:
        if request.content_type and 'multipart/form-data' in request.content_type:
            dry_run = request.form.get('dry_run', 'true').lower() != 'false'
            upload = request.files.get('file')
            if not upload or not upload.filename:
                return jsonify({"error": "Manjka Excel datoteka (file)."}), 400
            file_bytes = upload.read()
            result = restore_parfumi_names_from_excel(
                file_bytes=file_bytes,
                dry_run=dry_run,
            )
        else:
            data = request.get_json(silent=True) or {}
            dry_run = bool(data.get('dry_run', True))
            file_path = (data.get('file_path') or 'DEKLARACIJE_PARFUMOV_KOPER.xlsm').strip()
            import os
            if not os.path.isfile(file_path):
                return jsonify({
                    "error": (
                        f"Excel datoteka '{file_path}' ne obstaja na strežniku. "
                        "Naloži jo kot multipart/form-data (file)."
                    ),
                }), 400
            result = restore_parfumi_names_from_excel(
                file_path=file_path,
                dry_run=dry_run,
            )

        if result.get("error"):
            return jsonify({"error": result["error"], "result": result}), 500

        action = "Simulacija" if dry_run else "Obnovitev"
        message = (
            f"{action} imen končana. Excel vrstic: {result.get('excel_rows', 0)}, "
            f"posodobljenih: {result.get('updated', 0)}, "
            f"nespremenjenih: {result.get('unchanged', 0)}, "
            f"v bazi brez Excel zapisa: {result.get('not_in_excel', 0)}."
        )
        return jsonify({"message": message, "result": result})
    except Exception as e:
        current_app.logger.error(f"Napaka pri obnovi imen parfumov: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Prišlo je do napake na strežniku: {str(e)}"}), 500


@api_bp.route('/export-parfumi-imena', methods=['GET'])
def export_parfumi_imena():
    """Izvozi id, proizvajalec_id, ime_proizvajalca, ime_parfuma za vse parfume."""
    import csv
    import io
    from datetime import datetime

    fmt = (request.args.get("format") or "xlsx").strip().lower()

    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            """
            SELECT pr.ime AS proizvajalec, p.product_no, p.ime_parfuma
            FROM parfumi p
            JOIN proizvajalci pr ON pr.id = p.proizvajalec_id
            ORDER BY pr.ime, p.product_no
            """
        )
        rows = cursor.fetchall() or []
    finally:
        cursor.close()

    stamp = datetime.now().strftime("%Y%m%d")
    headers_ui = ["Proizvajalec", "Šifra izdelka (Product No)", "Ime parfuma"]

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers_ui)
        for row in rows:
            if isinstance(row, dict):
                writer.writerow(
                    [row["proizvajalec"], row["product_no"], row["ime_parfuma"]]
                )
            else:
                writer.writerow(row)
        return Response(
            buf.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="parfumi_imena_{stamp}.csv"',
            },
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Parfumi"
    ws.append(headers_ui)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        if isinstance(row, dict):
            ws.append([row["proizvajalec"], row["product_no"], row["ime_parfuma"]])
        else:
            ws.append(row)
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).number_format = "@"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 52
    ws.freeze_panes = "A2"

    xbuf = io.BytesIO()
    wb.save(xbuf)
    xbuf.seek(0)
    return send_file(
        xbuf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"parfumi_imena_{stamp}.xlsx",
    )


@api_bp.route('/import-parfumi-imena', methods=['POST'])
def import_parfumi_imena():
    """Uvozi popravljena ime_parfuma iz CSV (id, proizvajalec_id, ime_parfuma)."""
    dry_run = True
    try:
        if request.content_type and "multipart/form-data" in request.content_type:
            dry_run = request.form.get("dry_run", "true").lower() != "false"
            upload = request.files.get("file")
            if not upload or not upload.filename:
                return jsonify({"error": "Manjka CSV datoteka (file)."}), 400
            file_bytes = upload.read()
        else:
            data = request.get_json(silent=True) or {}
            dry_run = bool(data.get("dry_run", True))
            return jsonify({
                "error": "Naloži CSV kot multipart/form-data (polje file).",
            }), 400

        result = import_parfumi_names_from_csv(file_bytes, dry_run=dry_run)
        if result.get("error") and not result.get("ok"):
            return jsonify({"error": result["error"], "result": result}), 500

        action = "Simulacija" if dry_run else "Uvoz"
        message = (
            f"{action} imen končana. Vrstic v datoteki: {result.get('rows_in_file', 0)}, "
            f"posodobljenih: {result.get('updated', 0)}, "
            f"nespremenjenih: {result.get('unchanged', 0)}, "
            f"preskočenih: {result.get('skipped', 0)}."
        )
        return jsonify({"message": message, "result": result})
    except Exception as e:
        current_app.logger.error(f"Napaka pri uvozu imen parfumov: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Prišlo je do napake na strežniku: {str(e)}"}), 500


@api_bp.route('/preview-fix-amour-parfums-names', methods=['GET', 'POST'])
def preview_fix_amour_parfums_names_endpoint():
    """Predogled popravkov AMOUR PARFUMS - → deklaracije_vendor (ne piše v bazo)."""
    data = request.get_json(silent=True) or {}
    shop_domain = (data.get('shop_domain') or request.args.get('shop_domain') or 'amour-parfums-2.myshopify.com').strip()
    result = preview_fix_amour_parfums_names(shop_domain)
    if not result.get('ok'):
        return jsonify({"error": result.get("error", "Napaka")}), 400
    return jsonify(result)


@api_bp.route('/apply-fix-amour-parfums-names', methods=['POST'])
def apply_fix_amour_parfums_names_endpoint():
    """Uporabi popravke iz deklaracije_vendor (privzeto dry_run)."""
    data = request.get_json(silent=True) or {}
    shop_domain = (data.get('shop_domain') or 'amour-parfums-2.myshopify.com').strip()
    dry_run = bool(data.get('dry_run', True))
    result = apply_fix_amour_parfums_names(shop_domain, dry_run=dry_run)
    if not result.get('ok'):
        return jsonify({"error": result.get("error", "Napaka"), "result": result}), 400
    action = "Simulacija" if dry_run else "Popravek"
    message = (
        f"{action} končan. Kandidatov: {result.get('candidates', 0)}, "
        f"posodobitev: {result.get('would_update', 0) if dry_run else result.get('applied', 0)}."
    )
    return jsonify({"message": message, "result": result})


@api_bp.route('/auto-enable-shopify-sync', methods=['POST'])
def auto_enable_shopify_sync():
    """Avtomatsko vklopi sinhronizacijo s Shopify za vse parfume, ki obstajajo v Shopify."""
    db = get_db()
    cursor = db.cursor()
    try:
        current_app.logger.info("Starting auto-enable Shopify sync...")
        
        # Pridobi vse izdelke iz Shopify-ja (isti pristop kot sync-stock-status)
        shopify_products = get_all_shopify_products_with_metafields()
        if shopify_products is None:
            return jsonify({"error": "Napaka pri pridobivanju podatkov iz Shopify."}), 500
        current_app.logger.info(f"Fetched {len(shopify_products)} products from Shopify.")
        
        # Pridobi vse parfume iz baze
        cursor.execute("""
            SELECT p.id, p.product_no, pr.ime as ime_proizvajalca, p.sinhroniziraj_s_shopify
            FROM parfumi p 
            JOIN proizvajalci pr ON p.proizvajalec_id = pr.id
        """)
        local_perfumes = cursor.fetchall()
        
        enabled_count = 0
        already_enabled_count = 0
        not_found_count = 0
        
        for perfume in local_perfumes:
            # Uporabi isti lookup_key kot v sync-stock-status
            lookup_key = f"{perfume['product_no'].strip()}_{perfume['ime_proizvajalca'].strip().upper()}"
            shopify_data = shopify_products.get(lookup_key)
            
            if shopify_data:
                # Parfum obstaja v Shopify
                if not perfume['sinhroniziraj_s_shopify']:
                    # Vklopi sinhronizacijo
                    cursor.execute("UPDATE parfumi SET sinhroniziraj_s_shopify = TRUE WHERE id = %s", (perfume['id'],))
                    enabled_count += 1
                    current_app.logger.info(f"Enabled sync for perfume {perfume['product_no']} - {perfume['ime_proizvajalca']}")
                else:
                    already_enabled_count += 1
            else:
                not_found_count += 1
                current_app.logger.warning(f"Perfume not found in Shopify: {perfume['product_no']} - {perfume['ime_proizvajalca']}")
        
        db.commit()
        
        message = f"Avtomatska vklopitev sinhronizacije končana. Vklopljenih: {enabled_count}, že vklopljenih: {already_enabled_count}, ni najdenih v Shopify: {not_found_count}."
        current_app.logger.info(message)
        return jsonify({"message": message})

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri avtomatski vklopitvi sinhronizacije: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do napake na strežniku."}), 500
    finally:
        cursor.close()
def auto_disable_shopify_sync():
    """Avtomatsko izklopi sinhronizacijo s Shopify za vse parfume, ki ne obstajajo v Shopify-ju"""
    db = get_db()
    cursor = db.cursor()
    try:
        current_app.logger.info("Starting auto-disable Shopify sync...")
        # Pridobi vse izdelke iz Shopify-ja
        shopify_products = get_all_shopify_products_with_metafields()
        if shopify_products is None:
            return jsonify({"error": "Napaka pri pridobivanju podatkov iz Shopify."}), 500
        current_app.logger.info(f"Fetched {len(shopify_products)} products from Shopify.")
        # Pridobi vse parfume z vklopljeno sinhronizacijo
        cursor.execute("""
            SELECT p.id, p.product_no, pr.ime as ime_proizvajalca, p.sinhroniziraj_s_shopify
            FROM parfumi p 
            JOIN proizvajalci pr ON p.proizvajalec_id = pr.id
            WHERE p.sinhroniziraj_s_shopify = TRUE
        """)
        local_perfumes = cursor.fetchall()
        disabled_count = 0
        still_exist_count = 0
        errors = []
        for perfume in local_perfumes:
            try:
                # Uporabi isti lookup_key kot v sync-stock-status
                lookup_key = f"{perfume['product_no'].strip()}_{perfume['ime_proizvajalca'].strip().upper()}"
                shopify_data = shopify_products.get(lookup_key)
                if not shopify_data:
                    # Parfum ne obstaja v Shopify-ju, izklopi sinhronizacijo
                    cursor.execute("UPDATE parfumi SET sinhroniziraj_s_shopify = FALSE WHERE id = %s", (perfume['id'],))
                    disabled_count += 1
                    current_app.logger.info(f"Disabled sync for perfume {perfume['product_no']} - {perfume['ime_proizvajalca']} (not found in Shopify)")
                else:
                    still_exist_count += 1
                    current_app.logger.info(f"Perfume still exists in Shopify: {perfume['product_no']} - {perfume['ime_proizvajalca']}")
            except Exception as e:
                error_msg = f"Napaka pri parfumu {perfume['product_no']}: {str(e)}"
                errors.append(error_msg)
                current_app.logger.error(error_msg)
        db.commit()
        message = f"Avtomatska izklopitev sinhronizacije končana. Izklopljenih: {disabled_count}, še vedno obstajajo v Shopify: {still_exist_count}"
        if errors:
            message += f". Napake: {len(errors)}"
        current_app.logger.info(message)
        return jsonify({
            "message": message,
            "disabled_count": disabled_count,
            "still_exist_count": still_exist_count,
            "errors": errors
        })
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri avtomatski izklopitvi sinhronizacije: {e}")
        traceback.print_exc()
        return jsonify({"error": "Prišlo je do napake na strežniku."}), 500
    finally:
        cursor.close()

@api_bp.route('/parfum/<int:perfume_id>/check-shopify-exists', methods=['GET'])
def check_shopify_exists(perfume_id):
    """Preveri, ali parfum obstaja v Shopify-ju in vrni rezultat"""
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute("""
            SELECT p.*, pr.ime as ime_proizvajalca 
            FROM parfumi p 
            JOIN proizvajalci pr ON p.proizvajalec_id = pr.id 
            WHERE p.id = %s
        """, (perfume_id,))
        parfum = cursor.fetchone()
        if not parfum:
            return jsonify({"error": "Parfum ni najden."}), 404

        gid, find_msg = find_shopify_product_gid(parfum['product_no'], parfum['proizvajalec_id'])
        exists_in_shopify = gid is not None
        
        return jsonify({
            "exists_in_shopify": exists_in_shopify,
            "message": find_msg if not exists_in_shopify else "Parfum najden v Shopify-ju"
        })

    except Exception as e:
        current_app.logger.error(f"Error in check_shopify_exists for perfume_id {perfume_id}: {e}")
        traceback.print_exc()
        return jsonify({"error": "Napaka na strežniku pri preverjanju obstoja."}), 500
    finally:
        cursor.close()

@api_bp.route('/check-new-orders', methods=['GET'])
def check_new_orders():
    """Preveri, ali so nova naročila od zadnjega osveževanja"""
    global last_order_check
    
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Če je to prvi klic, nastavi trenutni čas
        if last_order_check is None:
            cursor.execute("SELECT MAX(created_at) as latest FROM orders")
            result = cursor.fetchone()
            last_order_check = result['latest'] if result and result['latest'] else time.time()
            return jsonify({"has_new_orders": False, "last_check": last_order_check})
        
        # Preveri, ali so nova naročila
        cursor.execute("SELECT COUNT(*) as count FROM orders WHERE created_at > %s", (last_order_check,))
        result = cursor.fetchone()
        new_orders_count = result['count'] if result else 0
        
        # Posodobi čas zadnjega preverjanja
        cursor.execute("SELECT MAX(created_at) as latest FROM orders")
        result = cursor.fetchone()
        last_order_check = result['latest'] if result and result['latest'] else last_order_check
        
        return jsonify({
            "has_new_orders": new_orders_count > 0,
            "new_orders_count": new_orders_count,
            "last_check": last_order_check
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri preverjanju novih naročil: {e}")
        return jsonify({"error": "Napaka pri preverjanju novih naročil"}), 500
    finally:
        cursor.close()

def notify_new_order(order_number):
    """Logiraj novo naročilo (za prihodnje uporabe)"""
    current_app.logger.info(f"Novo naročilo prejeto: {order_number}")

# --- API endpoints za upravljanje uporabnikov ---

@api_bp.route('/users', methods=['GET'])
def get_users():
    """Pridobi seznam vseh uporabnikov."""
    try:
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("""
            SELECT id, username, first_name, last_name, email, is_active, created_at,
                   role, permissions, kiosk_pin_plain, color_hex
            FROM users
            ORDER BY first_name, last_name
        """)
        
        users_raw = cursor.fetchall()

        # Parsiraj permissions za vsakega uporabnika
        users = []
        for user in users_raw:
            user_dict = dict(user)
            raw_permissions = user_dict.get('permissions')
            current_app.logger.info(f"Get users - User {user_dict['username']} raw permissions: {raw_permissions}")
            current_app.logger.info(f"Get users - User {user_dict['username']} permissions type: {type(raw_permissions)}")

            parsed_permissions = []
            if isinstance(raw_permissions, str) and raw_permissions.strip():
                try:
                    parsed_permissions = json.loads(raw_permissions)
                    current_app.logger.info(f"Get users - User {user_dict['username']} parsed permissions from string: {parsed_permissions}")
                except (json.JSONDecodeError, TypeError) as e:
                    current_app.logger.error(f"Get users - User {user_dict['username']} error parsing permissions string: {e}")
            elif isinstance(raw_permissions, list):
                parsed_permissions = raw_permissions
            elif raw_permissions is None:
                current_app.logger.info(f"Get users - User {user_dict['username']} no permissions in DB")

            user_dict['permissions'] = parsed_permissions
            current_app.logger.info(f"Get users - User {user_dict['username']} final permissions: {user_dict['permissions']}")
            users.append(user_dict)
        
        return jsonify(users)
    except Exception as e:
        current_app.logger.error(f"Napaka pri pridobivanju uporabnikov: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cursor.close()

@api_bp.route('/users', methods=['POST'])
def add_user():
    """Dodaj novega uporabnika."""
    try:
        data = request.get_json()
        username = data.get('username')
        first_name = data.get('first_name')
        last_name = data.get('last_name')
        email = data.get('email')
        password = data.get('password')  # Novo polje za geslo
        
        if not all([username, first_name, last_name]):
            return jsonify({'success': False, 'error': 'Manjkajo obvezni podatki'}), 400
        
        # Če geslo ni podano, ustvari privzeto geslo (username + "123")
        if not password:
            password = f"{username}123"
        
        # Hashiraj geslo
        password_hash = generate_password_hash(password)
        
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("""
            INSERT INTO users (username, first_name, last_name, email, password_hash)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id
        """, (username, first_name, last_name, email, password_hash))
        
        user_id = cursor.fetchone()['id']
        db.commit()

        # Pošlji welcome email (ne prekinjamo ob morebitni napaki pri pošiljanju)
        try:
            # Sestavi login URL (root aplikacije)
            app_base_url = current_app.config.get('APP_BASE_URL', request.host_url)
            send_new_user_welcome_email(username=username, recipient_email=email, password=password, login_url=app_base_url)
        except Exception as e:
            current_app.logger.error(f"Neuspešno pošiljanje welcome emaila za uporabnika {username}: {e}")

        return jsonify({
            'success': True,
            'user_id': user_id,
            'message': f'Uporabnik uspešno dodan z geslom: {password}',
            'password': password  # Vrni geslo za prikaz uporabniku
        })
    except Exception as e:
        # Če je UniqueViolation (psycopg), vrni jasno sporočilo; sicer generična napaka
        try:
            import psycopg
            if isinstance(e, psycopg.errors.UniqueViolation):
                return jsonify({'success': False, 'error': 'Uporabniško ime že obstaja'}), 400
        except Exception:
            pass
        current_app.logger.error(f"Napaka pri dodajanju uporabnika: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cursor.close()
@api_bp.route('/login', methods=['POST'])
def login():
    """Prijava uporabnika."""
    cursor = None
    try:
        data = request.get_json(silent=True) or {}
        username = data.get('username')
        password = data.get('password')

        if not username or not password:
            return jsonify({'success': False, 'error': 'Uporabniško ime in geslo sta obvezna'}), 400

        try:
            db = get_db()
            cursor = db.cursor()
        except Exception as e:
            current_app.logger.error(f"Login - DB ni dosegljiva: {e}")
            return jsonify({'success': False, 'error': 'Baza podatkov trenutno ni dosegljiva. Poskusi ponovno čez nekaj sekund.'}), 503
        
        # Poišči uporabnika
        cursor.execute("""
            SELECT id, username, first_name, last_name, email, password_hash, is_active, role, permissions
            FROM users WHERE username = %s
        """, (username,))
        
        user = cursor.fetchone()
        
        if not user:
            return jsonify({'success': False, 'error': 'Napačno uporabniško ime ali geslo'}), 401
        
        if not user['is_active']:
            return jsonify({'success': False, 'error': 'Uporabniški račun je deaktiviran'}), 401
        
        # Preveri geslo
        if not check_password_hash(user['password_hash'], password):
            return jsonify({'success': False, 'error': 'Napačno uporabniško ime ali geslo'}), 401
        
        # Parsiraj permissions (podpira JSONB kot list ali string JSON)
        permissions = []
        raw_permissions = user.get('permissions')
        current_app.logger.info(f"Login - Raw permissions from DB: {raw_permissions}")
        current_app.logger.info(f"Login - Permissions type: {type(raw_permissions)}")

        if isinstance(raw_permissions, str) and raw_permissions.strip():
            try:
                permissions = json.loads(raw_permissions)
                current_app.logger.info(f"Login - Parsed permissions from string: {permissions}")
            except (json.JSONDecodeError, TypeError) as e:
                current_app.logger.error(f"Login - Error parsing permissions string: {e}")
        elif isinstance(raw_permissions, list):
            permissions = raw_permissions
        elif raw_permissions is None:
            current_app.logger.info("Login - No permissions in DB")
        
        current_app.logger.info(f"Login - Final permissions: {permissions}")
        
        # Normaliziraj vlogo (za UI in backend bypass)
        role_normalized = str(user['role']).strip().lower() if user.get('role') is not None else ''

        # Nastavi Flask session
        session['logged_in'] = True
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['user'] = {
            'id': user['id'],
            'username': user['username'],
            'first_name': user['first_name'],
            'last_name': user['last_name'],
            'email': user['email'],
            'role': role_normalized,
            'permissions': permissions
        }
        
        return jsonify({
            'success': True,
            'user': {
                'id': user['id'],
                'username': user['username'],
                'first_name': user['first_name'],
                'last_name': user['last_name'],
                'email': user['email'],
                'role': role_normalized,
                'permissions': permissions
            },
            'message': 'Uspešna prijava'
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri prijavi: {e}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': 'Napaka pri prijavi'}), 500
    finally:
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
@api_bp.route('/current-user', methods=['GET'])
def get_current_user():
    """Pridobi podatke o trenutno prijavljenem uporabniku."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Uporabnik ni prijavljen'}), 401
    
    try:
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("""
            SELECT id, username, first_name, last_name, email, is_active, role, permissions
            FROM users WHERE id = %s
        """, (session['user_id'],))
        
        user = cursor.fetchone()
        
        if not user:
            return jsonify({'success': False, 'error': 'Uporabnik ne obstaja'}), 404
        
        if not user['is_active']:
            return jsonify({'success': False, 'error': 'Uporabniški račun je deaktiviran'}), 401
        
        # Parsiraj permissions (podpira JSONB kot list ali string JSON)
        permissions = []
        raw_permissions = user.get('permissions')
        if isinstance(raw_permissions, str) and raw_permissions.strip():
            try:
                permissions = json.loads(raw_permissions)
            except (json.JSONDecodeError, TypeError):
                permissions = []
        elif isinstance(raw_permissions, list):
            permissions = raw_permissions
        
        # Normaliziraj vlogo
        role_normalized = str(user['role']).strip().lower() if user.get('role') is not None else ''

        return jsonify({
            'success': True,
            'user': {
                'id': user['id'],
                'username': user['username'],
                'first_name': user['first_name'],
                'last_name': user['last_name'],
                'email': user['email'],
                'role': role_normalized,
                'permissions': permissions
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri pridobivanju podatkov o uporabniku: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri pridobivanju podatkov'}), 500
    finally:
        cursor.close()

@api_bp.route('/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    """Izbriši uporabnika."""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Preveri, ali uporabnik obstaja
        cursor.execute("SELECT username FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user:
            return jsonify({'success': False, 'error': 'Uporabnik ne obstaja'}), 404
        
        # Preveri, ali je admin (ne sme se izbrisati)
        if user['username'] == 'admin':
            return jsonify({'success': False, 'error': 'Admin uporabnika ni mogoče izbrisati'}), 400
        
        # Izbriši uporabnika
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        db.commit()
        
        return jsonify({
            'success': True,
            'message': 'Uporabnik uspešno izbrisan'
        })
    except Exception as e:
        current_app.logger.error(f"Napaka pri brisanju uporabnika: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cursor.close()

@api_bp.route('/users/<int:user_id>/permissions', methods=['PUT'])
def update_user_permissions(user_id):
    """Posodobi dovoljenja uporabnika."""
    if 'user_id' not in session:
        return jsonify({"error": "Uporabnik ni prijavljen"}), 401
    
    # Preveri, ali ima trenutni uporabnik dovoljenje za urejanje uporabnikov
    # Admin ima vedno vsa dovoljenja
    current_app.logger.info(f"Checking permissions for user ID: {session['user_id']}")
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("""
            SELECT username, role, permissions 
            FROM users 
            WHERE id = %s AND is_active = TRUE
        """, (session['user_id'],))
        
        user_data = cursor.fetchone()
        current_app.logger.info(f"User data from DB: {user_data}")
        
        if not user_data:
            return jsonify({"error": "Uporabnik ne obstaja"}), 404
        
        username = user_data['username']
        role = user_data['role']
        permissions = user_data['permissions']
        
        # Admin ima vedno vsa dovoljenja
        current_app.logger.info(f"Checking permissions - username: {username}, role: {role}, permissions: {permissions}")
        
        if str(role or '').strip().lower() == 'admin':
            has_edit_permission = True
            current_app.logger.info(f"Admin role detected: {username}, has_edit_permission: {has_edit_permission}")
        else:
            # Preveri, ali je permissions JSON string in ga parsiraj
            if isinstance(permissions, str):
                try:
                    permissions_parsed = json.loads(permissions)
                    current_app.logger.info(f"Parsed permissions: {permissions_parsed}")
                    has_edit_permission = 'edit_users' in permissions_parsed if permissions_parsed else False
                except json.JSONDecodeError:
                    current_app.logger.error(f"Failed to parse permissions JSON: {permissions}")
                    has_edit_permission = False
            else:
                has_edit_permission = 'edit_users' in permissions if permissions else False
            
            current_app.logger.info(f"Regular user: {username}, role: {role}, permissions: {permissions}, has_edit_permission: {has_edit_permission}")
        
        if not has_edit_permission:
            current_app.logger.warning(f"User {username} (ID: {session['user_id']}) denied permission to edit users")
            return jsonify({"error": "Nimate dovoljenja za urejanje uporabnikov"}), 403
            
    except Exception as e:
        current_app.logger.error(f"Napaka pri preverjanju dovoljenj: {e}")
        return jsonify({"error": "Napaka pri preverjanju dovoljenj"}), 500
    finally:
        cursor.close()


@api_bp.route('/users/<int:user_id>/kiosk-pin', methods=['PUT'])
@require_permission('edit_users')
def update_user_kiosk_pin(user_id):
    data = request.get_json(silent=True) or {}
    pin = str(data.get('pin', '')).strip()
    if pin:
        if not pin.isdigit() or len(pin) < 4 or len(pin) > 6:
            return jsonify({"error": "PIN mora biti 4–6 mestna številka"}), 400
        pin_hash = generate_password_hash(pin)
    else:
        pin_hash = None
        pin = None

    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            """
            UPDATE users
            SET kiosk_pin_plain = %s,
                kiosk_pin_hash = %s,
                kiosk_pin_updated_at = CASE WHEN %s IS NULL THEN NULL ELSE NOW() END
            WHERE id = %s
            """,
            (pin, pin_hash, pin, user_id)
        )
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri posodabljanju PIN: {e}")
        return jsonify({"error": "Napaka pri shranjevanju PIN"}), 500
    finally:
        cursor.close()


@api_bp.route('/users/<int:user_id>/color', methods=['PUT'])
@require_permission('edit_users')
def update_user_color(user_id):
    data = request.get_json(silent=True) or {}
    color = str(data.get('color_hex', '')).strip()
    if color:
        import re
        if not re.match(r'^#[0-9a-fA-F]{6}$', color):
            return jsonify({"error": "Barva mora biti v obliki #RRGGBB"}), 400
    else:
        color = None

    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            "UPDATE users SET color_hex = %s WHERE id = %s",
            (color, user_id)
        )
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri posodabljanju barve: {e}")
        return jsonify({"error": "Napaka pri shranjevanju barve"}), 500
    finally:
        cursor.close()
    
    data = request.get_json()
    permissions = data.get('permissions', [])
    role = data.get('role', 'user')
    
    if not isinstance(permissions, list):
        return jsonify({"error": "Dovoljenja morajo biti seznam"}), 400
    
    if role not in ['admin', 'user']:
        return jsonify({"error": "Neveljavna vloga"}), 400
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        # Preveri, ali uporabnik obstaja
        cursor.execute("SELECT username FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user:
            return jsonify({"error": "Uporabnik ne obstaja"}), 404
        
        # Preveri, ali se poskuša spremeniti admin
        # Dovoli admin-u, da ureja svoje dovoljenja
        if user['username'] == 'admin' and session['user_id'] != user_id:
            return jsonify({"error": "Admin uporabnika ni mogoče spremeniti"}), 403
        
        # Posodobi dovoljenja
        cursor.execute("""
            UPDATE users 
            SET permissions = %s, role = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (json.dumps(permissions), role, user_id))
        
        db.commit()
        
        return jsonify({
            "message": f"Dovoljenja za uporabnika {user['username']} uspešno posodobljena",
            "permissions": permissions,
            "role": role
        })
        
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri posodabljanju dovoljenj: {e}")
        return jsonify({"error": "Prišlo je do napake pri posodabljanju dovoljenj"}), 500
    finally:
        cursor.close()
@api_bp.route('/orders/list', methods=['GET'])
def list_all_order_numbers():
    """Seznam vseh order_number za izbiro v UI (max 1000 najnovejših)."""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            """
            SELECT order_number FROM orders 
            WHERE order_number IS NOT NULL AND order_number <> ''
            ORDER BY created_at DESC
            """
        )
        rows = cursor.fetchall()
        numbers = []
        for r in rows:
            if isinstance(r, dict):
                val = r.get('order_number')
            else:
                val = r[0]
            if val:
                numbers.append(val)
        return make_ok({'orders': numbers})
    except Exception as e:
        current_app.logger.error(f"Napaka pri list_all_order_numbers: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri pridobivanju seznamov naročil', status=500)

@api_bp.route('/order-by-number/<order_number>', methods=['GET'])
def get_order_by_number(order_number: str):
    try:
        # orders.order_number lahko vsebuje '#'
        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT id, order_number FROM orders WHERE order_number = %s", (f"#{order_number}" if not order_number.startswith('#') else order_number,))
        row = cursor.fetchone()
        if not row:
            # poskusi brez #
            cursor.execute("SELECT id, order_number FROM orders WHERE order_number = %s", (order_number.lstrip('#'),))
            row = cursor.fetchone()
        cursor.close()
        if not row:
            return jsonify({ 'error': 'Not found' }), 404
        return jsonify({ 'id': row['id'], 'order_number': row['order_number'] })
    except Exception as e:
        current_app.logger.error(f"Napaka pri get_order_by_number: {e}")
        return jsonify({ 'error': str(e) }), 500

@api_bp.route('/orders/search', methods=['GET'])
def search_orders():
    try:
        q_raw = request.args.get('q', '').strip()
        limit_raw = request.args.get('limit', 20)
        limit, err = validate_int(limit_raw, 'limit', min=1, max=50)
        if err:
            return err
        
        # If no query or too short, return recent orders
        if not q_raw or len(q_raw) < 2:
            db = get_db()
            cursor = db.cursor()
            cursor.execute(
                """
                SELECT order_number, COALESCE(customer_name, '') as customer_name, created_at
                FROM orders
                WHERE order_number IS NOT NULL AND order_number != ''
                ORDER BY created_at DESC
                LIMIT %s
                """,
                (limit,)
            )
            rows = cursor.fetchall()
            res = []
            for r in rows:
                if isinstance(r, dict):
                    order_number = r.get('order_number')
                    created_at = r.get('created_at')
                    customer_name = r.get('customer_name', '')
                else:
                    order_number = r[0]
                    customer_name = r[1] if len(r) > 1 else ''
                    created_at = r[2] if len(r) > 2 else None
                res.append({
                    'order_number': order_number,
                    'customer_name': customer_name or '',
                    'date': (created_at.date().isoformat() if hasattr(created_at, 'date') else str(created_at)) if created_at else ''
                })
            return make_ok({'results': res})
        
        # Validate query for search
        q, err = validate_str(q_raw, 'q', max_len=128, min_len=2)
        if err:
            return err
        db = get_db()
        cursor = db.cursor()
        like = f"%{q}%"
        res = []
        try:
            # Primarni poizvedba (če tabela ima customer_name)
            cursor.execute(
                """
                SELECT order_number, COALESCE(customer_name, '') as customer_name, created_at
                FROM orders
                WHERE order_number ILIKE %s 
                   OR REPLACE(order_number, '#','') ILIKE %s
                   OR customer_name ILIKE %s
                ORDER BY created_at DESC
                LIMIT %s
                """,
                (like, like, like, limit)
            )
        except Exception:
            # Fallback, če customer_name stolpec ne obstaja
            cursor.execute(
                """
                SELECT order_number, created_at
                FROM orders
                WHERE order_number ILIKE %s 
                   OR REPLACE(order_number, '#','') ILIKE %s
                ORDER BY created_at DESC
                LIMIT %s
                """,
                (like, like, limit)
            )
        rows = cursor.fetchall()
        for r in rows:
            if isinstance(r, dict):
                order_number = r.get('order_number')
                created_at = r.get('created_at')
                customer_name = r.get('customer_name', '')
            else:
                # Fallback tuple indeksi
                order_number = r[0]
                created_at = r[2] if len(r) > 2 else r[1]
                customer_name = r[1] if len(r) > 2 else ''
            res.append({
                'order_number': order_number,
                'customer_name': customer_name or '',
                'date': (created_at.date().isoformat() if hasattr(created_at, 'date') else str(created_at))
            })
        return make_ok({'results': res})
    except Exception as e:
        current_app.logger.error(f"Napaka pri search_orders: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri iskanju naročil', status=500)

@api_bp.route('/orders/<order_number>', methods=['GET'])
def get_order_details(order_number: str):
    """Pridobi podatke o naročilu po order_number"""
    try:
        current_app.logger.info(f"get_order_details called with order_number: {order_number}")
        
        # Počisti order_number - odstrani # znak, če obstaja
        if order_number.startswith('#'):
            order_number = order_number[1:]
        
        current_app.logger.info(f"Cleaned order_number: {order_number}")
        
        db = get_db()
        cursor = db.cursor(row_factory=dict_row)
        
        current_app.logger.info(f"Database connection established")
        
        # Pridobi osnovne podatke o naročilu
        query = '''
            SELECT o.*, 
                   CONCAT(u1.first_name, ' ', u1.last_name) as prepared_by_display,
                   CONCAT(u2.first_name, ' ', u2.last_name) as nalivalec_display
            FROM orders o
            LEFT JOIN users u1 ON o.prepared_by = u1.username
            LEFT JOIN users u2 ON o.nalivalec_id = u2.id
            WHERE o.order_number = %s
        '''
        
        current_app.logger.info(f"Executing query for order: {order_number}")
        
        # Poskusi najti naročilo brez # prefix
        cursor.execute(query, (order_number,))
        order = cursor.fetchone()
        
        # Če ni najdeno, poskusi z # prefix
        if not order:
            current_app.logger.info(f"Order {order_number} not found, trying with # prefix")
            cursor.execute(query, (f"#{order_number}",))
            order = cursor.fetchone()
        
        cursor.close()
        
        current_app.logger.info(f"Final query result for {order_number}: {order is not None}")
        
        if not order:
            current_app.logger.warning(f"Order {order_number} not found in both formats")
            return jsonify({"error": "Naročilo ni najdeno"}), 404
        
        # Convertaj order v običajen dict za JSON serialization
        order_dict = dict(order)
        
        # Če line_items obstajajo, jih parse in preveri product_type
        if order_dict.get('line_items'):
            try:
                line_items_raw = order_dict['line_items']
                line_items = json.loads(line_items_raw) if isinstance(line_items_raw, str) else line_items_raw
                
                # Preveri, ali line_items že vsebujejo product_type
                needs_update = any(item.get('product_type') is None for item in line_items if item)
                
                if needs_update:
                    current_app.logger.info(f"Posodabljam line_items za naročilo {order_number} z manjkajočimi product_type")
                    
                    # Pridobi product_type iz Shopify za vse izdelke
                    from services.shopify_service import get_bulk_product_details
                    product_ids = [str(item['product_id']) for item in line_items if item and item.get('product_id')]
                    
                    if product_ids:
                        shop_domain = order_dict.get('shopify_store_domain')
                        shopify_details = get_bulk_product_details(product_ids, shop_domain=shop_domain)
                        
                        # Posodobi line_items z product_type
                        for item in line_items:
                            if item and item.get('product_id'):
                                product_detail = shopify_details.get(str(item['product_id']))
                                if product_detail and 'product_type' in product_detail:
                                    item['product_type'] = product_detail['product_type']
                                    current_app.logger.info(f"Posodobljen product_type za {item.get('title')}: {product_detail['product_type']}")
                        
                        # Shrani posodobljene line_items nazaj v bazo
                        try:
                            db = get_db()
                            cursor_update = db.cursor()
                            cursor_update.execute("""
                                UPDATE orders 
                                SET line_items = %s 
                                WHERE order_number = %s OR order_number = %s
                            """, (json.dumps(line_items), order_number, f"#{order_number}"))
                            db.commit()
                            cursor_update.close()
                            current_app.logger.info(f"Uspešno posodobljeni line_items za naročilo {order_number}")
                        except Exception as e:
                            current_app.logger.error(f"Napaka pri posodabljanju line_items: {e}")
                
                # Posodobi order_dict z (morda posobljenimi) line_items
                order_dict['line_items'] = line_items
                
            except (json.JSONDecodeError, TypeError) as e:
                current_app.logger.warning(f"Could not parse line_items for order {order_number}: {e}")
        
        current_app.logger.info(f"Naročilo {order_number} pridobljeno successfully")
        
        return jsonify({"success": True, "data": order_dict})
        
    except Exception as e:
        current_app.logger.error(f"Exception in get_order_details for {order_number}: {str(e)}")
        current_app.logger.error(f"Exception type: {type(e).__name__}")
        import traceback
        current_app.logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({"error": f"Napaka pri pridobivanju naročila: {str(e)}"}), 500
@api_bp.route('/orders/<int:order_id>/set-nalivalec', methods=['POST'])
def set_nalivalec(order_id: int):
    try:
        data = request.get_json(silent=True) or {}
        nalivalec_id = data.get('nalivalec_id')
        if nalivalec_id is None:
            return jsonify({ 'success': False, 'error': 'Manjka nalivalec_id' }), 400
        db = get_db()
        cursor = db.cursor()
        # Kolona nalivalec_id se zagotovi v migracijah; brez runtime DDL
        # preveri, da naročilo vsebuje vsaj en izdelek tipa Parfumi
        cursor.execute("SELECT line_items FROM orders WHERE id = %s", (order_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({ 'success': False, 'error': 'Naročilo ne obstaja' }), 404
        line_items_raw = row['line_items'] if isinstance(row, dict) else row[0]
        try:
            items = json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])
        except Exception:
            items = []
        has_parfumi = False
        keywords = ['parfum', 'perfume', 'parfumi', 'eau de parfum', 'edp']
        def _norm(value: str) -> str:
            try:
                import unicodedata
                return unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
            except Exception:
                return value
        for it in items or []:
            try:
                ptype = _norm((it.get('product_type') or '').strip().lower())
                title = _norm((it.get('title') or '').strip().lower())
                if 'parfum' in ptype or 'perfume' in ptype:
                    has_parfumi = True
                    break
                if any(kw in title for kw in keywords):
                    has_parfumi = True
                    break
            except Exception:
                continue
        if not has_parfumi:
            return jsonify({ 'success': False, 'error': "Nalivalec je dovoljen samo za naročila z izdelki type='Parfumi'" }), 400
        cursor.execute("UPDATE orders SET nalivalec_id = %s WHERE id = %s", (nalivalec_id, order_id))
        db.commit()
        cursor.close()
        return jsonify({ 'success': True })
    except Exception as e:
        current_app.logger.error(f"Napaka pri set-nalivalec: {e}")
        return jsonify({ 'success': False, 'error': str(e) }), 500

@api_bp.route('/orders/<int:order_id>/set-prepared-by', methods=['POST'])
def set_prepared_by(order_id: int):
    try:
        data = request.get_json(silent=True) or {}
        user_id = data.get('user_id')
        if user_id is None:
            return jsonify({ 'success': False, 'error': 'Manjka user_id' }), 400
        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT username FROM users WHERE id = %s", (user_id,))
        user_row = cursor.fetchone()
        if not user_row:
            return jsonify({ 'success': False, 'error': 'Uporabnik ne obstaja' }), 404
        prepared_by = user_row['username'] if isinstance(user_row, dict) else user_row[0]

        cursor.execute(
            "UPDATE orders SET prepared_by = %s, prepared_at = NOW() WHERE id = %s",
            (prepared_by, order_id)
        )
        db.commit()
        cursor.close()
        return jsonify({ 'success': True })
    except Exception as e:
        current_app.logger.error(f"Napaka pri set-prepared-by: {e}")
        return jsonify({ 'success': False, 'error': str(e) }), 500
@api_bp.route('/change-password', methods=['POST'])
def change_password():
    """Spremeni geslo uporabnika."""
    try:
        data = request.get_json()
        username = data.get('username')
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        if not all([username, current_password, new_password]):
            return jsonify({'success': False, 'error': 'Vsi podatki so obvezni'}), 400
        
        if len(new_password) < 6:
            return jsonify({'success': False, 'error': 'Novo geslo mora biti dolgo vsaj 6 znakov'}), 400
        
        db = get_db()
        cursor = db.cursor()
        
        # Poišči uporabnika
        cursor.execute("""
            SELECT id, password_hash, is_active
            FROM users WHERE username = %s
        """, (username,))
        
        user = cursor.fetchone()
        
        if not user:
            return jsonify({'success': False, 'error': 'Uporabnik ne obstaja'}), 404
        
        if not user['is_active']:
            return jsonify({'success': False, 'error': 'Uporabniški račun je deaktiviran'}), 401
        
        # Preveri trenutno geslo
        if not check_password_hash(user['password_hash'], current_password):
            return jsonify({'success': False, 'error': 'Napačno trenutno geslo'}), 401
        
        # Posodobi geslo
        new_password_hash = generate_password_hash(new_password)
        cursor.execute("""
            UPDATE users SET password_hash = %s WHERE id = %s
        """, (new_password_hash, user['id']))
        
        db.commit()

        return jsonify({
            'success': True,
            'message': 'Geslo uspešno spremenjeno'
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri spreminjanju gesla: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri spreminjanju gesla'}), 500
    finally:
        cursor.close()

@api_bp.route('/admin/change-user-password', methods=['POST'])
def admin_change_user_password():
    """Admin spremeni geslo uporabnika"""
    if 'user_id' not in session:
        return jsonify({"error": "Ni prijavljen."}), 401
    
    # Preveri dovoljenje manage_users
    if not has_permission('manage_users'):
        return jsonify({"error": "Nimate dovoljenja za to akcijo."}), 403

    db = get_db()
    cursor = db.cursor()
    try:
        # nadaljuj z validacijami in spremembo gesla
        
        data = request.get_json()
        user_id = data.get('user_id')
        new_password = data.get('new_password')
        confirm_password = data.get('confirm_password')
        
        if not all([user_id, new_password, confirm_password]):
            return jsonify({"error": "Vsa polja so obvezna."}), 400
        
        if new_password != confirm_password:
            return jsonify({"error": "Novi gesli se ne ujemata."}), 400
        
        if len(new_password) < 6:
            return jsonify({"error": "Geslo mora biti dolgo vsaj 6 znakov."}), 400
        
        # Preveri, ali uporabnik obstaja
        cursor.execute("SELECT id FROM uporabniki WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({"error": "Uporabnik ni najden."}), 404
        
        # Spremeni geslo
        hashed_password = generate_password_hash(new_password)
        cursor.execute("UPDATE uporabniki SET geslo = %s WHERE id = %s", (hashed_password, user_id))
        db.commit()
        
        return jsonify({"message": "Geslo uporabnika uspešno spremenjeno."})
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri spreminjanju gesla uporabnika: {e}")
        return jsonify({"error": "Napaka na strežniku."}), 500
    finally:
        cursor.close()

@api_bp.route('/orders/<order_number>/reset-preparation', methods=['POST'])
def reset_order_preparation(order_number):
    """Ponastavi pripravil, prepared_at in nalivalec_id za naročilo"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Order number v order_images je brez '#', v orders pa z '#'
        order_number_clean = order_number.lstrip('#')
        order_number_with_hash = f"#{order_number_clean}"

        # Preverimo če ima naročilo kakšne slike (v order_images je brez '#')
        cursor.execute("SELECT COUNT(*) FROM order_images WHERE order_number = %s", (order_number_clean,))
        result = cursor.fetchone()
        image_count = result['count'] if isinstance(result, dict) else result[0]
        
        if image_count == 0:
            # Če ni slik, ponastavi polja
            cursor.execute("""
                UPDATE orders
                SET prepared_by = NULL, prepared_at = NULL, nalivalec_id = NULL
                WHERE order_number = %s
            """, (order_number_with_hash,))
            
            db.commit()
            current_app.logger.info(f"Auto-reset preparation fields for order {order_number} (no images)")
            
            return jsonify({
                "success": True,
                "message": "Polja so bila ponastavljena"
            })
        else:
            return jsonify({
                "success": False,
                "message": "Naročilo ima še vedno slike"
            })
    
    except Exception as e:
        current_app.logger.error(f"Napaka pri avtomatskem ponastavitvi naročila {order_number}: {e}")
        current_app.logger.error(f"Stack: {traceback.format_exc()}")
        return jsonify({"error": "Napaka pri ponastavitvi"}), 500
    finally:
        cursor.close()

# --- Seznam nalivalcev (minimalni podatki), dostopno vsem prijavljenim ---
@api_bp.route('/nalivalci', methods=['GET'])
def list_nalivalci():
    try:
        db = get_db()
        c = db.cursor()
        c.execute("SELECT id, first_name, last_name, username FROM users ORDER BY first_name NULLS FIRST, last_name NULLS FIRST")
        rows = c.fetchall()
        out = []
        for r in rows:
            d = dict(r)
            out.append({
                'id': d.get('id'),
                'first_name': d.get('first_name'),
                'last_name': d.get('last_name'),
                'username': d.get('username'),
            })
        return jsonify(out)
    except Exception as e:
        current_app.logger.error(f"Napaka pri list_nalivalci: {e}")
        return jsonify([])

@api_bp.route('/cleanup-orders-without-images', methods=['POST'])
def cleanup_orders_without_images():
    """Počisti naročila, ki nimajo slik ampak imajo nastavljena preparation polja"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Poišči naročila, ki imajo prepared_by ali nalivalec_id ampak nimajo slik
        cursor.execute("""
            SELECT o.order_number, o.prepared_by, o.nalivalec_id,
                   (SELECT COUNT(*) FROM order_images oi WHERE oi.order_number = o.order_number) as image_count
            FROM orders o
            WHERE (o.prepared_by IS NOT NULL OR o.nalivalec_id IS NOT NULL)
            AND (SELECT COUNT(*) FROM order_images oi WHERE oi.order_number = o.order_number) = 0
        """)
        
        orders_to_cleanup = cursor.fetchall()
        cleaned_count = 0
        
        for order_row in orders_to_cleanup:
            order_number, prepared_by, nalivalec_id, image_count = order_row
            
            # Ponastavi polja
            cursor.execute("""
                UPDATE orders
                SET prepared_by = NULL, prepared_at = NULL, nalivalec_id = NULL
                WHERE order_number = %s
            """, (order_number,))
            
            cleaned_count += 1
            current_app.logger.info(f"Cleaned up order {order_number} - removed preparation fields (no images)")
        
        db.commit()
        current_app.logger.info(f"Cleanup complete - processed {cleaned_count} orders")
        
        return jsonify({
            "success": True,
            "cleaned_count": cleaned_count,
            "message": f"Počiščenih {cleaned_count} naročil"
        })
    
    except Exception as e:
        current_app.logger.error(f"Napaka pri čiščenju naročil: {e}")
        current_app.logger.error(f"Stack: {traceback.format_exc()}")
        return jsonify({"error": "Napaka pri čiščenju"}), 500
    finally:
        cursor.close()
# --- API endpoint-i za slike naročil ---
@api_bp.route('/order-images/<order_number>', methods=['GET'])
def get_order_images_api(order_number):
    """Pridobi vse slike za določeno naročilo"""
    try:
        # Preveri, ali je order_number podan
        if not order_number:
            return jsonify({"error": "Manjka order_number"}), 400
        
        db = get_db()
        cursor = db.cursor()
        
        # Pridobi slike iz baze z imeni uporabnikov
        cursor.execute("""
            SELECT oi.id, oi.s3_key, oi.uploaded_by, oi.uploaded_at, oi.user_id,
                   u.first_name, u.last_name
            FROM order_images oi
            LEFT JOIN users u ON oi.user_id = u.id
            WHERE oi.order_number = %s 
            ORDER BY oi.uploaded_at DESC
        """, (order_number,))
        
        images_data = cursor.fetchall()
        
        # Proxy URL-je za vsako sliko, da se izognemo S3 podpisnim/CORS težavam
        from flask import url_for
        images = []
        cleaned_missing = 0
        # S3 client za validacijo obstoja objektov (self-healing)
        try:
            from services.s3_service import get_s3_client
            s3_client = get_s3_client()
            bucket_name = current_app.config['S3_BUCKET_NAME']
            from botocore.exceptions import ClientError as _S3ClientError
        except Exception:
            s3_client = None
            bucket_name = None
            _S3ClientError = Exception
        
        current_app.logger.info(f"Processing {len(images_data)} images for order {order_number}")
        
        current_user_id = session.get('user_id')
        for image_data in images_data:
            if isinstance(image_data, dict):
                image_id = image_data['id']
                s3_key = image_data['s3_key']
                uploaded_by = image_data['uploaded_by']
                uploaded_at = image_data['uploaded_at']
                uploader_user_id = image_data.get('user_id')
                first_name = image_data.get('first_name')
                last_name = image_data.get('last_name')
            else:
                image_id = image_data[0]
                s3_key = image_data[1]
                uploaded_by = image_data[2]
                uploaded_at = image_data[3]
                uploader_user_id = image_data[4] if len(image_data) > 4 else None
                first_name = image_data[5] if len(image_data) > 5 else None
                last_name = image_data[6] if len(image_data) > 6 else None
            
            current_app.logger.info(f"Processing image {image_id} with S3 key: {s3_key}")
            
            # Self-healing: preskoči in izbriši zapise, ki nimajo več objekta v S3
            if s3_client and bucket_name:
                try:
                    s3_client.head_object(Bucket=bucket_name, Key=s3_key)
                except _S3ClientError as e:
                    current_app.logger.warning(f"Missing S3 object for key {s3_key}, cleaning DB row id={image_id}: {e}")
                    try:
                        cleanup_cursor = db.cursor()
                        cleanup_cursor.execute("DELETE FROM order_images WHERE id = %s", (image_id,))
                        db.commit()
                        cleanup_cursor.close()
                        cleaned_missing += 1
                    except Exception as ce:
                        current_app.logger.error(f"Failed to cleanup missing image id={image_id}: {ce}")
                    continue
            
            # Generiraj proxy URL preko našega strežnika (kot pri returned/damaged)
            s3_url = url_for('api.proxy_order_image_v2', s3_key=s3_key, _external=True)
            
            if s3_url:
                current_app.logger.info(f"Generated proxy URL for image {image_id}")
            else:
                current_app.logger.error(f"Failed to generate proxy URL for image {image_id}")
            
            # Sestavi prikazno ime uporabnika
            if first_name and last_name:
                display_name = f"{first_name} {last_name}"
            elif first_name:
                display_name = first_name
            elif last_name:
                display_name = last_name
            else:
                display_name = uploaded_by  # Fallback na username
            
            can_delete = False
            try:
                if current_user_id and uploader_user_id and int(current_user_id) == int(uploader_user_id):
                    can_delete = True
                else:
                    # dovolimo brisanje tujih slik le s posebnim dovoljenjem
                    can_delete = has_permission('delete_any_images')
            except Exception:
                can_delete = False

            images.append({
                'id': image_id,
                's3_key': s3_key,
                's3_url': s3_url,
                'uploaded_by': display_name,
                'uploaded_at': uploaded_at.isoformat() if hasattr(uploaded_at, 'isoformat') else str(uploaded_at),
                'user_id': uploader_user_id,
                'can_delete': can_delete
            })
        
        # Če smo očistili manjkajoče slike in jih ni več, ponastavi pripravil/nalivalec
        if cleaned_missing > 0 and len(images) == 0:
            try:
                order_number_with_hash = f"#{order_number}" if not str(order_number).startswith('#') else str(order_number)
                order_number_without_hash = str(order_number).lstrip('#')
                cursor.execute("""
                    UPDATE orders 
                    SET prepared_by = NULL, prepared_at = NULL, nalivalec_id = NULL
                    WHERE order_number = %s OR order_number = %s
                """, (order_number_with_hash, order_number_without_hash))
                db.commit()
                current_app.logger.info(f"Auto-reset preparation fields for order {order_number_with_hash} after cleaning missing images")
            except Exception as e:
                current_app.logger.error(f"Failed to auto-reset after cleaning images for {order_number}: {e}")
                db.rollback()
        
        return jsonify({
            "success": True,
            "images": images,
            "cleaned_missing": cleaned_missing
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri pridobivanju slik za naročilo {order_number}: {e}")
        return jsonify({"error": "Napaka pri pridobivanju slik"}), 500
    finally:
        cursor.close()

@api_bp.route('/order-images/proxy/<path:s3_key>', methods=['GET'])
def proxy_order_image_v2(s3_key):
    """Proxy S3 slike za order_images. Isti princip kot vrnjeni/poškodovani paketi (direkten get_object)."""
    try:
        # Dovolimo samo order_photos/*
        if not s3_key.startswith('order_photos/'):
            return jsonify({'error': 'Neveljavna pot slike'}), 403
        from services.s3_service import get_s3_client
        s3_client = get_s3_client()
        bucket_name = current_app.config['S3_BUCKET_NAME']
        response = s3_client.get_object(Bucket=bucket_name, Key=s3_key)
        image_data = response['Body'].read()
        content_type = response.get('ContentType', 'image/jpeg')
        return Response(
            image_data,
            mimetype=content_type,
            headers={
                'Cache-Control': 'public, max-age=3600',
                'Content-Disposition': 'inline'
            }
        )
    except Exception as e:
        current_app.logger.error(f"Napaka pri proxy_order_image_v2 {s3_key}: {e}")
        return jsonify({'error': 'Slika ni dostopna'}), 404

def _insert_order_image_record(cursor, order_number, order_number_with_hash, user_id, s3_key):
    # Najprej poiščimo user_id iz users tabele
    cursor.execute("SELECT id FROM users WHERE username = %s", (user_id,))
    user_result = cursor.fetchone()

    if user_result:
        db_user_id = user_result['id']
    else:
        # Če uporabnik ne obstaja, uporabimo admin ID
        cursor.execute("SELECT id FROM users WHERE username = 'admin'")
        admin_result = cursor.fetchone()
        db_user_id = admin_result['id'] if admin_result else None

    cursor.execute(
        """
            INSERT INTO order_images (order_number, s3_key, s3_url, uploaded_by, user_id)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id
        """,
        (
            order_number,
            s3_key,
            '',  # URL se generira sproti
            user_id,  # Ohranimo za kompatibilnost
            db_user_id,
        ),
    )

    result = cursor.fetchone()

    # Označi naročilo kot pripravljeno
    current_app.logger.info(
        f"Poskušam posodobiti prepared_by za naročilo {order_number_with_hash} z uporabnikom {user_id}"
    )
    # prepared_by se nastavlja ročno prek toolbarja (ne ob uploadu slike)
    if result is None:
        current_app.logger.error("INSERT ni vrnil ID")
        raise RuntimeError("INSERT ni vrnil ID")

    return result['id'] if isinstance(result, dict) else result[0]
@api_bp.route('/order-images/<order_number>', methods=['POST'])
def upload_order_image_api(order_number):
    """Naloži sliko za naročilo"""
    try:
        # Enforce authenticated session
        current_user_id = session.get('user_id')
        current_username = session.get('username')
        if not current_user_id or not current_username:
            return jsonify({"error": "Uporabnik ni prijavljen"}), 401

        MAX_IMAGE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
        ALLOWED_MIME_TYPES = {"image/jpeg", "image/jpg", "image/png", "image/webp"}

        image_data = None
        image_bytes = None
        user_id = current_username  # always trust session username

        content_type = request.content_type or ''
        file_mime = None
        if 'multipart/form-data' in content_type:
            # Podpora za neposreden upload datoteke prek FormData
            file = request.files.get('image')
            if not file:
                return jsonify({"error": "Manjka datoteka 'image'"}), 400
            # MIME type validation
            file_mime = (file.mimetype or '').lower()
            if file_mime not in ALLOWED_MIME_TYPES:
                return jsonify({"error": "Napačen tip datoteke. Dovoljeni: JPG, PNG, WEBP"}), 415
            # Read bytes for size and later re-use
            file_bytes = file.read()
            if not file_bytes:
                return jsonify({"error": "Prazna datoteka"}), 400
            if len(file_bytes) > MAX_IMAGE_SIZE_BYTES:
                return jsonify({"error": "Datoteka je prevelika (max 10 MB)"}), 413
            image_bytes = file_bytes
        else:
            # JSON payload (obstoječa pot)
            data = request.get_json(silent=True)
            if not data:
                return jsonify({"error": "Neveljaven ali prazen JSON payload"}), 400
            image_data = data.get('image_data')  # Base64 encoded image
            # Validate base64 size and content
            if not image_data:
                return jsonify({"error": "Manjka slika"}), 400
            try:
                import base64, binascii
                decoded = base64.b64decode(image_data.split(',')[-1], validate=True)
                if len(decoded) > MAX_IMAGE_SIZE_BYTES:
                    return jsonify({"error": "Slika je prevelika (max 10 MB)"}), 413
                image_bytes = decoded
            except (binascii.Error, ValueError):
                return jsonify({"error": "Slika ni pravilno kodirana (base64)"}), 400
            except Exception:
                return jsonify({"error": "Datoteka ni veljavna slika"}), 400
        
        if not image_data and not image_bytes:
            return jsonify({"error": "Manjka slika"}), 400
            
        # Backend validacija: za naročila s parfumi je obvezen nalivalec
        db = get_db()
        cursor = db.cursor()
        order_number_with_hash = f"#{order_number}" if not order_number.startswith('#') else order_number
        cursor.execute("SELECT line_items, nalivalec_id, status FROM orders WHERE order_number = %s", (order_number_with_hash,))
        order_row = cursor.fetchone()
        if not order_row:
            return jsonify({"error": "Naročilo ne obstaja"}), 404
        try:
            line_items_raw = order_row.get('line_items') if isinstance(order_row, dict) else order_row[0]
            nalivalec_id = order_row.get('nalivalec_id') if isinstance(order_row, dict) else order_row[1]
            order_status = order_row.get('status') if isinstance(order_row, dict) else order_row[2]
            import json as _json
            items = _json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])
            keywords = ['parfum', 'perfume', 'parfumi', 'eau de parfum', 'edp']
            def _norm(value: str) -> str:
                try:
                    import unicodedata
                    return unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
                except Exception:
                    return value
            def _is_parfumi(it):
                try:
                    ptype = _norm(((it or {}).get('product_type') or '').strip().lower())
                    if 'parfum' in ptype or 'perfume' in ptype:
                        return True
                    title = _norm(((it or {}).get('title') or '').strip().lower())
                    return any(k in title for k in keywords)
                except Exception:
                    return False
            has_parfumi = any(_is_parfumi(item) for item in items)
            if order_status == 'brez_parfumov':
                has_parfumi = False
            if has_parfumi:
                cursor.execute(
                    "UPDATE orders SET nalivalec_id = %s WHERE order_number = %s",
                    (current_user_id, order_number_with_hash),
                )
        except Exception:
            # Če pride do napake pri parsiranju, ne blokiraj nalaganja, le zabeleži
            current_app.logger.warning(f"Neuspešno preverjanje Parfumi/nalivalec za naročilo {order_number}")
        
        # Naloži sliko v S3 (hitrejša binarna pot, brez base64)
        if image_bytes:
            client_optimized = request.headers.get('X-Client-Optimized') == '1'
            skip_processing = bool(client_optimized) and (file_mime == 'image/jpeg')
            s3_result = upload_order_image_bytes(
                image_bytes,
                order_number,
                user_id,
                content_type=file_mime or 'image/jpeg',
                skip_processing=skip_processing
            )
        else:
            s3_result = upload_order_image(image_data, order_number, user_id)
        
        image_id = _insert_order_image_record(
            cursor,
            order_number,
            order_number_with_hash,
            user_id,
            s3_result['public_id'],
        )
        db.commit()
        
        return jsonify({
            "success": True,
            "message": "Slika uspešno naložena",
            "image": {
                "id": image_id,
                "url": s3_result['secure_url'],
                "public_id": s3_result['public_id'],
                "uploaded_by": user_id,
                "uploaded_at": s3_result['uploaded_at']
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri nalaganju slike za naročilo {order_number}: {e}")
        current_app.logger.error(f"Stack trace: {traceback.format_exc()}")
        return jsonify({"error": f"Napaka pri nalaganju slike: {str(e)}"}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()

@api_bp.route('/order-images/<order_number>/presign', methods=['POST'])
def presign_order_image_upload(order_number):
    """Ustvari presigned POST za direkten upload slike v S3."""
    try:
        current_user_id = session.get('user_id')
        current_username = session.get('username')
        if not current_user_id or not current_username:
            return jsonify({"error": "Uporabnik ni prijavljen"}), 401

        data = request.get_json(silent=True) or {}
        content_type = (data.get('content_type') or '').lower()
        file_size = int(data.get('file_size') or 0)

        MAX_IMAGE_SIZE_BYTES = 10 * 1024 * 1024
        ALLOWED_MIME_TYPES = {"image/jpeg", "image/jpg", "image/png", "image/webp"}
        if content_type not in ALLOWED_MIME_TYPES:
            return jsonify({"error": "Napačen tip datoteke. Dovoljeni: JPG, PNG, WEBP"}), 415
        if file_size <= 0 or file_size > MAX_IMAGE_SIZE_BYTES:
            return jsonify({"error": "Datoteka je prevelika (max 10 MB)"}), 413

        db = get_db()
        cursor = db.cursor()
        order_number_with_hash = f"#{order_number}" if not order_number.startswith('#') else order_number
        cursor.execute(
            "SELECT line_items, nalivalec_id, status FROM orders WHERE order_number = %s",
            (order_number_with_hash,),
        )
        order_row = cursor.fetchone()
        if not order_row:
            return jsonify({"error": "Naročilo ne obstaja"}), 404

        try:
            line_items_raw = order_row.get('line_items') if isinstance(order_row, dict) else order_row[0]
            nalivalec_id = order_row.get('nalivalec_id') if isinstance(order_row, dict) else order_row[1]
            order_status = order_row.get('status') if isinstance(order_row, dict) else order_row[2]
            import json as _json
            items = _json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])
            keywords = ['parfum', 'perfume', 'parfumi', 'eau de parfum', 'edp']
            def _norm(value: str) -> str:
                try:
                    import unicodedata
                    return unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
                except Exception:
                    return value

            def _is_parfumi(it):
                try:
                    ptype = _norm(((it or {}).get('product_type') or '').strip().lower())
                    if 'parfum' in ptype or 'perfume' in ptype:
                        return True
                    title = _norm(((it or {}).get('title') or '').strip().lower())
                    return any(k in title for k in keywords)
                except Exception:
                    return False

            has_parfumi = any(_is_parfumi(item) for item in items)
            if order_status == 'brez_parfumov':
                has_parfumi = False
            if has_parfumi:
                cursor.execute(
                    "UPDATE orders SET nalivalec_id = %s WHERE order_number = %s",
                    (current_user_id, order_number_with_hash),
                )
        except Exception:
            current_app.logger.warning(
                f"Neuspešno preverjanje Parfumi/nalivalec za naročilo {order_number}"
            )

        import uuid as _uuid
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_id = str(_uuid.uuid4())[:8]
        filename = f"order_{order_number}_{timestamp}_{unique_id}.jpg"
        s3_key = f"order_photos/{filename}"

        presigned = generate_presigned_post_for_order_image(
            s3_key, content_type, MAX_IMAGE_SIZE_BYTES, expiration=600
        )

        return jsonify(
            {
                "success": True,
                "url": presigned["url"],
                "fields": presigned["fields"],
                "s3_key": s3_key,
            }
        )
    except Exception as e:
        current_app.logger.error(f"Napaka pri presign uploadu slike: {e}")
        current_app.logger.error(f"Stack trace: {traceback.format_exc()}")
        return jsonify({"error": "Napaka pri pripravi nalaganja slike"}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()

@api_bp.route('/order-images/<order_number>/finalize', methods=['POST'])
def finalize_order_image_upload(order_number):
    """Shrani zapis slike v bazo po direktnem S3 uploadu."""
    try:
        current_user_id = session.get('user_id')
        current_username = session.get('username')
        if not current_user_id or not current_username:
            return jsonify({"error": "Uporabnik ni prijavljen"}), 401

        data = request.get_json(silent=True) or {}
        s3_key = data.get('s3_key')
        if not s3_key or not s3_key.startswith('order_photos/'):
            return jsonify({"error": "Neveljaven ključ slike"}), 400

        MAX_IMAGE_SIZE_BYTES = 10 * 1024 * 1024
        ALLOWED_MIME_TYPES = {"image/jpeg", "image/jpg", "image/png", "image/webp"}

        s3_client = get_s3_client()
        bucket_name = current_app.config['S3_BUCKET_NAME']
        head = s3_client.head_object(Bucket=bucket_name, Key=s3_key)
        size = int(head.get('ContentLength') or 0)
        content_type = (head.get('ContentType') or '').lower()
        if size <= 0 or size > MAX_IMAGE_SIZE_BYTES:
            return jsonify({"error": "Slika je prevelika (max 10 MB)"}), 413
        if content_type not in ALLOWED_MIME_TYPES:
            return jsonify({"error": "Napačen tip datoteke. Dovoljeni: JPG, PNG, WEBP"}), 415

        db = get_db()
        cursor = db.cursor()
        order_number_with_hash = f"#{order_number}" if not order_number.startswith('#') else order_number
        cursor.execute(
            "SELECT id, line_items, nalivalec_id, status FROM orders WHERE order_number = %s",
            (order_number_with_hash,),
        )
        order_row = cursor.fetchone()
        if not order_row:
            return jsonify({"error": "Naročilo ne obstaja"}), 404
        try:
            line_items_raw = order_row.get('line_items') if isinstance(order_row, dict) else order_row[1]
            order_status = order_row.get('status') if isinstance(order_row, dict) else order_row[3]
            items = _json.loads(line_items_raw) if isinstance(line_items_raw, str) else (line_items_raw or [])
            keywords = ['parfum', 'perfume', 'parfumi', 'eau de parfum', 'edp']
            def _norm(value: str) -> str:
                try:
                    import unicodedata
                    return unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
                except Exception:
                    return value

            def _is_parfumi(it):
                try:
                    ptype = _norm(((it or {}).get('product_type') or '').strip().lower())
                    if 'parfum' in ptype or 'perfume' in ptype:
                        return True
                    title = _norm(((it or {}).get('title') or '').strip().lower())
                    return any(k in title for k in keywords)
                except Exception:
                    return False

            has_parfumi = any(_is_parfumi(item) for item in items)
            if order_status == 'brez_parfumov':
                has_parfumi = False
            if has_parfumi:
                cursor.execute(
                    "UPDATE orders SET nalivalec_id = %s WHERE order_number = %s",
                    (current_user_id, order_number_with_hash),
                )
        except Exception:
            current_app.logger.warning(
                f"Neuspešno preverjanje Parfumi/nalivalec za naročilo {order_number}"
            )

        image_id = _insert_order_image_record(
            cursor,
            order_number,
            order_number_with_hash,
            current_username,
            s3_key,
        )
        db.commit()

        secure_url = generate_presigned_url(s3_key, expiration=3600 * 24 * 7, operation='get_object')
        return jsonify(
            {
                "success": True,
                "message": "Slika uspešno naložena",
                "image": {
                    "id": image_id,
                    "url": secure_url,
                    "public_id": s3_key,
                    "uploaded_by": current_username,
                    "uploaded_at": datetime.now().isoformat(),
                },
            }
        )
    except Exception as e:
        current_app.logger.error(f"Napaka pri finalize uploadu slike: {e}")
        current_app.logger.error(f"Stack trace: {traceback.format_exc()}")
        return jsonify({"error": "Napaka pri shranjevanju slike"}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()

@api_bp.route('/order-images/<int:image_id>', methods=['DELETE'])
def delete_order_image_api(image_id):
    """Izbriši sliko naročila"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Pridobi podatke o sliki
        cursor.execute("SELECT s3_key, order_number, user_id FROM order_images WHERE id = %s", (image_id,))
        image_data = cursor.fetchone()
        
        if not image_data:
            return jsonify({"error": "Slika ni najdena"}), 404
        
        # psycopg vrača rezultate kot slovarje
        if isinstance(image_data, dict):
            s3_key = image_data['s3_key']
            order_number = image_data['order_number']
            uploader_user_id = image_data.get('user_id')
        else:
            s3_key, order_number, uploader_user_id = image_data

        # Dovoljenja: lastnik lahko briše; brisanje tujih slik samo z dovoljenjem 'delete_any_images'
        current_user_id = session.get('user_id')
        if not current_user_id:
            return jsonify({"error": "Uporabnik ni prijavljen"}), 401
        is_owner = uploader_user_id and int(uploader_user_id) == int(current_user_id)
        if not (is_owner or has_permission('delete_any_images')):
            return jsonify({"error": "Nimate dovoljenja za brisanje te slike"}), 403
        
        # Poskusi izbrisati iz S3, a ne blokiraj celotne operacije, če S3 delete pade
        from services.s3_service import delete_order_image
        s3_deleted = False
        try:
            s3_deleted = delete_order_image(s3_key)
        except Exception as e:
            current_app.logger.error(f"S3 delete exception for key {s3_key}: {e}")
            s3_deleted = False

        # V vsakem primeru izbriši zapis iz baze, da UI ne obstane blokiran
        cursor.execute("DELETE FROM order_images WHERE id = %s", (image_id,))

        # Preveri, ali obstajajo še druge slike za to naročilo
        order_number_clean = order_number.lstrip('#')
        cursor.execute("SELECT COUNT(*) as count FROM order_images WHERE order_number = %s", (order_number_clean,))
        remaining_images = cursor.fetchone()
        remaining_count = remaining_images['count'] if isinstance(remaining_images, dict) else remaining_images[0]

        # Če ni več slik za to naročilo, ponastavi "Pripravil" status in nalivalca
        if remaining_count == 0:
            order_with_hash = order_number if str(order_number).startswith('#') else f"#{order_number}"
            order_without_hash = str(order_number).lstrip('#')
            cursor.execute(
                """
                    UPDATE orders 
                    SET prepared_by = NULL, prepared_at = NULL, nalivalec_id = NULL
                    WHERE order_number = %s OR order_number = %s
                """,
                (order_with_hash, order_without_hash)
            )
            current_app.logger.info(
                f"Reset prepared_by and nalivalec_id for order {order_number} - no images remaining (both variants)"
            )

        db.commit()

        return jsonify({
            "success": True,
            "message": "Slika izbrisana",
            "remaining_images": remaining_count,
            "prepared_by_reset": remaining_count == 0,
            "s3_deleted": s3_deleted
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri brisanju slike {image_id}: {e}")
        return jsonify({"error": "Napaka pri brisanju slike"}), 500
    finally:
        cursor.close()
# --- Vrnjeni & poškodovani paketi ---
@api_bp.route('/returns', methods=['GET'])
def list_returns():
    """Vrne sezname vrnjenih in poškodovanih paketov s paginacijo in filtriranjem."""
    try:
        current_app.logger.info("list_returns: Starting function")
        db = get_db()
        current_app.logger.info(f"list_returns: Got database connection: {db}")
        cursor = db.cursor()
        current_app.logger.info(f"list_returns: Got cursor: {cursor}")
        current_app.logger.info("list_returns: About to execute CREATE TABLE")
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS returned_damaged (
                id SERIAL PRIMARY KEY,
                order_number VARCHAR(64) NOT NULL,
                type VARCHAR(16) NOT NULL CHECK (type IN ('returned','damaged')),
                note TEXT,
                image_urls JSONB DEFAULT '[]'::jsonb,
                created_by INTEGER,
                created_at TIMESTAMPTZ DEFAULT NOW(),
                UNIQUE(order_number, type)
            );
        """)
        current_app.logger.info("list_returns: CREATE TABLE executed successfully")
        
        # Query parametri za filtriranje in paginacijo
        current_app.logger.info("list_returns: Processing request parameters")
        page = max(1, int(request.args.get('page', 1)))
        per_page = min(50, max(1, int(request.args.get('per_page', 10))))
        search_query = request.args.get('search', '').strip()
        type_filter = request.args.get('type', '').strip()  # 'returned', 'damaged', ali prazno za vse
        current_app.logger.info(f"list_returns: Parameters - page={page}, per_page={per_page}, search='{search_query}', type='{type_filter}'")
        
        # Zgradimo WHERE clause
        where_conditions = []
        params = []
        
        if search_query:
            where_conditions.append("rd.order_number ILIKE %s")
            params.append(f"%{search_query}%")
        
        if type_filter in ('returned', 'damaged'):
            where_conditions.append("rd.type = %s")
            params.append(type_filter)
        
        where_clause = " WHERE " + " AND ".join(where_conditions) if where_conditions else ""
        
        # Preštej skupno število zapisov  
        count_sql = f"SELECT COUNT(*) FROM returned_damaged rd{where_clause}"
        current_app.logger.info(f"list_returns count_sql: {count_sql}, params: {params}")
        cursor.execute(count_sql, params)
        count_result = cursor.fetchone()
        current_app.logger.info(f"list_returns count_result: {count_result}")
        # With dict_row, the result is a dict, not a tuple
        total_count = count_result['count'] if count_result else 0
        
        # Pridobi zapise za trenutno stran
        offset = (page - 1) * per_page
        params_with_pagination = params + [per_page, offset]
        
        data_sql = f"""
            SELECT rd.id, rd.order_number, rd.type, rd.note, rd.image_urls, rd.created_at, rd.created_by,
                   u.first_name, u.last_name, u.username
            FROM returned_damaged rd
            LEFT JOIN users u ON rd.created_by = u.id
            {where_clause} 
            ORDER BY rd.created_at DESC 
            LIMIT %s OFFSET %s
        """
        current_app.logger.info(f"list_returns data_sql: {data_sql}, params_with_pagination: {params_with_pagination}")
        cursor.execute(data_sql, params_with_pagination)
        rows = cursor.fetchall()
        current_app.logger.info(f"list_returns rows fetched: {len(rows)}")
        
        # Razdeli na returned in damaged za kompatibilnost
        returned = []
        damaged = []
        all_items = []
        
        for r in rows:
            current_app.logger.info(f"list_returns processing row: {r}, type: {type(r)}")
            # Podpora za dict ali tuple rezultat
            try:
                rid = r['id'] if isinstance(r, dict) else r[0]
                order_number = r['order_number'] if isinstance(r, dict) else r[1]
                rtype = r['type'] if isinstance(r, dict) else r[2]
                note = r.get('note') if isinstance(r, dict) else r[3]
                image_urls = r.get('image_urls') if isinstance(r, dict) else r[4]
                created_at = r.get('created_at') if isinstance(r, dict) else r[5]
                created_by = r.get('created_by') if isinstance(r, dict) else r[6]
                first_name = r.get('first_name') if isinstance(r, dict) else r[7]
                last_name = r.get('last_name') if isinstance(r, dict) else r[8]
                username = r.get('username') if isinstance(r, dict) else r[9]
                
                # Pretvori S3 ključe v proxy URL-je
                s3_keys = image_urls if isinstance(image_urls, list) else (json.loads(image_urls) if image_urls else [])
                images = []
                for s3_key in s3_keys:
                    if isinstance(s3_key, str) and (s3_key.startswith('returned_packages/') or s3_key.startswith('damaged_packages/')):
                        # Ustvari proxy URL
                        from flask import url_for
                        proxy_url = url_for('api.proxy_returned_damaged_image', s3_key=s3_key, _external=True)
                        images.append(proxy_url)
                    elif s3_key.startswith('http'):
                        # Že je URL (za nazaj kompatibilnost), uporabi kar je
                        images.append(s3_key)
                    else:
                        # Neprepoznaven format, preskoči
                        continue
                
                # Sestavi prikazno ime uporabnika (podobno kot pri order images)
                if first_name and last_name:
                    display_name = f"{first_name} {last_name}"
                elif first_name:
                    display_name = first_name
                elif last_name:
                    display_name = last_name
                elif username:
                    display_name = username
                else:
                    display_name = 'Neznano'
                    
            except Exception as e:
                current_app.logger.error(f"list_returns row processing error: {e}")
                continue
            
            rec = {
                'id': rid,
                'order_number': order_number,
                'type': rtype,
                'note': note or '',
                'images': images,
                'created_at': created_at.isoformat() if hasattr(created_at, 'isoformat') else str(created_at),
                'created_by_name': display_name,
                'created_by': created_by,
                'first_name': first_name,
                'last_name': last_name,
                'username': username
            }
            
            all_items.append(rec)
            (returned if rtype == 'returned' else damaged).append(rec)
        
        # Računaj pagination info
        total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1
        has_next = page < total_pages
        has_prev = page > 1
        
        current_app.logger.info(f"list_returns response: total_count={total_count}, all_items_len={len(all_items)}, returned_len={len(returned)}, damaged_len={len(damaged)}")
        
        return make_ok({
            'returned': returned,
            'damaged': damaged,
            'all_items': all_items,  # Nova struktura za unified prikaz
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total_count': total_count,
                'total_pages': total_pages,
                'has_next': has_next,
                'has_prev': has_prev
            },
            'filters': {
                'search': search_query,
                'type': type_filter
            }
        })
    except Exception as e:
        current_app.logger.error(f"Napaka pri list_returns: {e}")
        return make_err('SERVER_ERROR', f'Napaka pri pridobivanju vrnitev: {str(e)}', status=500)
@api_bp.route('/returns', methods=['POST'])
def create_return():
    """Ustvari zapis vrnjenega/poškodovanega paketa in shrani slike."""
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': 'Uporabnik ni prijavljen'}), 401
        order_number = request.form.get('order_number')
        rtype = request.form.get('type', 'returned')
        note = request.form.get('note')
        if not order_number:
            return jsonify({'success': False, 'error': 'Manjka št. naročila'}), 400
        if rtype not in ('returned','damaged'):
            return jsonify({'success': False, 'error': 'Neveljaven tip'}), 400

        image_urls = []
        files = []
        try:
            files = request.files.getlist('images')
        except Exception:
            files = []
        if not files or not any(f and getattr(f, 'filename', None) for f in files):
            return jsonify({'success': False, 'error': 'Slika je obvezna'}), 400
        
        # Preveri če že obstaja zapis za to naročilo in tip
        db = get_db()
        cursor = db.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS returned_damaged (
                id SERIAL PRIMARY KEY,
                order_number VARCHAR(64) NOT NULL,
                type VARCHAR(16) NOT NULL CHECK (type IN ('returned','damaged')),
                note TEXT,
                image_urls JSONB DEFAULT '[]'::jsonb,
                created_by INTEGER,
                created_at TIMESTAMPTZ DEFAULT NOW(),
                UNIQUE(order_number, type)
            );
        """)
        
        cursor.execute("SELECT id FROM returned_damaged WHERE order_number = %s AND type = %s", (order_number, rtype))
        if cursor.fetchone():
            return jsonify({'success': False, 'error': f'Za naročilo {order_number} že obstaja zapis tipa "{rtype}"'}), 400
        
        for f in files:
            try:
                uploaded = upload_returned_damaged_image(f, order_number, rtype, session['user_id'])
                if uploaded and uploaded.get('key'):
                    # Shranjuj S3 ključ namesto URL-ja, da se izognemo CORS problemom
                    image_urls.append(uploaded['key'])
            except Exception as e:
                current_app.logger.warning(f"Napaka pri nalaganju slike za {rtype} paket: {e}")

        db = get_db()
        cursor = db.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS returned_damaged (
                id SERIAL PRIMARY KEY,
                order_number VARCHAR(64) NOT NULL,
                type VARCHAR(16) NOT NULL CHECK (type IN ('returned','damaged')),
                note TEXT,
                image_urls JSONB DEFAULT '[]'::jsonb,
                created_by INTEGER,
                created_at TIMESTAMPTZ DEFAULT NOW()
            );
        """)
        try:
            cursor.execute(
                """
                INSERT INTO returned_damaged (order_number, type, note, image_urls, created_by)
                VALUES (%s, %s, %s, %s::jsonb, %s)
                RETURNING id
                """,
                (order_number, rtype, note, json.dumps(image_urls), session['user_id'])
            )
            db.commit()
            return jsonify({'success': True})
        except Exception as insert_error:
            db.rollback()
            if "duplicate key" in str(insert_error).lower() or "unique constraint" in str(insert_error).lower():
                return jsonify({'success': False, 'error': f'Za naročilo {order_number} že obstaja zapis tipa "{rtype}"'}), 400
            else:
                raise insert_error
    except Exception as e:
        current_app.logger.error(f"Napaka pri create_return: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri shranjevanju'}), 500

@api_bp.route('/returns/images', methods=['GET'])
def get_returns_images():
    """Pridobi slike vrnjenih/poškodovanih paketov iz S3."""
    try:
        order_number = request.args.get('order_number')
        package_type = request.args.get('type')  # 'returned', 'damaged' ali None za vse
        
        images_data = get_returned_damaged_images(order_number=order_number, package_type=package_type)
        
        # Izvleci samo URL-je za enostavno uporabo v frontend-u
        image_urls = [img['secure_url'] for img in images_data if img.get('secure_url')]
        
        return jsonify({
            'success': True,
            'images': image_urls,
            'count': len(image_urls),
            'full_data': images_data  # Za debug ali dodatne potrebe
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri get_returns_images: {e}")
        return jsonify({
            'success': False, 
            'error': 'Napaka pri pridobivanju slik',
            'images': [],
            'count': 0
        }), 500

@api_bp.route('/returns/<int:record_id>', methods=['DELETE'])
def delete_return(record_id):
    """Izbriši zapis vrnjenega/poškodovanega paketa."""
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': 'Uporabnik ni prijavljen'}), 401
            
        db = get_db()
        cursor = db.cursor()
        
        # Preveri če zapis obstaja
        cursor.execute("SELECT id FROM returned_damaged WHERE id = %s", (record_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Zapis ni najden'}), 404
        
        # Izbriši zapis
        cursor.execute("DELETE FROM returned_damaged WHERE id = %s", (record_id,))
        db.commit()
        
        current_app.logger.info(f"Deleted returned/damaged record {record_id} by user {session['user_id']}")
        return jsonify({'success': True})
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri delete_return: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri brisanju'}), 500

@api_bp.route('/setup-s3-cors', methods=['POST'])
def setup_s3_cors_endpoint():
    """Nastavi CORS konfiguracijo za S3 bucket"""
    try:
        success = setup_s3_cors()
        if success:
            return jsonify({'success': True, 'message': 'CORS konfiguracija uspešno nastavljena'})
        else:
            return jsonify({'success': False, 'error': 'Napaka pri nastavljanju CORS konfiguracije'}), 500
    except Exception as e:
        current_app.logger.error(f"Napaka pri nastavljanju CORS: {e}")
        return jsonify({'success': False, 'error': 'Napaka pri nastavljanju CORS'}), 500

@api_bp.route('/returns/images/proxy/<path:s3_key>', methods=['GET'])
def proxy_returned_damaged_image(s3_key):
    """Proxy S3 slike da se izognemo CORS problemom"""
    try:
        # Preverimo, da je s3_key varen (samo returned_packages/ ali damaged_packages/)
        if not (s3_key.startswith('returned_packages/') or s3_key.startswith('damaged_packages/')):
            return jsonify({'error': 'Neveljavna pot slike'}), 403
            
        from services.s3_service import get_s3_client
        import io
        
        s3_client = get_s3_client()
        bucket_name = current_app.config['S3_BUCKET_NAME']
        
        # Pridobi sliko iz S3
        response = s3_client.get_object(Bucket=bucket_name, Key=s3_key)
        image_data = response['Body'].read()
        content_type = response.get('ContentType', 'image/jpeg')
        
        # Vrni sliko kot Response
        from flask import Response
        return Response(
            image_data,
            mimetype=content_type,
            headers={
                'Cache-Control': 'public, max-age=3600',
                'Content-Disposition': 'inline'
            }
        )
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri proxy slike {s3_key}: {e}")
        return jsonify({'error': 'Slika ni dostopna'}), 404

@api_bp.route('/health', methods=['GET', 'HEAD'])
def health_check():
    """Health check endpoint for connection testing"""
    return jsonify({"status": "ok"}), 200

# --- Navodila (Instructions) ---

@api_bp.route('/instruction-categories', methods=['GET'])
def list_instruction_categories():
    db = get_db()
    c = db.cursor()
    try:
        c.execute("SELECT id, name FROM instruction_categories ORDER BY name ASC")
        rows = c.fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        current_app.logger.error(f"Napaka pri branju kategorij navodil: {e}")
        return jsonify([])
    finally:
        c.close()
@api_bp.route('/instructions', methods=['GET'])
def list_instructions():
    db = get_db()
    c = db.cursor()
    try:
        category_id = request.args.get('category_id')
        if category_id:
            c.execute("SELECT id, category_id, title, content, created_by, updated_by, created_at, updated_at FROM instructions WHERE category_id = %s ORDER BY created_at DESC", (category_id,))
        else:
            c.execute("SELECT id, category_id, title, content, created_by, updated_by, created_at, updated_at FROM instructions ORDER BY created_at DESC")
        rows = c.fetchall()
        out = []
        for r in rows:
            d = dict(r)
            if d.get('created_at'):
                d['created_at'] = d['created_at'].isoformat()
            if d.get('updated_at'):
                d['updated_at'] = d['updated_at'].isoformat()
            out.append(d)
        return jsonify(out)
    except Exception as e:
        current_app.logger.error(f"Napaka pri branju navodil: {e}")
        return jsonify([])
    finally:
        c.close()

@api_bp.route('/instructions/<int:instruction_id>', methods=['PUT'])
def update_instruction(instruction_id):
    # Le admin (prek enforce_permissions -> edit_users)
    data = request.get_json() or {}
    title = data.get('title')
    content = data.get('content')
    category_id = data.get('category_id')
    db = get_db()
    c = db.cursor()
    try:
        # Defensive ensure table
        c.execute("CREATE TABLE IF NOT EXISTS instructions (id SERIAL PRIMARY KEY, category_id INTEGER, title VARCHAR(255) NOT NULL, content TEXT NOT NULL, created_by VARCHAR(255), updated_by VARCHAR(255), created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP WITH TIME ZONE)")
        fields = []
        params = []
        if title is not None:
            fields.append('title = %s')
            params.append(title)
        if content is not None:
            fields.append('content = %s')
            params.append(content)
        if category_id is not None:
            fields.append('category_id = %s')
            params.append(category_id)
        # Audit user
        current_user = session.get('user', {})
        updated_by = f"{current_user.get('first_name','')} {current_user.get('last_name','')}".strip() or session.get('username', 'admin')
        fields.append('updated_by = %s')
        params.append(updated_by)
        fields.append('updated_at = CURRENT_TIMESTAMP')
        set_clause = ', '.join(fields)
        params.append(instruction_id)
        c.execute(f"UPDATE instructions SET {set_clause} WHERE id = %s", params)
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri posodabljanju navodila: {e}")
        return jsonify({"error": "Napaka na strežniku"}), 500
    finally:
        c.close()

@api_bp.route('/instructions/<int:instruction_id>', methods=['DELETE'])
def delete_instruction(instruction_id):
    db = get_db()
    c = db.cursor()
    try:
        c.execute("DELETE FROM instructions WHERE id = %s", (instruction_id,))
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri brisanju navodila: {e}")
        return jsonify({"error": "Napaka na strežniku"}), 500
    finally:
        c.close()

@api_bp.route('/instructions/<int:instruction_id>/images', methods=['POST'])
def upload_instruction_image_api(instruction_id):
    # le admin (prek enforce_permissions)
    try:
        if 'image' not in request.files:
            return jsonify({"error": "Manjka slika (image)"}), 400
        file_storage = request.files['image']
        current_user_id = session.get('user_id')
        res = upload_instruction_image(file_storage, current_user_id)
        return jsonify({"success": True, **res})
    except Exception as e:
        current_app.logger.error(f"Napaka pri uploadu slike za navodila: {e}")
        return jsonify({"error": "Napaka pri nalaganju slike"}), 500

@api_bp.route('/instruction-categories', methods=['POST'])
def create_instruction_category():
    # Le admin (prek enforce_permissions -> edit_users)
    data = request.get_json() or {}
    name = data.get('name')
    if not name:
        return jsonify({"error": "Manjka ime kategorije"}), 400
    db = get_db()
    c = db.cursor()
    try:
        # Defensive: ensure table exists if migrations didn't run yet
        c.execute("""
            CREATE TABLE IF NOT EXISTS instruction_categories (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) UNIQUE NOT NULL,
                created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
            );
        """)
        c.execute("INSERT INTO instruction_categories (name) VALUES (%s) ON CONFLICT (name) DO NOTHING RETURNING id", (name,))
        row = c.fetchone()
        db.commit()
        return jsonify({"success": True, "id": (row['id'] if row else None)})
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri dodajanju kategorije: {e}")
        return jsonify({"error": "Napaka na strežniku"}), 500
    finally:
        c.close()
@api_bp.route('/instructions', methods=['POST'])
def create_instruction():
    # Le admin (prek enforce_permissions -> edit_users)
    data = request.get_json() or {}
    category_id = data.get('category_id')
    title = data.get('title')
    content = data.get('content')
    if not all([title, content]):
        return jsonify({"error": "Manjka naslov ali vsebina"}), 400
    db = get_db()
    c = db.cursor()
    try:
        # Defensive: ensure tables exist
        c.execute("""
            CREATE TABLE IF NOT EXISTS instruction_categories (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) UNIQUE NOT NULL,
                created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
            );
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS instructions (
                id SERIAL PRIMARY KEY,
                category_id INTEGER REFERENCES instruction_categories(id) ON DELETE SET NULL,
                title VARCHAR(255) NOT NULL,
                content TEXT NOT NULL,
                created_by VARCHAR(255),
                updated_by VARCHAR(255),
                created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP WITH TIME ZONE
            );
        """)
        current_user = session.get('user', {})
        created_by = f"{current_user.get('first_name','')} {current_user.get('last_name','')}".strip() or session.get('username', 'admin')
        c.execute("""
            INSERT INTO instructions (category_id, title, content, created_by)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (category_id, title, content, created_by))
        row = c.fetchone()
        db.commit()
        return jsonify({"success": True, "id": row['id']})
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri dodajanju navodila: {e}")
        return jsonify({"error": "Napaka na strežniku"}), 500
    finally:
        c.close()

@api_bp.route('/search-synonyms', methods=['GET'])
def list_search_synonyms():
    db = get_db()
    c = db.cursor()
    try:
        shop_domain = (request.args.get('shop_domain') or '').strip()
        q = (request.args.get('q') or '').strip()
        limit = request.args.get('limit')
        try:
            limit = int(limit) if limit else 200
        except Exception:
            limit = 200
        limit = max(1, min(limit, 1000))

        where = []
        params = []
        if shop_domain:
            where.append("s.shop_domain = %s")
            params.append(shop_domain)
        if q:
            like = f"%{q}%"
            where.append(
                "(s.phrase_norm ILIKE %s OR s.phrase_raw ILIKE %s OR s.target_code ILIKE %s OR t.product_handle ILIKE %s)"
            )
            params.extend([like, like, like, like])
        where_sql = f"WHERE {' AND '.join(where)}" if where else ""

        c.execute(
            f"""
            SELECT s.id, s.shop_domain, s.phrase_norm, s.phrase_raw, s.target_code,
                   t.product_handle, t.product_id, s.created_at, s.updated_at
            FROM search_synonyms s
            LEFT JOIN inspo_targets t
              ON t.shop_domain = s.shop_domain AND t.target_code = s.target_code
            {where_sql}
            ORDER BY s.updated_at DESC, s.id DESC
            LIMIT %s
            """,
            (*params, limit),
        )
        rows = [dict(r) for r in c.fetchall()]
        return make_ok(rows)
    except Exception as e:
        current_app.logger.error(f"Napaka pri branju search synonymov: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri branju sinonimov', status=500)
    finally:
        c.close()

@api_bp.route('/shopify-stores', methods=['GET'])
def list_shopify_stores():
    try:
        db = get_db()
        cursor = db.cursor()
        out = []
        seen: set[str] = set()
        try:
            cursor.execute(
                """
                SELECT shop_domain, is_active
                FROM shopify_stores
                ORDER BY shop_domain
                """
            )
            for row in cursor.fetchall() or []:
                domain = row['shop_domain'] if isinstance(row, dict) else row[0]
                is_active = row['is_active'] if isinstance(row, dict) else row[1]
                if not domain or domain in seen:
                    continue
                seen.add(domain)
                out.append({
                    'shop_domain': domain,
                    'is_active': bool(is_active),
                    'is_default': False,
                    'is_sync_default': domain == DEFAULT_SYNC_STORE,
                })
        finally:
            cursor.close()

        if not out:
            stores = get_all_shopify_stores(include_default=True)
            for store in stores:
                domain = store.get('shop_domain')
                if not domain or domain in seen:
                    continue
                seen.add(domain)
                out.append({
                    'shop_domain': domain,
                    'is_active': bool(store.get('is_active', True)),
                    'is_default': bool(store.get('is_default', False)),
                    'is_sync_default': domain == DEFAULT_SYNC_STORE,
                })

        out.sort(
            key=lambda s: (
                not s.get('is_sync_default', False),
                not s.get('is_active', True),
                s.get('shop_domain', ''),
            )
        )
        return make_ok(out)
    except Exception as e:
        current_app.logger.error(f"Napaka pri branju Shopify trgovin: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri branju Shopify trgovin', status=500)


@api_bp.route('/search-synonyms', methods=['POST'])
def upsert_search_synonym():
    data = request.get_json(silent=True) or {}
    shop_domain = (data.get('shop_domain') or '').strip()
    phrase = (data.get('phrase') or '').strip()
    target_code = (data.get('target_code') or '').strip()
    product_handle = (data.get('product_handle') or '').strip() or None
    product_id = data.get('product_id')
    try:
        product_id = int(product_id) if product_id not in (None, '') else None
    except Exception:
        return make_err('BAD_REQUEST', 'product_id mora biti celo število', status=400)

    if not shop_domain or not phrase or not target_code:
        return make_err('BAD_REQUEST', 'Manjkajo podatki (shop_domain, phrase, target_code)', status=400)

    phrase_norm = normalize_query(phrase)
    if not phrase_norm:
        return make_err('BAD_REQUEST', 'Neveljaven izraz za iskanje', status=400)

    try:
        synonym_id = upsert_synonym(shop_domain, phrase, target_code)
        upsert_inspo_target(shop_domain, target_code, product_handle=product_handle, product_id=product_id)
        return make_ok({'id': synonym_id, 'phrase_norm': phrase_norm})
    except Exception as e:
        current_app.logger.error(f"Napaka pri shranjevanju search synonyma: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri shranjevanju', status=500)


@api_bp.route('/search-synonyms/<int:syn_id>', methods=['DELETE'])
def delete_search_synonym(syn_id):
    db = get_db()
    c = db.cursor()
    try:
        c.execute("DELETE FROM search_synonyms WHERE id = %s", (syn_id,))
        db.commit()
        return make_ok({'deleted': True})
    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Napaka pri brisanju search synonyma: {e}")
        return make_err('SERVER_ERROR', 'Napaka pri brisanju', status=500)
    finally:
        c.close()

@api_bp.route('/run-migration', methods=['POST'])
def run_migration_endpoint():
    """Temporary endpoint to run database migration"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Add the shopify_fulfilled_at column
        cursor.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS shopify_fulfilled_at TIMESTAMP WITH TIME ZONE;")
        
        # Add index for better performance
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_orders_shopify_fulfilled_at ON orders(shopify_fulfilled_at);")
        
        # Add prepared_by and prepared_at columns
        cursor.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS prepared_by VARCHAR(255);")
        cursor.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS prepared_at TIMESTAMP;")
        
        # Add app_settings table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS app_settings (
                id SERIAL PRIMARY KEY,
                key VARCHAR(255) UNIQUE NOT NULL,
                value TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Add index for app_settings
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_app_settings_key ON app_settings(key)")
        
        # Add trigger for app_settings
        cursor.execute("""
            CREATE OR REPLACE FUNCTION update_updated_at_column() RETURNS TRIGGER AS $$
            BEGIN
                NEW.updated_at = CURRENT_TIMESTAMP;
                RETURN NEW;
            END;
            $$ language 'plpgsql'
        """)
        
        cursor.execute("""
            DROP TRIGGER IF EXISTS update_app_settings_updated_at ON app_settings
        """)
        
        cursor.execute("""
            CREATE TRIGGER update_app_settings_updated_at 
                BEFORE UPDATE ON app_settings 
                FOR EACH ROW 
                EXECUTE FUNCTION update_updated_at_column()
        """)
        
        # Add default email_test_mode setting
        cursor.execute("""
            INSERT INTO app_settings (key, value) 
            VALUES ('email_test_mode', 'true')
            ON CONFLICT (key) DO NOTHING
        """)
        
        db.commit()
        
        return jsonify({
            "success": True,
            "message": "Migration completed successfully - all columns and app_settings table added"
        })
        
    except Exception as e:
        current_app.logger.error(f"Migration failed: {e}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
    finally:
        cursor.close()
def _pridobi_podatke_za_deklaracijo_iz_shopify(line_items, db_cursor, shop_domain: str | None = None):
    """
    Pridobi podatke za deklaracijo iz line_items, ki imajo samo product_id.
    Pridobi product_no in proizvajalec_ime iz Shopify-ja.
    Združi line_items z istim parfumom (product_no + proizvajalec_ime).

    POMEMBNO: shop_domain MORA biti podan za multi-store setupe. Brez njega
    Shopify klic uporabi default trgovino in product_id-ji iz drugih
    trgovin (npr. SI vs EU) bodo silent fail -> prazne deklaracije.
    """
    from services.shopify_service import get_bulk_product_details, clear_product_cache

    declaration_items, missing_data_details, expiration_warnings = [], [], []

    # Pridobi vse product_id iz line_items
    product_ids = [str(item['product_id']) for item in line_items if item and item.get('product_id')]

    if not product_ids:
        missing_data_details.append("Ni product_id v line_items")
        return declaration_items, missing_data_details, expiration_warnings

    # Počisti cache in pridobi podatke iz Shopify-ja (z eksplicitnim shop_domain)
    clear_product_cache()
    shopify_details = get_bulk_product_details(product_ids, shop_domain=shop_domain)
    current_app.logger.info(
        f"Pridobil podatke iz Shopify-ja (shop={shop_domain or 'default'}) za "
        f"{len(product_ids)} izdelkov: {shopify_details}"
    )

    # Multi-store guard: če Shopify ni vrnil ničesar, je to skoraj zagotovo
    # napaka pri izbiri trgovine (klic pristal na default store, kjer
    # product_id-ji ne obstajajo). Vrnemo eksplicitno missing_data, da
    # admin v dnevnem reportu vidi vzrok in lahko popravi.
    if not shopify_details:
        missing_data_details.append(
            f"Shopify ni vrnil podatkov za nobenega od {len(product_ids)} izdelkov "
            f"(shop_domain='{shop_domain or 'default'}'). Preveri shopify_store_domain "
            f"na naročilu in Shopify access token za to trgovino."
        )
        return declaration_items, missing_data_details, expiration_warnings
    
    # Slovar za združevanje line_items z istim parfumom
    parfum_groups = {}
    
    for item in line_items:
        if not item or not item.get('product_id'):
            continue
            
        product_id_str = str(item.get('product_id'))
        details = shopify_details.get(product_id_str, {})
        current_app.logger.info(f"Obdelujem izdelek {product_id_str}: {details}")
        
        # Preveri, ali je to parfum
        product_type = details.get('product_type')
        if not product_type or product_type.strip().lower() != 'parfumi':
            current_app.logger.info(f"Izdelek {product_id_str} ni parfum (product_type: {product_type})")
            continue
        
        # Pridobi product_no in proizvajalec_ime iz Shopify podatkov
        product_no = details.get('product_no')
        proizvajalec_ime = details.get('proizvajalec_id')
        
        if not product_no or not proizvajalec_ime:
            missing_data_details.append(f"Manjka product_no ali proizvajalec_ime za '{item.get('title', 'N/A')}' (product_id: {product_id_str}, details: {details})")
            current_app.logger.warning(f"Manjka product_no ali proizvajalec_ime: product_no={product_no}, proizvajalec_ime={proizvajalec_ime}")
            continue
        
        # Ustvari ključ za združevanje (product_no + proizvajalec_ime)
        parfum_key = f"{product_no}_{proizvajalec_ime.upper()}"
        
        if parfum_key not in parfum_groups:
            parfum_groups[parfum_key] = {
                'product_no': product_no,
                'proizvajalec_ime': proizvajalec_ime.upper(),
                'items': [],
                'total_quantity': 0
            }
        
        # Dodaj item v skupino
        quantity = item.get('quantity', 1)
        parfum_groups[parfum_key]['items'].append(item)
        parfum_groups[parfum_key]['total_quantity'] += quantity
    
    # Obdelaj vsako skupino parfumov
    for parfum_key, group_data in parfum_groups.items():
        product_no = group_data['product_no']
        proizvajalec_ime = group_data['proizvajalec_ime']
        total_quantity = group_data['total_quantity']
        first_item = group_data['items'][0]  # Uporabimo prvi item za podatke
        
        current_app.logger.info(f"Obdelujem skupino parfumov: {parfum_key}, količina: {total_quantity}")
        
        # Pridobi podatke iz baze
        current_app.logger.info(f"Iščem parfum z product_no={product_no} in proizvajalec_ime={proizvajalec_ime}")
        db_cursor.execute("SELECT p.*, pr.ime as ime_proizvajalca FROM parfumi p JOIN proizvajalci pr ON p.proizvajalec_id = pr.id WHERE p.product_no = %s AND pr.ime = %s", (product_no, proizvajalec_ime))
        parfum_data = db_cursor.fetchone()
        
        if not parfum_data or not parfum_data.get('sestava_inci'):
            missing_data_details.append(f"Manjka INCI za '{first_item.get('title', 'N/A')}' (product_no: {product_no}, proizvajalec: {proizvajalec_ime})")
            current_app.logger.warning(f"Parfum ni najden ali nima INCI: product_no={product_no}, proizvajalec={proizvajalec_ime}, parfum_data={parfum_data}")
            continue
        
        current_app.logger.info(f"Našel parfum: {parfum_data}")
        
        db_cursor.execute("SELECT rok_uporabe, serijska_stevilka FROM serije WHERE parfum_id = %s AND rok_uporabe >= CURRENT_DATE ORDER BY id DESC LIMIT 1", (parfum_data['id'],))
        serija_data = db_cursor.fetchone()
        if not serija_data:
            missing_data_details.append(f"Ni razpoložljive serije (z veljavnim rokom) za '{first_item.get('title', 'N/A')}'")
            current_app.logger.warning(f"Ni razpoložljive serije: parfum_id={parfum_data['id']}")
            continue
            
        is_problem, problem_message = preveri_rok_uporabe(serija_data['rok_uporabe'])
        if is_problem: expiration_warnings.append(f"{first_item.get('title', 'N/A')}: {problem_message}")
        
        current_app.logger.info(f"Dodajam declaration_item: product_no={parfum_data['product_no']}, proizvajalec={parfum_data['ime_proizvajalca']}, količina={total_quantity}")
        
        declaration_items.append({
            "product_no": parfum_data['product_no'],
            "proizvajalec_ime": parfum_data['ime_proizvajalca'],
            "sestava_inci": parfum_data.get('sestava_inci', ''),
            "rok_uporabe": serija_data['rok_uporabe'].strftime('%d.%m.%Y') if hasattr(serija_data['rok_uporabe'], 'strftime') and serija_data['rok_uporabe'] else (str(serija_data['rok_uporabe']) if serija_data['rok_uporabe'] else None),
            "serijska_stevilka": serija_data.get("serijska_stevilka", "N/A"), 
            "title": first_item.get('title', 'N/A'), 
            "product_id": parfum_data['product_no'],
            "image_url": first_item.get('image_url', ''),
            "quantity": total_quantity  # Dodamo skupno količino
        })
    
    current_app.logger.info(f"Vrnil {len(declaration_items)} declaration_items, {len(missing_data_details)} missing, {len(expiration_warnings)} warnings")
    current_app.logger.info(f"Missing details: {missing_data_details}")
    current_app.logger.info(f"Expiration warnings: {expiration_warnings}")
    current_app.logger.info(f"Declaration items: {declaration_items}")
    return declaration_items, missing_data_details, expiration_warnings
@api_bp.route('/toggle-email-mode', methods=['POST'])
def toggle_email_mode():
    """Preklopi e-mail način med test in produkcija"""
    try:
        # Preveri, ali je trenutni uporabnik admin
        if not session.get('user_id'):
            return jsonify({"error": "Ni prijavljen"}), 401
        
        cursor = get_db().cursor()
        cursor.execute("SELECT role FROM users WHERE id = %s", (session['user_id'],))
        current_user = cursor.fetchone()
        
        if not current_user or current_user['role'] != 'admin':
            return jsonify({"error": "Nimate dovoljenja"}), 403
        
        # Preveri trenutni način
        cursor.execute("SELECT value FROM app_settings WHERE key = 'email_test_mode'")
        result = cursor.fetchone()
        
        if result:
            current_mode = result['value']
            # Ciklično preklopi: test -> produkcija -> oba -> test
            if current_mode == 'true':
                new_mode = 'false'  # test -> produkcija
            elif current_mode == 'false':
                new_mode = 'both'   # produkcija -> oba
            else:
                new_mode = 'true'   # oba -> test
        else:
            # Če ni nastavljen, privzeto je test način
            new_mode = 'false'
        
        # Posodobi nastavitev
        cursor.execute("""
            INSERT INTO app_settings (key, value, updated_at) 
            VALUES ('email_test_mode', %s, NOW())
            ON CONFLICT (key) DO UPDATE SET 
                value = EXCLUDED.value, 
                updated_at = NOW()
        """, (str(new_mode).lower(),))
        
        get_db().commit()
        
        # Določi besedilo za sporočilo
        if new_mode == 'true':
            mode_text = 'TEST'
        elif new_mode == 'false':
            mode_text = 'PRODUKCIJA'
        else:
            mode_text = 'OBA (admin + customer)'
        
        current_app.logger.info(f"Admin {session['user_id']} je preklopil e-mail način na {mode_text} (cache refresh)")
        
        return jsonify({
            "success": True, 
            "test_mode": new_mode,
            "message": f"E-mail način preklopljen na {mode_text}"
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri preklopu e-mail načina: {e}")
        return jsonify({"error": "Napaka pri preklopu e-mail načina"}), 500
    finally:
        cursor.close()

@api_bp.route('/email-mode', methods=['GET'])
def get_email_mode():
    """Pridobi trenutni e-mail način"""
    cursor = None
    try:
        current_app.logger.info("Pridobivam e-mail način iz baze...")
        cursor = get_db().cursor()
        
        # Poskusi direktno pridobiti nastavek
        cursor.execute("SELECT value FROM app_settings WHERE key = 'email_test_mode'")
        result = cursor.fetchone()
        current_app.logger.info(f"Rezultat iz baze: {result}")
        
        if result:
            test_mode = result['value']
        else:
            # Če ni nastavljen, privzeto je test način
            test_mode = 'true'
            current_app.logger.info("Ni najdenega nastavka, uporabljam privzeti način")
        
        current_app.logger.info(f"E-mail način: {test_mode}")
        
        return jsonify({
            "success": True, 
            "test_mode": test_mode
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri pridobivanju e-mail načina: {e}")
        current_app.logger.error(f"Traceback: {traceback.format_exc()}")
        # Vrni privzeti način v primeru napake
        return jsonify({
            "success": True, 
            "test_mode": "true"
        })
    finally:
        if cursor:
            cursor.close()

@api_bp.route('/email-logs', methods=['GET'])
def get_email_logs():
    """Pridobi seznam vseh poslanih e-mailov"""
    cursor = None
    try:
        # Preveri, ali je trenutni uporabnik admin
        if not session.get('user_id'):
            return jsonify({"error": "Ni prijavljen"}), 401
        
        cursor = get_db().cursor()
        cursor.execute("SELECT role FROM users WHERE id = %s", (session['user_id'],))
        current_user = cursor.fetchone()
        
        if not current_user or current_user['role'] != 'admin':
            return jsonify({"error": "Nimate dovoljenja"}), 403
        
        # Pridobi e-mail logi
        cursor.execute("""
            SELECT 
                o.order_number,
                o.email_recipient,
                o.email_sent_at,
                o.status,
                o.customer_email,
                o.shopify_order_id,
                o.country_code
            FROM orders o 
            WHERE o.email_sent_at IS NOT NULL 
            ORDER BY o.email_sent_at DESC 
            LIMIT 100
        """)
        
        logs = []
        for row in cursor.fetchall():
            logs.append({
                'order_number': row['order_number'],
                'email_recipient': row['email_recipient'],
                'email_sent_at': row['email_sent_at'].isoformat() if row['email_sent_at'] else None,
                'status': row['status'],
                'customer_email': row['customer_email'],
                'shopify_order_id': row['shopify_order_id'],
                'country_code': row['country_code']
            })
        
        current_app.logger.info(f"Pridobil {len(logs)} e-mail logov")
        
        return jsonify({
            "success": True,
            "logs": logs
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri pridobivanju e-mail logov: {e}")
        return jsonify({"error": "Napaka pri pridobivanju e-mail logov"}), 500
    finally:
        if cursor:
            cursor.close()


@api_bp.route('/email-details/<order_number>', methods=['GET'])
def get_email_details(order_number):
    """Pridobi podrobnosti emaila za določeno naročilo"""
    cursor = None
    try:
        # Preveri, ali je trenutni uporabnik admin
        if not session.get('user_id'):
            return jsonify({"error": "Ni prijavljen"}), 401
        
        cursor = get_db().cursor()
        cursor.execute("SELECT role FROM users WHERE id = %s", (session['user_id'],))
        current_user = cursor.fetchone()
        
        if not current_user or current_user['role'] != 'admin':
            return jsonify({"error": "Nimate dovoljenja"}), 403
        
        # Pridobi podrobnosti naročila
        cursor.execute("""
            SELECT 
                o.order_number,
                o.email_recipient,
                o.email_sent_at,
                o.status,
                o.customer_email,
                o.shopify_order_id,
                o.country_code,
                o.line_items,
                o.total_price,
                o.currency
            FROM orders o 
            WHERE o.order_number = %s AND o.email_sent_at IS NOT NULL
        """, (order_number,))
        
        order = cursor.fetchone()
        
        if not order:
            return jsonify({"error": "Naročilo ni bilo najdeno ali email ni bil poslan"}), 404
        
        # Pridobi podrobnosti izdelkov iz line_items
        line_items = []
        if order['line_items']:
            try:
                line_items_raw = json.loads(order['line_items']) if isinstance(order['line_items'], str) else order['line_items']
                for item in line_items_raw:
                    if item and item.get('product_id'):
                        # Zaenkrat uporabimo osnovne podatke iz line_items
                        line_items.append({
                            'title': item.get('title', 'N/A'),
                            'quantity': item.get('quantity', 1),
                            'price': item.get('price', 0.0),
                            'product_id': item.get('product_id'),
                            'product_type': 'N/A',  # TODO: Dodaj Shopify integracijo
                            'product_no': 'N/A',    # TODO: Dodaj Shopify integracijo
                            'proizvajalec': 'N/A'   # TODO: Dodaj Shopify integracijo
                        })
            except Exception as e:
                current_app.logger.error(f"Napaka pri parsiranju line_items: {e}")
                line_items = []
        
        email_details = {
            'order_number': order['order_number'],
            'email_recipient': order['email_recipient'],
            'email_sent_at': order['email_sent_at'].isoformat() if order['email_sent_at'] else None,
            'status': order['status'],
            'customer_email': order['customer_email'],
            'shopify_order_id': order['shopify_order_id'],
            'country_code': order['country_code'],
            'total_price': order['total_price'],
            'currency': order['currency'],
            'line_items': line_items
        }
        
        return jsonify({
            "success": True,
            "email_details": email_details
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri pridobivanju podrobnosti emaila: {e}")
        return jsonify({"error": "Napaka pri pridobivanju podrobnosti emaila"}), 500
    finally:
        if cursor:
            cursor.close()

@api_bp.route('/fix-user-permissions', methods=['POST'])
def fix_user_permissions_endpoint():
    """Popravi dovoljenja za obstoječe uporabnike."""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Osnovna dovoljenja za uporabnike
        user_permissions = [
            "view_global_actions", "view_orders", "view_perfumes", "view_proizvajalci", 
            "add_serije", "edit_serije", "generate_pdf", "send_email"
        ]
        
        # Admin dovoljenja
        admin_permissions = [
            "view_global_actions", "view_orders", "add_serije", "edit_serije", "delete_serije",
            "view_perfumes", "edit_perfumes", "add_perfumes", "delete_perfumes",
            "view_proizvajalci", "edit_proizvajalci", "add_proizvajalci", "delete_proizvajalci",
            "view_users", "edit_users", "add_users", "delete_users",
            "shopify_sync", "generate_pdf", "send_email"
        ]
        
        # Nastavi osnovna dovoljenja za vse uporabnike z role = 'user' in praznimi dovoljenji
        cursor.execute("""
            UPDATE users 
            SET permissions = %s
            WHERE role = 'user' 
              AND (permissions IS NULL OR permissions = '[]'::jsonb OR permissions = 'null'::jsonb)
        """, (json.dumps(user_permissions),))
        
        user_count = cursor.rowcount
        current_app.logger.info(f"Posodobljenih {user_count} uporabnikov z osnovnimi dovoljenji")
        
        # Nastavi osnovna dovoljenja za uporabnike brez role
        cursor.execute("""
            UPDATE users 
            SET role = 'user',
                permissions = %s
            WHERE role IS NULL 
              AND (permissions IS NULL OR permissions = '[]'::jsonb OR permissions = 'null'::jsonb)
        """, (json.dumps(user_permissions),))
        
        null_role_count = cursor.rowcount
        current_app.logger.info(f"Posodobljenih {null_role_count} uporabnikov brez role")
        
        # Posodobi admin uporabnika z vsemi dovoljenji
        cursor.execute("""
            UPDATE users 
            SET role = 'admin', 
                permissions = %s
            WHERE username = 'admin'
        """, (json.dumps(admin_permissions),))
        
        admin_count = cursor.rowcount
        current_app.logger.info(f"Posodobljenih {admin_count} admin uporabnikov")
        
        db.commit()
        
        return jsonify({
            'success': True,
            'message': f'Dovoljenja uspešno popravljena. Posodobljenih {user_count + null_role_count + admin_count} uporabnikov.',
            'updated_users': user_count,
            'updated_null_role': null_role_count,
            'updated_admin': admin_count
        })
        
    except Exception as e:
        current_app.logger.error(f"Napaka pri popravljanju dovoljenj: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cursor.close()