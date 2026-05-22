from pathlib import Path
from datetime import datetime
from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def generate_purchase_order_excel(po: dict, items: list[dict]) -> tuple[str | None, str]:
    """
    Ustvari XLSX (Excel) datoteko s povzetkom naročila dobavitelju.
    Vrne (path, message).
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Naročilo"

        # Header
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"Naročilo dobavitelju – {po.get('supplier','')} (PO #{po.get('id')})"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="left")

        created_at = po.get('submitted_at') or po.get('created_at') or datetime.now()
        created_at_str = created_at.strftime('%d.%m.%Y %H:%M') if hasattr(created_at, 'strftime') else str(created_at)
        ws["A2"].value = f"Datum: {created_at_str}"
        ws["A3"].value = f"Status: {po.get('status','')}"

        # Table header
        headers = [
            "Product No",
            "Proizvajalec",
            "Parfum",
            "Naročeno",
            "Prejeto",
            "Nismo prejeli",
        ]
        ws.append([])  # blank row before table
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        # Rows
        for it in items:
            requested = int(it.get('requested_qty', 0))
            received = int(it.get('received_qty', 0))
            backorder = max(0, requested - received)
            ws.append([
                it.get('product_no', ''),
                it.get('proizvajalec', ''),
                it.get('ime_parfuma', ''),
                requested,
                received,
                backorder,
            ])

        # Column widths
        widths = [16, 18, 40, 12, 12, 12]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width

        filename = f"PO_{po.get('id')}.xlsx"
        out_dir = Path(current_app.root_path) / "pdf"
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / filename
        wb.save(str(out_path))
        return str(out_path), "XLSX ustvarjen"
    except Exception as e:
        return None, str(e)


