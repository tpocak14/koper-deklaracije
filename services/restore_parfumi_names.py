"""Obnovi ime_parfuma iz Excel zvezka Parfumi (DEKLARACIJE_PARFUMOV_KOPER.xlsm)."""

from __future__ import annotations

import io
from typing import Any

from flask import current_app
from openpyxl import load_workbook

from database import get_db


def _load_excel_map(file_path: str | None = None, file_bytes: bytes | None = None) -> dict[tuple[str, str], str]:
    if file_bytes is not None:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    elif file_path:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    else:
        raise ValueError("Podaj file_path ali file_bytes.")

    if "Parfumi" not in workbook.sheetnames:
        raise ValueError("Zvezek 'Parfumi' ne obstaja v Excel datoteki.")

    sheet = workbook["Parfumi"]
    excel: dict[tuple[str, str], str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        product_no = str(row[0]).strip().replace(".0", "")
        ime_parfuma = str(row[1]).strip() if row[1] else ""
        proizvajalec = (
            str(row[4]).strip().upper() if len(row) > 4 and row[4] else ""
        )
        if product_no and proizvajalec and ime_parfuma:
            excel[(product_no.upper(), proizvajalec)] = ime_parfuma
    return excel


def restore_parfumi_names_from_excel(
    *,
    file_path: str | None = None,
    file_bytes: bytes | None = None,
    dry_run: bool = True,
) -> dict[str, Any]:
    """
    Posodobi samo ime_parfuma za zapise, ki se ujemajo po (product_no, proizvajalec).

    Ne briše parfumov. Ne spreminja INCI ali na_zalogi.
    """
    excel = _load_excel_map(file_path=file_path, file_bytes=file_bytes)

    db = get_db()
    cursor = db.cursor()
    result: dict[str, Any] = {
        "ok": False,
        "dry_run": dry_run,
        "excel_rows": len(excel),
        "updated": 0,
        "unchanged": 0,
        "not_in_excel": 0,
        "samples": [],
    }

    try:
        cursor.execute(
            """
            SELECT p.id, UPPER(TRIM(p.product_no)) AS product_no,
                   UPPER(TRIM(pr.ime)) AS proizvajalec, p.ime_parfuma
            FROM parfumi p
            JOIN proizvajalci pr ON pr.id = p.proizvajalec_id
            """
        )
        rows = cursor.fetchall() or []

        for row in rows:
            rid = row["id"] if isinstance(row, dict) else row[0]
            product_no = row["product_no"] if isinstance(row, dict) else row[1]
            proizvajalec = row["proizvajalec"] if isinstance(row, dict) else row[2]
            current_name = row["ime_parfuma"] if isinstance(row, dict) else row[3]

            key = (product_no, proizvajalec)
            target_name = excel.get(key)
            if not target_name:
                result["not_in_excel"] += 1
                continue
            if current_name == target_name:
                result["unchanged"] += 1
                continue

            if not dry_run:
                cursor.execute(
                    "UPDATE parfumi SET ime_parfuma = %s, updated_at = NOW() WHERE id = %s",
                    (target_name, rid),
                )

            result["updated"] += 1
            if len(result["samples"]) < 25:
                result["samples"].append(
                    {
                        "id": rid,
                        "product_no": product_no,
                        "proizvajalec": proizvajalec,
                        "from": current_name,
                        "to": target_name,
                    }
                )

        if not dry_run:
            db.commit()
        else:
            db.rollback()

        result["ok"] = True
        return result
    except Exception as exc:
        db.rollback()
        current_app.logger.error(f"restore_parfumi_names_from_excel failed: {exc}")
        result["error"] = str(exc)
        return result
    finally:
        cursor.close()
