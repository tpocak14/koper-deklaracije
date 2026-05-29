#!/usr/bin/env python3
"""Izvozi id, proizvajalec_id, ime_proizvajalca, ime_parfuma za vse parfume."""

from __future__ import annotations

import csv
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font

from app import create_app
from database import get_db

HEADERS = ["proizvajalec", "product_no", "ime_parfuma"]
HEADERS_UI = ["Proizvajalec", "Šifra izdelka (Product No)", "Ime parfuma"]

SQL = """
    SELECT pr.ime AS proizvajalec, p.product_no, p.ime_parfuma
    FROM parfumi p
    JOIN proizvajalci pr ON pr.id = p.proizvajalec_id
    ORDER BY pr.ime, p.product_no
"""


def _fetch_rows() -> list[dict]:
    app = create_app()
    with app.app_context():
        db = get_db()
        cursor = db.cursor()
        cursor.execute(SQL)
        rows = cursor.fetchall() or []
        cursor.close()
        return [dict(r) if not isinstance(r, dict) else r for r in rows]


def export_parfumi_imena_csv(output_path: str) -> int:
    rows = _fetch_rows()
    with open(output_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(HEADERS_UI)
        for row in rows:
            writer.writerow(
                [row["proizvajalec"], row["product_no"], row["ime_parfuma"]]
            )
    return len(rows)


def export_parfumi_imena_xlsx(output_path: str) -> int:
    rows = _fetch_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Parfumi"
    ws.append(HEADERS_UI)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row["proizvajalec"], row["product_no"], row["ime_parfuma"]])
    # Šifre (408N, 417P, …) morajo ostati besedilo — Excel jih sicer "poje".
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).number_format = "@"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 52
    ws.freeze_panes = "A2"
    wb.save(output_path)
    return len(rows)


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "parfumi_imena_export.xlsx"
    if out.lower().endswith(".csv"):
        count = export_parfumi_imena_csv(out)
    else:
        count = export_parfumi_imena_xlsx(out)
    print(f"Izvoženih {count} parfumov → {out}")
