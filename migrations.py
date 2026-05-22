import click
import openpyxl
import traceback
from flask.cli import with_appcontext
from database import get_db
from datetime import datetime
import psycopg
from flask import current_app
import os
import tempfile
import requests
import pandas as pd

# --- Pomožne funkcije ---

def parse_date(date_val):
    """
    Pretvori različne formate datumov iz Excela v Python date objekt.
    Podpira datetime objekte in nize v formatih 'dd.mm.yyyy' ali 'yyyy-mm-dd'.
    """
    if isinstance(date_val, datetime):
        return date_val
    if isinstance(date_val, str):
        try:
            return datetime.strptime(date_val, '%d.%m.%Y')
        except (ValueError, TypeError):
            try:
                return datetime.strptime(date_val, '%Y-%m-%d')
            except (ValueError, TypeError):
                return None
    return None

# --- Ukazi za migracijo ---

@click.command('migrate-catalog')
@with_appcontext
def migrate_catalog_command():
    """Migrira podatke iz zavihka 'Parfumi' v Excelu v bazo."""
    click.echo("⚙️  Začenjam migracijo kataloga iz zavihka 'Parfumi'...")
    try:
        wb = openpyxl.load_workbook('DEKLARACIJE_PARFUMOV_KOPER.xlsm', data_only=True, read_only=True)
        sheet_name = 'Parfumi'
        if sheet_name not in wb.sheetnames:
            click.secho(f"❌ NAPAKA: Zavihek '{sheet_name}' ne obstaja. Najdeni zavihki: {wb.sheetnames}", fg='red')
            return
        sheet = wb[sheet_name]
        
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("SELECT id, ime FROM proizvajalci")
        # --- POPRAVEK TUKAJ ---
        # Uporabimo imena ključev ('ime', 'id') namesto indeksov (1, 0)
        proizvajalci_map = {row['ime'].strip().upper(): row['id'] for row in cursor.fetchall()}
        click.echo(f"ℹ️ Najdenih {len(proizvajalci_map)} obstoječih proizvajalcev v bazi.")

        uspesno_vnesenih = 0
        preskocenih = 0
        vseh_vrstic = sheet.max_row - 1
        click.echo(f"  - Najdenih {vseh_vrstic} vrstic za obdelavo.")

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            # Pravilno branje stolpcev po vaši stari logiki
            product_no_raw = row[0]       # Stolpec A
            ime_parfuma_raw = row[1]      # Stolpec B
            sestava_inci = str(row[2]).strip() if row[2] else None # Stolpec C
            ime_proizvajalca_raw = row[4] # Stolpec E

            if not all([product_no_raw, ime_parfuma_raw, ime_proizvajalca_raw]):
                preskocenih += 1
                continue

            product_no = str(product_no_raw).strip()
            ime_parfuma = str(ime_parfuma_raw).strip()
            ime_proizvajalca = str(ime_proizvajalca_raw).strip().upper()

            if ime_proizvajalca not in proizvajalci_map:
                click.echo(f"⚠️  OPOZORILO: Proizvajalec '{ime_proizvajalca}' iz vrstice {i} ne obstaja v bazi. Preskačem.")
                preskocenih += 1
                continue
            
            proizvajalec_id = proizvajalci_map[ime_proizvajalca]

            cursor.execute(
                """
                INSERT INTO parfumi (product_no, proizvajalec_id, ime_parfuma, sestava_inci)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (product_no, proizvajalec_id) DO UPDATE SET
                    ime_parfuma = EXCLUDED.ime_parfuma,
                    sestava_inci = EXCLUDED.sestava_inci,
                    updated_at = CURRENT_TIMESTAMP;
                """,
                (product_no, proizvajalec_id, ime_parfuma, sestava_inci)
            )
            uspesno_vnesenih += 1
            if i % 100 == 0:
                click.echo(f"    ... obdelanih {i} / {vseh_vrstic} vrstic.")
        
        db.commit()
        cursor.close()
        click.secho(f"✅ Migracija kataloga končana. Uspešno vnesenih/posodobljenih: {uspesno_vnesenih}, preskočenih: {preskocenih}.", fg='green')

    except Exception as e:
        click.secho(f"❌ Kritična napaka med migracijo kataloga: {e}", fg='red')
        traceback.print_exc()

@click.command('migrate-stock')
@with_appcontext
def migrate_stock_command():
    """Migrira podatke o zalogi iz zavihka 'Database' v Excelu v bazo."""
    click.echo("⚙️  Začenjam migracijo zaloge iz zavihka 'Database'...")
    try:
        wb = openpyxl.load_workbook('DEKLARACIJE_PARFUMOV_KOPER.xlsm', data_only=True, read_only=True)
        sheet_name = 'Database'
        if sheet_name not in wb.sheetnames:
            click.secho(f"❌ NAPAKA: Zavihek '{sheet_name}' ne obstaja. Najdeni zavihki: {wb.sheetnames}", fg='red')
            return
        sheet = wb[sheet_name]
        
        db = get_db()
        cursor = db.cursor()
        
        # Pridobi vse parfume iz baze za hitro iskanje
        cursor.execute("SELECT p.id, p.product_no, pr.ime as ime_proizvajalca FROM parfumi p JOIN proizvajalci pr ON p.proizvajalec_id = pr.id")
        parfumi_map = {(p['product_no'].strip().upper(), p['ime_proizvajalca'].strip().upper()): p['id'] for p in cursor.fetchall()}
        click.echo(f"ℹ️ Najdenih {len(parfumi_map)} unikatnih parfumov v katalogu baze.")

        uspesno_vnesenih = 0
        preskocenih = 0
        vseh_vrstic = sheet.max_row - 1
        click.echo(f"  - Najdenih {vseh_vrstic} vrstic za obdelavo.")

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            # Pravilno branje stolpcev po vaši stari logiki
            excel_row_id_raw = row[0]       # Stolpec A
            product_no_raw = row[1]         # Stolpec B
            proizvajalec_raw = row[2]       # Stolpec C
            datum_odprtja_raw = row[4]      # Stolpec E
            rok_uporabe_raw = row[5]        # Stolpec F
            je_tester_raw = row[6]          # Stolpec G
            created_at_original_raw = row[7]# Stolpec H
            vnesel_uporabnik_raw = row[8]   # Stolpec I
            serijska_stevilka_raw = row[9]  # Stolpec J

            if not all([excel_row_id_raw, product_no_raw, proizvajalec_raw]):
                preskocenih += 1
                continue

            try:
                excel_row_id = int(excel_row_id_raw)
            except (ValueError, TypeError):
                click.echo(f"⚠️  OPOZORILO: Neveljaven ID vnosa '{excel_row_id_raw}' v vrstici {i}. Preskakujem.")
                preskocenih += 1
                continue

            product_no = str(product_no_raw).strip().upper()
            proizvajalec = str(proizvajalec_raw).strip().upper()
            
            parfum_id = parfumi_map.get((product_no, proizvajalec))
            if not parfum_id:
                click.echo(f"⚠️  OPOZORILO: Parfum z ID '{product_no}' in proizvajalcem '{proizvajalec}' (vrstica {i}) ni najden v katalogu. Preskakujem.")
                preskocenih += 1
                continue

            rok_uporabe = parse_date(rok_uporabe_raw)
            if not rok_uporabe:
                preskocenih += 1
                continue

            datum_odprtja = parse_date(datum_odprtja_raw)
            je_tester = str(je_tester_raw).strip().lower() == 'da' if je_tester_raw else False
            created_at_original = parse_date(created_at_original_raw)
            vnesel_uporabnik = str(vnesel_uporabnik_raw).strip() if vnesel_uporabnik_raw else None
            serijska_stevilka = str(serijska_stevilka_raw).strip() if serijska_stevilka_raw else None

            cursor.execute(
                """
                INSERT INTO serije (
                    excel_row_id, parfum_id, rok_uporabe, serijska_stevilka, stanje, 
                    datum_odprtja, je_tester, vnesel_uporabnik, created_at_original
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (excel_row_id) DO UPDATE SET
                    parfum_id = EXCLUDED.parfum_id, rok_uporabe = EXCLUDED.rok_uporabe,
                    serijska_stevilka = EXCLUDED.serijska_stevilka, stanje = EXCLUDED.stanje,
                    datum_odprtja = EXCLUDED.datum_odprtja, je_tester = EXCLUDED.je_tester,
                    vnesel_uporabnik = EXCLUDED.vnesel_uporabnik, created_at_original = EXCLUDED.created_at_original
                """,
                (
                    excel_row_id, parfum_id, rok_uporabe.date(), serijska_stevilka, 'NA ZALOGI',
                    datum_odprtja.date() if datum_odprtja else None,
                    je_tester, vnesel_uporabnik, created_at_original
                )
            )
            uspesno_vnesenih += 1
            if i % 100 == 0:
                click.echo(f"    ... obdelanih {i} / {vseh_vrstic} vrstic.")
        
        db.commit()
        cursor.close()
        click.secho(f"✅ Migracija zaloge končana. Uspešno vnesenih/posodobljenih: {uspesno_vnesenih}, preskočenih: {preskocenih}.", fg='green')

    except Exception as e:
        click.secho(f"❌ Kritična napaka med migracijo zaloge: {e}", fg='red')
        traceback.print_exc()

def run_migrations():
    """Zažene vse migracije."""
    db = psycopg.connect(current_app.config['DATABASE_URL'])
    cursor = db.cursor()
    
    try:
        # Preverimo, ali tabela migrations obstaja
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS migrations (
                id SERIAL PRIMARY KEY,
                version VARCHAR(50) UNIQUE NOT NULL,
                applied_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # Preverimo, katere migracije so že bile izvedene
        cursor.execute("SELECT version FROM migrations ORDER BY id")
        applied_migrations = {row[0] for row in cursor.fetchall()}
        
        # Seznam vseh migracij v vrstnem redu
        migrations = [
            ('001_initial_schema.sql', '001'),
            ('002_add_declarations_table.sql', '002'),
            ('003_update_declarations_table.sql', '003'),
            ('004_add_declarations_unique_constraint.sql', '004'),
            ('005_add_order_images_table.sql', '005'),
            ('006_add_users_table.sql', '006'),
            ('007_add_user_permissions.sql', '007'),
            ('011_fix_user_permissions.sql', '011'),
            ('012_fix_admin_permissions.sql', '012'),
            ('013_add_updated_by_to_serije.sql', '013'),
            ('014_create_instructions.sql', '014'),
            ('016_add_nalivalec_to_orders.sql', '016'),
            # Naročilo robe (procurement)
            ('017_create_procurement.sql', '017'),
            ('018_init_perfumes_stock.sql', '018'),
            ('019_add_min_on_hand.sql', '019'),
            ('020_add_shopify_stores.sql', '020'),
            ('021_add_search_rewrite.sql', '021'),
        ]
        
        for migration_file, version in migrations:
            if version not in applied_migrations:
                print(f"ℹ️  Izvajam migracijo: {migration_file}")
                
                # Preberemo SQL iz datoteke
                migration_path = os.path.join(current_app.root_path, 'migrations', 'versions', migration_file)
                if os.path.exists(migration_path):
                    with open(migration_path, 'r', encoding='utf-8') as f:
                        sql = f.read()

                    try:
                        # Izvedemo migracijo
                        cursor.execute(sql)
                        # Zabeležimo, da je bila migracija izvedena
                        cursor.execute("INSERT INTO migrations (version) VALUES (%s)", (version,))
                        db.commit()
                        print(f"✅ Migracija {migration_file} uspešno izvedena")
                    except Exception as e:
                        # Če migracija pade zaradi že obstoječih entitet, jo označimo kot izvedeno in nadaljujemo
                        msg = str(e).lower()
                        benign = any(keyword in msg for keyword in [
                            'already exists', 'exists', 'duplicate key', 'duplicate column', 'relation', 'index concurrently'
                        ])
                        if benign:
                            db.rollback()
                            try:
                                cursor.execute("INSERT INTO migrations (version) VALUES (%s) ON CONFLICT (version) DO NOTHING", (version,))
                                db.commit()
                                print(f"⚠️  Preskakujem {migration_file} (že obstaja); označeno kot izvedeno")
                            except Exception as ie:
                                db.rollback()
                                print(f"❌ Napaka pri označevanju migracije {migration_file} kot izvedene: {ie}")
                                raise
                        else:
                            db.rollback()
                            print(f"❌ Napaka pri izvajanju migracije {migration_file}: {e}")
                            raise
                else:
                    print(f"⚠️  Migracija {migration_file} ne obstaja")
            else:
                print(f"ℹ️  Migracija {migration_file} je že bila izvedena")
        
        print("✅ Vse migracije so bile uspešno izvedene")
        
    except Exception as e:
        db.rollback()
        print(f"❌ Napaka pri izvajanju migracij: {e}")
        raise
    finally:
        cursor.close()
        db.close()

def init_app(app):
    """Registrira ukaze za migracijo v Flask aplikaciji."""
    app.cli.add_command(migrate_catalog_command)
    app.cli.add_command(migrate_stock_command)

    @app.cli.command('migrate')
    def migrate_command():
        """Izvede vse migracije."""
        with app.app_context():
            run_migrations()
    
    @app.cli.command('migrate-status')
    def migrate_status_command():
        """Prikaže status migracij."""
        with app.app_context():
            show_migration_status()
    
    @app.cli.command('migrate-onedrive')
    def migrate_onedrive_command():
        """Popolna migracija iz OneDrive Excel datoteke - izprazni bazo in dodaj vse na novo."""
        with app.app_context():
            success = migrate_from_onedrive()
            if success:
                print("🎉 Popolna migracija iz OneDrive uspešno končana!")
            else:
                print("❌ Migracija iz OneDrive ni uspela!")
                exit(1)

    @app.cli.command('migrate-local-file')
    def migrate_local_file_command():
        """Migracija novih serij iz lokalne Excel datoteke (podobno kot prejšnje skripte)."""
        with app.app_context():
            success = migrate_from_local_excel_file()
            if success:
                print("✅ Migracija iz lokalne Excel datoteke uspešno končana")
            else:
                print("❌ Migracija iz lokalne Excel datoteke ni uspela")

def show_migration_status():
    """Prikaže status vseh migracij."""
    db = psycopg.connect(current_app.config['DATABASE_URL'])
    cursor = db.cursor()
    
    try:
        # Preverimo, ali tabela migrations obstaja
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = 'migrations'
            );
        """)
        migrations_table_exists = cursor.fetchone()[0]
        
        if not migrations_table_exists:
            print("ℹ️  Tabela migrations ne obstaja. Migracije še niso bile izvedene.")
            return
        
        # Pridobimo izvedene migracije
        cursor.execute("SELECT version, applied_at FROM migrations ORDER BY id")
        applied_migrations = cursor.fetchall()
        
        print("📋 Status migracij:")
        print("-" * 50)
        
        # Seznam vseh migracij
        all_migrations = [
            ('001_initial_schema.sql', '001'),
            ('002_add_declarations_table.sql', '002'),
            ('003_update_declarations_table.sql', '003'),
            ('004_add_declarations_unique_constraint.sql', '004')
        ]
        
        for migration_file, version in all_migrations:
            applied = any(row[0] == version for row in applied_migrations)
            if applied:
                applied_at = next(row[1] for row in applied_migrations if row[0] == version)
                print(f"✅ {migration_file} - izvedena {applied_at.strftime('%Y-%m-%d %H:%M:%S')}")
            else:
                print(f"⏳ {migration_file} - ni izvedena")
        
        print("-" * 50)
        
    except Exception as e:
        print(f"❌ Napaka pri preverjanju statusa migracij: {e}")
    finally:
        cursor.close()
        db.close()

def migrate_from_onedrive():
    """Popolna migracija iz OneDrive Excel datoteke - doda nove serije iz Database sheet-a."""
    print("🚀 Začenjam migracijo novih serij iz OneDrive Excel datoteke...")
    
    # Pridobimo OneDrive podatke iz konfiguracije
    client_id = current_app.config.get('CLIENT_ID')
    client_secret = current_app.config.get('CLIENT_SECRET')
    tenant_id = current_app.config.get('TENANT_ID')
    excel_file_id = current_app.config.get('EXCEL_FILE_ID')
    drive_id = current_app.config.get('DRIVE_ID')
    
    print(f"📋 OneDrive konfiguracija: CLIENT_ID={'*'*10 if client_id else 'MANJKA'}, TENANT_ID={'*'*10 if tenant_id else 'MANJKA'}")
    
    if not all([client_id, client_secret, tenant_id, excel_file_id, drive_id]):
        missing = []
        if not client_id: missing.append('CLIENT_ID')
        if not client_secret: missing.append('CLIENT_SECRET')
        if not tenant_id: missing.append('TENANT_ID')
        if not excel_file_id: missing.append('EXCEL_FILE_ID')
        if not drive_id: missing.append('DRIVE_ID')
        print(f"❌ Manjkajo OneDrive konfiguracijski podatki: {missing}")
        return False
    
    try:
        # 1. Pridobimo access token
        print("🔐 Pridobivam access token...")
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        token_data = {
            'client_id': client_id,
            'client_secret': client_secret,
            'scope': 'https://graph.microsoft.com/.default',
            'grant_type': 'client_credentials'
        }
        
        token_response = requests.post(token_url, data=token_data, timeout=10)
        if not token_response.ok:
            print(f"❌ Napaka pri pridobivanju tokena: {token_response.status_code}")
            return False
        
        access_token = token_response.json()['access_token']
        print("✅ Access token uspešno pridobljen")
        
        # 2. Prenesemo Excel datoteko iz OneDrive-a
        print("📥 Prenašam Excel datoteko iz OneDrive-a...")
        download_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{excel_file_id}/content"
        headers = {'Authorization': f'Bearer {access_token}'}
        
        excel_response = requests.get(download_url, headers=headers, timeout=60)
        if not excel_response.ok:
            print(f"❌ Napaka pri prenosu datoteke: {excel_response.status_code}")
            return False
        
        print(f"✅ Excel datoteka prenesena, velikost: {len(excel_response.content)} bajtov")
        
        # 3. Shranimo datoteko v temp direktorij
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as temp_file:
            temp_file.write(excel_response.content)
            temp_path = temp_file.name
        
        print(f"💾 Excel datoteka shranjena v: {temp_path}")
        
        # 4. Naložimo Excel datoteko z pandas
        print("📊 Nalagam Excel datoteko z pandas...")
        df = pd.read_excel(temp_path, sheet_name='Database', engine='openpyxl')
        print(f"✅ Excel datoteka naložena. Število vrstic: {len(df)}")
        
        # 5. Dodamo nove serije (ne izpraznimo baze)
        print("📝 Dodajam nove serije iz Excel-a...")
        db = get_db()
        cursor = db.cursor()
        
        # Uporabimo batch insert za hitrejše delovanje
        batch_data = []
        batch_size = 100
        
        for index, row in df.iterrows():
            try:
                # Preberemo podatke iz vrstice (A stolpec je prazen, začnemo z B)
                product_no = row.iloc[1]    # B stolpec (1) - številka izdelka
                proizvajalec_ime = row.iloc[2]  # C stolpec (2) - ime proizvajalca
                sestava_inci = row.iloc[3] if len(row) > 3 else None  # D stolpec (3) - sestava INCI
                rok_uporabe_raw = row.iloc[4] if len(row) > 4 else None  # E stolpec (4) - rok uporabe
                serijska_stevilka = row.iloc[5] if len(row) > 5 else None  # F stolpec (5) - serijska številka
                
                # Preverimo, ali so podatki veljavni (product_no in proizvajalec_ime so obvezni)
                if pd.isna(product_no) or pd.isna(proizvajalec_ime):
                    continue
                
                # Pretvorimo rok uporabe
                rok_uporabe = None
                if not pd.isna(rok_uporabe_raw):
                    try:
                        if isinstance(rok_uporabe_raw, str):
                            # Poskusimo različne formate datuma
                            for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
                                try:
                                    rok_uporabe = datetime.strptime(rok_uporabe_raw, fmt).date()
                                    break
                                except ValueError:
                                    continue
                        elif hasattr(rok_uporabe_raw, 'date'):
                            rok_uporabe = rok_uporabe_raw.date()
                        elif hasattr(rok_uporabe_raw, 'to_pydatetime'):
                            rok_uporabe = rok_uporabe_raw.to_pydatetime().date()
                    except:
                        pass
                
                # Dodamo v batch (order_number bo NULL, ker ga dobimo iz Shopify)
                batch_data.append((
                    None,  # order_number - NULL, ker ga dobimo iz Shopify
                    str(product_no),
                    str(proizvajalec_ime).upper(),
                    str(sestava_inci) if not pd.isna(sestava_inci) else None,
                    rok_uporabe,
                    str(serijska_stevilka) if not pd.isna(serijska_stevilka) else None
                ))
                
                # Izvedemo batch insert (dodajamo nove serije, ne posodabljamo)
                if len(batch_data) >= batch_size:
                    cursor.executemany(
                        """
                        INSERT INTO declarations (order_number, product_no, proizvajalec_ime, sestava_inci, rok_uporabe, serijska_stevilka)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        ON CONFLICT (product_no, proizvajalec_ime, rok_uporabe, serijska_stevilka) DO NOTHING
                        """,
                        batch_data
                    )
                    added_count += len(batch_data)
                    batch_data = []
                    print(f"📊 Obdelano {index + 1}/{len(df)} vrstic, dodano {added_count} novih serij")
            
            except Exception as e:
                error_count += 1
                print(f"⚠️  Napaka pri obdelavi vrstice {index}: {e}")
                continue
        
        # Obdelamo preostale podatke
        if batch_data:
            cursor.executemany(
                """
                INSERT INTO declarations (order_number, product_no, proizvajalec_ime, sestava_inci, rok_uporabe, serijska_stevilka)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (product_no, proizvajalec_ime, rok_uporabe, serijska_stevilka) DO NOTHING
                """,
                batch_data
            )
            added_count += len(batch_data)
        
        db.commit()
        print(f"✅ Migracija končana. Dodano {added_count} novih serij, napake: {error_count}")
        
        # Počistimo temp datoteko
        os.unlink(temp_path)
        
        return True
        
    except Exception as e:
        print(f"❌ Kritična napaka med migracijo: {e}")
        traceback.print_exc()
        return False

def migrate_from_local_excel_file():
    """Migracija novih serij iz lokalne Excel datoteke v tabelo serije (kot v starejši verziji)."""
    print("🚀 Začenjam migracijo novih serij iz lokalne Excel datoteke v tabelo serije...")
    
    try:
        import os
        from pathlib import Path
        
        # Pot do lokalne Excel datoteke (v root direktoriju)
        excel_path = Path("DEKLARACIJE_PARFUMOV_KOPER.xlsm")
        
        if not excel_path.is_file():
            print(f"❌ Excel datoteka ne obstaja: {excel_path}")
            print("💡 Prenesite Excel datoteko iz OneDrive in jo postavite v root direktorij aplikacije")
            return False
        
        print(f"📊 Nalagam Excel datoteko: {excel_path}")
        print(f"📏 Velikost datoteke: {excel_path.stat().st_size} bajtov")
        
        # Naložimo Excel z openpyxl (kot v prejšnjih skriptah)
        from openpyxl import load_workbook
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        
        if 'Database' not in workbook.sheetnames:
            print(f"❌ Zavihek 'Database' ne obstaja v Excel datoteki")
            return False
            
        sheet = workbook['Database']
        print(f"✅ Excel datoteka in zavihek 'Database' uspešno naložena")
        
        print("📝 Dodajam nove serije iz Excel-a v tabelo serije...")
        db = get_db()
        cursor = db.cursor()
        
        # Pridobi vse parfume iz baze za hitro iskanje
        cursor.execute("SELECT p.id, p.product_no, pr.ime as ime_proizvajalca FROM parfumi p JOIN proizvajalci pr ON p.proizvajalec_id = pr.id")
        parfumi_map = {(p['product_no'].strip().upper(), p['ime_proizvajalca'].strip().upper()): p['id'] for p in cursor.fetchall()}
        print(f"ℹ️ Najdenih {len(parfumi_map)} unikatnih parfumov v katalogu baze.")
        
        added_count = 0
        error_count = 0
        
        # Iteriraj čez vrstice (preskoči glavo)
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # --- LOGIKA BRANJA STOLPCEV (kot v starejši skripti) ---
                # A (indeks 0): ID vnosa (excel_row_id)
                # B (indeks 1): ID Parfuma (product_no)
                # C (indeks 2): Proizvajalec
                # D (indeks 3): Sestava INCI (če obstaja)
                # E (indeks 4): Datum odprtja
                # F (indeks 5): Rok uporabe
                # G (indeks 6): Tester (Da/Ne)
                # H (indeks 7): Čas vnosa
                # I (indeks 8): Vnesel
                # J (indeks 9): Serijska številka
                
                excel_row_id_raw = row[0]  # A stolpec
                product_no_raw = row[1]  # B stolpec
                proizvajalec_raw = row[2]  # C stolpec
                datum_odprtja_raw = row[4] if len(row) > 4 else None  # E stolpec
                rok_uporabe_raw = row[5] if len(row) > 5 else None  # F stolpec
                je_tester_raw = row[6] if len(row) > 6 else None  # G stolpec
                created_at_original_raw = row[7] if len(row) > 7 else None  # H stolpec
                vnesel_uporabnik_raw = row[8] if len(row) > 8 else None  # I stolpec
                serijska_stevilka_raw = row[9] if len(row) > 9 else None  # J stolpec
                
                if not all([excel_row_id_raw, product_no_raw, proizvajalec_raw]):
                    continue
                
                try:
                    excel_row_id = int(excel_row_id_raw)
                except (ValueError, TypeError):
                    print(f"⚠️  OPOZORILO: Neveljaven ID vnosa '{excel_row_id_raw}' v vrstici {i}. Preskakujem.")
                    error_count += 1
                    continue
                
                product_no = str(product_no_raw).strip().upper()
                proizvajalec = str(proizvajalec_raw).strip().upper()
                
                parfum_id = parfumi_map.get((product_no, proizvajalec))
                if not parfum_id:
                    print(f"⚠️  OPOZORILO: Parfum z ID '{product_no}' in proizvajalcem '{proizvajalec}' (vrstica {i}) ni najden v katalogu. Preskakujem.")
                    error_count += 1
                    continue
                
                rok_uporabe = parse_date(rok_uporabe_raw)
                if not rok_uporabe:
                    print(f"⚠️  OPOZORILO: Neveljaven ali manjkajoč rok uporabe v vrstici {i}. Preskakujem.")
                    error_count += 1
                    continue
                
                datum_odprtja = parse_date(datum_odprtja_raw)
                je_tester = str(je_tester_raw).strip().lower() == 'da' if je_tester_raw else False
                created_at_original = parse_date(created_at_original_raw)
                vnesel_uporabnik = str(vnesel_uporabnik_raw).strip() if vnesel_uporabnik_raw else None
                serijska_stevilka = str(serijska_stevilka_raw).strip() if serijska_stevilka_raw else None
                
                # Vstavi v serije tabelo
                cursor.execute(
                    """
                    INSERT INTO serije (
                        excel_row_id, parfum_id, rok_uporabe, serijska_stevilka, stanje, 
                        datum_odprtja, je_tester, vnesel_uporabnik, created_at_original
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (excel_row_id) DO UPDATE SET
                        parfum_id = EXCLUDED.parfum_id, rok_uporabe = EXCLUDED.rok_uporabe,
                        serijska_stevilka = EXCLUDED.serijska_stevilka, stanje = EXCLUDED.stanje,
                        datum_odprtja = EXCLUDED.datum_odprtja, je_tester = EXCLUDED.je_tester,
                        vnesel_uporabnik = EXCLUDED.vnesel_uporabnik, created_at_original = EXCLUDED.created_at_original
                    """,
                    (
                        excel_row_id, parfum_id, rok_uporabe.date(), serijska_stevilka, 'NA ZALOGI',
                        datum_odprtja.date() if datum_odprtja else None,
                        je_tester, vnesel_uporabnik, created_at_original
                    )
                )
                added_count += 1
                
                if added_count % 100 == 0:
                    print(f"📊 Obdelano {i} vrstic, dodano {added_count} novih serij")
            
            except Exception as e:
                error_count += 1
                print(f"⚠️  Napaka pri obdelavi vrstice {i}: {e}")
                continue
        
        db.commit()
        print(f"✅ Migracija končana. Dodano {added_count} novih serij, napake: {error_count}")
        
        return True
        
    except Exception as e:
        print(f"❌ Kritična napaka med migracijo: {e}")
        traceback.print_exc()
        return False

def migrate_perfumes_from_excel():
    """Migrira parfume iz Excel datoteke (zvezek 'Parfumi') v tabelo parfumi."""
    try:
        from openpyxl import load_workbook
        import os
        
        # Pot do Excel datoteke
        excel_file = 'DEKLARACIJE_PARFUMOV_KOPER.xlsm'
        
        if not os.path.exists(excel_file):
            print(f"❌ Excel datoteka {excel_file} ne obstaja!")
            return False
        
        print(f"📖 Odpiram Excel datoteko: {excel_file}")
        workbook = load_workbook(excel_file, data_only=True, read_only=True)
        
        # Preveri, če zvezek 'Parfumi' obstaja
        if 'Parfumi' not in workbook.sheetnames:
            print("❌ Zvezek 'Parfumi' ne obstaja v Excel datoteki!")
            print(f"   Razpoložljivi zvezki: {workbook.sheetnames}")
            return False
        
        sheet = workbook['Parfumi']
        print(f"✅ Zvezek 'Parfumi' najden. Vrstic: {sheet.max_row}")
        
        # Pridobi povezavo do baze
        from app import create_app
        from database import get_db
        
        app = create_app()
        with app.app_context():
            db = get_db()
            cursor = db.cursor()
            
            added_count = 0
            updated_count = 0
            skipped_count = 0
            error_count = 0
            
            # Iteriraj po vrsticah (preskoči naslovno vrstico)
            for row_num in range(2, sheet.max_row + 1):
                try:
                    # Pridobi podatke iz vrstice
                    product_no = sheet[f'A{row_num}'].value
                    ime_parfuma = sheet[f'B{row_num}'].value
                    sestava_inci = sheet[f'C{row_num}'].value
                    proizvajalec_ime = sheet[f'E{row_num}'].value
                    
                    # Preveri, če so obvezni podatki prisotni
                    if not all([product_no, proizvajalec_ime, ime_parfuma]):
                        print(f"⚠️  Vrstica {row_num}: Manjkajo obvezni podatki - preskačem")
                        skipped_count += 1
                        continue
                    
                    # Pridobi ali ustvari proizvajalca
                    cursor.execute("SELECT id FROM proizvajalci WHERE ime = %s", (proizvajalec_ime,))
                    proizvajalec = cursor.fetchone()
                    
                    if not proizvajalec:
                        print(f"➕ Dodajam proizvajalca: {proizvajalec_ime}")
                        cursor.execute("INSERT INTO proizvajalci (ime) VALUES (%s) ON CONFLICT (ime) DO NOTHING", (proizvajalec_ime,))
                        db.commit()
                        cursor.execute("SELECT id FROM proizvajalci WHERE ime = %s", (proizvajalec_ime,))
                        proizvajalec = cursor.fetchone()
                    
                    proizvajalec_id = proizvajalec['id']
                    
                    # Preveri, če parfum že obstaja
                    cursor.execute("""
                        SELECT id FROM parfumi 
                        WHERE product_no = %s AND proizvajalec_id = %s
                    """, (str(product_no), proizvajalec_id))
                    
                    existing_perfume = cursor.fetchone()
                    
                    if existing_perfume:
                        # Posodobi obstoječi parfum
                        cursor.execute("""
                            UPDATE parfumi 
                            SET ime_parfuma = %s, sestava_inci = %s
                            WHERE id = %s
                        """, (ime_parfuma, sestava_inci, existing_perfume['id']))
                        updated_count += 1
                        if row_num % 50 == 0:  # Izpisujemo manj pogosto
                            print(f"🔄 Posodobljen parfum: {product_no} - {ime_parfuma}")
                    else:
                        # Dodaj nov parfum
                        cursor.execute("""
                            INSERT INTO parfumi (product_no, proizvajalec_id, ime_parfuma, sestava_inci, sinhroniziraj_s_shopify)
                            VALUES (%s, %s, %s, %s, FALSE)
                        """, (product_no, proizvajalec_id, ime_parfuma, sestava_inci))
                        added_count += 1
                        if row_num % 50 == 0:  # Izpisujemo manj pogosto
                            print(f"➕ Dodan parfum: {product_no} - {ime_parfuma}")
                    
                    # Commit vsakih 20 vrstic (manj pogosto)
                    if row_num % 20 == 0:
                        db.commit()
                        print(f"💾 Commit vrstica {row_num}")
                
                except Exception as e:
                    print(f"❌ Napaka v vrstici {row_num}: {e}")
                    error_count += 1
                    continue
            
            # Končni commit
            db.commit()
            cursor.close()
            workbook.close()  # Zapri workbook za sprostitev pomnilnika
            
            print(f"\n✅ Migracija parfumov končana!")
            print(f"   Dodanih: {added_count}")
            print(f"   Posodobljenih: {updated_count}")
            print(f"   Preskočenih: {skipped_count}")
            print(f"   Napak: {error_count}")
            
            return True
            
    except Exception as e:
        print(f"❌ Napaka pri migraciji parfumov: {e}")
        import traceback
        traceback.print_exc()
        return False